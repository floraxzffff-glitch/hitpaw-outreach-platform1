"""
Excel 工作簿处理工具
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    openpyxl = None


QUEUE_HEADERS = ["频道名", "邮箱", "定制主题", "定制开头", "主页链接", "视频链接", "备注", "类型", "来源关键词"]


def safe_load_workbook(path, sheet_title="Sheet"):
    """
    稳健地打开工作簿。如果文件损坏，自动备份并创建新文件。
    """
    if not openpyxl:
        return None, None
    
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet_title] if sheet_title in wb.sheetnames else wb.active
            return wb, ws
        except Exception as exc:
            # 备份损坏的文件
            broken_path = path + f".broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                os.replace(path, broken_path)
                print(f"⚠️  {os.path.basename(path)} 损坏，已备份为 {os.path.basename(broken_path)}")
            except Exception:
                pass
    
    # 创建新工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def save_workbook_safe(wb, path, purpose="工作簿"):
    """安全保存工作簿"""
    if not openpyxl or not wb:
        return False
    
    try:
        # 先保存到临时文件，确保不会损坏原文件
        temp_path = path + ".tmp"
        wb.save(temp_path)
        
        # 原子化替换
        if os.path.exists(path):
            os.replace(path, path + ".bak")
        os.replace(temp_path, path)
        return True
    except Exception as exc:
        print(f"❌ 保存{purpose}失败: {exc}")
        return False


def ensure_queue_headers(ws):
    """确保发信名单的表头正确"""
    if not ws:
        return
    
    headers = {
        1: "频道名",
        2: "邮箱",
        3: "定制主题",
        4: "定制开头",
        5: "主页链接",
        6: "视频链接",
        7: "备注",
        8: "类型",
        9: "来源关键词",
    }
    
    for col, title in headers.items():
        cell = ws.cell(1, col)
        if not cell.value:
            cell.value = title


def append_queue_row(ws, name, email, note="", row_type="", source=""):
    """在队列表添加新行"""
    if not ws:
        return False
    
    row_num = ws.max_row + 1
    ws.cell(row_num, 1).value = name
    ws.cell(row_num, 2).value = email
    ws.cell(row_num, 7).value = note
    ws.cell(row_num, 8).value = row_type
    ws.cell(row_num, 9).value = source
    
    return True


def set_column_widths(ws, widths_dict):
    """设置列宽"""
    if not ws:
        return
    
    for col, width in widths_dict.items():
        ws.column_dimensions[col].width = width


def apply_fill_color(cell, color_hex):
    """应用填充颜色"""
    if not cell:
        return
    cell.fill = PatternFill("solid", start_color=color_hex)
