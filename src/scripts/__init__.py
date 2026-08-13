#!/usr/bin/env python3
"""
VikPea 项目初始化脚本 - 创建必要的配置和数据文件
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("❌ 需要 openpyxl 库，请先运行: pip install openpyxl")
    sys.exit(1)


PROJECT_ROOT = Path(__file__).parent.parent
CONFIG_DIR = PROJECT_ROOT / "config"
DATA_DIR = PROJECT_ROOT / "data"
LOGS_DIR = PROJECT_ROOT / "logs"


def create_config_template():
    """创建配置文件模板"""
    config_path = CONFIG_DIR / "config_template.xlsx"
    
    if config_path.exists():
        print(f"✓ {config_path} 已存在")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "配置"
    
    # 表头
    ws['A1'] = "配置项"
    ws['B1'] = "值"
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    
    # 配置项
    configs = [
        ("SMTP_SERVER", "smtp.qiye.aliyun.com"),
        ("SMTP_PORT", "465"),
        ("SMTP_TIMEOUT", "25"),
        ("SMTP_ALLOW_INSECURE_SSL", "True"),
        ("IMAP_SERVER", "imap.qiye.aliyun.com"),
        ("IMAP_PORT", "993"),
        ("FROM_EMAIL", "hannah@hitpaw.com"),
        ("FROM_NAME", "Hannah"),
        ("PASSWORD", "(邮箱密码)"),
        ("PRODUCT_NAME", "HitPaw VikPea"),
        ("PRODUCT_URL", "https://www.hitpaw.com/hitpaw-video-enhancer.html"),
        ("DELAY_SEC", "8"),
        ("DAILY_SEND_LIMIT", "80"),
        ("YOUTUBE_RESULTS_PER_KEYWORD", "35"),
        ("YOUTUBE_MIN_VIDEO_VIEWS", "800"),
        ("ARTICLE_RESULTS_PER_QUERY", "30"),
        ("ARTICLE_MIN_SITE_SCORE", "3"),
        ("YOUTUBE_API_KEY", "(可选)"),
    ]
    
    for idx, (key, value) in enumerate(configs, start=2):
        ws.cell(idx, 1).value = key
        ws.cell(idx, 2).value = value
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    wb.save(config_path)
    print(f"✓ 已创建配置模板: {config_path}")


def create_blacklist_template():
    """创建黑名单文件模板"""
    blacklist_path = DATA_DIR / "blacklist_template.xlsx"
    
    if blacklist_path.exists():
        print(f"✓ {blacklist_path} 已存在")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "黑名单"
    
    # 表头
    headers = ["类型", "值", "备注", "启用"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True)
    
    # 示例
    examples = [
        ("邮箱", "spam@example.com", "垃圾邮箱", "是"),
        ("邮箱", "noreply@gmail.com", "自动回复邮箱", "是"),
        ("域名", "amazon.com", "大平台无法投放", "是"),
        ("域名", "hitpaw.com", "竞品", "是"),
        ("名称", "Test Account", "测试账号", "是"),
    ]
    
    for idx, (row_type, value, note, enabled) in enumerate(examples, start=2):
        ws.cell(idx, 1).value = row_type
        ws.cell(idx, 2).value = value
        ws.cell(idx, 3).value = note
        ws.cell(idx, 4).value = enabled
    
    for i in range(1, 5):
        ws.column_dimensions[chr(64 + i)].width = 20
    
    wb.save(blacklist_path)
    print(f"✓ 已创建黑名单模板: {blacklist_path}")


def create_queue_template():
    """创建发信名单模板"""
    queue_path = DATA_DIR / "queue_template.xlsx"
    
    if queue_path.exists():
        print(f"✓ {queue_path} 已存在")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发信名单"
    
    # 表头
    headers = ["频道名", "邮箱", "定制主题", "定制开头", "主页链接", "视频链接", "备注", "类型", "来源关键词"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="366092")
    
    # 列宽
    widths = [20, 30, 30, 30, 40, 40, 30, 15, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    wb.save(queue_path)
    print(f"✓ 已创建发信名单模板: {queue_path}")


def create_tracker_template():
    """创建邮件追踪表模板"""
    tracker_path = DATA_DIR / "tracker_template.xlsx"
    
    if tracker_path.exists():
        print(f"✓ {tracker_path} 已存在")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "邮件追踪"
    
    # 表头
    headers = [
        "序号", "日期", "频道名", "邮箱", "类型", "频道/平台", "发信状态",
        "是否回复", "回复摘要", "链接", "当前状态", "分级", "跟进意向",
        "跟进1发送日期", "跟进2发送日期", "备注", "来源关键词", "最后更新"
    ]
    
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2F5496")
    
    # 列宽
    ws.column_dimensions['A'].width = 8
    for col in "BCDEFGHIJKLMNOPQR":
        ws.column_dimensions[col].width = 15
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    wb.save(tracker_path)
    print(f"✓ 已创建邮件追踪表模板: {tracker_path}")


def main():
    """初始化项目"""
    print("\n" + "=" * 50)
    print("VikPea 项目初始化")
    print("=" * 50 + "\n")
    
    # 创建目录
    CONFIG_DIR.mkdir(exist_ok=True)
    DATA_DIR.mkdir(exist_ok=True)
    LOGS_DIR.mkdir(exist_ok=True)
    
    print("✓ 目录结构已创建")
    print()
    
    # 创建模板
    create_config_template()
    create_blacklist_template()
    create_queue_template()
    create_tracker_template()
    
    print()
    print("=" * 50)
    print("✅ 初始化完成！")
    print("=" * 50)
    print()
    print("后续步骤:")
    print("1. 复制配置文件:")
    print(f"   cp config/config_template.xlsx config/config.xlsx")
    print("2. 编辑配置文件，填入邮箱信息")
    print("3. 运行工作台:")
    print("   python -m src.ui.cli_menu")
    print()


if __name__ == "__main__":
    main()
