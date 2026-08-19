"""
VikPea_补录回复状态.py — 从收件箱倒查历史回复，修复补录名单的回复率

用途：
  · 扫描近 N 天收件箱
  · 用发件人邮箱匹配 VikPea_邮件开发追踪.xlsx 的「邮件追踪」主表
  · 把漏掉的历史回复补回：是否回复 / 回复摘要 / 当前状态 / ABC分级

用法：
  先预览：
    python3 ~/Downloads/VikPea工作台/VikPea_补录回复状态.py --dry-run

  确认后写回：
    python3 ~/Downloads/VikPea工作台/VikPea_补录回复状态.py --apply
"""

import argparse
import email
import email.header
import email.utils
import imaplib
import importlib.util
import os
import re
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ 需要安装 openpyxl")
    sys.exit(1)

try:
    from VikPea_common import apply_config, log_event
except ImportError:
    apply_config = None
    log_event = None


IMAP_SERVER = "imap.qiye.aliyun.com"
IMAP_PORT = 993
EMAIL_ADDR = "hannah@hitpaw.com"
PASSWORD = ""  # 密码已迁移到 VikPea_配置.xlsx / 网页系统设置，不再硬编码
DAYS_BACK = 21

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_回复状态补录预览.xlsx")

T_NUM = 1
T_DATE = 2
T_NAME = 3
T_EMAIL = 4
T_REPLY = 9
T_SUM = 10
T_STATUS = 12
T_GRADE = 13

GREEN = PatternFill("solid", start_color="E2EFDA")
YELLOW = PatternFill("solid", start_color="FFF2CC")
ORANGE = PatternFill("solid", start_color="FCE4D6")


def decode_header_str(value):
    if not value:
        return ""
    out = ""
    for part, enc in email.header.decode_header(value):
        if isinstance(part, bytes):
            out += part.decode(enc or "utf-8", errors="replace")
        else:
            out += str(part)
    return out.strip()


def get_body(msg):
    bodies = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition") or "")
            if "attachment" in disp:
                continue
            if ctype in ("text/plain", "text/html"):
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                charset = part.get_content_charset() or "utf-8"
                bodies.append(payload.decode(charset, errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            bodies.append(payload.decode(charset, errors="replace"))
    text = "\n".join(bodies)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def strip_quoted_reply(body):
    text = body or ""
    cut_patterns = [
        r"\nOn .+ wrote:",
        r"\nFrom:\s",
        r"\nSent:\s",
        r"\n发件人：",
        r"\n在 .+ 写道：",
    ]
    cut_at = len(text)
    for pat in cut_patterns:
        m = re.search(pat, text, flags=re.I | re.S)
        if m:
            cut_at = min(cut_at, m.start())
    text = text[:cut_at]
    lines = [line for line in text.splitlines() if not line.strip().startswith(">")]
    return "\n".join(lines).strip() or body[:500].strip()


def grade_reply(subject, body):
    clean = strip_quoted_reply(body)
    text = f"{subject} {clean}".lower()
    c_keywords = [
        "not interested", "no thanks", "decline", "pass", "not a fit",
        "unsubscribe", "remove me", "stop emailing", "not relevant",
        "不感兴趣", "不合适",
    ]
    a_keywords = [
        "interested", "sounds good", "would love", "happy to", "how much",
        "pricing", "price", "rate", "send me", "tell me more", "yes",
        "可以", "感兴趣", "报价", "合作", "价格",
    ]
    b_keywords = [
        "what is", "can you", "could you", "more info", "question",
        "maybe", "consider", "currently", "busy", "later",
    ]
    for kw in c_keywords:
        if kw in text:
            return "C", clean[:180].replace("\n", " ").strip()
    for kw in a_keywords:
        if kw in text:
            return "A", clean[:180].replace("\n", " ").strip()
    for kw in b_keywords:
        if kw in text:
            return "B", clean[:180].replace("\n", " ").strip()
    return "B", clean[:180].replace("\n", " ").strip()


def get_tracker_sheet(wb):
    if "邮件追踪" in wb.sheetnames:
        wb.active = wb.sheetnames.index("邮件追踪")
        return wb["邮件追踪"]
    return wb.active


def load_tracker_email_rows(ws):
    rows = {}
    for row in ws.iter_rows(min_row=2):
        em = str(row[T_EMAIL - 1].value or "").strip().lower()
        if em:
            rows.setdefault(em, []).append(row[0].row)
    return rows


def parse_msg_date(msg):
    try:
        dt = email.utils.parsedate_to_datetime(msg.get("Date") or "")
    except Exception:
        dt = None
    return dt or datetime.now()


def is_likely_reply(subject, body):
    text = f"{subject} {body}".lower()
    return (
        subject.lower().startswith("re:")
        or "hitpaw" in text
        or "vikpea" in text
        or "hannah" in text
    )


def scan_inbox(days_back, tracker_emails):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(EMAIL_ADDR, PASSWORD)
    except Exception as exc:
        raise RuntimeError(f"连接邮箱失败: {exc}") from exc
    mail.select("INBOX")
    since = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    _, ids = mail.search(None, f"SINCE {since}")
    found = {}
    for eid in ids[0].split():
        try:
            _, data = mail.fetch(eid, "(RFC822)")
            msg = email.message_from_bytes(data[0][1])
        except Exception:
            continue
        subject = decode_header_str(msg.get("Subject"))
        body = get_body(msg)
        froms = email.utils.getaddresses(msg.get_all("From", []))
        from_email = ""
        for _, addr in froms:
            addr = (addr or "").strip().lower()
            if addr in tracker_emails:
                from_email = addr
                break
        if not from_email:
            continue
        if not is_likely_reply(subject, body):
            continue
        dt = parse_msg_date(msg)
        if from_email not in found or dt > found[from_email]["dt"]:
            grade, summary = grade_reply(subject, body)
            found[from_email] = {
                "email": from_email,
                "date": dt.strftime("%Y-%m-%d"),
                "subject": subject,
                "grade": grade,
                "summary": summary,
                "dt": dt,
            }
    mail.logout()
    return list(found.values())


def write_preview(rows, ws, email_rows):
    wb = openpyxl.Workbook()
    out = wb.active
    out.title = "待更新回复"
    out.append(["回复日期", "联系人/平台", "邮箱", "当前是否回复", "新ABC", "主题", "摘要"])
    for r in rows:
        rownum = email_rows[r["email"]][0]
        out.append([
            r["date"],
            ws.cell(rownum, T_NAME).value,
            r["email"],
            ws.cell(rownum, T_REPLY).value,
            r["grade"],
            r["subject"],
            r["summary"],
        ])
    for col, width in {"A": 12, "B": 28, "C": 34, "D": 14, "E": 8, "F": 45, "G": 70}.items():
        out.column_dimensions[col].width = width
    wb.save(OUTPUT_PATH)


def apply_updates(rows):
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = get_tracker_sheet(wb)
    email_rows = load_tracker_email_rows(ws)
    updated = 0
    for r in rows:
        for rownum in email_rows.get(r["email"], []):
            ws.cell(rownum, T_REPLY).value = "已回复"
            ws.cell(rownum, T_SUM).value = r["summary"]
            ws.cell(rownum, T_STATUS).value = "待处理" if r["grade"] in {"A", "B"} else "已拒绝"
            ws.cell(rownum, T_GRADE).value = r["grade"]
            color = {"A": GREEN, "B": YELLOW, "C": ORANGE}.get(r["grade"], YELLOW)
            for c in range(1, min(ws.max_column, T_GRADE) + 1):
                ws.cell(rownum, c).fill = color
            updated += 1
    wb.save(TRACKER_PATH)
    rebuild_summary_sheets()
    return updated


def rebuild_summary_sheets():
    """复用读取回复脚本里的汇总页生成逻辑。"""
    helper_path = os.path.join(SCRIPT_DIR, "VikPea_读取回复.py")
    if not os.path.exists(helper_path):
        return
    spec = importlib.util.spec_from_file_location("vikpea_read_replies", helper_path)
    helper = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(helper)
    wb = openpyxl.load_workbook(TRACKER_PATH)
    helper.rebuild_abc_sheet(wb)
    helper.rebuild_pending_sheet(wb)
    wb.active = wb.sheetnames.index("邮件追踪") if "邮件追踪" in wb.sheetnames else 0
    wb.save(TRACKER_PATH)


def main():
    if apply_config:
        apply_config(globals(), {
            "IMAP_SERVER": "IMAP_SERVER",
            "IMAP_PORT": "IMAP_PORT",
            "FROM_EMAIL": "EMAIL_ADDR",
            "PASSWORD": "PASSWORD",
        })
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=DAYS_BACK)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = get_tracker_sheet(wb)
    email_rows = load_tracker_email_rows(ws)
    try:
        replies = scan_inbox(args.days, set(email_rows))
    except RuntimeError as exc:
        print(f"❌ {exc}")
        print("请检查 IMAP 配置、邮箱授权码，以及当前网络是否可访问邮箱服务器。")
        return

    candidates = []
    for r in replies:
        rownums = email_rows.get(r["email"], [])
        if not rownums:
            continue
        already_replied = any(str(ws.cell(rownum, T_REPLY).value or "") in {"已回复", "是"} for rownum in rownums)
        if already_replied:
            continue
        candidates.append(r)

    candidates.sort(key=lambda x: (x["date"], x["email"]))
    write_preview(candidates, ws, email_rows)
    print(f"近 {args.days} 天匹配到历史回复: {len(replies)}")
    print(f"其中追踪表尚未标记回复: {len(candidates)}")
    print(f"预览表: {OUTPUT_PATH}")
    for r in candidates[:20]:
        rownum = email_rows[r["email"]][0]
        print(f"  {r['date']} | {ws.cell(rownum, T_NAME).value} | {r['email']} | {r['grade']} | {r['summary'][:80]}")

    if args.apply:
        updated = apply_updates(candidates)
        print(f"✅ 已更新回复状态 {updated} 行")
    elif not args.dry_run:
        print("如确认无误，运行：python3 ~/Downloads/VikPea工作台/VikPea_补录回复状态.py --apply")


if __name__ == "__main__":
    main()
