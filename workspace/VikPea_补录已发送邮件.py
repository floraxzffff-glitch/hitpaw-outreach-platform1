"""
VikPea_补录已发送邮件.py — 从已发送邮箱倒查漏录的开发信

用途：
  · 扫描阿里企业邮箱最近 N 天的已发送邮件
  · 只保留正文/主题里包含 HitPaw 或 VikPea 的外联邮件
  · 和 VikPea_邮件开发追踪.xlsx 的「邮件追踪」主表按邮箱去重
  · 预览后，把漏录的已发送对象补回追踪表

用法：
  先预览：
    python3 ~/Downloads/VikPea工作台/VikPea_补录已发送邮件.py --dry-run

  确认后补录：
    python3 ~/Downloads/VikPea工作台/VikPea_补录已发送邮件.py --apply
"""

import argparse
import email
import email.header
import email.utils
import imaplib
import os
import re
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl")
    sys.exit(1)

try:
    from VikPea_common import apply_config, log_event
except ImportError:
    apply_config = None
    log_event = None


IMAP_SERVER = "imap.qiye.aliyun.com"
IMAP_PORT = 993
EMAIL_ADDR = "hannah@hitpaw.com"
PASSWORD = ""  # 密码已迁移到 VikPea_配置.xlsx / 网页系统设置，不再硬编码
DAYS_BACK = 21

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_已发送补录预览.xlsx")

T_NUM = 1
T_DATE = 2
T_NAME = 3
T_EMAIL = 4
T_TYPE = 5
T_CHAN = 6
T_KIND = 8
T_REPLY = 9
T_SUM = 10
T_LINK = 11
T_STATUS = 12
T_SOURCE = 19

OUTREACH_WORDS = ("hitpaw", "vikpea")
IGNORE_EMAILS = {EMAIL_ADDR.lower()}


def decode_header_str(value):
    if not value:
        return ""
    out = ""
    for part, enc in email.header.decode_header(value):
        if isinstance(part, bytes):
            out += part.decode(enc or "utf-8", errors="replace")
        else:
            out += str(part)
    return out.strip()


def get_body(msg):
    bodies = []
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition") or "")
            if "attachment" in disp:
                continue
            if ctype in ("text/plain", "text/html"):
                payload = part.get_payload(decode=True)
                if not payload:
                    continue
                charset = part.get_content_charset() or "utf-8"
                text = payload.decode(charset, errors="replace")
                bodies.append(text)
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            charset = msg.get_content_charset() or "utf-8"
            bodies.append(payload.decode(charset, errors="replace"))
    text = "\n".join(bodies)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_mailbox_name(line):
    if not line:
        return None
    if isinstance(line, str):
        line = line.encode()
    m = re.search(rb' "([^"]+)"$', line)
    if m:
        return m.group(1).decode("ascii", errors="ignore")
    parts = line.rsplit(b" ", 1)
    if len(parts) == 2:
        return parts[1].strip(b'"').decode("ascii", errors="ignore")
    return None


def find_sent_folders(mail):
    _, rows = mail.list()
    names = []
    for row in rows or []:
        name = parse_mailbox_name(row)
        if not name:
            continue
        low = name.lower()
        raw = str(row).lower()
        if (
            "sent" in low
            or "\\sent" in raw
            or "已发送" in name
            or "sent messages" in low
            or "sent items" in low
        ):
            names.append(name)
    fallback = ["Sent", "Sent Messages", "Sent Items", "INBOX.Sent", "已发送", "已发送邮件"]
    for name in fallback:
        if name not in names:
            names.append(name)
    return names


def sent_date(msg):
    try:
        dt = email.utils.parsedate_to_datetime(msg.get("Date") or "")
    except Exception:
        dt = None
    if not dt:
        return datetime.now().strftime("%Y-%m-%d")
    return dt.strftime("%Y-%m-%d")


def extract_name(msg, body, to_email):
    m = re.search(r"\bHi\s+([^,\n]{1,60}),", body, flags=re.I)
    if m:
        return m.group(1).strip()
    tos = email.utils.getaddresses(msg.get_all("To", []))
    for display, addr in tos:
        if addr.lower() == to_email and display:
            return decode_header_str(display).strip('" ')
    local = to_email.split("@", 1)[0]
    return re.sub(r"[._-]+", " ", local).strip().title()


def classify_type(subject, body):
    text = f"{subject} {body}".lower()
    if "article" in text or "guest post" in text or "including hitpaw vikpea in your article" in text:
        return "外链网站", "文章插链接开发"
    return "YouTube KOL", "视频插链接开发"


def existing_tracker_emails(ws):
    emails = set()
    for row in ws.iter_rows(min_row=2):
        em = str(row[T_EMAIL - 1].value or "").strip().lower()
        if em:
            emails.add(em)
    return emails


def get_tracker_sheet(wb):
    if "邮件追踪" in wb.sheetnames:
        wb.active = wb.sheetnames.index("邮件追踪")
        return wb["邮件追踪"]
    return wb.active


def scan_sent_messages(days_back):
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(EMAIL_ADDR, PASSWORD)
    except Exception as exc:
        raise RuntimeError(f"连接邮箱失败: {exc}") from exc

    since = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    folders = find_sent_folders(mail)
    found = {}
    used_folder = ""

    for folder in folders:
        try:
            status, _ = mail.select(f'"{folder}"')
        except Exception:
            continue
        if status != "OK":
            continue
        used_folder = folder
        _, ids = mail.search(None, f"SINCE {since}")
        for eid in ids[0].split():
            try:
                _, data = mail.fetch(eid, "(RFC822)")
                msg = email.message_from_bytes(data[0][1])
            except Exception:
                continue

            subject = decode_header_str(msg.get("Subject"))
            body = get_body(msg)
            text = f"{subject} {body}".lower()
            if not any(word in text for word in OUTREACH_WORDS):
                continue

            tos = email.utils.getaddresses(msg.get_all("To", []))
            for _, addr in tos:
                to_email = (addr or "").strip().lower()
                if "@" not in to_email or to_email in IGNORE_EMAILS:
                    continue
                if to_email in found:
                    continue
                name = extract_name(msg, body, to_email)
                contact_type, kind = classify_type(subject, body)
                found[to_email] = {
                    "date": sent_date(msg),
                    "name": name,
                    "email": to_email,
                    "subject": subject,
                    "type": contact_type,
                    "kind": kind,
                }
        if found:
            break

    mail.logout()
    return used_folder, list(found.values())


def write_preview(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "待补录"
    ws.append(["日期", "联系人/平台", "邮箱", "类型", "邮件类型", "主题"])
    for r in rows:
        ws.append([r["date"], r["name"], r["email"], r["type"], r["kind"], r["subject"]])
    for col, width in {"A": 12, "B": 28, "C": 34, "D": 14, "E": 18, "F": 45}.items():
        ws.column_dimensions[col].width = width
    wb.save(OUTPUT_PATH)


def append_to_tracker(rows):
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = get_tracker_sheet(wb)
    if not ws.cell(1, T_SOURCE).value:
        ws.cell(1, T_SOURCE).value = "来源关键词"

    existing = existing_tracker_emails(ws)
    appended = 0
    for r in rows:
        if r["email"] in existing:
            continue
        last_num = ws.cell(ws.max_row, T_NUM).value
        new_num = (last_num + 1) if isinstance(last_num, int) else ws.max_row
        row = [
            new_num,
            r["date"],
            r["name"],
            r["email"],
            r["type"],
            "",
            "",  # 视频链接，补录时不知道，留空
            r["kind"] + "（已发送补录）",
            "未回复",
            None,
            "",
            "已发送",
        ]
        while len(row) < T_SOURCE - 1:
            row.append(None)
        row.append("已发送邮箱补录")
        row.append("")  # 频道标签，先留空
        ws.append(row)
        existing.add(r["email"])
        appended += 1

    wb.active = wb.sheetnames.index("邮件追踪") if "邮件追踪" in wb.sheetnames else 0
    wb.save(TRACKER_PATH)
    return appended


def main():
    if apply_config:
        apply_config(globals(), {
            "IMAP_SERVER": "IMAP_SERVER",
            "IMAP_PORT": "IMAP_PORT",
            "FROM_EMAIL": "EMAIL_ADDR",
            "PASSWORD": "PASSWORD",
        })
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=DAYS_BACK)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(TRACKER_PATH):
        print(f"❌ 找不到追踪表: {TRACKER_PATH}")
        return

    try:
        folder, sent_rows = scan_sent_messages(args.days)
    except RuntimeError as exc:
        print(f"❌ {exc}")
        print("请检查 IMAP 配置、邮箱授权码，以及当前网络是否可访问邮箱服务器。")
        return
    wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True)
    ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
    existing = existing_tracker_emails(ws)
    missing = [r for r in sent_rows if r["email"] not in existing]
    missing.sort(key=lambda x: (x["date"], x["email"]))

    write_preview(missing)
    print(f"已扫描已发送文件夹: {folder or '未识别'}")
    print(f"近 {args.days} 天 VikPea/HitPaw 已发送收件人: {len(sent_rows)}")
    print(f"追踪表已存在: {len(sent_rows) - len(missing)}")
    print(f"待补录: {len(missing)}")
    print(f"预览表: {OUTPUT_PATH}")
    for r in missing[:20]:
        print(f"  {r['date']} | {r['name']} | {r['email']} | {r['subject']}")

    if args.apply:
        appended = append_to_tracker(missing)
        print(f"✅ 已补录 {appended} 条到邮件追踪主表")
    elif not args.dry_run:
        print("如确认无误，运行：python3 ~/Downloads/VikPea工作台/VikPea_补录已发送邮件.py --apply")


if __name__ == "__main__":
    main()
