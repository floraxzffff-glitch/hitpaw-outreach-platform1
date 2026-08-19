"""
VikPea_导出团队格式.py — 把 VikPea 的联络记录导出成跟团队"合作跟踪表"一样的表头格式，
方便谈成合作之后直接整行复制粘贴过去，不用再手动对表头。

思路：
  · 数据来源是 VikPea_邮件开发追踪.xlsx（真实联络记录：联系过谁、什么时候、有没有回复）
  · 输出的表头、列顺序完全照抄团队"合作跟踪表.xlsx"的"已合作" sheet（27列）
  · VikPea 能填的就填（开发日期/类型/主页链接/视频链接/网站视频名称/联系方式），
    其余谈合作时才会有的字段（费用/品牌/负责人/流量数据等）留空，人工填
  · 视频链接这一列追踪表本身没有，会尝试拿邮箱去发信名单/待确认邮箱里找一下补上
  · 频道标签这一列团队现在还没有标签体系，先留空

用法：
  python3 VikPea_导出团队格式.py
  输出文件：VikPea_联络记录_团队格式.xlsx（每次运行都会覆盖重新生成，是个快照，不是主表）
"""

import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
QUEUE_PATH = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
PENDING_PATH = os.path.join(SCRIPT_DIR, "VikPea_待确认邮箱.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_联络记录_团队格式.xlsx")

# 完全照抄团队"合作跟踪表.xlsx" -> "已合作" sheet 的表头顺序（27列）
TEAM_HEADERS = [
    "开发日期", "类型", "主页链接", "视频链接", "网站、视频名称", "跟踪链接",
    "关键词、排名", "分级", "个人评价", "频道标签", "合作主题", "联系方式",
    "支付账户", "费用", "品牌", "合作产品", "合作方式", "语言", "国家",
    "负责人", "网站流量", "权重", "vph", "曝光", "流量", "下载", "跳出率",
]


def norm(value):
    return str(value or "").strip()


def norm_email(value):
    return norm(value).lower()


def read_rows(path, sheet_name=None):
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws.max_row < 1:
        return []
    headers = [norm(c.value) for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def build_video_link_lookup():
    """邮箱 -> 视频链接，从发信名单 + 待确认邮箱里找，谁先出现用谁的。"""
    lookup = {}
    for row in read_rows(QUEUE_PATH):
        email = norm_email(row.get("邮箱"))
        link = norm(row.get("视频链接"))
        if email and link and email not in lookup:
            lookup[email] = link
    for row in read_rows(PENDING_PATH):
        email = norm_email(row.get("候选邮箱"))
        link = norm(row.get("视频链接"))
        if email and link and email not in lookup:
            lookup[email] = link
    return lookup


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return norm(value)


def main():
    print("\n" + "=" * 64)
    print("  VikPea 导出团队格式")
    print("=" * 64)

    tracker_rows = read_rows(TRACKER_PATH, sheet_name="邮件追踪")
    if not tracker_rows:
        print(f"没有联络记录可导出：{TRACKER_PATH} 里还是空的")
        return

    video_link_lookup = build_video_link_lookup()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "联络记录"
    ws.append(TEAM_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")

    count = 0
    for row in tracker_rows:
        email = norm_email(row.get("邮箱"))
        team_row = {h: "" for h in TEAM_HEADERS}
        team_row["开发日期"] = format_date(row.get("日期"))
        team_row["类型"] = norm(row.get("类型"))
        team_row["主页链接"] = norm(row.get("主页链接"))
        team_row["视频链接"] = video_link_lookup.get(email, "")
        team_row["网站、视频名称"] = norm(row.get("联系人/平台"))
        team_row["联系方式"] = row.get("邮箱") or ""
        ws.append([team_row[h] for h in TEAM_HEADERS])
        count += 1

    for col, width in zip("ABCDEFGHIJKLMNOPQRSTUVWXYZ", [12, 10, 40, 40, 24, 24, 16, 8, 20, 14, 20, 26,
                                                            14, 10, 14, 16, 14, 10, 10, 10, 12, 8, 8, 8, 8, 8, 8]):
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_PATH)
    print(f"✅ 已导出 {count} 条联络记录")
    print(f"输出文件: {OUTPUT_PATH}")
    print("表头跟团队「合作跟踪表.xlsx -> 已合作」完全一致，谈成了直接整行复制过去就行。")
    print("「频道标签」列先留空——团队还没定标签体系，定了之后可以再加自动打标。")


if __name__ == "__main__":
    main()
