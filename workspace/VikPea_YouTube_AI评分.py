"""
VikPea_YouTube_AI评分.py - 对 YouTube 搜索结果进行 AI 适配度判断

用法：
  python3 VikPea_YouTube_AI评分.py

功能：
1. 读取 VikPea_发信名单.xlsx 中的 YouTube 候选人
2. 按频道聚合（同一频道通过多个关键词命中时合并）
3. AI 轻量初筛
4. 拉取频道完整信息
5. AI 综合适配度判断
6. 导出新的 Excel：VikPea_YouTube_AI评分结果.xlsx
"""

import os
import sys
import time
import logging
from datetime import datetime
from typing import Dict, List, Any

# 添加 api 目录到 path，以便 import ai_kol_scorer
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
API_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "api")
if API_DIR not in sys.path:
    sys.path.insert(0, API_DIR)

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ pip3 install openpyxl")
    sys.exit(1)

try:
    import ai_kol_scorer
except ImportError:
    print("❌ 找不到 ai_kol_scorer 模块")
    print("   请确保 api/ai_kol_scorer.py 存在")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ======================== 配置 ========================

QUEUE_PATH = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_YouTube_AI评分结果.xlsx")
PRODUCT_FEATURE = "AI 视频增强"  # 当前推广的产品功能

# Excel 样式
HEADER_FILL = PatternFill("solid", start_color="4472C4")
RECOMMENDED_FILL = PatternFill("solid", start_color="C6EFCE")
UNCERTAIN_FILL = PatternFill("solid", start_color="FFEB9C")
NOT_RECOMMENDED_FILL = PatternFill("solid", start_color="FFC7CE")


# ======================== 读取现有候选人 ========================

def load_youtube_candidates_from_queue(queue_path: str) -> List[Dict[str, Any]]:
    """
    从 VikPea_发信名单.xlsx 读取 YouTube 候选人
    
    Returns:
        候选人列表，每条包含：
        - channel_id (如果有)
        - channel_name
        - channel_url
        - video_title
        - keyword
        - video_url
        - view_count
        - subscriber_count
    """
    if not os.path.exists(queue_path):
        logger.error(f"找不到发信名单: {queue_path}")
        return []
    
    wb = openpyxl.load_workbook(queue_path, data_only=True)
    ws = wb.active
    
    if ws.max_row < 2:
        logger.warning("发信名单为空")
        return []
    
    # 读取表头
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    
    candidates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        
        # 只处理 YouTube 来源的候选人
        source = str(row_dict.get("来源") or "").strip()
        if "youtube" not in source.lower():
            continue
        
        # 提取必要字段
        channel_name = str(row_dict.get("频道名称") or row_dict.get("姓名") or "").strip()
        if not channel_name:
            continue
        
        channel_url = str(row_dict.get("频道链接") or row_dict.get("链接") or "").strip()
        video_url = str(row_dict.get("命中视频") or "").strip()
        keyword = str(row_dict.get("关键词") or row_dict.get("来源") or "").strip()
        
        # 尝试从 URL 提取 channel_id
        channel_id = ""
        if "/channel/" in channel_url:
            channel_id = channel_url.split("/channel/")[1].split("/")[0].split("?")[0]
        elif "/@" in channel_url:
            # @handle 形式，暂时用 handle 作为标识
            channel_id = channel_url.split("/@")[1].split("/")[0].split("?")[0]
        
        candidate = {
            "channel_id": channel_id or channel_url,  # 没有 ID 就用 URL 作为唯一标识
            "channel_name": channel_name,
            "channel_url": channel_url,
            "video_title": str(row_dict.get("视频标题") or "").strip(),
            "keyword": keyword,
            "video_url": video_url,
            "view_count": 0,  # 发信名单里可能没有播放量
            "subscriber_count": parse_int(row_dict.get("粉丝数")),
        }
        
        candidates.append(candidate)
    
    logger.info(f"从发信名单读取到 {len(candidates)} 条 YouTube 候选人")
    return candidates


def parse_int(value: Any) -> int:
    """解析整数"""
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value or "").strip()
    text = text.replace(",", "").replace("K", "000").replace("M", "000000")
    try:
        return int(float(text))
    except:
        return 0


# ======================== 拉取频道完整信息 ========================

def fetch_channel_info(channel_id: str) -> Dict[str, Any]:
    """
    拉取频道完整信息（第4层）
    
    这里简化实现，实际项目中应该调用 YouTube API 或抓取频道页面
    
    Returns:
        {
            "description": "频道简介",
            "recent_videos": [
                {"title": "xxx", "views": 1000},
                ...
            ]
        }
    """
    # TODO: 实际实现需要调用 YouTube API 或网页抓取
    # 这里返回空数据，AI 判断会基于现有信息进行
    
    logger.debug(f"拉取频道信息: {channel_id}")
    return {
        "description": "",
        "recent_videos": []
    }


# ======================== 导出结果到 Excel ========================

def export_results_to_excel(results: List[Dict[str, Any]], output_path: str):
    """导出 AI 判断结果到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI评分结果"
    
    # 表头
    headers = [
        "频道名称", "频道链接", "订阅数", 
        "适配度评分", "判断结果", "理由", "建议合作角度",
        "命中关键词", "命中视频数", "命中视频链接"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = HEADER_FILL
        from openpyxl.styles import Font
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 数据行
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row_idx, 1, result["channel_name"])
        ws.cell(row_idx, 2, result["channel_url"])
        ws.cell(row_idx, 3, result.get("subscriber_count") or 0)
        ws.cell(row_idx, 4, result["fit_score"])
        ws.cell(row_idx, 5, result["verdict"])
        ws.cell(row_idx, 6, result["reason"])
        ws.cell(row_idx, 7, result["suggested_angle"])
        ws.cell(row_idx, 8, result["hit_keywords"])
        ws.cell(row_idx, 9, result["hit_video_count"])
        ws.cell(row_idx, 10, result["hit_video_urls"])
        
        # 根据判断结果设置行颜色
        verdict = result["verdict"]
        if verdict == "推荐":
            fill = RECOMMENDED_FILL
        elif verdict == "待确认":
            fill = UNCERTAIN_FILL
        else:
            fill = NOT_RECOMMENDED_FILL
        
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row_idx, col_idx).fill = fill
    
    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 50
    
    wb.save(output_path)
    logger.info(f"结果已导出到: {output_path}")


# ======================== 主流程 ========================

def main():
    print(f"\n{'═'*70}")
    print(f"  🤖 VikPea YouTube KOL AI 适配度判断")
    print(f"{'═'*70}\n")
    
    # 检查 API Key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("❌ 未配置 ANTHROPIC_API_KEY")
        print("   请在环境变量或 VikPea_配置.xlsx 中设置")
        sys.exit(1)
    
    print(f"📊 推广产品功能: {PRODUCT_FEATURE}\n")
    
    # 1. 读取候选人
    print("📖 读取候选人...")
    candidates = load_youtube_candidates_from_queue(QUEUE_PATH)
    if not candidates:
        print("⚠️  没有找到 YouTube 候选人")
        return
    
    print(f"   共 {len(candidates)} 条原始候选人\n")
    
    # 2. AI 判断流程
    print("🤖 开始 AI 适配度判断...\n")
    start_time = time.time()
    
    results = ai_kol_scorer.process_candidates_with_ai_scoring(
        raw_candidates=candidates,
        product_feature=PRODUCT_FEATURE,
        fetch_channel_info_func=fetch_channel_info
    )
    
    elapsed = time.time() - start_time
    print(f"\n✅ AI 判断完成，耗时 {elapsed:.1f} 秒\n")
    
    # 3. 统计结果
    recommended = sum(1 for r in results if r["verdict"] == "推荐")
    uncertain = sum(1 for r in results if r["verdict"] == "待确认")
    not_recommended = sum(1 for r in results if r["verdict"] == "不推荐")
    not_screened = sum(1 for r in results if r["verdict"] == "未筛选")
    
    print("📈 判断结果统计:")
    print(f"   推荐:      {recommended} 个")
    print(f"   待确认:    {uncertain} 个 (需人工复核)")
    print(f"   不推荐:    {not_recommended} 个")
    print(f"   未筛选:    {not_screened} 个 (AI调用失败)")
    print(f"   总计:      {len(results)} 个频道\n")
    
    # 4. 导出结果
    print("💾 导出结果...")
    export_results_to_excel(results, OUTPUT_PATH)
    
    print(f"\n{'═'*70}")
    print(f"  ✅ 完成！结果已保存到:")
    print(f"     {OUTPUT_PATH}")
    print(f"{'═'*70}\n")


if __name__ == "__main__":
    main()
