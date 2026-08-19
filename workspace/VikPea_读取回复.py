"""
VikPea_读取回复.py — 自动读取邮箱回复 + 更新追踪表
  · 连接阿里企业邮箱 IMAP
  · 找最近7天所有回复邮件（Subject含 Re:）
  · 与 VikPea_邮件开发追踪.xlsx 匹配
  · 自动更新：是否回复 / 回复摘要 / 当前状态 / ABC分级
  · ABC分级规则：
      A = 感兴趣/报价/愿意合作
      B = 回复但中性（问问题/暂不考虑/需要更多信息）
      C = 拒绝/不感兴趣

用法: python3 ~/Downloads/VikPea_读取回复.py
"""

import imaplib, email, email.header, os, sys, re, json, subprocess, shutil, time
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("❌ pip3 install openpyxl")
    sys.exit(1)

try:
    from VikPea_common import apply_config, log_event
except ImportError:
    apply_config = None
    log_event = None

# ── 配置 ─────────────────────────────────────────────────────
IMAP_SERVER = "imap.qiye.aliyun.com"
IMAP_PORT   = 993
EMAIL_ADDR  = "hannah@hitpaw.com"
PASSWORD    = ""  # 密码已迁移到 VikPea_配置.xlsx / 网页系统设置，不再硬编码
DAYS_BACK   = 7   # 往前查几天

TRACKER_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "VikPea_邮件开发追踪.xlsx")
PROCESSED_IDS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              ".processed_message_ids.json")

# 列索引（1-based）
T_NUM    = 1
T_DATE   = 2
T_NAME   = 3
T_EMAIL  = 4
T_TYPE   = 5
T_CHAN   = 6
T_KIND   = 8
T_REPLY  = 9
T_SUM    = 10
T_LINK   = 11
T_STATUS = 12
T_GRADE  = 13   # ABC分级

EXTRA_HEADERS = [
    "回复意向",
    "账号粉丝数",
    "近热视频数",
    "平均播放",
    "VPH",
    "相关度",
    "综合分",
    "账号分析",
    "建议动作",
    "最后分析时间",
]

ANALYSIS_SHEET = "ABC分级表"
RECENT_VIDEO_LIMIT = 8
ACCOUNT_FETCH_DELAY = 1.2

# 填色
GREEN  = PatternFill("solid", start_color="E2EFDA")   # A级
YELLOW = PatternFill("solid", start_color="FFF2CC")   # B级
ORANGE = PatternFill("solid", start_color="FCE4D6")   # C级
BLUE   = PatternFill("solid", start_color="DDEBF7")   # 已更新标记
# ─────────────────────────────────────────────────────────────


def decode_header_str(s):
    if not s:
        return ""
    parts = email.header.decode_header(s)
    result = ""
    for part, enc in parts:
        if isinstance(part, bytes):
            result += part.decode(enc or "utf-8", errors="replace")
        else:
            result += str(part)
    return result.strip()


def get_body(msg) -> str:
    """提取邮件正文（纯文本）"""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            cd = str(part.get("Content-Disposition", ""))
            if ct == "text/plain" and "attachment" not in cd:
                charset = part.get_content_charset() or "utf-8"
                body = part.get_payload(decode=True).decode(charset, errors="replace")
                break
    else:
        charset = msg.get_content_charset() or "utf-8"
        body = msg.get_payload(decode=True).decode(charset, errors="replace")
    return body.strip()


def find_ytdlp():
    try:
        subprocess.run([sys.executable, "-m", "yt_dlp", "--version"],
                       capture_output=True, check=True)
        return [sys.executable, "-m", "yt_dlp"]
    except Exception:
        pass
    p = shutil.which("yt-dlp")
    if p:
        return [p]
    ver = f"{sys.version_info.major}.{sys.version_info.minor}"
    user_bin = os.path.expanduser(f"~/Library/Python/{ver}/bin/yt-dlp")
    if os.path.exists(user_bin):
        return [user_bin]
    return []


YTDLP_CMD = find_ytdlp()


def run_ytdlp_json(url: str, timeout: int = 90) -> dict:
    if not YTDLP_CMD or not url:
        return {}
    try:
        r = subprocess.run(
            YTDLP_CMD + [
                "--no-warnings", "--no-color", "--skip-download",
                "--playlist-end", str(RECENT_VIDEO_LIMIT),
                "--dump-single-json", url,
            ],
            capture_output=True, text=True, timeout=timeout
        )
        if r.returncode != 0 or not r.stdout.strip():
            return {}
        return json.loads(r.stdout)
    except Exception:
        return {}


def run_ytdlp_text(*args, timeout: int = 60) -> str:
    if not YTDLP_CMD:
        return ""
    try:
        r = subprocess.run(
            YTDLP_CMD + ["--no-warnings", "--no-color"] + list(args),
            capture_output=True, text=True, timeout=timeout
        )
        return r.stdout.strip()
    except Exception:
        return ""


def strip_quoted_reply(body: str) -> str:
    """尽量只保留对方新写的内容，避免把我们上一封邮件的关键词算进去。"""
    text = body or ""
    cut_patterns = [
        r"\nOn .+ wrote:",
        r"\n在 .+ 写道：",
        r"\nFrom:\s",
        r"\nSent:\s",
        r"\n发件人：",
        r"\n>+",
    ]
    cut_at = len(text)
    for pat in cut_patterns:
        m = re.search(pat, text, flags=re.I | re.S)
        if m:
            cut_at = min(cut_at, m.start())
    cleaned = text[:cut_at]
    cleaned = "\n".join(line for line in cleaned.splitlines() if not line.strip().startswith(">"))
    return cleaned.strip() or text[:500].strip()


def grade_reply(subject: str, body: str) -> tuple:
    """
    根据主题+正文判断ABC级别
    返回 (grade, summary)
    """
    clean_body = strip_quoted_reply(body)
    text = (subject + " " + clean_body).lower()

    # A级关键词：明确感兴趣
    a_keywords = [
        "interested", "sounds good", "love to", "would love", "happy to",
        "let's do", "let us know", "how much", "pricing", "price", "rate",
        "how does it work", "tell me more", "send me", "i'm in", "great idea",
        "looking forward", "可以", "感兴趣", "报价", "合作", "多少钱", "价格"
    ]
    # B级关键词：中性/问问题
    b_keywords = [
        "what is", "what are", "can you", "could you", "more info",
        "question", "wondering", "curious", "not sure", "maybe",
        "currently", "at the moment", "right now", "busy", "consider",
        "what kind", "how many", "viewers", "audience", "niche",
        "了解", "问一下", "什么", "怎么", "考虑"
    ]
    # C级关键词：拒绝
    c_keywords = [
        "not interested", "no thanks", "decline", "pass", "won't be",
        "don't think", "not a fit", "doesn't fit", "not relevant",
        "unsubscribe", "remove", "stop", "不感兴趣", "不合适", "谢谢但是"
    ]

    for kw in c_keywords:
        if kw in text:
            summary = clean_body[:150].replace("\n", " ").strip()
            return "C", summary
    for kw in a_keywords:
        if kw in text:
            summary = clean_body[:150].replace("\n", " ").strip()
            return "A", summary
    for kw in b_keywords:
        if kw in text:
            summary = clean_body[:150].replace("\n", " ").strip()
            return "B", summary

    # 默认B（至少回复了）
    summary = clean_body[:150].replace("\n", " ").strip()
    return "B", summary


def ensure_extra_headers(ws) -> dict:
    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        title = str(ws.cell(1, c).value or "").strip()
        if title:
            header_to_col[title] = c
    next_col = ws.max_column + 1
    for title in EXTRA_HEADERS:
        if title not in header_to_col:
            ws.cell(1, next_col).value = title
            header_to_col[title] = next_col
            next_col += 1
    return header_to_col


def first_youtube_url(*texts) -> str:
    for text in texts:
        s = str(text or "")
        m = re.search(r"https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^\s|,，]+", s, re.I)
        if m:
            return m.group(0).strip()
    return ""


def resolve_youtube_url(name: str, channel_field: str, link_field: str) -> str:
    url = first_youtube_url(link_field, channel_field)
    if url:
        return url
    raw = run_ytdlp_text(
        f"ytsearch1:{name} YouTube",
        "--skip-download",
        "--print", "%(channel_url)s",
        timeout=60
    )
    return raw.splitlines()[0].strip() if raw else ""


def parse_upload_hours(item: dict) -> float:
    ts = item.get("timestamp")
    if ts:
        hours = (datetime.now().timestamp() - float(ts)) / 3600
        return max(hours, 1)
    upload_date = str(item.get("upload_date") or "")
    if len(upload_date) == 8 and upload_date.isdigit():
        try:
            dt = datetime.strptime(upload_date, "%Y%m%d")
            return max((datetime.now() - dt).total_seconds() / 3600, 1)
        except Exception:
            return 0
    return 0


def fetch_account_metrics(name: str, channel_field: str, link_field: str) -> dict:
    url = resolve_youtube_url(name, channel_field, link_field)
    if not url:
        return {"url": "", "error": "缺少可查询的YouTube主页/视频链接"}

    data = run_ytdlp_json(url)
    if not data:
        return {"url": url, "error": "yt-dlp未抓到账号数据"}

    if data.get("_type") == "playlist":
        entries = [e for e in data.get("entries") or [] if isinstance(e, dict)]
        channel_url = data.get("channel_url") or data.get("webpage_url") or url
        subscribers = data.get("channel_follower_count") or 0
        channel_title = data.get("channel") or data.get("uploader") or name
    else:
        entries = [data]
        channel_url = data.get("channel_url") or data.get("uploader_url") or url
        subscribers = data.get("channel_follower_count") or 0
        channel_title = data.get("channel") or data.get("uploader") or name
        if channel_url and "watch?" in url:
            channel_data = run_ytdlp_json(channel_url)
            if channel_data:
                entries = [e for e in channel_data.get("entries") or [] if isinstance(e, dict)] or entries
                subscribers = channel_data.get("channel_follower_count") or subscribers
                channel_title = channel_data.get("channel") or channel_data.get("uploader") or channel_title

    views = []
    vph_values = []
    titles = []
    for item in entries[:RECENT_VIDEO_LIMIT]:
        title = str(item.get("title") or "").strip()
        if title:
            titles.append(title)
        view_count = item.get("view_count")
        if isinstance(view_count, int):
            views.append(view_count)
            hours = parse_upload_hours(item)
            if hours:
                vph_values.append(view_count / hours)

    avg_views = round(sum(views) / len(views)) if views else 0
    avg_vph = round(sum(vph_values) / len(vph_values), 1) if vph_values else 0
    return {
        "url": channel_url,
        "channel_title": channel_title,
        "subscribers": int(subscribers or 0),
        "video_count": len(entries[:RECENT_VIDEO_LIMIT]),
        "avg_views": avg_views,
        "vph": avg_vph,
        "titles": titles,
        "error": "",
    }


def relevance_score(name: str, channel_field: str, titles: list, summary: str) -> tuple:
    text = " ".join([name or "", channel_field or "", summary or ""] + (titles or [])).lower()
    primary = [
        "video enhancer", "video enhancement", "video upscaler", "video upscaling",
        "upscale video", "enhance video", "video quality", "video restoration",
        "topaz video", "hitpaw", "vikpea", "ai video", "4k video", "noise reduction",
    ]
    secondary = [
        "video editing", "editing software", "content creator", "youtube tutorial",
        "ai tools", "ai software", "generative ai", "photo editing", "creator tools",
    ]
    negative = ["meditation", "music", "movie recap", "gaming only", "asmr", "lyrics"]
    score = 0
    hits = []
    for kw in primary:
        if kw in text:
            score += 12
            hits.append(kw)
    for kw in secondary:
        if kw in text:
            score += 5
            hits.append(kw)
    for kw in negative:
        if kw in text:
            score -= 20
            hits.append(f"不相关:{kw}")
    score = max(0, min(100, score))
    reason = "、".join(hits[:6]) if hits else "未看到明显视频增强/AI视频相关词"
    return score, reason


def score_account(intent_grade: str, metrics: dict, rel_score: int) -> tuple:
    if intent_grade == "C":
        return "C", 0, "对方明确拒绝/退订，不建议继续推进"

    subs = metrics.get("subscribers", 0) or 0
    avg_views = metrics.get("avg_views", 0) or 0
    vph = metrics.get("vph", 0) or 0

    score = 0
    score += {"A": 25, "B": 12}.get(intent_grade, 8)

    if 5_000 <= subs <= 200_000:
        score += 22
    elif 200_000 < subs <= 500_000:
        score += 17
    elif 1_000 <= subs < 5_000:
        score += 8
    elif subs > 500_000:
        score += 8
    else:
        score += 4

    if avg_views >= 20_000:
        score += 20
    elif avg_views >= 8_000:
        score += 16
    elif avg_views >= 3_000:
        score += 12
    elif avg_views >= 1_000:
        score += 8
    elif avg_views > 0:
        score += 4

    if vph >= 300:
        score += 18
    elif vph >= 100:
        score += 14
    elif vph >= 30:
        score += 10
    elif vph >= 10:
        score += 6
    elif vph > 0:
        score += 3

    score += round(rel_score * 0.25)

    if rel_score < 20:
        final = "C" if score < 70 else "B"
        reason = "账号/标题与VikPea相关度偏低"
    elif score >= 72:
        final = "A"
        reason = "回复有推进空间，账号数据和相关度适合重点沟通"
    elif score >= 45:
        final = "B"
        reason = "可沟通，但优先级低于A级"
    else:
        final = "C"
        reason = "账号规模/播放/相关度不足，不建议投入太多时间"
    return final, score, reason


def analyze_tracker_row(ws, rownum: int, header_cols: dict, intent_grade: str, summary: str) -> str:
    name = str(ws.cell(rownum, T_NAME).value or "").strip()
    channel_field = str(ws.cell(rownum, T_CHAN).value or "").strip()
    link_field = str(ws.cell(rownum, T_LINK).value or "").strip()
    metrics = fetch_account_metrics(name, channel_field, link_field)
    rel, rel_reason = relevance_score(name, channel_field, metrics.get("titles", []), summary)

    if metrics.get("error"):
        final, score, score_reason = ("B", 35, metrics["error"]) if intent_grade != "C" else ("C", 0, "对方明确拒绝")
    else:
        final, score, score_reason = score_account(intent_grade, metrics, rel)

    analysis = (
        f"{score_reason}；相关度: {rel}({rel_reason})；"
        f"粉丝: {metrics.get('subscribers', 0)}；"
        f"近{metrics.get('video_count', 0)}条均播: {metrics.get('avg_views', 0)}；"
        f"VPH: {metrics.get('vph', 0)}"
    )
    action = {
        "A": "重点沟通：优先人工看主页/报价，适合推进合作",
        "B": "可沟通：回复即可，但控制时间和预算",
        "C": "不建议沟通：只记录，除非领导指定",
    }.get(final, "待人工判断")

    values = {
        "回复意向": intent_grade,
        "账号粉丝数": metrics.get("subscribers", 0),
        "近热视频数": metrics.get("video_count", 0),
        "平均播放": metrics.get("avg_views", 0),
        "VPH": metrics.get("vph", 0),
        "相关度": rel,
        "综合分": score,
        "账号分析": analysis,
        "建议动作": action,
        "最后分析时间": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    for title, value in values.items():
        ws.cell(rownum, header_cols[title]).value = value
    if metrics.get("url") and not link_field:
        ws.cell(rownum, T_LINK).value = metrics["url"]
    ws.cell(rownum, T_GRADE).value = final
    ws.cell(rownum, T_STATUS).value = "待处理" if final in ("A", "B") else "不建议沟通"
    color = {"A": GREEN, "B": YELLOW, "C": ORANGE}.get(final, YELLOW)
    for c in range(1, max(T_GRADE, header_cols["最后分析时间"]) + 1):
        ws.cell(rownum, c).fill = color
    return final


def collect_unanalyzed_reply_rows(ws, header_cols: dict, used_rows: set) -> list:
    rows = []
    last_col = header_cols.get("最后分析时间")
    for row in ws.iter_rows(min_row=2):
        rownum = row[0].row
        if rownum in used_rows:
            continue
        replied = str(row[T_REPLY - 1].value or "")
        if replied not in {"已回复", "是"}:
            continue
        if last_col and str(ws.cell(rownum, last_col).value or "").strip():
            continue
        summary = str(row[T_SUM - 1].value or "")
        old_grade = str(row[T_GRADE - 1].value or "").strip()
        intent_grade = grade_reply("", summary)[0] if summary else (old_grade if old_grade in {"A", "B", "C"} else "B")
        rows.append((rownum, intent_grade, summary))
    return rows


def main():
    if apply_config:
        apply_config(globals(), {
            "IMAP_SERVER": "IMAP_SERVER",
            "IMAP_PORT": "IMAP_PORT",
            "FROM_EMAIL": "EMAIL_ADDR",
            "PASSWORD": "PASSWORD",
        })
    print(f"\n🔌 连接 {IMAP_SERVER}...")
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(EMAIL_ADDR, PASSWORD)
        print("✅ 登录成功")
        if log_event:
            log_event("读取回复", f"登录成功，读取最近 {DAYS_BACK} 天")
    except Exception as e:
        print(f"❌ 连接失败: {e}")
        sys.exit(1)

    # 加载已处理过的 Message-ID（防止重复摘取）
    if os.path.exists(PROCESSED_IDS):
        with open(PROCESSED_IDS, "r") as f:
            processed_set = set(json.load(f))
    else:
        processed_set = set()
    print(f"📋 已处理过 {len(processed_set)} 封（本次跳过）")

    mail.select("INBOX")
    since = (datetime.now() - timedelta(days=DAYS_BACK)).strftime("%d-%b-%Y")
    _, ids = mail.search(None, f'SINCE {since}')
    all_ids = ids[0].split()
    print(f"📬 近{DAYS_BACK}天共 {len(all_ids)} 封邮件，筛选回复中...")

    # 读取所有邮件，找回复
    replies = []
    new_processed = []
    for eid in all_ids:
        try:
            _, data = mail.fetch(eid, "(RFC822)")
            msg = email.message_from_bytes(data[0][1])
            msg_id = (msg["Message-ID"] or "").strip()
            subj = decode_header_str(msg["Subject"] or "")
            sender = decode_header_str(msg["From"] or "")
            date_str = msg["Date"] or ""
            # 只要回复邮件（Re:开头）
            if not subj.lower().startswith("re:"):
                continue
            # 跳过已处理的
            if msg_id and msg_id in processed_set:
                continue
            body = get_body(msg)
            # 提取发件人邮箱
            email_match = re.search(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}", sender)
            from_email = email_match.group(0).lower() if email_match else ""
            replies.append({
                "subject": subj,
                "from_email": from_email,
                "from_name": sender,
                "date": date_str,
                "body": body,
                "msg_id": msg_id,
            })
            if msg_id:
                new_processed.append(msg_id)
        except Exception:
            continue

    mail.logout()
    print(f"✉️  找到 {len(replies)} 封新回复（已跳过重复）")

    # 打印所有回复
    if replies:
        print(f"\n{'─'*60}")
        for r in replies:
            print(f"  发件人: {r['from_name']}")
            print(f"  邮箱:   {r['from_email']}")
            print(f"  主题:   {r['subject']}")
            print(f"  时间:   {r['date']}")
            print(f"  正文:   {r['body'][:200]}")
            print()
    else:
        print("✅ 没有新回复；继续检查旧回复是否缺少账号分析")

    # 加载追踪表
    if not os.path.exists(TRACKER_PATH):
        print(f"❌ 找不到追踪表: {TRACKER_PATH}")
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = get_tracker_sheet(wb)

    # 确保T_GRADE列有表头
    if ws.cell(1, T_GRADE).value is None:
        ws.cell(1, T_GRADE).value = "ABC分级"
    header_cols = ensure_extra_headers(ws)

    # 建立邮箱→行号映射
    email_to_row = {}
    for row in ws.iter_rows(min_row=2):
        tracked_email = str(row[T_EMAIL - 1].value or "").strip().lower()
        if tracked_email:
            email_to_row[tracked_email] = row[0].row

    updated = 0
    new_rows = []
    rows_to_analyze = []

    for r in replies:
        intent_grade, summary = grade_reply(r["subject"], r["body"])
        color = {"A": GREEN, "B": YELLOW, "C": ORANGE}[intent_grade]

        if r["from_email"] in email_to_row:
            rownum = email_to_row[r["from_email"]]
            ws.cell(rownum, T_REPLY).value  = "已回复"
            ws.cell(rownum, T_SUM).value    = summary
            ws.cell(rownum, T_STATUS).value = "待分析"
            ws.cell(rownum, T_GRADE).value  = intent_grade
            ws.cell(rownum, header_cols["回复意向"]).value = intent_grade
            for c in range(1, max(T_GRADE, header_cols["回复意向"]) + 1):
                ws.cell(rownum, c).fill = color
            rows_to_analyze.append((rownum, intent_grade, summary))
            print(f"  ✅ 已更新回复: {r['from_email']} → 回复意向{intent_grade} | {ws.cell(rownum, T_NAME).value}")
            updated += 1
        else:
            # 在追踪表里找不到，新增一行
            last_num = ws.cell(ws.max_row, T_NUM).value
            new_num = (last_num + 1) if isinstance(last_num, int) else ws.max_row
            ws.append([
                new_num,
                datetime.now().strftime("%Y-%m-%d"),
                r["from_name"],
                r["from_email"],
                "未知",
                "",
                "",  # 视频链接，读回复时不知道，留空
                "收到回复",
                "已回复",
                summary,
                "",
                "待分析",
                intent_grade
            ])
            rownum = ws.max_row
            ws.cell(rownum, header_cols["回复意向"]).value = intent_grade
            for c in range(1, max(T_GRADE, header_cols["回复意向"]) + 1):
                ws.cell(ws.max_row, c).fill = color
            rows_to_analyze.append((rownum, intent_grade, summary))
            print(f"  ➕ 新增回复行: {r['from_email']} → 回复意向{intent_grade}（追踪表无记录）")
            new_rows.append(r["from_email"])

    used_rows = {rownum for rownum, _, _ in rows_to_analyze}
    old_rows = collect_unanalyzed_reply_rows(ws, header_cols, used_rows)
    if old_rows:
        rows_to_analyze.extend(old_rows)
        print(f"  另外发现 {len(old_rows)} 条历史已回复行缺少账号分析，也会一起补上")

    if rows_to_analyze:
        print(f"\n🔎 开始复查回复账号质量：{len(rows_to_analyze)} 个")
        for i, (rownum, intent_grade, summary) in enumerate(rows_to_analyze, 1):
            name = ws.cell(rownum, T_NAME).value
            print(f"  [{i}/{len(rows_to_analyze)}] {name}")
            try:
                final_grade = analyze_tracker_row(ws, rownum, header_cols, intent_grade, summary)
                print(f"      → 综合分级 {final_grade}")
            except KeyboardInterrupt:
                raise
            except Exception as exc:
                ws.cell(rownum, header_cols["账号分析"]).value = f"账号分析失败：{exc}"
                ws.cell(rownum, header_cols["建议动作"]).value = "待人工复查主页和数据"
                ws.cell(rownum, header_cols["最后分析时间"]).value = datetime.now().strftime("%Y-%m-%d %H:%M")
                if not ws.cell(rownum, T_GRADE).value:
                    ws.cell(rownum, T_GRADE).value = "B" if intent_grade != "C" else "C"
                if not ws.cell(rownum, T_STATUS).value:
                    ws.cell(rownum, T_STATUS).value = "待处理" if intent_grade != "C" else "不建议沟通"
                print(f"      ⚠️ 账号分析失败，已跳过并保留后续流程: {exc}")
            time.sleep(ACCOUNT_FETCH_DELAY)

    wb.save(TRACKER_PATH)

    # 记录本次已处理的 Message-ID，防止下次重复
    if new_processed:
        processed_set.update(new_processed)
        with open(PROCESSED_IDS, "w") as f:
            json.dump(list(processed_set), f)
        print(f"💾 已记录 {len(new_processed)} 个 Message-ID（下次自动跳过）")

    print(f"\n{'─'*60}")
    print(f"✅ 完成: 更新 {updated} 条 | 新增 {len(new_rows)} 条")
    print(f"   综合A级（重点沟通）: {sum(1 for rownum, _, _ in rows_to_analyze if ws.cell(rownum, T_GRADE).value=='A')}")
    print(f"   综合B级（可沟通）:   {sum(1 for rownum, _, _ in rows_to_analyze if ws.cell(rownum, T_GRADE).value=='B')}")
    print(f"   综合C级（不建议）:   {sum(1 for rownum, _, _ in rows_to_analyze if ws.cell(rownum, T_GRADE).value=='C')}")

    # ── 输出「待跟进」汇总 ───────────────────────────────────────
    print_pending_summary(ws)

    # ── 刷新追踪表中「待跟进」筛选页 ────────────────────────────
    wb2 = openpyxl.load_workbook(TRACKER_PATH)
    rebuild_abc_sheet(wb2)
    rebuild_pending_sheet(wb2)
    wb2.save(TRACKER_PATH)
    print(f"\n追踪表已保存（含「ABC分级表」「待跟进」页）: {TRACKER_PATH}")


def print_pending_summary(ws):
    """打印所有状态=待处理的行"""
    pending = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[T_STATUS - 1] or "")
        if status == "待处理":
            pending.append(row)

    if not pending:
        print("\n✅ 当前没有待处理项目")
        return

    print(f"\n{'═'*70}")
    print(f"  📋 待跟进清单（共 {len(pending)} 条）")
    print(f"{'═'*70}")
    for p in pending:
        grade  = str(p[T_GRADE - 1] or "?")
        name   = str(p[T_NAME  - 1] or "")
        em     = str(p[T_EMAIL - 1] or "")
        summ   = str(p[T_SUM   - 1] or "（无摘要）")[:80]
        kind   = str(p[T_KIND  - 1] or "")
        tag    = "🔥" if grade == "A" else "💬"
        print(f"  {tag} [{grade}级] {name:<25} {em}")
        print(f"       类型: {kind}")
        print(f"       摘要: {summ}")
        print()
    print(f"{'═'*70}")
    print(f"  → 请优先回复A级，再处理B级")


def get_header_cols(ws) -> dict:
    cols = {}
    for c in range(1, ws.max_column + 1):
        title = str(ws.cell(1, c).value or "").strip()
        if title:
            cols[title] = c
    return cols


def get_tracker_sheet(wb):
    """追踪工作簿里必须明确使用主表，不能依赖当前激活页。"""
    if "邮件追踪" in wb.sheetnames:
        wb.active = wb.sheetnames.index("邮件追踪")
        return wb["邮件追踪"]
    return wb.active


def rebuild_abc_sheet(wb):
    """重建账号综合ABC分级表。"""
    if ANALYSIS_SHEET in wb.sheetnames:
        del wb[ANALYSIS_SHEET]

    ws_main = get_tracker_sheet(wb)
    header_cols = get_header_cols(ws_main)
    ws_out = wb.create_sheet(ANALYSIS_SHEET)

    headers = [
        "ABC分级", "建议动作", "联系人/平台", "邮箱", "类型", "主页链接",
        "回复意向", "账号粉丝数", "近热视频数", "平均播放", "VPH",
        "相关度", "综合分", "回复摘要", "账号分析",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws_out.cell(1, c)
        cell.value = h
        cell.fill = PatternFill("solid", start_color="2F5496")
        cell.font = Font(color="FFFFFF", bold=True)

    rows = []
    for row in ws_main.iter_rows(min_row=2):
        replied = str(row[T_REPLY - 1].value or "")
        if replied not in {"已回复", "是"}:
            continue
        grade = str(row[T_GRADE - 1].value or "B")
        score_col = header_cols.get("综合分")
        score = ws_main.cell(row[0].row, score_col).value if score_col else 0
        try:
            score_num = float(score or 0)
        except Exception:
            score_num = 0
        order = {"A": 1, "B": 2, "C": 3}.get(grade, 4)
        rows.append((order, -score_num, row[0].row))

    rows.sort()
    out_row = 2
    for _, _, rownum in rows:
        grade = str(ws_main.cell(rownum, T_GRADE).value or "B")
        color = {"A": GREEN, "B": YELLOW, "C": ORANGE}.get(grade, YELLOW)
        values = [
            grade,
            ws_main.cell(rownum, header_cols.get("建议动作", T_STATUS)).value,
            ws_main.cell(rownum, T_NAME).value,
            ws_main.cell(rownum, T_EMAIL).value,
            ws_main.cell(rownum, T_TYPE).value,
            ws_main.cell(rownum, T_LINK).value,
            ws_main.cell(rownum, header_cols.get("回复意向", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("账号粉丝数", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("近热视频数", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("平均播放", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("VPH", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("相关度", T_GRADE)).value,
            ws_main.cell(rownum, header_cols.get("综合分", T_GRADE)).value,
            ws_main.cell(rownum, T_SUM).value,
            ws_main.cell(rownum, header_cols.get("账号分析", T_SUM)).value,
        ]
        for c, value in enumerate(values, 1):
            cell = ws_out.cell(out_row, c)
            cell.value = value
            cell.fill = color
        out_row += 1

    widths = {
        "A": 10, "B": 38, "C": 24, "D": 30, "E": 16, "F": 36,
        "G": 10, "H": 12, "I": 10, "J": 12, "K": 10, "L": 10,
        "M": 10, "N": 45, "O": 65,
    }
    for col, width in widths.items():
        ws_out.column_dimensions[col].width = width
    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:O{max(out_row - 1, 1)}"


def rebuild_pending_sheet(wb):
    """在追踪表中重建「待跟进」工作页"""
    SHEET_NAME = "待跟进"
    # 删除旧页
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    ws_main = get_tracker_sheet(wb)
    header_cols = get_header_cols(ws_main)
    ws_pend = wb.create_sheet(SHEET_NAME)

    # 表头
    headers = ["#", "日期", "联系人/平台", "邮箱", "类型", "频道/文章",
               "邮件类型", "是否回复", "回复摘要", "主页链接", "当前状态", "ABC分级",
               "综合分", "相关度", "账号分析", "建议动作"]
    for c, h in enumerate(headers, 1):
        cell = ws_pend.cell(1, c)
        cell.value = h
        cell.fill = PatternFill("solid", start_color="2F5496")
        cell.font = Font(color="FFFFFF", bold=True)

    # 筛选待处理行
    row_out = 2
    for row in ws_main.iter_rows(min_row=2):
        status = str(row[T_STATUS - 1].value or "")
        if status == "待处理":
            grade = str(row[T_GRADE - 1].value or "B")
            color = {"A": GREEN, "B": YELLOW, "C": ORANGE}.get(grade, YELLOW)
            for c, cell in enumerate(row[:12], 1):
                new_cell = ws_pend.cell(row_out, c)
                new_cell.value = cell.value
                new_cell.fill = color
            extra_values = [
                ws_main.cell(row[0].row, header_cols.get("综合分", T_GRADE)).value,
                ws_main.cell(row[0].row, header_cols.get("相关度", T_GRADE)).value,
                ws_main.cell(row[0].row, header_cols.get("账号分析", T_SUM)).value,
                ws_main.cell(row[0].row, header_cols.get("建议动作", T_STATUS)).value,
            ]
            for offset, value in enumerate(extra_values, 13):
                new_cell = ws_pend.cell(row_out, offset)
                new_cell.value = value
                new_cell.fill = color
            row_out += 1

    # 列宽
    ws_pend.column_dimensions["C"].width = 25
    ws_pend.column_dimensions["D"].width = 30
    ws_pend.column_dimensions["I"].width = 40
    ws_pend.column_dimensions["O"].width = 65
    ws_pend.column_dimensions["P"].width = 36


if __name__ == "__main__":
    main()
