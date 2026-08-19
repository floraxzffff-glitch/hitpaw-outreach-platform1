"""
VikPea 读表发信脚本 v2 — 联动版
  • 从 VikPea_发信名单.xlsx 读取待发目标（绿色行 = 有邮箱）
  • 发送成功 → 自动写入 VikPea_邮件开发追踪.xlsx → 从名单中删除该行
  • 发送失败 → 在名单备注列标记"发送失败"，保留行
  • 无邮箱行（黄色）直接跳过

文件路径（与本脚本同目录）:
  VikPea_发信名单.xlsx         ← 待发队列
  VikPea_邮件开发追踪.xlsx     ← 历史总追踪表

用法: python3 ~/Downloads/VikPea_读表发信.py
"""

import smtplib, ssl, time, os, sys, json, shutil, subprocess, re, socket
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from VikPea_common import (
        apply_config, load_blacklist, safety_reason, export_send_preview,
        log_event, PREVIEW_PATH, add_no_email_candidate, ensure_queue_headers,
        CONFIG_PATH, CONFIG_VKP_PATH, CONFIG_FP_PATH, append_send_queue_row,
        save_workbook_safe, PENDING_EMAIL_REVIEW_PATH,
    )
except ImportError:
    apply_config = None
    load_blacklist = None
    safety_reason = None
    export_send_preview = None
    log_event = None
    PREVIEW_PATH = ""
    add_no_email_candidate = None
    ensure_queue_headers = None
    CONFIG_PATH = ""
    CONFIG_VKP_PATH = ""
    CONFIG_FP_PATH = ""
    append_send_queue_row = None
    save_workbook_safe = None
    PENDING_EMAIL_REVIEW_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                             "VikPea_待确认邮箱.xlsx")

# ── SMTP 配置 ───────────────────────────────────────────
SMTP_SERVER = "smtp.qiye.aliyun.com"
SMTP_PORT   = 465
FROM_EMAIL  = "hannah@hitpaw.com"
FROM_NAME   = "Hannah"
PASSWORD    = ""  # 密码已迁移到 VikPea_配置.xlsx / 网页系统设置，不再硬编码
PRODUCT_NAME = "HitPaw VikPea"
PRODUCT_TEAM = "HitPaw Team"
PRODUCT_URL = "https://www.hitpaw.com/hitpaw-video-enhancer.html"
OUTREACH_SUBJECT_YOUTUBE = "Quick collaboration idea for your channel"
OUTREACH_SUBJECT_ARTICLE = "Quick collaboration idea for your article"
OUTREACH_TEMPLATE_YOUTUBE = """{opening}

I'm {from_name} from {product_name}.

We're looking for simple creator collaborations:
1. adding {product_name} as a relevant link/tool mention in a video description, or
2. testing {product_name} for a dedicated video if it fits your content.

Would you be open to either option? If yes, please feel free to send your rate.

Best,
{from_name}
{product_team}
{product_url}"""
OUTREACH_TEMPLATE_ARTICLE = """{opening}

I'm {from_name} from {product_name}.

We're looking for simple collaborations:
1. adding {product_name} as a relevant link/tool mention in an existing article, or
2. testing {product_name} for a dedicated review/tutorial if it fits your content.

Would you be open to either option? If yes, please feel free to send your rate.

Best,
{from_name}
{product_team}
{product_url}"""
DELAY_SEC   = 8
USE_LIGHT_OUTREACH = True
PERSONALIZE_RECENT_VIDEOS = True
RECENT_VIDEO_COUNT = 4
PERSONALIZE_TIMEOUT = 12
PERSONALIZE_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 ".outreach_recent_video_cache.json")
DAILY_SEND_LIMIT = 80
REQUIRE_SEND_CODE = "SEND"
SMTP_TIMEOUT = 25
SMTP_ALLOW_INSECURE_SSL = True
EXTRA_DEDUPE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "VikPea_额外已联络去重.xlsx")
NO_EMAIL_POOL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  "VikPea_无邮箱候选.xlsx")
# ──────────────────────────────────────────────────────

SEND_CONFIRM_WORDS = {"y", "yes", "send", "发送", "确认"}
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")

# 发信名单列（1-based）
Q_NAME  = 1   # 频道名/网站名
Q_EMAIL = 2   # 邮箱
Q_SUBJ  = 3   # 定制主题
Q_OPEN  = 4   # 定制开头
Q_HOME  = 5   # 主页链接
Q_VIDEO = 6   # 命中视频链接/文章链接
Q_NOTE  = 7   # 备注
Q_TYPE  = 8   # 类型: 留空/"YouTube" = YouTube视频插链; "文章外链" = 网站文章插链
Q_SOURCE = 9  # 来源关键词

# 追踪表列（1-based）
T_NUM    = 1   # #
T_DATE   = 2   # 日期
T_NAME   = 3   # 联系人/平台
T_EMAIL  = 4   # 邮箱
T_TYPE   = 5   # 类型
T_CHAN   = 6   # 频道/文章/平台
T_VIDEO  = 7   # 视频链接
T_KIND   = 8   # 邮件类型
T_REPLY  = 9   # 是否回复
T_SUM    = 10  # 回复摘要
T_LINK   = 11  # 主页链接
T_STATUS = 12  # 当前状态
T_SOURCE = 19  # 来源关键词（14-18列留给自动跟进脚本）
T_TAG    = 20  # 频道标签

P_DEEP_STATUS = 12
P_DEEP_DATE = 13

def get_path(filename):
    d = os.path.dirname(os.path.abspath(__file__))
    p = os.path.join(d, filename)
    if not os.path.exists(p):
        print(f"❌ 找不到文件: {p}")
        sys.exit(1)
    return p


def normalize_email_type(value, link=""):
    text = str(value or "").strip()
    if text:
        return text
    if "youtube.com" in str(link or "").lower() or "youtu.be" in str(link or "").lower():
        return "YouTube"
    return "YouTube"


def sync_manual_emails_from_pool(ws_q, ws_t, sent_keys):
    """把无邮箱候选表里人工补好的邮箱自动转入主发信表。"""
    if not os.path.exists(NO_EMAIL_POOL_PATH):
        return 0
    wb_p = openpyxl.load_workbook(NO_EMAIL_POOL_PATH)
    ws_p = wb_p.active

    queue_links = set()
    queue_emails = set()
    queue_names = set()
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        queue_names.add(norm_text(row[Q_NAME - 1] if len(row) >= Q_NAME else ""))
        queue_emails.add(norm_text(row[Q_EMAIL - 1] if len(row) >= Q_EMAIL else ""))
        queue_links.add(norm_link(row[Q_HOME - 1] if len(row) >= Q_HOME else ""))

    moved = 0
    today = datetime.now().strftime("%Y-%m-%d")
    for rownum in range(2, ws_p.max_row + 1):
        name = str(ws_p.cell(rownum, Q_NAME).value or "").strip()
        email = str(ws_p.cell(rownum, Q_EMAIL).value or "").strip()
        subj = str(ws_p.cell(rownum, Q_SUBJ).value or "").strip()
        opening = str(ws_p.cell(rownum, Q_OPEN).value or "").strip()
        home_link = str(ws_p.cell(rownum, Q_HOME).value or "").strip()
        video_link = str(ws_p.cell(rownum, Q_VIDEO).value or "").strip() if ws_p.max_column >= Q_VIDEO else ""
        link = video_link or home_link
        note = str(ws_p.cell(rownum, Q_NOTE).value or "").strip()
        etype = normalize_email_type(ws_p.cell(rownum, Q_TYPE).value, link)
        source_kw = str(ws_p.cell(rownum, Q_SOURCE).value or "").strip() or "无邮箱候选补录"
        status = str(ws_p.cell(rownum, P_DEEP_STATUS).value or "").strip()

        if not name or not email or "@" not in email:
            continue
        if status in {"已转主表", "已发送"}:
            continue

        email_key = norm_text(email)
        link_key = norm_link(home_link or link)
        name_key = norm_text(name)
        if email_key in queue_emails or link_key in queue_links or name_key in queue_names:
            ws_p.cell(rownum, P_DEEP_STATUS).value = "已转主表"
            if "补录邮箱已在主表/队列" not in note:
                ws_p.cell(rownum, Q_NOTE).value = (note + " | " if note else "") + "补录邮箱已在主表/队列"
            ws_p.cell(rownum, P_DEEP_DATE).value = today
            continue

        sent_reason = already_sent_reason(name, email, link, sent_keys)
        if sent_reason:
            ws_p.cell(rownum, P_DEEP_STATUS).value = "已发送"
            if sent_reason not in note:
                ws_p.cell(rownum, Q_NOTE).value = (note + " | " if note else "") + sent_reason
            ws_p.cell(rownum, P_DEEP_DATE).value = today
            continue

        if not subj:
            subj = make_fallback_subject(name, etype, source_kw)
        if not opening:
            opening = make_fallback_opening(name, link, etype, source_kw)

        append_send_queue_row(
            ws_q, name, email, subj, opening, home_link,
            (note + " | " if note else "") + "无邮箱候选中的补录邮箱自动转入主表",
            etype, source_kw, video_link=video_link, home_link=home_link
        )
        ws_p.cell(rownum, P_DEEP_STATUS).value = "已转主表"
        ws_p.cell(rownum, P_DEEP_DATE).value = today
        ws_p.cell(rownum, Q_NOTE).value = (note + " | " if note else "") + "无邮箱候选中的补录邮箱已自动转入主发信表"
        queue_emails.add(email_key)
        queue_links.add(link_key)
        queue_names.add(name_key)
        moved += 1

    if moved:
        save_book(wb_p, NO_EMAIL_POOL_PATH, "无邮箱候选")
    else:
        # 即使没有新增转入，也要保存状态修正
        save_book(wb_p, NO_EMAIL_POOL_PATH, "无邮箱候选")
    return moved


def sync_pending_email_review(ws_q, ws_t, sent_keys):
    """把待确认邮箱表里人工确认过的邮箱自动转入主发信表。"""
    if not os.path.exists(PENDING_EMAIL_REVIEW_PATH):
        return 0
    wb_p = openpyxl.load_workbook(PENDING_EMAIL_REVIEW_PATH)
    ws_p = wb_p.active

    queue_links = set()
    queue_emails = set()
    queue_names = set()
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        queue_names.add(norm_text(row[Q_NAME - 1] if len(row) >= Q_NAME else ""))
        queue_emails.add(norm_text(row[Q_EMAIL - 1] if len(row) >= Q_EMAIL else ""))
        queue_links.add(norm_link(row[Q_HOME - 1] if len(row) >= Q_HOME else ""))

    moved = 0
    today = datetime.now().strftime("%Y-%m-%d")
    for rownum in range(2, ws_p.max_row + 1):
        action = str(ws_p.cell(rownum, 11).value or "").strip().lower()
        if action in {"放弃", "不要", "不发", "skip", "no"}:
            continue
        if action not in {"转主表", "转入主表", "转发信名单", "转入发信名单", "send", "y", "yes"}:
            continue

        name = str(ws_p.cell(rownum, 1).value or "").strip()
        candidate_email = str(ws_p.cell(rownum, 2).value or "").strip()
        confirmed_email = str(ws_p.cell(rownum, 12).value or "").strip()
        email = confirmed_email or candidate_email
        confidence = str(ws_p.cell(rownum, 3).value or "").strip()
        source_label = str(ws_p.cell(rownum, 4).value or "").strip()
        home_link = str(ws_p.cell(rownum, 5).value or "").strip()
        video_link = str(ws_p.cell(rownum, 6).value or "").strip()
        source_kw = str(ws_p.cell(rownum, 7).value or "").strip() or "待确认邮箱人工确认"
        etype = normalize_email_type(ws_p.cell(rownum, 8).value, home_link or video_link)
        note = str(ws_p.cell(rownum, 9).value or "").strip()
        manual_note = str(ws_p.cell(rownum, 13).value or "").strip()

        if not name or not email or not EMAIL_RE.match(email):
            ws_p.cell(rownum, 13).value = (manual_note + " | " if manual_note else "") + "未转入：邮箱为空或格式不对"
            continue

        email_key = norm_text(email)
        link_key = norm_link(home_link or video_link)
        name_key = norm_text(name)
        if email_key in queue_emails or link_key in queue_links or name_key in queue_names:
            ws_p.cell(rownum, 11).value = "已转主表"
            ws_p.cell(rownum, 13).value = (manual_note + " | " if manual_note else "") + "已存在于发信名单"
            continue

        sent_reason = already_sent_reason(name, email, home_link or video_link, sent_keys)
        if sent_reason:
            ws_p.cell(rownum, 11).value = "已发送"
            ws_p.cell(rownum, 13).value = (manual_note + " | " if manual_note else "") + sent_reason
            continue

        subj = make_fallback_subject(name, etype, source_kw)
        opening = make_fallback_opening(name, home_link or video_link, etype, source_kw)
        review_note = "待确认邮箱人工确认后转入主表"
        if confidence or source_label or note:
            review_note += f" | 置信度:{confidence or '-'} | 来源:{source_label or '-'} | {note}"
        append_send_queue_row(
            ws_q, name, email, subj, opening, home_link,
            review_note, etype, source_kw, video_link=video_link, home_link=home_link
        )
        ws_p.cell(rownum, 11).value = "已转主表"
        ws_p.cell(rownum, 12).value = email
        ws_p.cell(rownum, 13).value = (manual_note + " | " if manual_note else "") + f"{today} 已转入发信名单"
        queue_emails.add(email_key)
        queue_links.add(link_key)
        queue_names.add(name_key)
        moved += 1

    save_book(wb_p, PENDING_EMAIL_REVIEW_PATH, "待确认邮箱")
    return moved


def migrate_no_email_rows(ws_q):
    """把主发信表里的无邮箱候选移到副表，避免主表越来越乱。"""
    if not add_no_email_candidate:
        return 0
    to_move = []
    keep_rows = []
    for row in ws_q.iter_rows(min_row=2):
        rownum = row[0].row
        values = [cell.value for cell in row]
        name = str(row[Q_NAME - 1].value or "").strip()
        email = str(row[Q_EMAIL - 1].value or "").strip()
        home_link = str(row[Q_HOME - 1].value or "").strip()
        video_link = str(row[Q_VIDEO - 1].value or "").strip() if ws_q.max_column >= Q_VIDEO else ""
        link = video_link or home_link
        if not name and not link:
            continue
        if email and "@" in email:
            keep_rows.append(values)
            continue
        subj = str(row[Q_SUBJ - 1].value or "").strip()
        opening = str(row[Q_OPEN - 1].value or "").strip()
        note = str(row[Q_NOTE - 1].value or "").strip()
        etype = normalize_email_type(row[Q_TYPE - 1].value if ws_q.max_column >= Q_TYPE else "", link)
        source = str(row[Q_SOURCE - 1].value or "").strip() if ws_q.max_column >= Q_SOURCE else ""
        to_move.append((rownum, name, home_link, video_link, note, etype, source, subj, opening))

    for rownum, name, home_link, video_link, note, etype, source, subj, opening in to_move:
        add_no_email_candidate(
            name=name,
            link=home_link,
            note=note or "主发信表迁移：无邮箱",
            email_type=etype,
            source=source or "历史主表迁移",
            subject=subj,
            opening=opening,
            source_row=rownum,
            status="待深度查邮箱",
            video_link=video_link,
        )

    if to_move:
        max_cols = max(ws_q.max_column, Q_SOURCE)
        ws_q.delete_rows(2, max(1, ws_q.max_row - 1))
        for values in keep_rows:
            ws_q.append((values + [None] * max_cols)[:max_cols])
    return len(to_move)


def is_send_confirm(value):
    text = str(value or "").strip().lower()
    required = str(REQUIRE_SEND_CODE or "").strip().lower()
    return text in SEND_CONFIRM_WORDS or (required and text == required)


def is_ssl_cert_error(exc) -> bool:
    text = str(exc or "")
    return isinstance(exc, ssl.SSLCertVerificationError) or "CERTIFICATE_VERIFY_FAILED" in text or "certificate verify failed" in text.lower()


def smtp_connect():
    """
    优先正常 SSL 连接；若当前 Python 证书链异常，则自动回退到不校验证书模式，
    避免因为本机 Python 证书问题导致整批无法发送。
    """
    secure_ctx = ssl.create_default_context()
    try:
        server = smtplib.SMTP_SSL(SMTP_SERVER, int(SMTP_PORT), context=secure_ctx, timeout=SMTP_TIMEOUT)
        return server, "secure"
    except Exception as exc:
        if SMTP_ALLOW_INSECURE_SSL and is_ssl_cert_error(exc):
            print("⚠️ SMTP 证书校验失败，已自动启用兼容模式继续连接。")
            print("   ↳ 这通常是本机 Python 证书链异常，不是邮箱账号本身有问题。")
            insecure_ctx = ssl._create_unverified_context()
            server = smtplib.SMTP_SSL(SMTP_SERVER, int(SMTP_PORT), context=insecure_ctx, timeout=SMTP_TIMEOUT)
            return server, "insecure"
        raise


def load_personalize_cache():
    if not os.path.exists(PERSONALIZE_CACHE):
        return {}
    try:
        with open(PERSONALIZE_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_personalize_cache(cache):
    try:
        with open(PERSONALIZE_CACHE, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def find_ytdlp():
    try:
        subprocess.run([sys.executable, "-m", "yt_dlp", "--version"],
                       capture_output=True, check=True, timeout=10)
        return [sys.executable, "-m", "yt_dlp"]
    except Exception:
        pass
    p = shutil.which("yt-dlp")
    if p:
        return [p]
    ver = f"{sys.version_info.major}.{sys.version_info.minor}"
    user_bin = os.path.expanduser(f"~/Library/Python/{ver}/bin/yt-dlp")
    if os.path.exists(user_bin):
        return [user_bin]
    return []


YTDLP_CMD = find_ytdlp()


def run_ytdlp(args, timeout=PERSONALIZE_TIMEOUT):
    if not YTDLP_CMD:
        return ""
    try:
        r = subprocess.run(
            YTDLP_CMD + ["--no-warnings", "--no-color"] + list(args),
            capture_output=True, text=True, timeout=timeout
        )
        if r.returncode != 0:
            return ""
        return r.stdout.strip()
    except Exception:
        return ""


def normalize_channel_url(url):
    text = (url or "").strip()
    if not text or "youtube.com" not in text and "youtu.be" not in text:
        return ""

    channel_url = run_ytdlp(["--skip-download", "--print", "%(channel_url)s", text], timeout=35)
    channel_url = channel_url.splitlines()[0].strip() if channel_url else ""
    if channel_url and channel_url != "NA":
        return channel_url.rstrip("/")
    if "/@" in text or "/channel/" in text or "/c/" in text:
        return text.rstrip("/")
    return ""


def get_recent_video_titles(link, cache):
    if not PERSONALIZE_RECENT_VIDEOS or not link:
        return []
    cache_key = link.strip().rstrip("/")
    if cache_key in cache:
        return cache[cache_key].get("titles", [])

    channel_url = normalize_channel_url(link)
    urls_to_try = []
    if channel_url:
        urls_to_try.extend([channel_url + "/videos", channel_url])
    urls_to_try.append(link)

    titles = []
    for url in urls_to_try:
        raw = run_ytdlp([
            "--skip-download",
            "--flat-playlist",
            "--playlist-end", str(RECENT_VIDEO_COUNT),
            "--print", "%(title)s",
            url,
        ])
        for line in raw.splitlines():
            title = line.strip()
            if title and title != "NA" and title not in titles:
                titles.append(title)
            if len(titles) >= RECENT_VIDEO_COUNT:
                break
        if titles:
            break

    cache[cache_key] = {"channel_url": channel_url, "titles": titles}
    save_personalize_cache(cache)
    return titles


def trim_title(title, limit=78):
    title = " ".join(str(title or "").split())
    if len(title) <= limit:
        return title
    return title[:limit - 3].rstrip() + "..."


def describe_recent_content(titles, source_keyword=""):
    text = " ".join(titles or []).lower()
    if any(w in text for w in ["upscale", "enhance", "quality", "4k", "1080p", "blurry", "restore", "video enhancer"]):
        return "video quality improvement and enhancement tutorials"
    if any(w in text for w in ["ai", "tool", "tools", "generator", "automation", "chatgpt"]):
        return "practical AI tool and creator workflow content"
    if any(w in text for w in ["capcut", "premiere", "davinci", "editing", "edit", "editor"]):
        return "hands-on video editing tutorials"
    if any(w in text for w in ["android", "iphone", "mobile", "phone"]):
        return "clear mobile video and app tutorials"
    if source_keyword:
        return f'content around "{source_keyword}"'
    return "practical creator-focused videos"


def make_personalized_opening(name, link="", email_type="", source_keyword="", titles=None):
    titles = [trim_title(t) for t in (titles or []) if str(t or "").strip()]
    if str(email_type).strip() == "文章外链":
        return make_fallback_opening(name, link, email_type, source_keyword)
    if not titles:
        return make_fallback_opening(name, link, email_type, source_keyword)
    content_desc = describe_recent_content(titles, source_keyword)
    return (
        f"I came across your recent channel content. I liked how your channel focuses on "
        f"{content_desc}, which feels very relevant for viewers comparing ways to improve video quality."
    )


def render_template_text(template, **values):
    text = str(template or "")
    for key, value in values.items():
        text = text.replace("{" + key + "}", str(value if value is not None else ""))
    return text


def make_email(name, to_email, subject, opening, email_type=""):
    template = OUTREACH_TEMPLATE_ARTICLE if str(email_type).strip() == "文章外链" else OUTREACH_TEMPLATE_YOUTUBE
    body_core = render_template_text(
        template,
        name=name,
        opening=opening,
        from_name=FROM_NAME,
        from_email=FROM_EMAIL,
        product_name=PRODUCT_NAME,
        product_team=PRODUCT_TEAM,
        product_url=PRODUCT_URL,
    )
    body = f"Hi {name},\n\n" + body_core
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"]      = to_email
    msg.attach(MIMEText(body, "plain", "utf-8"))
    return msg, body


def choose_config_profile():
    options = [
        ("1", "通用配置", CONFIG_PATH, "default"),
        ("2", "VKP 配置", CONFIG_VKP_PATH, "vkp"),
        ("3", "FP 配置", CONFIG_FP_PATH, "fp"),
    ]
    existing = [item for item in options if item[2] and os.path.exists(item[2])]
    if len(existing) <= 1:
        return (existing[0][2], existing[0][3], existing[0][1]) if existing else ("", "default", "内置默认")
    print("\n请选择本次发信使用的配置表：")
    for key, label, path, _profile in existing:
        print(f"  {key}. {label} -> {os.path.basename(path)}")
    choice = input("输入 1 / 2 / 3，直接回车默认 1: ").strip() or "1"
    matched = [item for item in existing if item[0] == choice]
    if not matched:
        matched = [existing[0]]
    _, label, path, profile = matched[0]
    return path, profile, label


def subject_topic_hint(source_keyword="", titles=None):
    text = " ".join([str(source_keyword or "")] + [str(t or "") for t in (titles or [])]).lower()
    if any(w in text for w in ["capcut", "premiere", "davinci", "editing", "editor", "autocut"]):
        return "video editing"
    if any(w in text for w in ["upscale", "upscaler", "4k", "resolution", "quality enhancer", "enhance"]):
        return "video quality"
    if any(w in text for w in ["restore", "restoration", "denoise", "deblur", "blurry", "noise"]):
        return "video restoration"
    if any(w in text for w in ["ai tool", "chatgpt", "automation", "workflow", "generator"]):
        return "AI tool"
    if source_keyword:
        short_kw = " ".join(str(source_keyword).strip().split()[:4])
        return short_kw
    return "video content"


def pick_subject_template(name="", email_type="", source_keyword="", titles=None):
    etype = str(email_type).strip()
    config_subject = OUTREACH_SUBJECT_ARTICLE if etype == "文章外链" else OUTREACH_SUBJECT_YOUTUBE
    config_subject = str(config_subject or "").strip()
    if config_subject:
        return render_template_text(
            config_subject,
            name=name,
            from_name=FROM_NAME,
            from_email=FROM_EMAIL,
            product_name=PRODUCT_NAME,
            product_team=PRODUCT_TEAM,
            product_url=PRODUCT_URL,
            source_keyword=source_keyword or "",
        ).strip()
    topic = subject_topic_hint(source_keyword, titles)
    key = (str(name or "") + "|" + str(source_keyword or "") + "|" + topic).lower()
    templates_article = [
        f"Quick collaboration idea for your {topic} content",
        f"Possible VikPea mention for your {topic} article?",
        f"Small partnership idea for {topic} content",
        f"Question about your {topic} article",
    ]
    templates_youtube = [
        f"Quick collaboration idea for your {topic} videos",
        f"Possible VikPea fit for your {topic} content?",
        f"A small idea for your {topic} audience",
        f"Question about your recent {topic} videos",
    ]
    pool = templates_article if etype == "文章外链" else templates_youtube
    idx = sum(ord(ch) for ch in key) % len(pool)
    return pool[idx]


def make_fallback_subject(name, email_type="", source_keyword="", titles=None):
    return pick_subject_template(name, email_type, source_keyword, titles)


def make_fallback_opening(name, link="", email_type="", source_keyword=""):
    if str(email_type).strip() == "文章外链":
        return (
            f"I came across your article on {name} while researching AI video enhancement tools — "
            "it looks like a helpful resource for readers comparing options."
        )
    if source_keyword:
        return (
            f"I came across your content while researching \"{source_keyword}\" — "
            "it seems relevant for viewers exploring AI video enhancement tools."
        )
    if link:
        return (
            "I came across your video while researching AI video enhancement content — "
            "it seems relevant for viewers comparing tools to improve video quality."
        )
    return (
        f"I came across your channel, {name}, while researching AI video enhancement content — "
        "it seems relevant for creators comparing video quality tools."
    )


def ensure_tracker_headers(ws):
    if not ws.cell(1, T_SOURCE).value:
        ws.cell(1, T_SOURCE).value = "来源关键词"


def norm_text(value):
    return " ".join(str(value or "").strip().lower().split())


def norm_link(value):
    return str(value or "").strip().rstrip("/").lower()


def header_aliases():
    return {
        "name": {
            "频道名", "名称", "联系人", "联系人/平台", "账号名", "博主名", "channel", "name",
            "channel name", "creator", "频道/文章/平台", "平台"
        },
        "email": {
            "邮箱", "联系邮箱", "email", "e-mail", "mail"
        },
        "link": {
            "主页链接", "链接", "主页", "主页/视频链接", "视频链接", "频道链接", "url",
            "channel url", "link", "page url"
        },
    }


def simplify_header(value):
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-:/]+", "", text)


def find_header_columns(ws):
    aliases = header_aliases()
    headers = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(1, col).value
        key = simplify_header(raw)
        if not key:
            continue
        for field, candidates in aliases.items():
            if field in headers:
                continue
            normalized = {simplify_header(x) for x in candidates}
            if key in normalized:
                headers[field] = col
    return headers


def merge_row_to_sent_keys(name, email, link, sent_keys):
    name = norm_text(name)
    email = norm_text(email)
    link = norm_link(link)
    if email and "@" in email:
        sent_keys["emails"].add(email)
    if name:
        sent_keys["names"].add(name)
    if link:
        sent_keys["links"].add(link)


def load_extra_sent_keys(extra_path, sent_keys):
    """允许把同事历史已联络表改名后直接并入去重。"""
    if not os.path.exists(extra_path):
        return 0
    try:
        wb = openpyxl.load_workbook(extra_path, data_only=True)
    except Exception as exc:
        print(f"⚠️ 额外去重表读取失败，已忽略: {exc}")
        return 0

    added = 0
    for sheet in wb.worksheets:
        cols = find_header_columns(sheet)
        if not cols:
            continue
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, cols.get("name", 0)).value if cols.get("name") else ""
            email = sheet.cell(row, cols.get("email", 0)).value if cols.get("email") else ""
            link = sheet.cell(row, cols.get("link", 0)).value if cols.get("link") else ""
            if not any([name, email, link]):
                continue
            before = (len(sent_keys["emails"]), len(sent_keys["names"]), len(sent_keys["links"]))
            merge_row_to_sent_keys(name, email, link, sent_keys)
            after = (len(sent_keys["emails"]), len(sent_keys["names"]), len(sent_keys["links"]))
            if after != before:
                added += 1
    if added:
        print(f"📚 额外去重表已生效：{os.path.basename(extra_path)}，载入 {added} 条历史联络键")
    else:
        print(f"📚 检测到额外去重表：{os.path.basename(extra_path)}，但未识别到可用列")
    return added


def load_sent_keys(tracker_ws, extra_path=""):
    """从追踪总表读取已经联系过的人，发信前强制去重。"""
    emails = set()
    names = set()
    links = set()
    sent_keys = {"emails": emails, "names": names, "links": links}
    for row in tracker_ws.iter_rows(min_row=2, values_only=True):
        name = norm_text(row[T_NAME - 1] if len(row) >= T_NAME else "")
        email = norm_text(row[T_EMAIL - 1] if len(row) >= T_EMAIL else "")
        chan = norm_link(row[T_CHAN - 1] if len(row) >= T_CHAN else "")
        link = norm_link(row[T_LINK - 1] if len(row) >= T_LINK else "")
        status = norm_text(row[T_STATUS - 1] if len(row) >= T_STATUS else "")
        kind = norm_text(row[T_KIND - 1] if len(row) >= T_KIND else "")
        if not (email or name or chan or link):
            continue
        # 只要进过追踪表，就默认不再首封重复触达；跟进由自动跟进脚本处理。
        merge_row_to_sent_keys(name, email, chan, sent_keys)
        merge_row_to_sent_keys(name, email, link, sent_keys)
    if extra_path:
        load_extra_sent_keys(extra_path, sent_keys)
    return sent_keys


def already_sent_reason(name, email, link, sent_keys):
    email_key = norm_text(email)
    name_key = norm_text(name)
    link_key = norm_link(link)
    if email_key and email_key in sent_keys["emails"]:
        return "已在邮件追踪表发过"
    if link_key and link_key in sent_keys["links"]:
        return "主页/视频链接已在邮件追踪表"
    if name_key and name_key in sent_keys["names"]:
        return "名称已在邮件追踪表"
    return ""


def should_remove_from_queue_as_sent_duplicate(reason):
    reason = str(reason or "").strip()
    return reason in {
        "已在邮件追踪表发过",
        "主页/视频链接已在邮件追踪表",
        "名称已在邮件追踪表",
    }


def save_book(wb, path, purpose):
    if save_workbook_safe:
        return save_workbook_safe(wb, path, purpose)
    try:
        wb.save(path)
        return True
    except PermissionError:
        print(f"⚠️ {os.path.basename(path)} 正被 Excel/WPS 打开，保存失败。请关闭表格后重跑。")
        return False


def cleanup_queue_rows(ws_q, rownums, queue_path, reason="已在追踪表/去重表"):
    rownums = sorted(set(int(r) for r in rownums if r), reverse=True)
    if not rownums:
        return 0
    for rownum in rownums:
        ws_q.delete_rows(rownum)
    if save_book(ws_q.parent, queue_path, "发信名单"):
        print(f"🧹 已从发信名单清理 {len(rownums)} 条{reason}的行")
    else:
        print(f"⚠️ 原计划清理 {len(rownums)} 条{reason}的行，但发信名单保存失败")
    return len(rownums)


def housekeeping_queue(ws_q, queue_path, sent_keys):
    """主发信表只保留可发送对象，旧的空邮箱/已追踪残留先清掉。"""
    stale_rows = []
    duplicate_rows = []
    for row in ws_q.iter_rows(min_row=2):
        rownum = row[0].row
        name = str(row[Q_NAME - 1].value or "").strip()
        email = str(row[Q_EMAIL - 1].value or "").strip()
        home_link = str(row[Q_HOME - 1].value or "").strip()
        video_link = str(row[Q_VIDEO - 1].value or "").strip() if ws_q.max_column >= Q_VIDEO else ""
        link = video_link or home_link
        if not any(str(cell.value or "").strip() for cell in row):
            stale_rows.append(rownum)
            continue
        if not email or not EMAIL_RE.match(email):
            stale_rows.append(rownum)
            continue
        if already_sent_reason(name, email, link, sent_keys):
            duplicate_rows.append(rownum)
    cleaned = 0
    if stale_rows:
        cleaned += cleanup_queue_rows(ws_q, stale_rows, queue_path, "不可发送/空邮箱残留")
    if duplicate_rows:
        cleaned += cleanup_queue_rows(ws_q, duplicate_rows, queue_path, "已在追踪表")
    return cleaned


def get_tracker_sheet(wb):
    """追踪工作簿里必须明确写主表，不能依赖当前激活页。"""
    if "邮件追踪" in wb.sheetnames:
        wb.active = wb.sheetnames.index("邮件追踪")
        return wb["邮件追踪"]
    return wb.active


def append_to_tracker(tracker_ws, name, email, link, email_type="", source_keyword="", home_link="", video_link=""):
    """写一行到追踪表。home_link/video_link 不传时退回用 link（兼容老调用方式）。"""
    last_row = tracker_ws.max_row
    last_num = tracker_ws.cell(last_row, T_NUM).value
    new_num  = (last_num + 1) if isinstance(last_num, int) else last_row
    today    = datetime.now().strftime("%Y-%m-%d")

    chan = link if link else ""
    home_link = home_link or link
    if str(email_type).strip() == "文章外链":
        contact_type = "外链网站"
        kind = "文章插链接开发"
    else:
        contact_type = "YouTube KOL"
        kind = "视频插链接开发"

    row = [
        new_num, today, name, email,
        contact_type, chan, video_link,
        kind, "未回复", None, home_link, "已发送"
    ]
    while len(row) < T_SOURCE - 1:
        row.append(None)
    row.append(source_keyword)
    row.append("")  # 频道标签，先留空
    tracker_ws.append(row)


def apply_send_config(config_path=None, profile=""):
    """给网页端用：不走 choose_config_profile() 的交互选择，直接套用指定/默认配置。"""
    if not apply_config:
        return {}
    return apply_config(globals(), {
        "SMTP_SERVER": "SMTP_SERVER",
        "SMTP_PORT": "SMTP_PORT",
        "SMTP_TIMEOUT": "SMTP_TIMEOUT",
        "SMTP_ALLOW_INSECURE_SSL": "SMTP_ALLOW_INSECURE_SSL",
        "FROM_EMAIL": "FROM_EMAIL",
        "FROM_NAME": "FROM_NAME",
        "PASSWORD": "PASSWORD",
        "PRODUCT_NAME": "PRODUCT_NAME",
        "PRODUCT_TEAM": "PRODUCT_TEAM",
        "PRODUCT_URL": "PRODUCT_URL",
        "OUTREACH_SUBJECT_YOUTUBE": "OUTREACH_SUBJECT_YOUTUBE",
        "OUTREACH_SUBJECT_ARTICLE": "OUTREACH_SUBJECT_ARTICLE",
        "OUTREACH_TEMPLATE_YOUTUBE": "OUTREACH_TEMPLATE_YOUTUBE",
        "OUTREACH_TEMPLATE_ARTICLE": "OUTREACH_TEMPLATE_ARTICLE",
        "DELAY_SEC": "DELAY_SEC",
        "DAILY_SEND_LIMIT": "DAILY_SEND_LIMIT",
        "REQUIRE_SEND_CODE": "REQUIRE_SEND_CODE",
        "PERSONALIZE_RECENT_VIDEOS": "PERSONALIZE_RECENT_VIDEOS",
    }, config_path=config_path, profile=profile)


def build_send_targets(should_personalize=False):
    """
    准备本次可发送目标：读表、判重、安全拦截、生成主题/开头。
    跟原来 main() 发信前的逻辑完全一致，只是把"要不要抓最近视频标题"从 input() 改成参数，
    这样桌面版和网页版可以共用同一套逻辑，不会出现两边判断不一致。
    """
    queue_path   = get_path("VikPea_发信名单.xlsx")
    tracker_path = get_path("VikPea_邮件开发追踪.xlsx")

    wb_q = openpyxl.load_workbook(queue_path)
    ws_q = wb_q.active
    if ensure_queue_headers:
        ensure_queue_headers(ws_q)
    # 主发信表现在只作为“可发送队列”。无邮箱候选由搜索/深度找邮箱脚本写入副表。

    wb_t = openpyxl.load_workbook(tracker_path)
    ws_t = get_tracker_sheet(wb_t)
    ensure_tracker_headers(ws_t)
    sent_keys = load_sent_keys(ws_t, EXTRA_DEDUPE_PATH)
    if housekeeping_queue(ws_q, queue_path, sent_keys):
        return {
            "retry": True,
            "message": "✅ 发信名单已先清理残留行。为避免行号错位，本次不继续发信，请重新运行一次发信步骤。",
        }
    moved_from_pool = sync_manual_emails_from_pool(ws_q, ws_t, sent_keys)
    if moved_from_pool:
        save_book(wb_q, queue_path, "发信名单")
        print(f"📥 已从无邮箱候选表自动转入 {moved_from_pool} 条补录邮箱到主发信表")
    moved_from_pending = sync_pending_email_review(ws_q, ws_t, sent_keys)
    if moved_from_pending:
        save_book(wb_q, queue_path, "发信名单")
        print(f"📥 已从待确认邮箱表自动转入 {moved_from_pending} 条确认邮箱到主发信表")
    personalize_cache = load_personalize_cache()
    if PERSONALIZE_RECENT_VIDEOS and not YTDLP_CMD:
        print("⚠️ 未找到 yt-dlp，本次会使用普通开头；如需抓最近视频标题，请先安装 yt-dlp")
    blacklist = load_blacklist() if load_blacklist else {"emails": set(), "domains": set(), "names": set()}

    # 收集待发行（有邮箱的行）
    targets = []
    blocked_targets = []
    duplicate_sent_rows = []
    for row in ws_q.iter_rows(min_row=2):
        name  = str(row[Q_NAME  - 1].value or "").strip()
        email    = str(row[Q_EMAIL - 1].value or "").strip()
        subj     = str(row[Q_SUBJ  - 1].value or "").strip()
        op       = str(row[Q_OPEN  - 1].value or "").strip()
        home_link = str(row[Q_HOME - 1].value or "").strip()
        video_link = str(row[Q_VIDEO - 1].value or "").strip() if ws_q.max_column >= Q_VIDEO else ""
        link     = video_link or home_link
        etype    = normalize_email_type(row[Q_TYPE - 1].value if ws_q.max_column >= Q_TYPE else "", link)
        source_kw = str(row[Q_SOURCE - 1].value or "").strip() if ws_q.max_column >= Q_SOURCE else ""
        if not email or "@" not in email:
            continue
        if etype and not str(row[Q_TYPE - 1].value or "").strip():
            ws_q.cell(row[0].row, Q_TYPE).value = etype
        if not source_kw:
            source_kw = "发信名单手动补录"
            ws_q.cell(row[0].row, Q_SOURCE).value = source_kw
        if not str(row[Q_NOTE - 1].value or "").strip():
            ws_q.cell(row[0].row, Q_NOTE).value = "手动补录：脚本自动补齐邮件内容"
        reason = safety_reason(name, email, link, blacklist) if safety_reason else ""
        sent_reason = already_sent_reason(name, email, link, sent_keys)
        if sent_reason:
            reason = sent_reason
        if reason:
            blocked_targets.append({
                "rownum": row[0].row, "name": name, "email": email, "subject": subj,
                "opening": op, "link": link, "type": etype, "source": source_kw,
                "status": "已拦截", "reason": reason,
            })
            note = str(row[Q_NOTE - 1].value or "").strip()
            if reason not in note:
                ws_q.cell(row[0].row, Q_NOTE).value = (note + " | " if note else "") + f"发信前拦截: {reason}"
            if should_remove_from_queue_as_sent_duplicate(reason):
                duplicate_sent_rows.append(row[0].row)
            continue
        if USE_LIGHT_OUTREACH:
            subj = make_fallback_subject(name, etype, source_kw)
            ws_q.cell(row[0].row, Q_SUBJ).value = subj
        if not subj:
            subj = make_fallback_subject(name, etype, source_kw)
            ws_q.cell(row[0].row, Q_SUBJ).value = subj
            print(f"  🧩 {name} ({email}) 自动补主题")

        titles = []
        if should_personalize and etype != "文章外链":
            print(f"  ⏳ 正在抓最近视频标题: {name} ({email})", flush=True)
            titles = get_recent_video_titles(link, personalize_cache)
            subj = pick_subject_template(name, etype, source_kw, titles)
            ws_q.cell(row[0].row, Q_SUBJ).value = subj
            personalized = make_personalized_opening(name, link, etype, source_kw, titles)
            if titles:
                op = personalized
                ws_q.cell(row[0].row, Q_OPEN).value = op
                title_preview = " / ".join(titles[:2])
                print(f"  ✨ {name} ({email}) 已按最近视频定制开头: {title_preview}")
            else:
                print(f"  · {name} 未抓到标题，使用普通开头")

        if not op:
            op = make_fallback_opening(name, link, etype, source_kw)
            ws_q.cell(row[0].row, Q_OPEN).value = op
            print(f"  🧩 {name} ({email}) 自动补开头")
        targets.append((row[0].row, name, email, subj, op, link, etype, source_kw, titles, home_link, video_link))

    if duplicate_sent_rows:
        cleanup_queue_rows(ws_q, duplicate_sent_rows, queue_path, "已在追踪表")
        return {
            "retry": True,
            "message": "✅ 已先清理旧重复行。为避免行号错位，本次不继续发信，请重新运行一次发信步骤。",
        }
    else:
        save_book(wb_q, queue_path, "发信名单")
    preview_rows = [
        {
            "rownum": rownum, "name": name, "email": email, "subject": subj,
            "opening": op, "link": link, "type": etype, "source": source_kw,
            "status": "可发送",
        }
        for rownum, name, email, subj, op, link, etype, source_kw, titles, home_link, video_link in targets
    ] + blocked_targets
    if export_send_preview and preview_rows:
        preview_path = export_send_preview(preview_rows)
        print(f"📋 已生成本次发信预览: {preview_path}")
    if blocked_targets:
        print(f"🛑 发信前安全拦截 {len(blocked_targets)} 条，已写入预览表和发信名单备注")

    if DAILY_SEND_LIMIT and len(targets) > int(DAILY_SEND_LIMIT):
        print(f"⚠️ 当前可发送 {len(targets)} 封，超过单次上限 {DAILY_SEND_LIMIT}，本次只发送前 {DAILY_SEND_LIMIT} 封。")
        targets = targets[:int(DAILY_SEND_LIMIT)]

    return {
        "retry": False,
        "targets": targets,
        "blocked_targets": blocked_targets,
        "_wb_q": wb_q, "_ws_q": ws_q, "_wb_t": wb_t, "_ws_t": ws_t,
        "_sent_keys": sent_keys, "_blacklist": blacklist,
        "_queue_path": queue_path, "_tracker_path": tracker_path,
    }


def send_targets(session, selected_rownums=None):
    """
    真正建立 SMTP 连接并发送 session["targets"] 里的邮件，跟原来 main() 发信循环完全一致。
    selected_rownums 不传 = 发 session 里全部目标；传一个行号集合可以只发其中一部分
    （网页端用来支持"先勾选再发"）。
    返回 (sent_rows, failed)。
    """
    targets = session["targets"]
    if selected_rownums is not None:
        targets = [t for t in targets if t[0] in selected_rownums]

    ws_q = session["_ws_q"]
    wb_q = session["_wb_q"]
    ws_t = session["_ws_t"]
    wb_t = session["_wb_t"]
    sent_keys = session["_sent_keys"]
    blacklist = session["_blacklist"]
    queue_path = session["_queue_path"]
    tracker_path = session["_tracker_path"]

    sent_rows = []
    failed = []
    server, smtp_mode = smtp_connect()
    with server:
        if smtp_mode == "insecure":
            print("⚠️ 当前 SMTP 使用兼容 SSL 模式。")
        server.login(FROM_EMAIL, PASSWORD)
        print(f"\n✅ SMTP 登录成功，开始发送...\n")
        if log_event:
            log_event("发信开始", f"计划发送 {len(targets)} 封")

        for rownum, name, email, subj, opening, link, etype, source_kw, titles, home_link, video_link in targets:
            reason = safety_reason(name, email, link, blacklist) if safety_reason else ""
            sent_reason = already_sent_reason(name, email, link, sent_keys)
            if sent_reason:
                reason = sent_reason
            if reason:
                ws_q.cell(row=rownum, column=Q_NOTE).value = f"发送前二次拦截: {reason}"
                save_book(wb_q, queue_path, "发信名单")
                print(f"  ⏭️ {name} → {email} 已跳过: {reason}")
                continue
            try:
                msg, _ = make_email(name, email, subj, opening, etype)
                server.sendmail(FROM_EMAIL, email, msg.as_string())
                sent_rows.append(rownum)
                append_to_tracker(ws_t, name, email, link, etype, source_kw, home_link=home_link, video_link=video_link)
                sent_keys["emails"].add(norm_text(email))
                sent_keys["names"].add(norm_text(name))
                if link:
                    sent_keys["links"].add(norm_link(link))
                save_book(wb_t, tracker_path, "邮件追踪表")
                print(f"  ✅ [{len(sent_rows)}/{len(targets)}] {name} → {email}")
                if log_event:
                    log_event("发信成功", f"{name} <{email}>")
                time.sleep(DELAY_SEC)
            except Exception as e:
                failed.append((name, email, str(e)))
                ws_q.cell(row=rownum, column=Q_NOTE).value = f"发送失败: {str(e)}"
                save_book(wb_q, queue_path, "发信名单")
                print(f"  ❌ {name} → {email}  错误: {e}")
                if log_event:
                    log_event("发信失败", f"{name} <{email}> {e}")

    cleanup_queue_rows(ws_q, sent_rows, queue_path, "已成功发送")
    return sent_rows, failed


def main():
    config_path, config_profile, config_label = choose_config_profile()
    apply_send_config(config_path=config_path, profile=config_profile)
    print(f"📄 本次发信配置: {config_label}" + (f" ({config_path})" if config_path else ""))

    should_personalize = False
    if PERSONALIZE_RECENT_VIDEOS and YTDLP_CMD:
        print("\n提示：抓最近视频标题会更定制，但一批几十封时会比较慢。")
        ans = input("是否现在抓最近视频标题生成定制开头？输入 p 开启，直接回车跳过: ").strip().lower()
        should_personalize = ans == "p"
        if not should_personalize:
            print("已跳过抓标题，本次使用已有开头/普通开头，速度更快。")

    session = build_send_targets(should_personalize=should_personalize)
    if session.get("retry"):
        print(session["message"])
        return

    targets = session["targets"]

    if not targets:
        print("✅ 发信名单里没有待发邮件（全部为空邮箱行）")
        print("   → 请在 VikPea_发信名单.xlsx 的黄色行填入邮箱后再运行")
        return

    print(f"\n待发 {len(targets)} 封\n{'─'*60}")
    for _, name, em, subj, _, _, etype, source_kw, titles, _, _ in targets:
        tag = "[文章]" if etype == "文章外链" else "[YT]  "
        src = f" | {source_kw}" if source_kw else ""
        recent = f" | 最近: {trim_title(titles[0], 42)}" if titles else ""
        print(f"  {tag} {name:<26} → {em}{src}{recent}")
    print(f"{'─'*60}")

    confirm = input(f"\n确认发送以上 {len(targets)} 封？先输入 y 进入内容预览: ").strip().lower()
    if confirm != "y":
        print("已取消")
        return

    # 预览
    r0, n0, e0, s0, o0, _, et0, _, _, _, _ = targets[0]
    _, preview = make_email(n0, e0, s0, o0, et0)
    print(f"\n【预览】Subject: {s0}\nTo: {e0}\n")
    print(preview[:500] + "...\n")
    go = input("确认内容无误，输入 y 或 send 才会真正发送: ").strip()
    if not is_send_confirm(go):
        print("已取消")
        return

    try:
        sent_rows, failed = send_targets(session)
    except smtplib.SMTPAuthenticationError as e:
        print(f"❌ SMTP 登录失败：账号或授权码不对。{e}")
        print("   ↳ 请检查配置表里的 FROM_EMAIL / PASSWORD。")
        return
    except (socket.gaierror, TimeoutError) as e:
        print(f"❌ SMTP 连接失败：服务器地址无法访问或超时。{e}")
        print("   ↳ 请检查网络、VPN、防火墙，或确认 SMTP_SERVER / SMTP_PORT 是否正确。")
        return
    except ssl.SSLError as e:
        print(f"❌ SMTP SSL 建连失败：{e}")
        print("   ↳ 大概率是本机 Python 证书链或企业邮箱 SSL 兼容问题。")
        return
    except Exception as e:
        print(f"❌ SMTP 建连失败：{e}")
        return

    print(f"\n{'─'*60}")
    print(f"完成: ✅ {len(sent_rows)} 成功  ❌ {len(failed)} 失败")
    if failed:
        for name, em, err in failed:
            print(f"  {name} ({em}): {err}")
    print(f"\n✅ 已发送记录 → 写入 VikPea_邮件开发追踪.xlsx")
    print(f"✅ 已发送行   → 从 VikPea_发信名单.xlsx 中删除")


if __name__ == "__main__":
    main()
