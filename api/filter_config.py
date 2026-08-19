"""
KOL 过滤配置管理
支持视频负关键词、竞品站点、竞品邮箱后缀、Affiliate名单等过滤规则
"""

import os
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple
import openpyxl

WORKSPACE_DIR = os.environ.get(
    "VIKPEA_WORKSPACE_DIR",
    "/Users/xuzifu/Downloads/VikPea工作台/VikPea工作台_Mac试用包_2026-08-12/01_请在这个文件夹里操作",
)

# 过滤配置文件路径
NEGATIVE_KEYWORDS_PATH = os.path.join(WORKSPACE_DIR, "VikPea_视频负关键词.xlsx")
COMPETITOR_SITES_PATH = os.path.join(WORKSPACE_DIR, "VikPea_竞品站点黑名单.xlsx")
COMPETITOR_EMAILS_PATH = os.path.join(WORKSPACE_DIR, "VikPea_竞品邮箱后缀.xlsx")
AFFILIATE_BLACKLIST_PATH = os.path.join(WORKSPACE_DIR, "VikPea_Affiliate黑名单.xlsx")
LONGTERM_PARTNERS_PATH = os.path.join(WORKSPACE_DIR, "VikPea_长期合作名单.xlsx")
CONTACTED_HISTORY_PATH = os.path.join(WORKSPACE_DIR, "VikPea_已联络历史.xlsx")


def _load_excel_column(path: str, column_name: str, sheet: Optional[str] = None) -> List[str]:
    """从Excel文件加载指定列的所有值"""
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    if ws.max_row < 1:
        return []

    # 找到列索引
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    if column_name not in headers:
        return []

    col_idx = headers.index(column_name)
    values = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_idx < len(row) and row[col_idx]:
            val = str(row[col_idx]).strip()
            if val:
                values.append(val)

    return values


def _load_excel_rows(path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    """加载Excel所有行为字典列表"""
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    if ws.max_row < 1:
        return []

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)

    return rows


class FilterConfig:
    """过滤配置加载器"""

    def __init__(self):
        self.reload()

    def reload(self):
        """重新加载所有过滤配置"""
        self._load_negative_keywords()
        self._load_competitor_sites()
        self._load_competitor_email_suffixes()
        self._load_blacklists()
        self._load_contacted_history()

    def _load_negative_keywords(self):
        """加载视频负关键词配置
        Excel格式：关键词 | 合作时间阈值(天)
        """
        self.negative_keywords = {}
        rows = _load_excel_rows(NEGATIVE_KEYWORDS_PATH)

        for row in rows:
            keyword = str(row.get("关键词", "")).strip().lower()
            threshold_days = row.get("合作时间阈值(天)", 90)

            if keyword:
                try:
                    self.negative_keywords[keyword] = int(threshold_days)
                except (ValueError, TypeError):
                    self.negative_keywords[keyword] = 90  # 默认90天

    def _load_competitor_sites(self):
        """加载竞品站点黑名单
        Excel格式：站点域名
        """
        self.competitor_sites = set()
        sites = _load_excel_column(COMPETITOR_SITES_PATH, "站点域名")
        self.competitor_sites = {s.lower() for s in sites if s}

    def _load_competitor_email_suffixes(self):
        """加载竞品邮箱后缀
        Excel格式：邮箱后缀
        """
        self.competitor_email_suffixes = set()
        suffixes = _load_excel_column(COMPETITOR_EMAILS_PATH, "邮箱后缀")
        self.competitor_email_suffixes = {s.lower() for s in suffixes if s}

    def _load_blacklists(self):
        """加载Affiliate黑名单和长期合作名单
        Excel格式：频道名 | 邮箱 | 备注
        """
        self.affiliate_blacklist = set()
        self.longterm_partners = set()

        # 加载Affiliate黑名单
        rows = _load_excel_rows(AFFILIATE_BLACKLIST_PATH)
        for row in rows:
            channel = str(row.get("频道名", "")).strip()
            email = str(row.get("邮箱", "")).strip().lower()
            if channel:
                self.affiliate_blacklist.add(channel)
            if email:
                self.affiliate_blacklist.add(email)

        # 加载长期合作名单
        rows = _load_excel_rows(LONGTERM_PARTNERS_PATH)
        for row in rows:
            channel = str(row.get("频道名", "")).strip()
            email = str(row.get("邮箱", "")).strip().lower()
            if channel:
                self.longterm_partners.add(channel)
            if email:
                self.longterm_partners.add(email)

    def _load_contacted_history(self):
        """加载已联络历史名单
        Excel格式：频道名 | 邮箱 | 联络日期 | 备注
        """
        self.contacted_history = {}  # key: (频道名, 邮箱), value: {date: 联络日期, note: 备注}

        rows = _load_excel_rows(CONTACTED_HISTORY_PATH)
        for row in rows:
            channel = str(row.get("频道名", "")).strip()
            email = str(row.get("邮箱", "")).strip().lower()
            contact_date = row.get("联络日期", "")
            note = str(row.get("备注", "")).strip()

            if not channel and not email:
                continue

            # 尝试解析日期
            try:
                if isinstance(contact_date, datetime):
                    date_obj = contact_date
                elif contact_date:
                    date_obj = datetime.strptime(str(contact_date).split()[0], "%Y-%m-%d")
                else:
                    date_obj = None
            except (ValueError, TypeError):
                date_obj = None

            # 存储到字典（支持按频道名或邮箱查询）
            if channel:
                key = ("channel", channel)
                if key not in self.contacted_history or (date_obj and self.contacted_history[key].get("date") and date_obj > self.contacted_history[key]["date"]):
                    self.contacted_history[key] = {"date": date_obj, "note": note}
            if email:
                key = ("email", email)
                if key not in self.contacted_history or (date_obj and self.contacted_history[key].get("date") and date_obj > self.contacted_history[key]["date"]):
                    self.contacted_history[key] = {"date": date_obj, "note": note}

    def check_negative_keywords(
        self,
        title: str,
        description: str,
        last_collab_date: Optional[str] = None
    ) -> Tuple[bool, Optional[str], Optional[int]]:
        """检查视频标题和简介中的负关键词

        Args:
            title: 视频标题
            description: 视频简介
            last_collab_date: 最后合作日期 (YYYY-MM-DD格式)

        Returns:
            (should_exclude, matched_keyword, days_since_collab)
            - should_exclude: 是否应该排除
            - matched_keyword: 匹配到的关键词
            - days_since_collab: 距离上次合作的天数
        """
        text = (title + " " + description).lower()

        for keyword, threshold_days in self.negative_keywords.items():
            if keyword in text:
                # 匹配到负关键词
                if not last_collab_date:
                    # 没有合作日期，直接排除
                    return (True, keyword, None)

                try:
                    collab_date = datetime.strptime(last_collab_date, "%Y-%m-%d")
                    days_diff = (datetime.now() - collab_date).days

                    if days_diff < threshold_days:
                        # 合作时间小于阈值，排除
                        return (True, keyword, days_diff)
                    else:
                        # 合作时间大于阈值，不排除但需标注
                        return (False, keyword, days_diff)
                except (ValueError, TypeError):
                    # 日期格式错误，直接排除
                    return (True, keyword, None)

        return (False, None, None)

    def check_contacted_history(
        self,
        channel_name: str,
        email: str,
        threshold_days: int = 90
    ) -> Tuple[bool, Optional[int], Optional[str]]:
        """检查已联络历史

        Args:
            channel_name: 频道名
            email: 邮箱地址
            threshold_days: 时间阈值（天）

        Returns:
            (should_exclude, days_since_contact, note)
            - should_exclude: 是否应该排除（小于阈值天数）
            - days_since_contact: 距离上次联络的天数
            - note: 备注信息
        """
        email_lower = email.lower() if email else ""

        # 检查频道名
        channel_info = self.contacted_history.get(("channel", channel_name))
        email_info = self.contacted_history.get(("email", email_lower))

        # 取最近的联络记录
        latest_info = None
        if channel_info and email_info:
            if channel_info.get("date") and email_info.get("date"):
                latest_info = channel_info if channel_info["date"] > email_info["date"] else email_info
            else:
                latest_info = channel_info if channel_info.get("date") else email_info
        elif channel_info:
            latest_info = channel_info
        elif email_info:
            latest_info = email_info

        if not latest_info or not latest_info.get("date"):
            return (False, None, None)

        days_diff = (datetime.now() - latest_info["date"]).days
        note = latest_info.get("note", "")

        if days_diff < threshold_days:
            # 小于阈值，排除
            return (True, days_diff, note)
        else:
            # 大于阈值，不排除但需标注
            return (False, days_diff, note)

    def check_competitor_site(self, video_url: str, channel_url: str) -> Tuple[bool, Optional[str]]:
        """检查是否推广竞品站点

        Args:
            video_url: 视频链接
            channel_url: 频道主页链接

        Returns:
            (is_competitor, matched_site)
        """
        text = (video_url + " " + channel_url).lower()

        for site in self.competitor_sites:
            if site in text:
                return (True, site)

        return (False, None)

    def check_competitor_email(self, email: str) -> Tuple[bool, Optional[str]]:
        """检查是否为竞品邮箱

        Args:
            email: 邮箱地址

        Returns:
            (is_competitor, matched_suffix)
        """
        if not email:
            return (False, None)

        email_lower = email.lower()

        for suffix in self.competitor_email_suffixes:
            if email_lower.endswith(suffix):
                return (True, suffix)

        return (False, None)

    def check_blacklist(self, channel_name: str, email: str) -> Tuple[bool, str]:
        """检查是否在黑名单中

        Args:
            channel_name: 频道名
            email: 邮箱地址

        Returns:
            (is_blacklisted, reason)
            reason: "affiliate" 或 "longterm"
        """
        email_lower = email.lower() if email else ""

        if channel_name in self.affiliate_blacklist or email_lower in self.affiliate_blacklist:
            return (True, "affiliate")

        if channel_name in self.longterm_partners or email_lower in self.longterm_partners:
            return (True, "longterm")

        return (False, "")


# 全局配置实例
_filter_config: Optional[FilterConfig] = None


def get_filter_config() -> FilterConfig:
    """获取全局过滤配置实例"""
    global _filter_config
    if _filter_config is None:
        _filter_config = FilterConfig()
    return _filter_config


def reload_filter_config():
    """重新加载过滤配置"""
    global _filter_config
    _filter_config = FilterConfig()


def get_config_stats() -> Dict[str, int]:
    """获取配置统计信息"""
    config = get_filter_config()
    return {
        "negative_keywords_count": len(config.negative_keywords),
        "competitor_sites_count": len(config.competitor_sites),
        "competitor_email_suffixes_count": len(config.competitor_email_suffixes),
        "affiliate_blacklist_count": len(config.affiliate_blacklist),
        "longterm_partners_count": len(config.longterm_partners),
        "contacted_history_count": len(config.contacted_history),
    }


def upload_contacted_history(records: List[Dict[str, str]]) -> Dict[str, Any]:
    """上传已联络历史记录

    Args:
        records: 记录列表，每条记录包含: 频道名, 邮箱, 联络日期, 备注

    Returns:
        处理结果统计
    """
    if not records:
        raise ValueError("没有提供任何记录")

    # 读取现有数据
    existing_rows = []
    if os.path.exists(CONTACTED_HISTORY_PATH):
        wb = openpyxl.load_workbook(CONTACTED_HISTORY_PATH)
        ws = wb.active
        headers = [str(cell.value or "").strip() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            existing_rows.append(row_dict)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["频道名", "邮箱", "联络日期", "备注"])
        headers = ["频道名", "邮箱", "联络日期", "备注"]

    # 建立去重索引 (频道名或邮箱)
    existing_keys = set()
    for row in existing_rows:
        channel = str(row.get("频道名", "")).strip()
        email = str(row.get("邮箱", "")).strip().lower()
        if channel:
            existing_keys.add(("channel", channel))
        if email:
            existing_keys.add(("email", email))

    # 处理新记录
    added = 0
    skipped = 0

    for record in records:
        channel = str(record.get("频道名", "")).strip()
        email = str(record.get("邮箱", "")).strip()
        contact_date = str(record.get("联络日期", "")).strip()
        note = str(record.get("备注", "")).strip()

        if not channel and not email:
            skipped += 1
            continue

        # 检查是否已存在
        is_duplicate = False
        if channel and ("channel", channel) in existing_keys:
            is_duplicate = True
        if email and ("email", email.lower()) in existing_keys:
            is_duplicate = True

        if is_duplicate:
            skipped += 1
            continue

        # 添加新记录
        ws.append([channel, email, contact_date, note])
        if channel:
            existing_keys.add(("channel", channel))
        if email:
            existing_keys.add(("email", email.lower()))
        added += 1

    # 保存文件
    wb.save(CONTACTED_HISTORY_PATH)

    # 重新加载配置
    reload_filter_config()

    return {
        "added": added,
        "skipped": skipped,
        "total": len(records),
        "message": f"成功添加 {added} 条记录，跳过 {skipped} 条重复记录"
    }


def update_negative_keyword_threshold(keyword: str, threshold_days: int) -> Dict[str, Any]:
    """更新负关键词的时间阈值

    Args:
        keyword: 关键词
        threshold_days: 阈值天数

    Returns:
        更新结果
    """
    if not os.path.exists(NEGATIVE_KEYWORDS_PATH):
        raise FileNotFoundError("负关键词配置文件不存在")

    wb = openpyxl.load_workbook(NEGATIVE_KEYWORDS_PATH)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    # 确保有"合作时间阈值(天)"列
    if "合作时间阈值(天)" not in headers:
        ws.cell(1, len(headers) + 1).value = "合作时间阈值(天)"
        threshold_col = len(headers) + 1
    else:
        threshold_col = headers.index("合作时间阈值(天)") + 1

    # 查找关键词所在行
    keyword_col = headers.index("关键词") + 1
    found = False

    for row in ws.iter_rows(min_row=2):
        if str(row[keyword_col - 1].value or "").strip().lower() == keyword.lower():
            ws.cell(row[0].row, threshold_col).value = threshold_days
            found = True
            break

    if not found:
        raise ValueError(f"找不到关键词: {keyword}")

    wb.save(NEGATIVE_KEYWORDS_PATH)

    # 重新加载配置
    reload_filter_config()

    return {
        "status": "success",
        "keyword": keyword,
        "threshold_days": threshold_days,
        "message": f"已更新关键词 '{keyword}' 的时间阈值为 {threshold_days} 天"
    }

