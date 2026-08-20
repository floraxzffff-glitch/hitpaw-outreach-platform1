"""
Contact Threshold Configuration
Reads follow-up timing configuration from VikPea_配置.xlsx
Bridges between the VikPea_common.py config system and the web API
"""
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VIKPEA_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "01_请在这个文件夹里操作")
sys.path.insert(0, VIKPEA_DIR)

try:
    import openpyxl
except ImportError:
    openpyxl = None

CONFIG_PATH = os.path.join(VIKPEA_DIR, "VikPea_配置.xlsx")

def get_contact_threshold():
    """
    Get contact threshold from VikPea_配置.xlsx
    Returns a dict with followup timing settings
    """
    if not openpyxl or not os.path.exists(CONFIG_PATH):
        return {
            "followup1_after_days": 5,
            "followup2_after_days": 7,
        }

    try:
        wb = openpyxl.load_workbook(CONFIG_PATH, data_only=True)
        ws = wb.active

        config = {
            "followup1_after_days": 5,
            "followup2_after_days": 7,
        }

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            key = str(row[0] or "").strip()
            value = row[1]

            if key == "FOLLOWUP1_AFTER_DAYS":
                try:
                    config["followup1_after_days"] = int(value)
                except (ValueError, TypeError):
                    pass
            elif key == "FOLLOWUP2_AFTER_DAYS":
                try:
                    config["followup2_after_days"] = int(value)
                except (ValueError, TypeError):
                    pass

        return config
    except Exception:
        return {
            "followup1_after_days": 5,
            "followup2_after_days": 7,
        }

def set_contact_threshold(threshold_days):
    """
    Update follow-up timing in VikPea_配置.xlsx
    threshold_days should be a dict with followup1_after_days and/or followup2_after_days
    """
    if not openpyxl or not os.path.exists(CONFIG_PATH):
        return {"success": False, "error": "Config file not found"}

    try:
        wb = openpyxl.load_workbook(CONFIG_PATH)
        ws = wb.active

        updated = {}

        if isinstance(threshold_days, dict):
            followup1 = threshold_days.get("followup1_after_days")
            followup2 = threshold_days.get("followup2_after_days")
        else:
            followup1 = threshold_days
            followup2 = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            key = str(row[0].value or "").strip()

            if key == "FOLLOWUP1_AFTER_DAYS" and followup1 is not None:
                ws.cell(row_idx, 2).value = int(followup1)
                updated["followup1_after_days"] = int(followup1)
            elif key == "FOLLOWUP2_AFTER_DAYS" and followup2 is not None:
                ws.cell(row_idx, 2).value = int(followup2)
                updated["followup2_after_days"] = int(followup2)

        wb.save(CONFIG_PATH)
        return {"success": True, "updated": updated}
    except Exception as e:
        return {"success": False, "error": str(e)}
