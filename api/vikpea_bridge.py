"""
把网页 API 接到 VikPea 原始工作台的真实脚本/数据上。

原则：
- 不修改原工作台脚本（同事仍在用桌面版），只 import 其中的纯函数/常量。
- 所有读操作直接读原工作台目录下的真实 xlsx 文件，不复制、不建镜像数据。
- 重活（深度找邮箱批量跑、SEO 实时扫描）不在这里做同步调用，留给后续的后台任务方案。
"""

import os
import sys
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

WORKSPACE_DIR = os.environ.get(
    "VIKPEA_WORKSPACE_DIR",
    "/Users/xuzifu/Downloads/VikPea工作台/VikPea工作台_Mac试用包_2026-08-12/01_请在这个文件夹里操作",
)

if not os.path.isdir(WORKSPACE_DIR):
    raise RuntimeError(f"找不到 VikPea 工作台目录: {WORKSPACE_DIR}")

if WORKSPACE_DIR not in sys.path:
    sys.path.insert(0, WORKSPACE_DIR)

import openpyxl  # noqa: E402

import VikPea_common as vikpea_common  # noqa: E402
from VikPea_关键词聚类 import (  # noqa: E402
    load_keywords as _load_article_keywords,
    cluster_keywords as _cluster_keywords,
    ARTICLE_KEYWORD_PATH,
)
import VikPea_关键词复盘 as keyword_review  # noqa: E402
import VikPea_读表发信 as outreach_sender  # noqa: E402

import job_runner  # noqa: E402

SEO_SCAN_OUTPUT = os.path.join(WORKSPACE_DIR, "VikPea_SEO渠道机会扫描.xlsx")


def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _workbook_rows(path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    """按第一行表头把 xlsx 读成 list[dict]，文件不存在时返回空列表。"""
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    if ws.max_row < 1:
        return []
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for i, values in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row = dict(zip(headers, values))
        if any(v not in (None, "") for v in row.values()):
            row["_rownum"] = 2 + i  # 对应 Excel 里的实际行号，编辑/删除时要用
            rows.append(row)
    return rows


def _parse_date_to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return datetime.now().isoformat()
    try:
        return datetime.strptime(text, "%Y-%m-%d").isoformat()
    except ValueError:
        return datetime.now().isoformat()


# ======================== 关键词分析（历史搜索/发信复盘数据） ========================

def get_keyword_review_data() -> Dict[str, dict]:
    data = keyword_review.load_search_metrics()
    data = keyword_review.load_tracker_metrics(data)
    return data


def lookup_keyword(keyword: str) -> Dict[str, Any]:
    keyword = keyword.strip()
    data = get_keyword_review_data()
    item = data.get(keyword)
    if not item:
        return {
            "found": False,
            "total_found": 0,
            "eligible_count": 0,
            "email_found_count": 0,
            "email_rate": 0.0,
            "details": {"说明": "该关键词在历史搜索/发信记录里还没有数据（不是实时抓取，只复盘已跑过的关键词）"},
        }
    added = item.get("new_email", 0) + item.get("new_no_email", 0)
    return {
        "found": True,
        "total_found": item.get("found_channels", 0),
        "eligible_count": item.get("eligible_channels", 0),
        "email_found_count": item.get("new_email", 0),
        "email_rate": keyword_review.pct(item.get("new_email", 0), added),
        "details": {
            "搜索次数": item.get("search_runs", 0),
            "已发送": item.get("sent", 0),
            "已回复": item.get("replied", 0),
            "A级": item.get("grade_a", 0),
            "B级": item.get("grade_b", 0),
            "C级": item.get("grade_c", 0),
            "建议动作": keyword_review.suggest(item),
            "最近搜索日期": item.get("last_search_date", ""),
        },
    }


def list_keyword_reviews() -> List[Dict[str, Any]]:
    data = get_keyword_review_data()
    rows = []
    for keyword, item in data.items():
        added = item.get("new_email", 0) + item.get("new_no_email", 0)
        rows.append({
            "keyword": keyword,
            "search_runs": item.get("search_runs", 0),
            "found_channels": item.get("found_channels", 0),
            "eligible_channels": item.get("eligible_channels", 0),
            "new_email": item.get("new_email", 0),
            "email_rate": keyword_review.pct(item.get("new_email", 0), added),
            "sent": item.get("sent", 0),
            "replied": item.get("replied", 0),
            "reply_rate": keyword_review.pct(item.get("replied", 0), item.get("sent", 0)),
            "grade_a": item.get("grade_a", 0),
            "suggestion": keyword_review.suggest(item),
            "last_search_date": item.get("last_search_date", ""),
        })
    rows.sort(key=lambda r: (-r["grade_a"], -r["sent"], -r["new_email"], r["keyword"]))
    return rows


def cluster_article_keywords() -> List[Dict[str, Any]]:
    rows = _load_article_keywords(ARTICLE_KEYWORD_PATH)
    clusters = []
    for theme, representative, count, top_tokens, items in _cluster_keywords(rows):
        clusters.append({
            "theme": theme,
            "representative": representative,
            "count": count,
            "top_tokens": top_tokens,
            "keywords": [item[0] for item in items],
        })
    return clusters


# ======================== 邮箱验证 ========================

def validate_email_address(email: str, check_blacklist: bool = True) -> Dict[str, Any]:
    format_reason = vikpea_common.classify_bad_email(email)
    is_blacklisted = False
    blacklist_reason_text = ""
    if check_blacklist and not format_reason:
        blacklist_reason_text = vikpea_common.blacklist_reason("", email, "")
        is_blacklisted = bool(blacklist_reason_text)

    reason = format_reason or blacklist_reason_text or None
    is_valid = not format_reason
    confidence_score = 1.0 if (is_valid and not is_blacklisted) else (0.5 if is_blacklisted else 0.0)

    return {
        "email": email,
        "is_valid": is_valid,
        "is_blacklisted": is_blacklisted,
        "confidence_score": confidence_score,
        "reason": reason,
    }


# ======================== SEO 机会（读取上一次扫描结果） ========================

def get_seo_opportunities(keyword: Optional[str] = None, min_score: float = 0) -> List[Dict[str, Any]]:
    rows = _workbook_rows(SEO_SCAN_OUTPUT)
    results = []
    for row in rows:
        url = str(row.get("URL") or "").strip()
        title = str(row.get("标题") or "").strip()
        if not url and not title:
            continue  # 跳过占位/说明行（表头版本和脚本实际输出不一致时会出现）
        raw_score = row.get("站点分") or 0
        try:
            score = float(raw_score)
        except (TypeError, ValueError):
            score = 0.0
        if score < min_score:
            continue
        if keyword and keyword.strip().lower() not in str(row.get("关键词") or "").lower():
            continue
        results.append({
            "url": row.get("URL") or "",
            "title": row.get("标题") or "",
            "relevance_score": score,
            "level": row.get("机会等级") or "C",
            "opportunity_type": row.get("机会类型") or "",
            "action": row.get("建议动作") or "",
            "timestamp": _parse_date_to_iso(row.get("扫描日期")),
            "keyword": row.get("关键词") or "",
        })
    return results


def has_seo_scan_data() -> bool:
    return os.path.exists(SEO_SCAN_OUTPUT)


# ======================== 报告 / 仪表盘统计 ========================

def get_dashboard_stats() -> Dict[str, Any]:
    today = _today()

    keyword_data = get_keyword_review_data()
    total_keywords_analyzed = len(keyword_data)
    keywords_today = sum(
        1 for item in keyword_data.values() if str(item.get("last_search_date") or "") == today
    )

    tracker_rows = _workbook_rows(vikpea_common.TRACKER_PATH, "邮件追踪")
    total_emails_validated = len(tracker_rows)
    emails_today = sum(1 for row in tracker_rows if str(row.get("日期") or "").strip() == today)

    seo_rows = get_seo_opportunities()
    total_opportunities_found = len(seo_rows)
    opportunities_today = sum(
        1 for row in seo_rows if str(row.get("timestamp") or "").startswith(today)
    )

    return {
        "total_keywords_analyzed": total_keywords_analyzed,
        "total_emails_validated": total_emails_validated,
        "total_opportunities_found": total_opportunities_found,
        "today_activities": {
            "keywords_analyzed": keywords_today,
            "emails_validated": emails_today,
            "opportunities_found": opportunities_today,
        },
    }


def get_keyword_review_report() -> Dict[str, Any]:
    rows = list_keyword_reviews()
    return {
        "keyword_count": len(rows),
        "top_keywords": rows[:20],
    }


def get_seo_analysis_report() -> Dict[str, Any]:
    rows = get_seo_opportunities()
    level_counts = {"A": 0, "B": 0, "C": 0}
    for row in rows:
        level_counts[row["level"]] = level_counts.get(row["level"], 0) + 1
    return {
        "total_opportunities": len(rows),
        "level_counts": level_counts,
        "has_scan_data": has_seo_scan_data(),
        "top_opportunities": sorted(rows, key=lambda r: -r["relevance_score"])[:20],
    }


# ======================== YouTube KOL 搜索（重活，走后台任务） ========================

YOUTUBE_KEYWORD_PATH = vikpea_common.YOUTUBE_KEYWORD_PATH


def _is_enabled_flag(value: Any) -> bool:
    text = str(value or "是").strip().lower()
    return text not in {"停用", "禁用", "否", "no", "n", "0", "false"}


def list_youtube_keywords() -> List[Dict[str, Any]]:
    if not os.path.exists(YOUTUBE_KEYWORD_PATH):
        return []
    wb = openpyxl.load_workbook(YOUTUBE_KEYWORD_PATH, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        keyword = str((row[0] if row else "") or "").strip()
        if not keyword or keyword.lower() in {"keyword", "keywords", "关键词"}:
            continue
        note = str((row[2] if len(row) > 2 else "") or "").strip()
        rows.append({
            "keyword": keyword,
            "enabled": _is_enabled_flag(row[1] if len(row) > 1 else "是"),
            "note": note,
        })
    return rows


def set_youtube_keyword(keyword: str, enabled: bool, note: Optional[str] = None) -> bool:
    """开/关一个已有关键词；如果关键词不存在就新增一行（默认启用）。"""
    keyword = keyword.strip()
    if not keyword:
        raise ValueError("关键词不能为空")
    if not os.path.exists(YOUTUBE_KEYWORD_PATH):
        raise FileNotFoundError(YOUTUBE_KEYWORD_PATH)

    wb = openpyxl.load_workbook(YOUTUBE_KEYWORD_PATH)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2):
        cell_kw = str(row[0].value or "").strip()
        if cell_kw == keyword:
            row[1].value = "是" if enabled else "否"
            if note is not None:
                if len(row) > 2:
                    row[2].value = note
            found = True
            break
    if not found:
        ws.append([keyword, "是" if enabled else "否", note or ""])
    return vikpea_common.save_workbook_safe(wb, YOUTUBE_KEYWORD_PATH, "YouTube搜索关键词")


def add_youtube_keywords_batch(keywords: List[str]) -> Dict[str, Any]:
    """一次加一批关键词（比如从 Excel 整列粘贴过来），全部默认启用。"""
    if not os.path.exists(YOUTUBE_KEYWORD_PATH):
        raise FileNotFoundError(YOUTUBE_KEYWORD_PATH)

    cleaned = []
    seen = set()
    for kw in keywords:
        kw = str(kw or "").strip()
        if kw and kw not in seen:
            seen.add(kw)
            cleaned.append(kw)
    if not cleaned:
        raise ValueError("没有有效的关键词")

    wb = openpyxl.load_workbook(YOUTUBE_KEYWORD_PATH)
    ws = wb.active
    existing = {}
    for row in ws.iter_rows(min_row=2):
        cell_kw = str(row[0].value or "").strip()
        if cell_kw:
            existing[cell_kw] = row

    added, updated = 0, 0
    for kw in cleaned:
        if kw in existing:
            existing[kw][1].value = "是"
            updated += 1
        else:
            ws.append([kw, "是", ""])
            added += 1

    vikpea_common.save_workbook_safe(wb, YOUTUBE_KEYWORD_PATH, "YouTube搜索关键词")
    return {"added": added, "updated": updated, "keywords": list_youtube_keywords()}


def delete_youtube_keyword(keyword: str) -> Dict[str, Any]:
    keyword = keyword.strip()
    if not os.path.exists(YOUTUBE_KEYWORD_PATH):
        raise FileNotFoundError(YOUTUBE_KEYWORD_PATH)
    wb = openpyxl.load_workbook(YOUTUBE_KEYWORD_PATH)
    ws = wb.active
    target_row = None
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip() == keyword:
            target_row = row[0].row
            break
    if target_row is None:
        raise ValueError("找不到这个关键词")
    ws.delete_rows(target_row)
    vikpea_common.save_workbook_safe(wb, YOUTUBE_KEYWORD_PATH, "YouTube搜索关键词")
    return {"keywords": list_youtube_keywords()}


# 网页上暴露出来可调的搜索参数：(配置表里的 key, 类型转换函数)
YOUTUBE_SETTINGS_FIELDS = {
    "YOUTUBE_RESULTS_PER_KEYWORD": int,
    "YOUTUBE_MIN_VIDEO_VIEWS": int,
    "YOUTUBE_MIN_SHORTS_VIEWS": int,
    "YOUTUBE_MIN_RECENT_AVG_VIEWS": int,
    "YOUTUBE_RECENT_VIDEO_COUNT": int,
    "YOUTUBE_ACTIVE_WITHIN_DAYS": int,
    "YOUTUBE_SUB_MIN": int,
    "YOUTUBE_SUB_MAX": int,
    "YOUTUBE_MARKET_SCORE_MIN": int,
}


def get_youtube_search_settings() -> Dict[str, Any]:
    """当前生效的搜索参数（配置表里有就用配置表的，没有就用脚本默认值——跟桌面版跑起来看到的是同一份）"""
    config = vikpea_common.load_config()
    return {key: config.get(key) for key in YOUTUBE_SETTINGS_FIELDS}


def _upsert_config_rows(updates: Dict[str, Any]) -> None:
    """把 {配置项: 值} 写进 VikPea_配置.xlsx（已有就改，没有就新增一行）"""
    if os.path.exists(vikpea_common.CONFIG_PATH):
        wb = openpyxl.load_workbook(vikpea_common.CONFIG_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.active.append(["配置项", "值", "说明"])
    ws = wb["配置"] if "配置" in wb.sheetnames else wb.active

    existing_keys = {}
    for row in ws.iter_rows(min_row=2):
        key = str(row[0].value or "").strip()
        if key:
            existing_keys[key] = row

    for key, value in updates.items():
        if key in existing_keys:
            existing_keys[key][1].value = value
        else:
            ws.append([key, value, ""])

    vikpea_common.save_workbook_safe(wb, vikpea_common.CONFIG_PATH, "配置")


def set_youtube_search_settings(updates: Dict[str, Any]) -> Dict[str, Any]:
    """把改动写回 VikPea_配置.xlsx（新增/更新对应行），桌面版和网页版会一起生效。"""
    unknown = set(updates) - set(YOUTUBE_SETTINGS_FIELDS)
    if unknown:
        raise ValueError(f"未知的设置项: {', '.join(sorted(unknown))}")
    casted = {key: YOUTUBE_SETTINGS_FIELDS[key](value) for key, value in updates.items()}
    _upsert_config_rows(casted)
    return get_youtube_search_settings()


# ======================== 邮件模板 / 产品信息 ========================
# 只管"写什么内容"，不碰 SMTP 密码/发信这块（那是高风险的真发信功能，要单独确认）

EMAIL_TEMPLATE_FIELDS = {
    "PRODUCT_NAME": str,
    "PRODUCT_TEAM": str,
    "PRODUCT_URL": str,
    "FROM_NAME": str,
    "OUTREACH_SUBJECT_YOUTUBE": str,
    "OUTREACH_TEMPLATE_YOUTUBE": str,
    "OUTREACH_SUBJECT_ARTICLE": str,
    "OUTREACH_TEMPLATE_ARTICLE": str,
}


def get_email_template_settings() -> Dict[str, Any]:
    config = vikpea_common.load_config()
    return {key: config.get(key, "") for key in EMAIL_TEMPLATE_FIELDS}


def set_email_template_settings(updates: Dict[str, Any]) -> Dict[str, Any]:
    unknown = set(updates) - set(EMAIL_TEMPLATE_FIELDS)
    if unknown:
        raise ValueError(f"未知的设置项: {', '.join(sorted(unknown))}")
    casted = {key: str(value) for key, value in updates.items()}
    _upsert_config_rows(casted)
    return get_email_template_settings()


# ======================== 系统设置（SMTP/IMAP/搜索引擎API/发信节奏） ========================
# 这里只是"存密码/存参数"，不会触发任何真实发信或调用外部 API——
# 真正会用到这些密钥的功能（发信、自动跟进、SerpAPI搜索）还没接进网页。

SYSTEM_SETTINGS_FIELDS = {
    "SMTP_SERVER": str,
    "SMTP_PORT": int,
    "SMTP_TIMEOUT": int,
    "SMTP_ALLOW_INSECURE_SSL": bool,
    "FROM_EMAIL": str,
    "IMAP_SERVER": str,
    "IMAP_PORT": int,
    "DAILY_SEND_LIMIT": int,
    "FOLLOWUP_DAILY_LIMIT": int,
    "FOLLOWUP1_AFTER_DAYS": int,
    "FOLLOWUP2_AFTER_DAYS": int,
    "DELAY_SEC": int,
    "SERP_PROVIDER": str,
    "DATAFORSEO_LOGIN": str,
    "ANTHROPIC_API_BASE": str,
    "ANTHROPIC_TAG_MODEL": str,
}

# 密钥类字段：读的时候只返回"有没有设置"，绝不把明文传回前端；
# 写的时候留空 = 不修改，填了新值才覆盖。
SYSTEM_SECRET_FIELDS = ["PASSWORD", "SERPER_API_KEY", "SERPAPI_KEY", "DATAFORSEO_PASSWORD", "YOUTUBE_API_KEY", "ANTHROPIC_API_KEY"]


def get_system_settings() -> Dict[str, Any]:
    config = vikpea_common.load_config()
    result = {key: config.get(key) for key in SYSTEM_SETTINGS_FIELDS}
    for key in SYSTEM_SECRET_FIELDS:
        result[f"{key}_SET"] = bool(str(config.get(key) or "").strip())
    return result


def set_system_settings(updates: Dict[str, Any]) -> Dict[str, Any]:
    known = set(SYSTEM_SETTINGS_FIELDS) | set(SYSTEM_SECRET_FIELDS)
    unknown = set(updates) - known
    if unknown:
        raise ValueError(f"未知的设置项: {', '.join(sorted(unknown))}")

    casted: Dict[str, Any] = {}
    for key, value in updates.items():
        if key in SYSTEM_SECRET_FIELDS:
            text = str(value or "").strip()
            if text:  # 空值 = 不修改，避免误清空已保存的密钥
                casted[key] = text
        else:
            caster = SYSTEM_SETTINGS_FIELDS[key]
            casted[key] = caster(value)

    if casted:
        _upsert_config_rows(casted)
    return get_system_settings()


# ======================== 发送开发信（高风险，两段式：先预览，再明确确认才真发） ========================
# 复用 VikPea_读表发信.py 里跟桌面版完全一样的 build_send_targets/send_targets，
# 不是重写一遍逻辑——安全拦截、去重、主题/开头生成都是同一套代码。

SEND_SESSIONS: Dict[str, Dict[str, Any]] = {}


def _serialize_target(t) -> Dict[str, Any]:
    rownum, name, email, subj, opening, link, etype, source_kw, titles, home_link, video_link = t
    return {
        "rownum": rownum,
        "name": name,
        "email": email,
        "subject": subj,
        "opening": opening,
        "link": link,
        "type": etype,
        "source": source_kw,
        "home_link": home_link,
        "video_link": video_link,
    }


def build_send_preview(should_personalize: bool = False) -> Dict[str, Any]:
    if not str(vikpea_common.load_config().get("PASSWORD") or "").strip():
        raise ValueError("还没配置发信密码，先去「系统设置」填邮箱授权码")

    outreach_sender.apply_send_config()
    session = outreach_sender.build_send_targets(should_personalize=should_personalize)
    if session.get("retry"):
        return {"preview_id": None, "message": session["message"], "targets": [], "blocked": []}

    preview_id = f"send_{uuid.uuid4().hex[:8]}"
    SEND_SESSIONS.clear()  # 同一时间只保留最新一次预览，避免拿着过期 session 发信
    SEND_SESSIONS[preview_id] = session
    return {
        "preview_id": preview_id,
        "message": None,
        "targets": [_serialize_target(t) for t in session["targets"]],
        "blocked": session["blocked_targets"],
    }


def start_send_job(preview_id: str, selected_rownums: List[int]) -> str:
    session = SEND_SESSIONS.get(preview_id)
    if not session:
        raise ValueError("这次预览已经过期（可能被新的预览覆盖了），重新点一次「预览」")
    if not selected_rownums:
        raise ValueError("一个收件人都没选，没什么可发的")

    def run():
        sent_rows, failed = outreach_sender.send_targets(session, selected_rownums=set(selected_rownums))
        print(f"\n完成：成功 {len(sent_rows)} 封，失败 {len(failed)} 封")

    job_id = job_runner.start_job("email_send", run, label=f"发送 {len(selected_rownums)} 封")
    SEND_SESSIONS.pop(preview_id, None)  # session 交给后台任务用完就丢，防止被重复提交
    return job_id


# ======================== 候选库（搜索结果） ========================

def get_confirmed_candidates() -> List[Dict[str, Any]]:
    """VikPea_发信名单.xlsx —— 高置信度、可以直接发信的候选人"""
    rows = _workbook_rows(vikpea_common.QUEUE_PATH)
    return [r for r in rows if str(r.get("频道名") or "").strip()]


def get_pending_candidates() -> List[Dict[str, Any]]:
    """VikPea_待确认邮箱.xlsx —— 置信度不够高，需要人工看一眼的候选人"""
    rows = _workbook_rows(vikpea_common.PENDING_EMAIL_REVIEW_PATH)
    return [r for r in rows if str(r.get("频道名") or "").strip()]


def get_no_email_candidates() -> List[Dict[str, Any]]:
    """VikPea_无邮箱候选.xlsx —— 暂时没找到邮箱的候选人"""
    rows = _workbook_rows(vikpea_common.NO_EMAIL_POOL_PATH)
    return [r for r in rows if str(r.get("频道名") or "").strip()]


CONFIRMED_CANDIDATE_FIELDS = ["频道名", "邮箱", "主页链接", "视频链接", "备注", "类型", "来源关键词", "频道标签"]


def add_confirmed_candidate(data: Dict[str, Any]) -> Dict[str, Any]:
    """人工往 VikPea_发信名单.xlsx 里加一条——比如你自己搜到的、置信度已经很高的博主。"""
    name = str(data.get("频道名") or "").strip()
    email = str(data.get("邮箱") or "").strip()
    if not name or not email:
        raise ValueError("频道名和邮箱必填")

    if os.path.exists(vikpea_common.QUEUE_PATH):
        wb = openpyxl.load_workbook(vikpea_common.QUEUE_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.active.append(["频道名", "邮箱", "定制主题", "定制开头", "主页链接", "视频链接", "备注", "类型", "来源关键词", "频道标签"])
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]

    row_values = {h: "" for h in headers}
    row_values["频道名"] = name
    row_values["邮箱"] = email
    row_values["主页链接"] = str(data.get("主页链接") or "").strip()
    row_values["视频链接"] = str(data.get("视频链接") or "").strip()
    row_values["备注"] = str(data.get("备注") or "人工添加")
    row_values["类型"] = str(data.get("类型") or "YouTube")
    row_values["来源关键词"] = str(data.get("来源关键词") or "人工添加")
    ws.append([row_values.get(h, "") for h in headers])

    vikpea_common.save_workbook_safe(wb, vikpea_common.QUEUE_PATH, "发信名单")
    return {"candidates": get_confirmed_candidates()}


def update_confirmed_candidate(rownum: int, data: Dict[str, Any]) -> Dict[str, Any]:
    """编辑 VikPea_发信名单.xlsx 里已有的一行（比如补全/改主页链接、视频链接）。"""
    if not os.path.exists(vikpea_common.QUEUE_PATH):
        raise FileNotFoundError(vikpea_common.QUEUE_PATH)
    wb = openpyxl.load_workbook(vikpea_common.QUEUE_PATH)
    ws = wb.active
    if rownum < 2 or rownum > ws.max_row:
        raise ValueError("找不到这一行")
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}

    for field in CONFIRMED_CANDIDATE_FIELDS:
        if field in data and field in col_index:
            ws.cell(rownum, col_index[field]).value = data[field]

    vikpea_common.save_workbook_safe(wb, vikpea_common.QUEUE_PATH, "发信名单")
    return {"candidates": get_confirmed_candidates()}


def delete_confirmed_candidate(rownum: int) -> Dict[str, Any]:
    if not os.path.exists(vikpea_common.QUEUE_PATH):
        raise FileNotFoundError(vikpea_common.QUEUE_PATH)
    wb = openpyxl.load_workbook(vikpea_common.QUEUE_PATH)
    ws = wb.active
    if rownum < 2 or rownum > ws.max_row:
        raise ValueError("找不到这一行")
    ws.delete_rows(rownum)
    vikpea_common.save_workbook_safe(wb, vikpea_common.QUEUE_PATH, "发信名单")
    return {"candidates": get_confirmed_candidates()}


# ======================== 发送追踪（已发邮件 + 回复进度） ========================

TRACKER_EDITABLE_FIELDS = [
    "是否回复", "回复摘要", "当前状态", "ABC分级",
    "跟进1日期", "跟进1状态", "跟进2日期", "跟进2状态", "最近回复日期", "频道标签",
]


def get_tracker_rows() -> List[Dict[str, Any]]:
    """VikPea_邮件开发追踪.xlsx —— 已经联络过的人，以及回复/跟进进度"""
    rows = _workbook_rows(vikpea_common.TRACKER_PATH, "邮件追踪")
    return [r for r in rows if str(r.get("邮箱") or "").strip()]


def update_tracker_row(rownum: int, data: Dict[str, Any]) -> Dict[str, Any]:
    """更新一条追踪记录的回复/跟进状态（是否回复、当前状态、ABC分级等）。"""
    if not os.path.exists(vikpea_common.TRACKER_PATH):
        raise FileNotFoundError(vikpea_common.TRACKER_PATH)
    wb = openpyxl.load_workbook(vikpea_common.TRACKER_PATH)
    ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
    if rownum < 2 or rownum > ws.max_row:
        raise ValueError("找不到这一行")
    headers = [str(c.value or "").strip() for c in ws[1]]
    col_index = {h: i + 1 for i, h in enumerate(headers)}

    for field in TRACKER_EDITABLE_FIELDS:
        if field in data and field in col_index:
            ws.cell(rownum, col_index[field]).value = data[field]

    vikpea_common.save_workbook_safe(wb, vikpea_common.TRACKER_PATH, "邮件追踪表")
    return {"rows": get_tracker_rows()}


def is_youtube_search_running() -> bool:
    return job_runner.is_resource_busy("youtube_search")


def start_youtube_search_job() -> str:
    """
    跑真正的 VikPea_YouTube批量搜索.py（真子进程，跟桌面工作台用 subprocess 拉起脚本
    是同一种方式），不是模拟。用真子进程是为了能真正"停止"（杀进程），而不是假装停止。
    """
    enabled_keywords = [k for k in list_youtube_keywords() if k["enabled"]]
    if not enabled_keywords:
        raise ValueError("没有启用的关键词，先在关键词表里启用至少一个")
    script_path = os.path.join(WORKSPACE_DIR, "VikPea_YouTube批量搜索.py")
    return job_runner.start_subprocess_job(
        "youtube_search",
        [sys.executable, "-u", script_path],  # -u：子进程标准输出不缓冲，日志才能实时冒出来
        cwd=WORKSPACE_DIR,
        label=f"YouTube搜索 {len(enabled_keywords)} 个关键词",
    )


def stop_job(job_id: str) -> bool:
    return job_runner.stop_job(job_id)


def get_job_status(job_id: str) -> Optional[Dict[str, Any]]:
    return job_runner.get_job(job_id)


def list_jobs(resource: Optional[str] = None) -> List[Dict[str, Any]]:
    return job_runner.list_jobs(resource)


def get_email_validation_report() -> Dict[str, Any]:
    queue_rows = _workbook_rows(vikpea_common.QUEUE_PATH)
    emails = [str(row.get("邮箱") or "").strip() for row in queue_rows]
    emails = [e for e in emails if e]
    results = [validate_email_address(e) for e in emails]
    return {
        "total_checked": len(results),
        "valid": sum(1 for r in results if r["is_valid"] and not r["is_blacklisted"]),
        "blacklisted": sum(1 for r in results if r["is_blacklisted"]),
        "invalid": sum(1 for r in results if not r["is_valid"]),
        "source": "VikPea_发信名单.xlsx",
    }
