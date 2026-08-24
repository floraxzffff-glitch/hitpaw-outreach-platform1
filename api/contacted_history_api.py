"""
Contacted History API
Manages records of previously contacted channels/sites by reading from Excel files
Bridges between the VikPea_common.py Excel-based system and the web API
"""
import os
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
VIKPEA_DIR = os.path.join(os.path.dirname(SCRIPT_DIR), "01_请在这个文件夹里操作")
sys.path.insert(0, VIKPEA_DIR)

try:
    import openpyxl
except ImportError:
    openpyxl = None

EXTRA_DEDUPE_PATH = os.path.join(VIKPEA_DIR, "VikPea_额外已联络去重.xlsx")
TRACKER_PATH = os.path.join(VIKPEA_DIR, "VikPea_邮件开发追踪.xlsx")

def get_contacted_history_records():
    """
    Get contacted history records from VikPea_邮件开发追踪.xlsx
    Returns list of dicts with keys: name, email, link, note
    """
    if not openpyxl or not os.path.exists(TRACKER_PATH):
        return []

    try:
        wb = openpyxl.load_workbook(TRACKER_PATH, data_only=True)
        ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
        records = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            records.append({
                "name": str(row[2] or "").strip(),
                "email": str(row[3] or "").strip(),
                "link": str(row[5] or "").strip() if len(row) > 5 else "",
                "note": str(row[6] or "").strip() if len(row) > 6 else "",
            })

        return records
    except Exception as e:
        return []

def add_contacted_history_record(record):
    """
    Add a new record to VikPea_额外已联络去重.xlsx
    record should have keys: name, email, link, note (optional)
    """
    if not openpyxl:
        return {"success": False, "error": "openpyxl not available"}

    try:
        if os.path.exists(EXTRA_DEDUPE_PATH):
            wb = openpyxl.load_workbook(EXTRA_DEDUPE_PATH)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["名称", "邮箱", "主页链接", "备注"])

        ws.append([
            record.get("name", ""),
            record.get("email", ""),
            record.get("link", ""),
            record.get("note", f"Web添加 {datetime.now().strftime('%Y-%m-%d')}"),
        ])

        wb.save(EXTRA_DEDUPE_PATH)
        return {"success": True, "record": record}
    except Exception as e:
        return {"success": False, "error": str(e)}

def delete_contacted_history_record(row_index):
    """
    Delete a record from VikPea_额外已联络去重.xlsx by row index
    row_index is 0-based (row 0 = first data row after header)
    """
    if not openpyxl or not os.path.exists(EXTRA_DEDUPE_PATH):
        return {"success": False, "error": "File not found"}

    try:
        wb = openpyxl.load_workbook(EXTRA_DEDUPE_PATH)
        ws = wb.active

        actual_row = row_index + 2

        if actual_row < 2 or actual_row > ws.max_row:
            return {"success": False, "error": "Invalid row index"}

        deleted = {
            "name": str(ws.cell(actual_row, 1).value or ""),
            "email": str(ws.cell(actual_row, 2).value or ""),
            "link": str(ws.cell(actual_row, 3).value or ""),
            "note": str(ws.cell(actual_row, 4).value or ""),
        }

        ws.delete_rows(actual_row)
        wb.save(EXTRA_DEDUPE_PATH)

        return {"success": True, "deleted": deleted}
    except Exception as e:
        return {"success": False, "error": str(e)}
