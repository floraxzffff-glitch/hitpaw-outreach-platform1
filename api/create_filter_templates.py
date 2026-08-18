"""
创建过滤配置Excel模板文件
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKSPACE_DIR = os.environ.get(
    "VIKPEA_WORKSPACE_DIR",
    "/Users/xuzifu/Downloads/VikPea工作台/VikPea工作台_Mac试用包_2026-08-12/01_请在这个文件夹里操作",
)


def create_negative_keywords_template():
    """创建视频负关键词配置模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "负关键词配置"

    # 设置表头
    headers = ["关键词", "合作时间阈值(天)"]
    ws.append(headers)

    # 样式设置
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["hitpaw", 90],
        ["tenorshare", 90],
        ["imyfone", 90],
        ["wondershare", 90],
    ]

    for row in examples:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    path = os.path.join(WORKSPACE_DIR, "VikPea_视频负关键词.xlsx")
    wb.save(path)
    print(f"✅ 创建: {path}")


def create_competitor_sites_template():
    """创建竞品站点黑名单模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品站点"

    # 设置表头
    headers = ["站点域名", "备注"]
    ws.append(headers)

    # 样式设置
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["competitor1.com", "竞品A官网"],
        ["competitor2.com", "竞品B官网"],
    ]

    for row in examples:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30

    path = os.path.join(WORKSPACE_DIR, "VikPea_竞品站点黑名单.xlsx")
    wb.save(path)
    print(f"✅ 创建: {path}")


def create_competitor_emails_template():
    """创建竞品邮箱后缀模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品邮箱"

    # 设置表头
    headers = ["邮箱后缀", "备注"]
    ws.append(headers)

    # 样式设置
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["@competitor1.com", "竞品A公司邮箱"],
        ["@competitor2.com", "竞品B公司邮箱"],
    ]

    for row in examples:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30

    path = os.path.join(WORKSPACE_DIR, "VikPea_竞品邮箱后缀.xlsx")
    wb.save(path)
    print(f"✅ 创建: {path}")


def create_affiliate_blacklist_template():
    """创建Affiliate黑名单模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Affiliate黑名单"

    # 设置表头
    headers = ["频道名", "邮箱", "备注"]
    ws.append(headers)

    # 样式设置
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["示例频道1", "example1@email.com", "Affiliate用户"],
        ["示例频道2", "example2@email.com", "Affiliate用户"],
    ]

    for row in examples:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30

    path = os.path.join(WORKSPACE_DIR, "VikPea_Affiliate黑名单.xlsx")
    wb.save(path)
    print(f"✅ 创建: {path}")


def create_longterm_partners_template():
    """创建长期合作名单模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "长期合作名单"

    # 设置表头
    headers = ["频道名", "邮箱", "备注"]
    ws.append(headers)

    # 样式设置
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 添加示例数据
    examples = [
        ["长期合作频道1", "partner1@email.com", "已建立长期合作关系"],
        ["长期合作频道2", "partner2@email.com", "已建立长期合作关系"],
    ]

    for row in examples:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30

    path = os.path.join(WORKSPACE_DIR, "VikPea_长期合作名单.xlsx")
    wb.save(path)
    print(f"✅ 创建: {path}")


if __name__ == "__main__":
    print("开始创建过滤配置模板文件...")
    create_negative_keywords_template()
    create_competitor_sites_template()
    create_competitor_emails_template()
    create_affiliate_blacklist_template()
    create_longterm_partners_template()
    print("\n✅ 所有模板文件创建完成！")
