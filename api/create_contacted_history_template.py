"""
创建已联络历史名单模板文件
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKSPACE_DIR = os.environ.get(
    "WORKSPACE_DIR",
    "/Users/xuzifu/Downloads/VikPea_项目改进版/workspace"
)

OUTPUT_PATH = os.path.join(WORKSPACE_DIR, "VikPea_已联络历史.xlsx")


def create_contacted_history_template():
    """创建已联络历史名单模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "已联络历史"

    # 设置表头
    headers = ["频道名", "邮箱", "联络日期", "备注"]
    ws.append(headers)

    # 美化表头
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    ws.column_dimensions["A"].width = 30  # 频道名
    ws.column_dimensions["B"].width = 35  # 邮箱
    ws.column_dimensions["C"].width = 15  # 联络日期
    ws.column_dimensions["D"].width = 40  # 备注

    # 添加示例数据
    ws.append([
        "示例频道名",
        "example@example.com",
        "2024-01-15",
        "团队成员A联络，已回复"
    ])

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(OUTPUT_PATH)
    print(f"已联络历史模板已创建: {OUTPUT_PATH}")


if __name__ == "__main__":
    create_contacted_history_template()
