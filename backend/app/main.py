import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware


APP_NAME = "HitPaw Outreach Backend"


def get_cors_origins() -> list[str]:
    raw = os.getenv("CORS_ORIGINS", "*")
    if raw.strip() == "*":
        return ["*"]
    return [item.strip() for item in raw.split(",") if item.strip()]


app = FastAPI(title=APP_NAME)

app.add_middleware(
    CORSMiddleware,
    allow_origins=get_cors_origins(),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "ok": True,
        "service": APP_NAME,
        "time": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }


@app.get("/config/check")
def config_check() -> dict[str, Any]:
    keys = [
        "YOUTUBE_API_KEY",
        "SERPER_API_KEY",
        "SERPAPI_KEY",
        "OPENAI_API_KEY",
        "SMTP_SERVER",
        "SMTP_PORT",
        "SMTP_USER",
        "SMTP_PASSWORD",
    ]
    return {
        "configured": {key: bool(os.getenv(key)) for key in keys},
        "note": "Secrets should be configured in Render environment variables, not committed to GitHub.",
    }


@app.post("/files/inspect-excel")
async def inspect_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx or .xlsm file.")

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File is too large. Max size is 10 MB for V1.")

    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / filename
        path.write_bytes(data)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for ws in wb.worksheets:
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
                headers.append(cell.value)
            sheets.append(
                {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "headers": headers,
                }
            )

    return {"filename": filename, "sheets": sheets}
