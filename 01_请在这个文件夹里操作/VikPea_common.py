"""
VikPea_common.py — shared safety helpers for the VikPea outreach workspace.

This module is intentionally small and dependency-light so existing scripts can
keep their current behavior while gaining config, blacklist, preview, and logs.
"""

import os
import re
import urllib.parse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(SCRIPT_DIR, "VikPea_配置.xlsx")
CONFIG_VKP_PATH = os.path.join(SCRIPT_DIR, "VikPea_配置_VKP.xlsx")
CONFIG_FP_PATH = os.path.join(SCRIPT_DIR, "VikPea_配置_FP.xlsx")
BLACKLIST_PATH = os.path.join(SCRIPT_DIR, "VikPea_黑名单.xlsx")
LOG_PATH = os.path.join(SCRIPT_DIR, "VikPea_运行日志.txt")
PREVIEW_PATH = os.path.join(SCRIPT_DIR, "VikPea_本次发信预览.xlsx")
NO_EMAIL_POOL_PATH = os.path.join(SCRIPT_DIR, "VikPea_无邮箱候选.xlsx")
MANUAL_EMAIL_PATH = os.path.join(SCRIPT_DIR, "VikPea_人工补充邮箱.xlsx")
YOUTUBE_KEYWORD_PATH = os.path.join(SCRIPT_DIR, "VikPea_搜索关键词.xlsx")
ARTICLE_KEYWORD_PATH = os.path.join(SCRIPT_DIR, "VikPea_文章搜索关键词.xlsx")
EXTRA_DEDUPE_PATH = os.path.join(SCRIPT_DIR, "VikPea_额外已联络去重.xlsx")
QUEUE_PATH = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
KEYWORD_METRICS_PATH = os.path.join(SCRIPT_DIR, "VikPea_关键词搜索记录.xlsx")
PENDING_EMAIL_REVIEW_PATH = os.path.join(SCRIPT_DIR, "VikPea_待确认邮箱.xlsx")

EMAIL_RE = re.compile(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}")
URL_ENCODED_FRAGMENT_RE = re.compile(r"%(?:22|20|2b|40|2e|2d|5f|[0-9a-f]{2})", re.I)


DEFAULT_CONFIG = {
    "SMTP_SERVER": "smtp.qiye.aliyun.com",
    "SMTP_PORT": 465,
    "SMTP_TIMEOUT": 25,
    "SMTP_ALLOW_INSECURE_SSL": True,
    "IMAP_SERVER": "imap.qiye.aliyun.com",
    "IMAP_PORT": 993,
    "FROM_EMAIL": "hannah@hitpaw.com",
    "FROM_NAME": "Hannah",
    "PASSWORD": "",
    "PRODUCT_NAME": "HitPaw VikPea",
    "PRODUCT_TEAM": "HitPaw Team",
    "PRODUCT_URL": "https://www.hitpaw.com/hitpaw-video-enhancer.html",
    "OUTREACH_SUBJECT_YOUTUBE": "Possible collaboration with {product_name}",
    "OUTREACH_SUBJECT_ARTICLE": "Possible collaboration with {product_name}",
    "OUTREACH_TEMPLATE_YOUTUBE": "{opening}\n\nI'm {from_name} from {product_name}.\n\nWe'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n1. adding {product_name} as a relevant tool or link mention in your existing content, or\n2. collaborating on a dedicated review or tutorial.\n\nIf you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\nBest,\n{from_name}\n{product_team}\n{product_url}",
    "OUTREACH_TEMPLATE_ARTICLE": "{opening}\n\nI'm {from_name} from {product_name}.\n\nWe'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n1. adding {product_name} as a relevant tool or link mention in an existing article, or\n2. collaborating on a dedicated review or tutorial.\n\nIf you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\nBest,\n{from_name}\n{product_team}\n{product_url}",
    "DELAY_SEC": 8,
    "DAILY_SEND_LIMIT": 80,
    "FOLLOWUP_DAILY_LIMIT": 50,
    "FOLLOWUP1_AFTER_DAYS": 5,
    "FOLLOWUP2_AFTER_DAYS": 7,
    "DEEP_EMAIL_MAX_ROWS": 80,
    "ARTICLE_RESULTS_PER_QUERY": 30,
    "ARTICLE_MIN_SITE_SCORE": 3,
    "RESPECT_ROBOTS_TXT": True,
    "CRAWL_DELAY_PER_DOMAIN": 1.5,
    "SEO_OPPORTUNITY_RESULTS_PER_KEYWORD": 30,
    "SEO_OPPORTUNITY_MIN_SCORE": 3,
    "SERP_PROVIDER": "",
    "SERPER_API_KEY": "",
    "SERPAPI_KEY": "",
    "DATAFORSEO_LOGIN": "",
    "DATAFORSEO_PASSWORD": "",
    "YOUTUBE_RESULTS_PER_KEYWORD": 35,
    "YOUTUBE_MIN_VIDEO_VIEWS": 800,
    "YOUTUBE_MIN_SHORTS_VIEWS": 2500,
    "YOUTUBE_MIN_RECENT_AVG_VIEWS": 1000,
    "YOUTUBE_RECENT_VIDEO_COUNT": 5,
    "YOUTUBE_ACTIVE_WITHIN_DAYS": 30,
    "YOUTUBE_SUB_MIN": 1000,
    "YOUTUBE_SUB_MAX": 250000,
    "YOUTUBE_MARKET_SCORE_MIN": 2,
    "YOUTUBE_VIDEO_METRICS_TIMEOUT": 15,
    "YOUTUBE_RECENT_CHANNEL_TIMEOUT": 20,
    "YOUTUBE_API_KEY": "",
    "YOUTUBE_API_DELAY_SEC": 0.5,
    "YOUTUBE_API_RETRY_TIMES": 3,
    "YOUTUBE_API_429_COOLDOWN": 10,
    "YTDLP_COOKIES_FROM_BROWSER": "",
    "YTDLP_RETRY_TIMES": 2,
    "USE_SEMRUSH_FOR_YOUTUBE": False,
    "YOUTUBE_KEYWORD_FILTER": True,
    "COLLAB_TRACKER_PATH": "",  # 图谱扩展搜索用的种子来源；留空则用默认的桌面路径
    "REQUIRE_SEND_CODE": "SEND",
    "PERSONALIZE_RECENT_VIDEOS": True,
}


BAD_DOMAIN_FRAGMENTS = {
    "random", "duckduckgo", "google", "youtube", "sentry", "schema", "cloudflare",
    "googleapis", "googleusercontent", "googlevideo", "gstatic", "ytimg",
    "window.wiz", "doubleclick", "googlesyndication", "notegpt",
    "hitpaw", "tenorshare", "topaz", "topazlabs", "vanceai", "avclabs",
    "wondershare", "aiarty", "patreon", "invideo", "openshot", "envato", "pxf.io",
    "o.market", "wix.com", "beacons.ai", "linktr.ee",
    "amazon.com", "newspapers.com", "whitepages.com", "ziffdavis.com",
    "audicable.com", "any-video-converter.com",
    "w3schools.com", "geniuslink.com", "payhip.com", "opensourcealternatives.to",
    "bilibili.com", "deepseek.com", "deepseek.ai", "baidu.com", "bytedance.com",
    "tencent.com", "alibaba.com", "xiaomi.com", "microsoft.com", "openai.com",
}

BAD_TLDS = {"png", "jpg", "jpeg", "gif", "webp", "svg", "avif", "css", "js", "json", "xml"}
BAD_LOCAL_FRAGMENTS = {"www.", "1.env", "6-months-l", "tunacoretalentcre", "ai-category-card"}
LOW_VALUE_EMAILS = {
    "error-lite@duckduckgo.com", "support@wix.com", "abuse@cloudflare.com",
    "privacy@youtube.com", "support@linktr.ee", "support@beacons.ai", "support@notegpt.io",
}
PLACEHOLDER_LOCAL_PARTS = {
    "john.smith", "jane.smith", "test", "testing", "example", "demo",
}
PUBLIC_EMAIL_DOMAINS = {
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com",
    "aol.com", "proton.me", "protonmail.com",
}


def coerce_value(value):
    if isinstance(value, str):
        text = value.strip()
        if text.lower() in {"true", "yes", "y", "是"}:
            return True
        if text.lower() in {"false", "no", "n", "否"}:
            return False
        if text.isdigit():
            return int(text)
    return value


def build_default_config(profile=""):
    config = dict(DEFAULT_CONFIG)
    profile = str(profile or "").strip().lower()
    if profile == "vkp":
        config["PRODUCT_NAME"] = "HitPaw VikPea"
        config["PRODUCT_URL"] = "https://www.hitpaw.com/hitpaw-video-enhancer.html"
        config["OUTREACH_SUBJECT_YOUTUBE"] = "Possible collaboration with HitPaw VikPea"
        config["OUTREACH_SUBJECT_ARTICLE"] = "Possible collaboration with HitPaw VikPea"
        config["OUTREACH_TEMPLATE_YOUTUBE"] = (
            "{opening}\n\n"
            "I'm {from_name} from {product_name}.\n\n"
            "We'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n"
            "1. adding {product_name} as a relevant tool or link mention in your existing video content, or\n"
            "2. collaborating on a dedicated review or tutorial around video quality enhancement.\n\n"
            "If you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\n"
            "Best,\n{from_name}\n{product_team}\n{product_url}"
        )
        config["OUTREACH_TEMPLATE_ARTICLE"] = (
            "{opening}\n\n"
            "I'm {from_name} from {product_name}.\n\n"
            "We'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n"
            "1. adding {product_name} as a relevant tool or link mention in an existing article, or\n"
            "2. collaborating on a dedicated review or tutorial around video quality enhancement.\n\n"
            "If you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\n"
            "Best,\n{from_name}\n{product_team}\n{product_url}"
        )
    elif profile == "fp":
        config["PRODUCT_NAME"] = "FotorPea"
        config["PRODUCT_URL"] = "https://www.hitpaw.com/fotorpea-photo-enhancer.html"
        config["OUTREACH_SUBJECT_YOUTUBE"] = "Possible collaboration with FotorPea"
        config["OUTREACH_SUBJECT_ARTICLE"] = "Possible collaboration with FotorPea"
        config["OUTREACH_TEMPLATE_YOUTUBE"] = (
            "{opening}\n\n"
            "I'm {from_name} from {product_name}.\n\n"
            "We'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n"
            "1. adding {product_name} as a relevant tool or link mention in your existing content, or\n"
            "2. collaborating on a dedicated review or tutorial around photo enhancement, image quality, or creator workflows.\n\n"
            "If you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\n"
            "Best,\n{from_name}\n{product_team}\n{product_url}"
        )
        config["OUTREACH_TEMPLATE_ARTICLE"] = (
            "{opening}\n\n"
            "I'm {from_name} from {product_name}.\n\n"
            "We'd love to explore a collaboration with you. If it feels like a fit, we're open to either:\n"
            "1. adding {product_name} as a relevant tool or link mention in an existing article, or\n"
            "2. collaborating on a dedicated review or tutorial around photo enhancement, image quality, or creator workflows.\n\n"
            "If you're open to this, please feel free to share your rate card or usual collaboration pricing.\n\n"
            "Best,\n{from_name}\n{product_team}\n{product_url}"
        )
    return config


def config_path_for_profile(profile=""):
    profile = str(profile or "").strip().lower()
    if profile == "vkp":
        return CONFIG_VKP_PATH
    if profile == "fp":
        return CONFIG_FP_PATH
    return CONFIG_PATH


def load_config(config_path=None, profile=""):
    config = build_default_config(profile)
    path = config_path or config_path_for_profile(profile)
    if not openpyxl or not os.path.exists(path):
        return config
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str((row[0] if row else "") or "").strip()
            if not key:
                continue
            value = row[1] if len(row) > 1 else None
            if value is not None and str(value).strip() != "":
                config[key] = coerce_value(value)
    except Exception as exc:
        log_event("配置", f"读取配置失败，使用脚本默认值: {exc}")
    return config


def apply_config(globals_dict, mapping, config_path=None, profile=""):
    config = load_config(config_path=config_path, profile=profile)
    for config_key, global_name in mapping.items():
        value = config.get(config_key)
        if value is not None and value != "":
            globals_dict[global_name] = value
    return config


def log_event(action, message):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {action}: {message}\n")
    except Exception:
        pass


def normalize_email(email):
    return str(email or "").strip().lower()


# 常见 TLD，长的排前面（比如要先试 "info" 再试 "in"，避免截错）。
# 用来修正"网页正文里邮箱和后面的单词没加空格，被正则一起吞进域名后缀"的问题，
# 比如 "artturi@codingem.com" 后面紧跟 "Hi Artturi" 被抓成了 "artturi@codingem.comhi"。
_COMMON_TLDS = sorted({
    "com", "net", "org", "io", "co", "ai", "info", "biz", "tv", "me", "us",
    "uk", "ca", "de", "fr", "cn", "jp", "edu", "gov", "xyz", "app", "dev",
    "cloud", "live", "online", "store", "tech", "site", "shop", "blog",
    "email", "name", "pro", "club", "world", "media", "studio", "agency",
    "company", "group", "team", "link", "page", "design",
}, key=len, reverse=True)


def fix_glued_tld(email):
    """邮箱域名后缀如果比常见 TLD 长，且前缀正好是个常见 TLD，就截断掉多余部分。"""
    email = str(email or "")
    if "@" not in email:
        return email
    local, _, domain = email.rpartition("@")
    parts = domain.split(".")
    if len(parts) < 2:
        return email
    tld = parts[-1].lower()
    if tld in _COMMON_TLDS:
        return email
    for known in _COMMON_TLDS:
        if tld.startswith(known) and len(tld) > len(known):
            parts[-1] = tld[:len(known)]
            return f"{local}@{'.'.join(parts)}"
    return email


def domain_of_url(url):
    try:
        from urllib.parse import urlparse
        return urlparse(str(url or "")).netloc.lower().replace("www.", "")
    except Exception:
        return ""


def root_domain(value: str) -> str:
    domain = str(value or "").strip().lower().replace("www.", "")
    if "@" in domain:
        domain = domain.split("@", 1)[1]
    parts = [part for part in domain.split(".") if part]
    if len(parts) <= 2:
        return domain
    special_suffixes = {
        "co.uk", "org.uk", "ac.uk", "com.au", "net.au", "org.au",
        "co.jp", "com.cn", "com.hk", "com.sg", "com.br", "co.in",
    }
    suffix = ".".join(parts[-2:])
    suffix3 = ".".join(parts[-3:])
    if suffix in special_suffixes and len(parts) >= 3:
        return ".".join(parts[-3:])
    if suffix3 in special_suffixes and len(parts) >= 4:
        return ".".join(parts[-4:])
    return ".".join(parts[-2:])


def normalize_name_tokens(value: str) -> list:
    text = re.sub(r"[^a-zA-Z0-9]+", " ", str(value or "").lower())
    tokens = []
    for token in text.split():
        if len(token) >= 4 and token not in {"video", "videos", "channel", "official"}:
            tokens.append(token)
    return tokens


def channel_identity_tokens(channel_name: str = "", channel_url: str = "") -> set:
    tokens = set(normalize_name_tokens(channel_name))
    m = re.search(r"/@([^/?#]+)", str(channel_url or ""))
    if m:
        tokens.update(normalize_name_tokens(m.group(1)))
    root = root_domain(domain_of_url(channel_url))
    if root and "." in root:
        tokens.update(normalize_name_tokens(root.split(".", 1)[0]))
    return {token for token in tokens if len(token) >= 4}


def linked_root_domains_from_text(text: str) -> set:
    urls = re.findall(r"https?://[^\s<>\")'\]]+", str(text or ""), flags=re.I)
    domains = set()
    for url in urls:
        root = root_domain(domain_of_url(url))
        if root:
            domains.add(root)
    return domains


def email_relevance_reason(email: str, channel_name: str = "", channel_url: str = "",
                           source_label: str = "", text: str = "", page_url: str = "") -> str:
    bad = classify_bad_email(email)
    if bad:
        return bad
    e = normalize_email(email)
    if "@" not in e:
        return "邮箱格式异常"
    local, domain = e.split("@", 1)
    root = root_domain(domain)
    page_root = root_domain(domain_of_url(page_url))
    linked_roots = linked_root_domains_from_text(text)
    tokens = channel_identity_tokens(channel_name, channel_url)
    source_label = str(source_label or "")

    if root in {"gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com", "proton.me", "protonmail.com", "aol.com"}:
        if source_label.startswith("搜索"):
            if not any(token in local for token in tokens):
                return "搜索结果里的公共邮箱与频道标识不匹配"
        return ""

    text_lower = str(text or "").lower()
    if page_root and root == page_root:
        if source_label.startswith("外链") or source_label.startswith("聚合页"):
            # 外链页必须更像“这个频道自己的站”，不能只是随便一个带邮箱的工具页。
            if any(token in text_lower for token in tokens):
                return ""
            if any(token in root or token in local for token in tokens):
                return ""
            return "外链页面邮箱与频道标识不匹配"
        return ""
    if linked_roots and root in linked_roots:
        return ""
    if any(token in root or token in local for token in tokens):
        return ""

    if source_label.startswith("YouTube") or "描述" in source_label:
        return ""
    if source_label.startswith("外链") or source_label.startswith("聚合页"):
        return "外链邮箱域名与频道/站点不匹配"
    if source_label.startswith("搜索"):
        return "搜索结果邮箱域名与频道不匹配"
    return ""


def classify_bad_email(email):
    e = normalize_email(email)
    if not e:
        return "空邮箱"
    if e in LOW_VALUE_EMAILS:
        return "平台/错误邮箱"
    if e.count("@") != 1 or not EMAIL_RE.fullmatch(e):
        return "邮箱格式异常"
    local, domain = e.split("@", 1)
    parts = domain.split(".")
    tld = parts[-1] if parts else ""
    if len(local) < 2 or len(local) > 64 or len(domain) > 80:
        return "邮箱格式可疑"
    if any(ch in e for ch in {'"', "'", " "}) or URL_ENCODED_FRAGMENT_RE.search(e):
        return "URL编码碎片误判邮箱"
    if "%" in e or local.startswith("-") or local.startswith("."):
        return "邮箱格式可疑"
    if re.fullmatch(r"u00[0-9a-f]{2,}", local):
        return "编码碎片误判邮箱"
    if local in PLACEHOLDER_LOCAL_PARTS and domain in PUBLIC_EMAIL_DOMAINS:
        return "占位测试邮箱"
    if tld in BAD_TLDS or re.search(r"\d+x\.", domain) or "compressed" in domain:
        return "资源文件误判"
    if any(fragment in local for fragment in BAD_LOCAL_FRAGMENTS):
        return "单词被误切成邮箱"
    if any(fragment in domain for fragment in BAD_DOMAIN_FRAGMENTS):
        return "平台/竞品/追踪域名邮箱"
    return ""


def load_blacklist():
    data = {"emails": set(), "domains": set(), "names": set()}
    if not openpyxl or not os.path.exists(BLACKLIST_PATH):
        return data
    try:
        wb = openpyxl.load_workbook(BLACKLIST_PATH, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            kind = str((row[0] if row else "") or "").strip().lower()
            value = str((row[1] if len(row) > 1 else "") or "").strip().lower()
            enabled = str((row[3] if len(row) > 3 else "是") or "是").strip()
            if not value or enabled in {"否", "no", "n", "0"}:
                continue
            if kind in {"邮箱", "email"}:
                data["emails"].add(value)
            elif kind in {"域名", "domain"}:
                data["domains"].add(value.replace("www.", ""))
            elif kind in {"名称", "name"}:
                data["names"].add(value)
    except Exception as exc:
        log_event("黑名单", f"读取黑名单失败: {exc}")
    return data


def blacklist_reason(name, email, link, blacklist=None):
    blacklist = blacklist or load_blacklist()
    email = normalize_email(email)
    name_text = str(name or "").strip().lower()
    domain = email.split("@", 1)[1] if "@" in email else ""
    link_domain = domain_of_url(link)

    if email and email in blacklist["emails"]:
        return "邮箱在黑名单"
    if domain and any(domain == d or domain.endswith("." + d) for d in blacklist["domains"]):
        return "邮箱域名在黑名单"
    if link_domain and any(link_domain == d or link_domain.endswith("." + d) for d in blacklist["domains"]):
        return "链接域名在黑名单"
    if name_text and name_text in blacklist["names"]:
        return "名称在黑名单"
    return ""


def safety_reason(name, email, link, blacklist=None):
    bad = classify_bad_email(email)
    if bad:
        return bad
    blocked = blacklist_reason(name, email, link, blacklist)
    if blocked:
        return blocked
    return ""


def export_send_preview(targets, preview_path=PREVIEW_PATH):
    """targets: iterable of dicts with rownum/name/email/subject/opening/link/type/source/reason."""
    if not openpyxl:
        return ""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "本次发信预览"
    headers = ["队列表行", "状态", "名称", "邮箱", "类型", "来源关键词", "链接", "主题", "开头/备注"]
    ws.append(headers)
    for item in targets:
        ws.append([
            item.get("rownum"),
            item.get("status", "可发送"),
            item.get("name"),
            item.get("email"),
            item.get("type"),
            item.get("source"),
            item.get("link"),
            item.get("subject"),
            item.get("opening") or item.get("reason", ""),
        ])
    header_fill = PatternFill("solid", fgColor="2F5496")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = {"A": 10, "B": 14, "C": 26, "D": 32, "E": 12, "F": 28, "G": 48, "H": 34, "I": 80}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(preview_path)
    return preview_path


QUEUE_HEADERS = ["频道名", "邮箱", "定制主题", "定制开头", "主页链接", "视频链接", "备注", "类型", "来源关键词"]
NO_EMAIL_HEADERS = QUEUE_HEADERS + ["入池日期", "来源表行", "处理状态", "最后处理日期"]
PENDING_EMAIL_HEADERS = [
    "频道名", "候选邮箱", "置信度", "邮箱来源", "主页链接", "视频链接", "来源关键词",
    "类型", "候选备注", "入池日期", "处理动作", "人工确认邮箱", "人工备注",
]


def get_sheet(wb, title, create=False):
    if title in wb.sheetnames:
        return wb[title]
    if create:
        return wb.create_sheet(title)
    return None


def safe_load_workbook(path, sheet_title="Sheet"):
    """
    尽量稳地打开工作簿。
    如果旧文件损坏，就自动旁路备份并重建一个空簿，避免工作台启动直接崩。
    """
    if os.path.exists(path):
        try:
            wb = openpyxl.load_workbook(path)
            ws = get_sheet(wb, sheet_title) or wb.active
            return wb, ws
        except Exception as exc:
            broken_path = path + f".broken_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            try:
                os.replace(path, broken_path)
                log_event("工作簿修复", f"{os.path.basename(path)} 损坏，已备份为 {os.path.basename(broken_path)}；原因: {exc}")
            except Exception as move_exc:
                log_event("工作簿修复", f"{os.path.basename(path)} 损坏，但备份失败: {move_exc}; 原始原因: {exc}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def save_workbook_safe(wb, path, purpose="工作簿"):
    try:
        wb.save(path)
        return True
    except PermissionError:
        basename = os.path.basename(path)
        print(f"⚠️ {basename} 正被 Excel/WPS 占用，已跳过自动保存，不影响继续运行。")
        print(f"   请先关闭这个表，再重新运行需要写入 {purpose} 的步骤。")
        log_event("表格保存跳过", f"{purpose}: {path} 被占用")
        return False


def set_column_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_instruction_sheet(wb, lines, title="填写说明"):
    ws = get_sheet(wb, title, create=True)
    ws.delete_rows(1, ws.max_row or 1)
    ws["A1"] = "请先看这里"
    ws["A1"].font = Font(bold=True, size=13)
    for idx, line in enumerate(lines, start=3):
        ws.cell(idx, 1).value = line
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    return ws


def append_example_rows(ws, example_rows):
    if ws.max_row > 1:
        return
    for row in example_rows:
        ws.append(row)


def ensure_queue_headers(ws):
    if str(ws.cell(1, 6).value or "").strip() == "备注":
        ws.insert_cols(6, 1)
    if not str(ws.cell(1, 6).value or "").strip() and str(ws.cell(1, 7).value or "").strip() == "备注":
        ws.cell(1, 6).value = "视频链接"
    for idx, title in enumerate(QUEUE_HEADERS, 1):
        ws.cell(1, idx).value = title


def ensure_no_email_pool(path=NO_EMAIL_POOL_PATH):
    if not openpyxl:
        return None, None
    wb, ws = safe_load_workbook(path, "无邮箱候选")
    if str(ws.cell(1, 6).value or "").strip() == "备注":
        ws.insert_cols(6, 1)
    if not str(ws.cell(1, 6).value or "").strip() and str(ws.cell(1, 7).value or "").strip() == "备注":
        ws.cell(1, 6).value = "视频链接"
    for idx, title in enumerate(NO_EMAIL_HEADERS, 1):
        ws.cell(1, idx).value = title
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 28, "B": 28, "C": 34, "D": 54, "E": 52, "F": 52,
        "G": 54, "H": 14, "I": 24, "J": 14, "K": 12, "L": 16, "M": 16,
    }
    set_column_widths(ws, widths)
    write_instruction_sheet(wb, [
        "这个表是“暂时没找到邮箱”的副表，搜索脚本和深度找邮箱脚本会自动往这里写。",
        "你不用手工搬迁数据；如果你后来人工找到了邮箱，直接在本表第 B 列“邮箱”补进去即可。",
        "建议人工只改这些列：A 频道名、B 邮箱、E 主页链接、F 视频链接、G 备注、H 类型、I 来源关键词。",
        "第 L 列“处理状态”和第 M 列“最后处理日期”由脚本维护，尽量不要手改。",
        "一旦你补了邮箱，发信脚本和深度找邮箱脚本会自动把这一行转入主发信表。",
    ])
    return wb, ws


def existing_no_email_keys(path=NO_EMAIL_POOL_PATH):
    keys = set()
    if not openpyxl or not os.path.exists(path):
        return keys
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str((row[0] if row else "") or "").strip().lower()
            link = str((row[4] if len(row) > 4 else "") or "").strip().rstrip("/").lower()
            if name or link:
                keys.add((name, link))
    except Exception:
        pass
    return keys


def normalize_contact_text(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_contact_link(value):
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.split("#", 1)[0].rstrip("/")
    return text.lower()


def workbook_contact_keys(paths):
    keys = {"emails": set(), "links": set(), "names": set()}
    if not openpyxl:
        return keys
    for path in paths:
        if not path or not os.path.exists(path):
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for cell in row:
                        text = str(cell or "").strip()
                        if not text:
                            continue
                        if "@" in text:
                            match = EMAIL_RE.search(text)
                            if match:
                                keys["emails"].add(match.group(0).lower())
                        if "http" in text:
                            parsed = urllib.parse.urlparse(text)
                            if parsed.netloc:
                                keys["links"].add(normalize_contact_link(text))
                                keys["links"].add((parsed.netloc + parsed.path).rstrip("/").lower())
                    for idx in (0, 2, 5):
                        if idx < len(row):
                            name = normalize_contact_text(row[idx])
                            if name and "http" not in name and "@" not in name:
                                keys["names"].add(name)
        except Exception:
            continue
    return keys


def contact_already_known(name="", email="", home_link="", video_link="", paths=None):
    paths = paths or [TRACKER_PATH, QUEUE_PATH, EXTRA_DEDUPE_PATH]
    keys = workbook_contact_keys(paths)
    email_key = normalize_contact_text(email)
    name_key = normalize_contact_text(name)
    link_values = [home_link, video_link]
    link_keys = {normalize_contact_link(x) for x in link_values if x}
    link_hosts = set()
    for item in link_values:
        parsed = urllib.parse.urlparse(str(item or ""))
        if parsed.netloc:
            link_hosts.add((parsed.netloc + parsed.path).rstrip("/").lower())

    if email_key and email_key in keys["emails"]:
        return "邮箱已在邮件追踪/发信名单"
    if name_key and name_key in keys["names"]:
        return "名称已在邮件追踪/发信名单"
    if any(x and x in keys["links"] for x in link_keys | link_hosts):
        return "主页/视频链接已在邮件追踪/发信名单"
    return ""


def add_no_email_candidate(name, link="", note="", email_type="", source="", subject="", opening="",
                           source_row="", status="待深度查邮箱", path=NO_EMAIL_POOL_PATH, video_link=""):
    """把无邮箱候选放到副表，主发信表保持只放可发送对象。"""
    if not openpyxl:
        return ""
    wb, ws = ensure_no_email_pool(path)
    home_link = link
    key = (str(name or "").strip().lower(), str(home_link or "").strip().rstrip("/").lower())
    for row in ws.iter_rows(min_row=2):
        old_key = (
            str(row[0].value or "").strip().lower(),
            str(row[4].value or "").strip().rstrip("/").lower(),
        )
        if old_key == key:
            if video_link and not str(row[5].value or "").strip():
                row[5].value = video_link
            if note and note not in str(row[6].value or ""):
                row[6].value = (str(row[6].value or "") + " | " if row[6].value else "") + note
            if source and not row[8].value:
                row[8].value = source
            existing_status = str(row[11].value or "").strip()
            if status and not existing_status:
                row[11].value = status
            elif existing_status in {"待深度查邮箱", "新线索待查"} and any(
                marker in str(row[6].value or "") for marker in ["深度未找到", "深度找到", "已清除", "假邮箱", "竞品/平台邮箱"]
            ):
                # 旧行重新被搜索命中时，保留真正的历史处理状态，不要被新的待查状态冲掉。
                if "深度找到" in str(row[6].value or ""):
                    row[11].value = "已找到/后续清理"
                elif any(marker in str(row[6].value or "") for marker in ["已清除", "假邮箱", "竞品/平台邮箱"]):
                    row[11].value = "已清除"
                else:
                    row[11].value = "已查无邮箱"
            save_workbook_safe(wb, path, "无邮箱候选池")
            return path
    ws.append([
        name, None, subject, opening, home_link, video_link, note, email_type, source,
        datetime.now().strftime("%Y-%m-%d"), source_row, status, None,
    ])
    save_workbook_safe(wb, path, "无邮箱候选池")
    return path


def append_send_queue_row(ws, name, email, subject="", opening="", link="", note="",
                          email_type="", source="", fill=None, video_link="", home_link=None):
    ensure_queue_headers(ws)
    if home_link is None:
        home_link = link
    r = ws.max_row + 1
    values = [name, email, subject, opening, home_link, video_link, note, email_type, source]
    for idx, value in enumerate(values, 1):
        ws.cell(r, idx).value = value
        if fill:
            ws.cell(r, idx).fill = fill
    return r


def ensure_config_workbook(path, profile=""):
    default_config = build_default_config(profile)
    notes = {
        "PASSWORD": "邮箱授权码；留空时继续使用脚本里原有授权码",
        "SMTP_TIMEOUT": "SMTP 连接超时秒数；网络慢时可适当调大",
        "SMTP_ALLOW_INSECURE_SSL": "若本机 Python 证书链异常，是否允许自动回退到兼容 SSL 模式；建议保持 True",
        "PRODUCT_NAME": "产品名；如 HitPaw VikPea / FotorPea",
        "PRODUCT_TEAM": "邮件签名中的团队/公司名",
        "PRODUCT_URL": "产品落地页链接",
        "OUTREACH_SUBJECT_YOUTUBE": "YouTube 默认邮件标题；同事可直接粘贴自己整套固定标题模板",
        "OUTREACH_SUBJECT_ARTICLE": "文章站默认邮件标题；同事可直接粘贴自己整套固定标题模板",
        "OUTREACH_TEMPLATE_YOUTUBE": "YouTube 正文模板；建议直接粘贴完整正文，只在需要的位置保留 {opening}。可用变量：{name} {opening} {from_name} {from_email} {product_name} {product_team} {product_url}",
        "OUTREACH_TEMPLATE_ARTICLE": "文章站正文模板；建议直接粘贴完整正文，只在需要的位置保留 {opening}。可用变量：{name} {opening} {from_name} {from_email} {product_name} {product_team} {product_url}",
        "DAILY_SEND_LIMIT": "单次首封最多发送数量",
        "FOLLOWUP_DAILY_LIMIT": "单次跟进最多发送数量",
        "ARTICLE_RESULTS_PER_QUERY": "文章站搜索：每个关键词最多读取多少条搜索结果",
        "ARTICLE_MIN_SITE_SCORE": "文章站搜索：最低站点质量分；越高越严格，建议 3，结果太少可改 2",
        "RESPECT_ROBOTS_TXT": "站点爬取是否遵守 robots.txt；建议保持 True",
        "CRAWL_DELAY_PER_DOMAIN": "同一个域名两次抓取之间等待秒数；避免访问太快",
        "SEO_OPPORTUNITY_RESULTS_PER_KEYWORD": "SEO机会扫描：每个关键词最多读取多少条搜索结果",
        "SEO_OPPORTUNITY_MIN_SCORE": "SEO机会扫描：最低机会分；越高越严格，建议 3，结果太少可改 2",
        "SERP_PROVIDER": "搜索结果API来源；可填 serper / serpapi / dataforseo。留空则用旧版 DuckDuckGo+Bing 兜底",
        "SERPER_API_KEY": "Serper.dev API Key；用于更稳定获取Google搜索结果",
        "SERPAPI_KEY": "SerpApi API Key；用于更稳定获取Google搜索结果",
        "DATAFORSEO_LOGIN": "DataForSEO 登录邮箱/账号；预留给后续排名与SERP监控",
        "DATAFORSEO_PASSWORD": "DataForSEO API 密码；预留给后续排名与SERP监控",
        "YOUTUBE_MIN_VIDEO_VIEWS": "YouTube搜索：命中的那条常规视频至少多少播放；Shorts 会直接跳过",
        "YOUTUBE_MIN_SHORTS_VIEWS": "历史兼容项：现在 Shorts 不进入筛选，这个值不再用于主筛选",
        "YOUTUBE_MIN_RECENT_AVG_VIEWS": "YouTube搜索：最近几条视频平均播放至少多少",
        "YOUTUBE_RECENT_VIDEO_COUNT": "YouTube搜索：计算平均播放时看最近几条视频",
        "YOUTUBE_ACTIVE_WITHIN_DAYS": "YouTube搜索：最近多少天内必须更新过",
        "YOUTUBE_API_DELAY_SEC": "YouTube API 请求间隔秒数；429 多时可调大",
        "YOUTUBE_API_RETRY_TIMES": "YouTube API 单次失败最多重试几次",
        "YOUTUBE_API_429_COOLDOWN": "YouTube API 命中 429 后冷却几秒再试",
        "REQUIRE_SEND_CODE": "正式发送确认码；现在输入 y / yes / send / SEND 都可确认",
    }
    profile_name = {"": "通用", "vkp": "VikPea", "fp": "FotorPea"}.get(str(profile or "").strip().lower(), "通用")
    wb, ws = safe_load_workbook(path, "配置")

    if not ws.cell(1, 1).value:
        ws.append(["配置项", "值", "说明"])
    existing_rows = {}
    for r in range(2, ws.max_row + 1):
        key = str(ws.cell(r, 1).value or "").strip()
        if key:
            existing_rows[key] = r
    for key, value in default_config.items():
        if key in existing_rows:
            row = existing_rows[key]
            if ws.cell(row, 3).value in ("", None):
                ws.cell(row, 3).value = notes.get(key, "")
        else:
            ws.append([key, value, notes.get(key, "")])
    set_column_widths(ws, {"A": 28, "B": 72, "C": 54})
    write_instruction_sheet(wb, [
        f"这是 {profile_name} 配置表。只需要改第 B 列“值”，不要改第 A 列配置项名称。",
        "最常改的是：发信邮箱、授权码、产品名、产品链接、正文模板、YouTube筛选阈值。",
        "OUTREACH_SUBJECT_YOUTUBE / OUTREACH_SUBJECT_ARTICLE 是默认标题模板：主发信表没填标题时就用这里。",
        "OUTREACH_TEMPLATE_YOUTUBE / OUTREACH_TEMPLATE_ARTICLE 是正文模板：建议直接粘贴完整正文，不用按原脚本句式写。",
        "正文模板可用变量：{name} {opening} {from_name} {from_email} {product_name} {product_team} {product_url}",
        "YOUTUBE_RESULTS_PER_KEYWORD = 每个关键词抓多少个结果；数越大越慢，但候选更多。",
        "YOUTUBE_MIN_VIDEO_VIEWS = 命中的那条常规视频至少多少播放；Shorts 会直接跳过。",
        "YOUTUBE_MIN_RECENT_AVG_VIEWS = 最近几条长视频平均播放至少多少。",
        "YOUTUBE_RECENT_VIDEO_COUNT = 计算平均播放时看最近几条长视频。",
        "YOUTUBE_ACTIVE_WITHIN_DAYS = 最近多少天内必须发过一条非 Shorts 的长视频。",
        "YOUTUBE_API_KEY 建议填写；填了以后播放量、最新发布时间会更稳。",
        "SERP_PROVIDER / SERPER_API_KEY / SERPAPI_KEY 用于文章站点与SEO机会搜索；填了会比直接抓搜索网页稳定。",
        "RESPECT_ROBOTS_TXT / CRAWL_DELAY_PER_DOMAIN 控制爬站礼貌规则，建议不要关闭。",
        "SEO_OPPORTUNITY_RESULTS_PER_KEYWORD = SEO机会扫描每个关键词看多少条搜索结果。",
        "SEO_OPPORTUNITY_MIN_SCORE = SEO机会扫描最低机会分；结果太少可降到 2。",
    ])
    save_workbook_safe(wb, path, f"{profile_name}配置表")


def ensure_blacklist_workbook(path=BLACKLIST_PATH):
    if not openpyxl:
        return
    defaults = [
        ["域名", "hitpaw.com", "自家域名", "是"],
        ["域名", "tenorshare.com", "自家域名", "是"],
        ["域名", "topazlabs.com", "竞品", "是"],
        ["域名", "vanceai.com", "竞品", "是"],
        ["域名", "avclabs.com", "竞品", "是"],
        ["域名", "wondershare.com", "竞品", "是"],
        ["邮箱", "error-lite@duckduckgo.com", "搜索引擎错误邮箱", "是"],
    ]
    wb, ws = safe_load_workbook(path, "黑名单")
    if not ws.cell(1, 1).value:
        ws.append(["类型", "值", "原因", "启用"])
    existing = {
        (
            str(ws.cell(r, 1).value or "").strip(),
            str(ws.cell(r, 2).value or "").strip().lower(),
        )
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value or ws.cell(r, 2).value
    }
    for row in defaults:
        key = (row[0], str(row[1]).lower())
        if key not in existing:
            ws.append(row)
    set_column_widths(ws, {"A": 12, "B": 34, "C": 30, "D": 10})
    write_instruction_sheet(wb, [
        "这个表用于禁止重复联系或禁止联系某些对象。",
        "A 列类型只填：域名 / 邮箱 / 名称。",
        "B 列值示例：topazlabs.com、abc@gmail.com、某频道名。",
        "D 列启用只填：是 / 否。填“否”表示先保留记录但暂时不生效。",
        "建议把竞品、平台错误邮箱、自家域名、明确说过不合作的人都放进来。",
    ])
    save_workbook_safe(wb, path, "黑名单")


def ensure_keyword_workbook(path, title, intro_lines, examples):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, title)
    header_a = str(ws.cell(1, 1).value or "").strip()
    header_b = str(ws.cell(1, 2).value or "").strip()
    header_c = str(ws.cell(1, 3).value or "").strip()
    old_status_headers = {"状态", "启用", "是否启用"}
    if header_b not in old_status_headers and header_c in old_status_headers:
        for row in range(2, ws.max_row + 1):
            old_note = ws.cell(row, 2).value
            old_status = ws.cell(row, 3).value
            ws.cell(row, 2).value = old_status
            ws.cell(row, 3).value = old_note
    ws.cell(1, 1).value = "关键词"
    ws.cell(1, 2).value = "是否启用"
    ws.cell(1, 3).value = "备注"
    append_example_rows(ws, examples)
    set_column_widths(ws, {"A": 52, "B": 12, "C": 36})
    write_instruction_sheet(wb, intro_lines)
    save_workbook_safe(wb, path, title)


def ensure_manual_email_workbook(path=MANUAL_EMAIL_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "人工补充邮箱")
    if not ws.cell(1, 1).value:
        ws.append(["频道名", "主页链接", "补充邮箱", "备注", "启用"])
    append_example_rows(ws, [
        ["Sandeep Goswami", "https://www.youtube.com/@SandeepGoswami_24", "demo@example.com", "人工确认可用邮箱", "否"],
    ])
    set_column_widths(ws, {"A": 28, "B": 52, "C": 34, "D": 36, "E": 10})
    write_instruction_sheet(wb, [
        "这个表用于人工补邮箱，优先级高于脚本自动抓到的邮箱。",
        "推荐填写方式：优先填主页链接；同名频道很多时，只写频道名容易撞错。",
        "A 频道名、B 主页链接、C 补充邮箱 至少要填其中两项，最好三项都填。",
        "E 列启用填“是”才会生效；填“否”相当于先备忘，不参与脚本。",
        "如果这个频道本来在“无邮箱候选”表里，你不需要再手动搬迁，脚本会自动转入主发信表。",
    ])
    save_workbook_safe(wb, path, "人工补充邮箱表")


def ensure_extra_dedupe_workbook(path=EXTRA_DEDUPE_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "额外去重")
    if not any(ws.cell(1, c).value for c in range(1, 5)) and ws.max_row >= 2:
        ws.delete_rows(1, 1)
    ws.cell(1, 1).value = "名称"
    ws.cell(1, 2).value = "邮箱"
    ws.cell(1, 3).value = "主页链接"
    ws.cell(1, 4).value = "备注"
    if all(
        str(ws.cell(2, c).value or "").strip() == header
        for c, header in enumerate(["名称", "邮箱", "主页链接", "备注"], start=1)
    ):
        ws.delete_rows(2, 1)
    has_data = any(
        any(ws.cell(r, c).value not in ("", None) for c in range(1, 5))
        for r in range(2, ws.max_row + 1)
    )
    if not has_data:
        ws.append(["历史已联络示例", "demo@example.com", "https://www.youtube.com/@demo", "同事旧表迁移示例"])
    set_column_widths(ws, {"A": 28, "B": 32, "C": 54, "D": 36})
    write_instruction_sheet(wb, [
        "这个表是“额外去重池”，用于并入同事以前已经联络过的人，避免重复发首封。",
        "最稳的列是：名称、邮箱、主页链接。三项不必全有，但至少填一项。",
        "如果你有旧表，不用改成脚本原格式；把有效的名称/邮箱/主页链接贴到这里就行。",
        "这个表只用于去重，不会直接参与发信。",
    ])
    save_workbook_safe(wb, path, "额外去重表")


def ensure_queue_workbook(path=QUEUE_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "发信名单")
    ensure_queue_headers(ws)
    set_column_widths(ws, {
        "A": 28, "B": 30, "C": 34, "D": 54, "E": 52, "F": 52, "G": 54, "H": 14, "I": 24,
    })
    write_instruction_sheet(wb, [
        "这是主发信表。这里只保留“已经有邮箱、可以发送”的对象。",
        "通常由搜索脚本、深度找邮箱脚本、人工补邮箱自动写入；不建议大批量手填。",
        "如果你要手动新增，请至少填写：A 频道名、B 邮箱、E 主页链接、H 类型。",
        "C 定制主题、D 定制开头可以留空；发信脚本会优先用现成值，没有时按逻辑补。",
        "F 视频链接建议保留命中的那条视频；E 主页链接和 F 视频链接最好都保留。",
        "H 类型建议填：YouTube / 文章外链 / 视频插链接开发 / 初次开发+跟进1 等现有类型。",
    ])
    save_workbook_safe(wb, path, "发信名单")


def ensure_pending_email_review_workbook(path=PENDING_EMAIL_REVIEW_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "待确认邮箱")
    if str(ws.cell(1, 6).value or "").strip() == "来源关键词":
        ws.insert_cols(6, 1)
    if not str(ws.cell(1, 6).value or "").strip() and str(ws.cell(1, 7).value or "").strip() == "来源关键词":
        ws.cell(1, 6).value = "视频链接"
    for idx, title in enumerate(PENDING_EMAIL_HEADERS, 1):
        ws.cell(1, idx).value = title
    set_column_widths(ws, {
        "A": 28, "B": 30, "C": 10, "D": 24, "E": 52, "F": 52, "G": 24,
        "H": 14, "I": 48, "J": 14, "K": 16, "L": 30, "M": 36,
    })
    write_instruction_sheet(wb, [
        "这个表放的是‘脚本找到候选邮箱，但不够稳，不应该直接发’的人。",
        "只有 A 级高置信邮箱才会自动进主发信表；B/C 级都会先进这里。",
        "你人工确认后，建议填写第 L 列“人工确认邮箱”，并在第 K 列“处理动作”写：转主表 / 放弃。",
        "如果候选邮箱本身就是正确的，也可以直接复制到第 L 列，便于后续统一处理。",
        "这个表本轮先作为人工审核池，不自动发信。",
    ])
    save_workbook_safe(wb, path, "待确认邮箱表")


def add_pending_email_review_row(name, email, confidence, source_label, link="", source="", email_type="", note="", path=PENDING_EMAIL_REVIEW_PATH, video_link="", home_link=None):
    if not openpyxl:
        return ""
    wb, ws = safe_load_workbook(path, "待确认邮箱")
    if home_link is None:
        home_link = link
    known_reason = contact_already_known(name=name, email=email, home_link=home_link, video_link=video_link)
    if known_reason:
        return f"SKIP:{known_reason}"
    key = (
        str(name or "").strip().lower(),
        str(home_link or "").strip().rstrip("/").lower(),
        str(email or "").strip().lower(),
    )
    for rownum in range(2, ws.max_row + 1):
        old_key = (
            str(ws.cell(rownum, 1).value or "").strip().lower(),
            str(ws.cell(rownum, 5).value or "").strip().rstrip("/").lower(),
            str(ws.cell(rownum, 2).value or "").strip().lower(),
        )
        if old_key == key:
            if video_link and not str(ws.cell(rownum, 6).value or "").strip():
                ws.cell(rownum, 6).value = video_link
            old_note = str(ws.cell(rownum, 9).value or "").strip()
            if note and note not in old_note:
                ws.cell(rownum, 9).value = (old_note + " | " if old_note else "") + note
            save_workbook_safe(wb, path, "待确认邮箱表")
            return path
    ws.append([
        name, email, confidence, source_label, home_link, video_link, source,
        email_type, note, datetime.now().strftime("%Y-%m-%d"),
        "", "", "",
    ])
    save_workbook_safe(wb, path, "待确认邮箱表")
    return path


def cleanup_pending_email_review_known(path=PENDING_EMAIL_REVIEW_PATH):
    """清理待确认邮箱表里已经联系过或已经在发信名单里的历史残留。"""
    if not openpyxl or not os.path.exists(path):
        return 0
    wb, ws = safe_load_workbook(path, "待确认邮箱")
    changed = 0
    queue_keys = workbook_contact_keys([QUEUE_PATH])
    sent_keys = workbook_contact_keys([TRACKER_PATH, EXTRA_DEDUPE_PATH])
    for rownum in range(2, ws.max_row + 1):
        action = str(ws.cell(rownum, 11).value or "").strip()
        if action in {"已发送", "已转主表", "放弃"}:
            continue
        name = str(ws.cell(rownum, 1).value or "").strip()
        email = str(ws.cell(rownum, 2).value or "").strip()
        home_link = str(ws.cell(rownum, 5).value or "").strip()
        video_link = str(ws.cell(rownum, 6).value or "").strip()
        email_key = normalize_contact_text(email)
        name_key = normalize_contact_text(name)
        link_values = [normalize_contact_link(home_link), normalize_contact_link(video_link)]

        in_sent = (
            (email_key and email_key in sent_keys["emails"])
            or (name_key and name_key in sent_keys["names"])
            or any(link and link in sent_keys["links"] for link in link_values)
        )
        in_queue = (
            (email_key and email_key in queue_keys["emails"])
            or (name_key and name_key in queue_keys["names"])
            or any(link and link in queue_keys["links"] for link in link_values)
        )
        if in_sent:
            ws.cell(rownum, 11).value = "已发送"
            ws.cell(rownum, 13).value = "主页/视频链接已在邮件追踪表"
            changed += 1
        elif in_queue:
            ws.cell(rownum, 11).value = "已转主表"
            ws.cell(rownum, 13).value = "已存在于发信名单"
            changed += 1
    if changed:
        save_workbook_safe(wb, path, "待确认邮箱表")
    return changed


def archive_processed_pending_email_rows(path=PENDING_EMAIL_REVIEW_PATH):
    """把待确认邮箱主表里的已处理行移到工作簿内的已处理记录页。"""
    if not openpyxl or not os.path.exists(path):
        return 0
    wb, ws = safe_load_workbook(path, "待确认邮箱")
    archive_name = "已处理记录"
    if archive_name in wb.sheetnames:
        ws_archive = wb[archive_name]
    else:
        ws_archive = wb.create_sheet(archive_name)
    if not ws_archive.cell(1, 1).value:
        for col in range(1, ws.max_column + 1):
            ws_archive.cell(1, col).value = ws.cell(1, col).value
        ws_archive.cell(1, ws.max_column + 1).value = "归档日期"

    processed = {"已发送", "已转主表", "放弃"}
    rows_to_delete = []
    for rownum in range(2, ws.max_row + 1):
        action = str(ws.cell(rownum, 11).value or "").strip()
        if action not in processed:
            continue
        values = [ws.cell(rownum, col).value for col in range(1, ws.max_column + 1)]
        values.append(datetime.now().strftime("%Y-%m-%d"))
        ws_archive.append(values)
        rows_to_delete.append(rownum)

    for rownum in reversed(rows_to_delete):
        ws.delete_rows(rownum)
    if rows_to_delete:
        save_workbook_safe(wb, path, "待确认邮箱表")
    return len(rows_to_delete)


def ensure_tracker_workbook(path=TRACKER_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "邮件追踪")
    headers = [
        "#", "日期", "联系人/平台", "邮箱", "类型", "频道/文章/平台", "视频链接",
        "邮件类型", "是否回复", "回复摘要", "主页链接", "当前状态",
        "ABC分级", "跟进1日期", "跟进1状态", "跟进2日期", "跟进2状态",
        "最近回复日期", "来源关键词", "频道标签",
    ]
    for idx, title in enumerate(headers, 1):
        ws.cell(1, idx).value = title
    set_column_widths(ws, {
        "A": 8, "B": 14, "C": 28, "D": 30, "E": 16, "F": 30, "G": 40,
        "H": 22, "I": 12, "J": 36, "K": 54, "L": 18, "M": 10,
        "N": 14, "O": 14, "P": 14, "Q": 14, "R": 14, "S": 24, "T": 16,
    })
    write_instruction_sheet(wb, [
        "这是邮件追踪主表。首封发信、补录已发送、读取回复、自动跟进都会依赖它。",
        "正常情况下不要人工批量改主表结构；尤其不要改第 1 行表头。",
        "如果是第一次试用，这个表为空是正常的；等真正发信后脚本会自动写入。",
        "ABC分级、回复摘要、跟进状态这些列会由回复/分析脚本逐步补齐。",
    ])
    save_workbook_safe(wb, path, "邮件追踪表")


def ensure_keyword_metrics_workbook(path=KEYWORD_METRICS_PATH):
    if not openpyxl:
        return
    wb, ws = safe_load_workbook(path, "关键词搜索记录")
    headers = ["日期", "关键词", "搜到频道数", "入池候选数", "新增有邮箱", "新增无邮箱", "跳过数"]
    for idx, title in enumerate(headers, 1):
        ws.cell(1, idx).value = title
    set_column_widths(ws, {"A": 14, "B": 48, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12})
    write_instruction_sheet(wb, [
        "这个表由搜索脚本自动追加记录，用来做关键词复盘。",
        "不需要手填；空表是正常的。",
        "如果你想保留历史复盘数据，直接保留这个文件即可。",
    ])
    save_workbook_safe(wb, path, "关键词复盘表")


def load_manual_email_rows(path=MANUAL_EMAIL_PATH):
    rows = []
    if not openpyxl or not os.path.exists(path):
        return rows
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str((row[0] if row else "") or "").strip()
            link = str((row[1] if len(row) > 1 else "") or "").strip().rstrip("/")
            email = normalize_email(row[2] if len(row) > 2 else "")
            note = str((row[3] if len(row) > 3 else "") or "").strip()
            enabled = str((row[4] if len(row) > 4 else "是") or "是").strip().lower()
            if not email or enabled in {"否", "no", "n", "0", "false"}:
                continue
            rows.append({
                "name": name,
                "link": link,
                "email": email,
                "note": note,
            })
    except Exception as exc:
        log_event("人工补充邮箱", f"读取失败: {exc}")
    return rows


def find_manual_email(channel_name="", channel_url="", path=MANUAL_EMAIL_PATH):
    rows = load_manual_email_rows(path)
    if not rows:
        return "", ""
    target_name = str(channel_name or "").strip().lower()
    target_link = str(channel_url or "").strip().rstrip("/").lower()
    for row in rows:
        row_link = str(row["link"] or "").strip().rstrip("/").lower()
        row_name = str(row["name"] or "").strip().lower()
        if row_link and target_link and row_link == target_link:
            return row["email"], row["note"] or "人工补充邮箱"
    for row in rows:
        row_name = str(row["name"] or "").strip().lower()
        if row_name and target_name and row_name == target_name:
            return row["email"], row["note"] or "人工补充邮箱"
    return "", ""


def create_default_workbooks():
    if not openpyxl:
        return
    ensure_config_workbook(CONFIG_PATH, "")
    ensure_config_workbook(CONFIG_VKP_PATH, "vkp")
    ensure_config_workbook(CONFIG_FP_PATH, "fp")
    ensure_blacklist_workbook(BLACKLIST_PATH)
    ensure_keyword_workbook(
        YOUTUBE_KEYWORD_PATH,
        "YouTube关键词",
        [
            "这个表给 YouTube 搜索 KOL 用。A 列填关键词，B 列填是否启用。",
            "B 列只建议填：是 / 否。填“是”的才会参与搜索。",
            "关键词建议填用户真实会搜的主题，不要只写产品名；例如 review / tutorial / fix / compare。",
            "每行一个关键词，不要把多个关键词塞进同一个单元格。",
        ],
        [
            ["topaz video ai review", "否", "示例：评测型；先关闭，按自己需要启用"],
            ["how to improve video quality", "否", "示例：先备选，不启用"],
        ],
    )
    ensure_keyword_workbook(
        ARTICLE_KEYWORD_PATH,
        "文章关键词",
        [
            "这个表给文章/网站搜索用。A 列填关键词，B 列填是否启用。",
            "建议放 review / best / alternative / comparison 这种更容易找到外链文章站的词。",
            "如果某个关键词效果差，直接把 B 列改成“否”，不用删。",
        ],
        [
            ["best ai video enhancer software 2026", "否", "示例：榜单文章；先关闭，按自己需要启用"],
            ["topaz video ai alternative", "否", "示例：替代型文章；先关闭，按自己需要启用"],
        ],
    )
    ensure_extra_dedupe_workbook(EXTRA_DEDUPE_PATH)
    ensure_queue_workbook(QUEUE_PATH)
    ensure_tracker_workbook(TRACKER_PATH)
    ensure_keyword_metrics_workbook(KEYWORD_METRICS_PATH)
    ensure_pending_email_review_workbook(PENDING_EMAIL_REVIEW_PATH)
    wb_pool, _ws_pool = ensure_no_email_pool(NO_EMAIL_POOL_PATH)
    if wb_pool:
        save_workbook_safe(wb_pool, NO_EMAIL_POOL_PATH, "无邮箱候选池")
