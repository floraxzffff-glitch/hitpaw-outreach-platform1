#!/usr/bin/env python3
"""
主程序 - YouTube KOL关键词拓展+视频搜索工具
"""

import os
import sys
import argparse
from datetime import datetime
from typing import List, Dict

# 导入自定义模块
from keyword_expansion import KeywordExpander, estimate_dataforseo_cost
from video_search import YouTubeSearcher, QuotaManager, search_with_both_orders
from scoring import (
    filter_videos, score_and_rank, deduplicate_by_channel,
    format_video_for_output, aggregate_by_channel, print_video_stats
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ 缺少依赖: pip install openpyxl")
    sys.exit(1)


def load_env():
    """加载.env文件"""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return {}

    env_vars = {}
    with open(env_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                env_vars[key.strip()] = value.strip()

    return env_vars


def save_to_excel(videos_by_keyword: Dict[str, List[Dict]],
                  channel_stats: List[Dict],
                  output_path: str):
    """
    保存到Excel，两个sheet:
    1. 按视频维度
    2. 按频道维度汇总
    """
    wb = openpyxl.Workbook()

    # Sheet 1: 按视频维度
    ws_videos = wb.active
    ws_videos.title = "视频列表"

    # 表头
    headers = ["关键词来源", "视频标题", "视频链接", "频道名", "频道链接",
               "观看数", "发布时间", "VPH", "得分"]
    ws_videos.append(headers)

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_videos.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据
    for keyword, videos in videos_by_keyword.items():
        for video in videos:
            formatted = format_video_for_output(video, keyword)
            ws_videos.append([
                formatted["关键词来源"],
                formatted["视频标题"],
                formatted["视频链接"],
                formatted["频道名"],
                formatted["频道链接"],
                formatted["观看数"],
                formatted["发布时间"],
                formatted["VPH"],
                formatted["得分"]
            ])

    # 调整列宽
    ws_videos.column_dimensions['A'].width = 20
    ws_videos.column_dimensions['B'].width = 50
    ws_videos.column_dimensions['C'].width = 40
    ws_videos.column_dimensions['D'].width = 30
    ws_videos.column_dimensions['E'].width = 40
    ws_videos.column_dimensions['F'].width = 12
    ws_videos.column_dimensions['G'].width = 18
    ws_videos.column_dimensions['H'].width = 10
    ws_videos.column_dimensions['I'].width = 10

    # Sheet 2: 按频道维度汇总
    ws_channels = wb.create_sheet("频道汇总")

    channel_headers = ["频道名", "频道链接", "命中视频数", "最高VPH", "平均VPH"]
    ws_channels.append(channel_headers)

    for col_idx, header in enumerate(channel_headers, 1):
        cell = ws_channels.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for stat in channel_stats:
        ws_channels.append([
            stat["频道名"],
            stat["频道链接"],
            stat["命中视频数"],
            stat["最高VPH"],
            stat["平均VPH"]
        ])

    ws_channels.column_dimensions['A'].width = 30
    ws_channels.column_dimensions['B'].width = 40
    ws_channels.column_dimensions['C'].width = 12
    ws_channels.column_dimensions['D'].width = 12
    ws_channels.column_dimensions['E'].width = 12

    # 保存
    wb.save(output_path)
    print(f"\n✅ 结果已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="YouTube KOL关键词拓展+视频搜索工具")
    parser.add_argument("--seed", required=True, nargs="+", help="种子关键词（可多个）")
    parser.add_argument("--vph-threshold", type=float, default=20, help="VPH过滤阈值（默认20）")
    parser.add_argument("--max-keywords", type=int, default=50, help="拓展关键词数量上限（默认50）")
    parser.add_argument("--output", default="results.xlsx", help="输出文件路径（默认results.xlsx）")
    parser.add_argument("--depth", type=int, default=1, help="DataForSEO拓展深度1-3（默认1）")
    parser.add_argument("--skip-expansion", action="store_true", help="跳过关键词拓展，直接搜索种子词")
    parser.add_argument("--resume", action="store_true", help="从断点继续")

    args = parser.parse_args()

    print(f"\n{'='*70}")
    print(f"  🚀 YouTube KOL 关键词拓展 + 视频搜索工具")
    print(f"{'='*70}\n")

    # 加载环境变量
    env = load_env()
    dataforseo_login = env.get("DATAFORSEO_LOGIN", "")
    dataforseo_password = env.get("DATAFORSEO_PASSWORD", "")
    youtube_api_key = env.get("YOUTUBE_API_KEY", "")

    if not youtube_api_key:
        print("❌ 缺少YOUTUBE_API_KEY，请在.env文件中配置")
        sys.exit(1)

    # 初始化
    expander = KeywordExpander(dataforseo_login, dataforseo_password)
    searcher = YouTubeSearcher(youtube_api_key)
    quota_mgr = QuotaManager()

    # 种子关键词
    seed_keywords = args.seed
    print(f"📌 种子关键词: {', '.join(seed_keywords)}")

    # 关键词拓展
    all_keywords = []

    if args.skip_expansion:
        print("\n⏭️  跳过关键词拓展，直接使用种子词")
        all_keywords = [{"keyword": kw, "sources": ["seed"]} for kw in seed_keywords]
    else:
        print("\n" + "="*70)
        print("  第一步：关键词拓展")
        print("="*70)

        # 预估DataForSEO成本
        if dataforseo_login and dataforseo_password:
            num_tasks = len(seed_keywords) * 2  # related + ideas
            estimated_cost = estimate_dataforseo_cost(num_tasks)
            print(f"\n💰 预估DataForSEO成本: ${estimated_cost:.4f} ({num_tasks} tasks)")

        expanded_map = {}
        for seed in seed_keywords:
            results = expander.expand_all(seed, max_keywords=args.max_keywords, depth=args.depth)
            expanded_map[seed] = results

        # 合并所有拓展结果
        all_results = {"youtube": [], "related": [], "ideas": []}
        for seed, results in expanded_map.items():
            for source, keywords in results.items():
                all_results[source].extend(keywords)

        merged = expander.merge_and_deduplicate(all_results)
        all_keywords = merged[:args.max_keywords]

        print(f"\n✓ 最终关键词列表: {len(all_keywords)} 个")
        for i, item in enumerate(all_keywords[:10], 1):
            sources = ", ".join(item["sources"])
            print(f"  {i}. {item['keyword']} ({sources})")
        if len(all_keywords) > 10:
            print(f"  ... 还有 {len(all_keywords) - 10} 个")

    # 检查配额
    print("\n" + "="*70)
    print("  第二步：YouTube视频搜索")
    print("="*70)

    estimated_quota = searcher.estimate_search_quota(len(all_keywords))
    remaining_quota = searcher.get_remaining_quota()

    print(f"\n📊 配额状态:")
    print(f"  预估需要: {estimated_quota} units")
    print(f"  剩余配额: {remaining_quota} units")

    if estimated_quota > remaining_quota:
        print(f"\n⚠️  配额不足！将在达到上限后停止并保存断点。")

    # 断点续跑
    keywords_to_search = [item["keyword"] for item in all_keywords]
    processed_keywords = []

    if args.resume:
        checkpoint = quota_mgr.load_checkpoint()
        if checkpoint:
            processed_keywords = checkpoint.get("processed", [])
            keywords_to_search = checkpoint.get("remaining", keywords_to_search)
            print(f"  ↳ 从断点继续，跳过已完成的 {len(processed_keywords)} 个关键词")

    # 搜索
    videos_by_keyword = {}
    all_videos = []

    for i, keyword in enumerate(keywords_to_search, 1):
        print(f"\n[{i}/{len(keywords_to_search)}] ", end="")

        # 检查配额
        if not searcher.check_quota_available(100):
            print(f"\n\n⚠️  配额已用完！已完成 {i-1}/{len(keywords_to_search)} 个关键词")
            print(f"  → 保存断点，明天继续...")
            quota_mgr.save_checkpoint(
                processed_keywords + keywords_to_search[:i-1],
                keywords_to_search[i-1:]
            )
            break

        videos = search_with_both_orders(searcher, keyword, max_results=25)

        if videos:
            videos_by_keyword[keyword] = videos
            all_videos.extend(videos)

        processed_keywords.append(keyword)

    # 筛选打分
    print("\n" + "="*70)
    print("  第三步：筛选打分")
    print("="*70)

    print(f"\n  原始视频数: {len(all_videos)}")

    # 过滤VPH
    filtered = filter_videos(all_videos, vph_threshold=args.vph_threshold)
    print(f"  VPH>={args.vph_threshold}筛选后: {len(filtered)}")

    # 打分排序
    scored = score_and_rank(filtered, "")
    print(f"  打分排序完成")

    # 频道去重
    unique = deduplicate_by_channel(scored)
    print(f"  按频道去重后: {len(unique)}")

    # 统计
    print_video_stats(unique)

    # 频道汇总
    channel_stats = aggregate_by_channel(unique)
    print(f"\n  频道汇总: {len(channel_stats)} 个独特频道")

    # 保存结果
    print("\n" + "="*70)
    print("  第四步：导出结果")
    print("="*70)

    # 按关键词重新分组（仅保留过滤后的）
    filtered_by_keyword = {}
    for video in unique:
        # 找到这个视频来自哪个关键词
        for kw, vids in videos_by_keyword.items():
            if any(v["video_id"] == video["video_id"] for v in vids):
                if kw not in filtered_by_keyword:
                    filtered_by_keyword[kw] = []
                filtered_by_keyword[kw].append(video)
                break

    save_to_excel(filtered_by_keyword, channel_stats, args.output)

    # 清理断点
    if not args.resume or len(keywords_to_search) == len(processed_keywords):
        quota_mgr.clear_checkpoint()

    print(f"\n{'='*70}")
    print(f"  ✅ 完成！")
    print(f"  → 找到 {len(unique)} 个符合条件的视频")
    print(f"  → 涉及 {len(channel_stats)} 个独特频道")
    print(f"  → 今日配额使用: {searcher.quota_used}/{searcher.quota_limit}")
    print(f"{'='*70}\n")


if __name__ == "__main__":
    main()
