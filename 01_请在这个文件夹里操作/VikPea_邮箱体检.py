"""
VikPea_邮箱体检.py — 发信前检查并清理明显假邮箱/平台邮箱/竞品邮箱

用法:
  python3 ~/Downloads/VikPea工作台/VikPea_邮箱体检.py
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from VikPea_common import classify_bad_email
except ImportError:
    classify_bad_email = None


QUEUE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_发信名单.xlsx")

Q_NAME = 1
Q_EMAIL = 2
Q_NOTE = 7

ORANGE = PatternFill("solid", start_color="FCE4D6")

if classify_bad_email is None:
    def classify_bad_email(email: str) -> str:
        return ""


def main():
    if not os.path.exists(QUEUE_PATH):
        print(f"❌ 找不到发信名单: {QUEUE_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(QUEUE_PATH)
    ws = wb.active
    cleared = []
    checked = 0
    for row in ws.iter_rows(min_row=2):
        email = str(row[Q_EMAIL - 1].value or "").strip()
        if not email:
            continue
        checked += 1
        reason = classify_bad_email(email)
        if not reason:
            continue
        cleared.append((row[0].row, row[Q_NAME - 1].value, email, reason))
        row[Q_EMAIL - 1].value = None
        note = str(row[Q_NOTE - 1].value or "").strip()
        row[Q_NOTE - 1].value = (note + " | " if note else "") + f"{reason}已清除，需人工确认或重新查找"
        for cell in row:
            cell.fill = ORANGE

    wb.save(QUEUE_PATH)
    print(f"✅ 邮箱体检完成：检查 {checked} 个邮箱，清理 {len(cleared)} 个可疑邮箱")
    for _, name, email, reason in cleared[:80]:
        print(f"  - {name}: {email} ({reason})")


if __name__ == "__main__":
    main()
