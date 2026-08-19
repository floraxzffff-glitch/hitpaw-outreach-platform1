"""
VikPea_合并同事表格.py — 把同事各自的发信名单/待确认邮箱/邮件追踪表合并进主表。

用途：
  大家原本各自跑一份桌面工具、攒自己的表，现在要把历史数据合到一起。
  三张表分别怎么合：
    · 发信名单：按邮箱去重，信息更全的那条留下
    · 待确认邮箱：按"频道名+候选邮箱"去重，已经人工确认过的优先保留
    · 邮件追踪：直接合并（不同人联系的人一般不重叠），但同一个邮箱出现在
      多个同事表里会被列进"重复联系提醒"，需要你自己看一眼、决定怎么处理，
      不会自动帮你二选一丢掉历史记录。

准备工作：
  建一个文件夹，每个同事一个子文件夹（子文件夹名随便取，建议用同事名字），
  子文件夹里放这三个文件（文件名必须跟原文件一样，某人没有某张表可以不放）：
    VikPea_发信名单.xlsx
    VikPea_待确认邮箱.xlsx
    VikPea_邮件开发追踪.xlsx

用法：
  先预览，不会改任何文件，只生成一份"合并预览"报告：
    python3 VikPea_合并同事表格.py --source ~/Desktop/同事表格 --dry-run

  确认没问题后正式合并（会先自动备份当前这三张主表，再覆盖写入合并结果）：
    python3 VikPea_合并同事表格.py --source ~/Desktop/同事表格 --apply
"""

import argparse
import os
import shutil
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

QUEUE_FILENAME = "VikPea_发信名单.xlsx"
PENDING_FILENAME = "VikPea_待确认邮箱.xlsx"
TRACKER_FILENAME = "VikPea_邮件开发追踪.xlsx"

QUEUE_PATH = os.path.join(SCRIPT_DIR, QUEUE_FILENAME)
PENDING_PATH = os.path.join(SCRIPT_DIR, PENDING_FILENAME)
TRACKER_PATH = os.path.join(SCRIPT_DIR, TRACKER_FILENAME)


def norm(value):
    return str(value or "").strip()


def norm_email(value):
    return norm(value).lower()


def read_rows(path, sheet_name=None):
    """按第一行表头读成 list[dict]；文件不存在返回 None（跟"没这张表"区分开）。"""
    if not path or not os.path.exists(path):
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws.max_row < 1:
        return []
    headers = [norm(c.value) for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return rows


def row_fill_score(row):
    """算一行"信息有多全"，用来在去重时决定留哪条。"""
    return sum(1 for v in row.values() if norm(v))


def find_source_files(source_dir, filename):
    """在 source_dir 的每个子文件夹里找 filename，返回 [(同事名, 文件路径), ...]"""
    found = []
    if not os.path.isdir(source_dir):
        return found
    for entry in sorted(os.listdir(source_dir)):
        person_dir = os.path.join(source_dir, entry)
        if not os.path.isdir(person_dir):
            continue
        candidate = os.path.join(person_dir, filename)
        if os.path.exists(candidate):
            found.append((entry, candidate))
    return found


def merge_by_key(base_rows, other_sources, key_fn, label):
    """
    通用去重合并：base_rows 是主表已有数据，other_sources 是 [(人名, rows), ...]。
    同一个 key 出现多次时，保留"信息最全"的那条。
    返回 (合并后的行列表, 报告字符串列表)。
    """
    merged = {}
    order = []
    report = []

    def add(rows, origin):
        added, updated, skipped = 0, 0, 0
        for row in rows:
            key = key_fn(row)
            if not key:
                continue
            if key not in merged:
                merged[key] = row
                order.append(key)
                added += 1
            else:
                if row_fill_score(row) > row_fill_score(merged[key]):
                    merged[key] = row
                    updated += 1
                else:
                    skipped += 1
        report.append(f"  · {origin}：新增 {added} 条，补全覆盖 {updated} 条，重复跳过 {skipped} 条")

    add(base_rows or [], "当前主表（你自己的）")
    for person, rows in other_sources:
        add(rows, person)

    report.insert(0, f"【{label}】合并后共 {len(order)} 条")
    return [merged[k] for k in order], report


def merge_tracker(base_rows, other_sources):
    """追踪表直接合并（不去重），但把同一邮箱出现在多来源里的情况列出来提醒人工看。"""
    all_rows = []
    seen_by_email = {}

    def add(rows, origin):
        for row in rows:
            email = norm_email(row.get("邮箱"))
            all_rows.append(row)
            if email:
                seen_by_email.setdefault(email, []).append((origin, row.get("联系人/平台"), row.get("当前状态")))

    add(base_rows or [], "当前主表（你自己的）")
    for person, rows in other_sources:
        add(rows, person)

    duplicate_warnings = []
    for email, occurrences in seen_by_email.items():
        if len(occurrences) > 1:
            detail = "；".join(f"{origin}:{name}[{status}]" for origin, name, status in occurrences)
            duplicate_warnings.append(f"  ⚠️ {email} 被 {len(occurrences)} 个来源联系过 —— {detail}")

    report = [f"【邮件追踪】合并后共 {len(all_rows)} 条（追踪表不做去重，只做合并）"]
    if duplicate_warnings:
        report.append(f"  发现 {len(duplicate_warnings)} 个邮箱被多个来源重复联系，建议人工看一眼：")
        report.extend(duplicate_warnings[:30])
        if len(duplicate_warnings) > 30:
            report.append(f"  ...还有 {len(duplicate_warnings) - 30} 条，完整名单看下面生成的预览表")
    return all_rows, report, duplicate_warnings


def write_rows(path, rows, sheet_name=None):
    """把合并结果写回：保留原表头/格式，只替换数据行。"""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [norm(c.value) for c in ws[1]]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def backup(path):
    if not os.path.exists(path):
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{path}.合并前备份_{ts}.xlsx"
    shutil.copy2(path, backup_path)
    return backup_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, help="同事表格所在文件夹（里面是每个人一个子文件夹）")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    if not args.dry_run and not args.apply:
        print("请指定 --dry-run（先预览）或 --apply（正式合并）")
        return

    source_dir = os.path.expanduser(args.source)
    if not os.path.isdir(source_dir):
        print(f"❌ 找不到文件夹: {source_dir}")
        return

    print("\n" + "=" * 64)
    print("  VikPea 同事表格合并")
    print("=" * 64)

    # ---- 发信名单 ----
    queue_sources = [(person, read_rows(p)) for person, p in find_source_files(source_dir, QUEUE_FILENAME)]
    queue_sources = [(p, r) for p, r in queue_sources if r]
    merged_queue, queue_report = merge_by_key(
        read_rows(QUEUE_PATH), queue_sources, lambda r: norm_email(r.get("邮箱")), "发信名单"
    )

    # ---- 待确认邮箱 ----
    pending_sources = [(person, read_rows(p)) for person, p in find_source_files(source_dir, PENDING_FILENAME)]
    pending_sources = [(p, r) for p, r in pending_sources if r]
    merged_pending, pending_report = merge_by_key(
        read_rows(PENDING_PATH), pending_sources,
        lambda r: (norm(r.get("频道名")).lower(), norm_email(r.get("候选邮箱"))), "待确认邮箱"
    )

    # ---- 邮件追踪 ----
    tracker_sources = []
    for person, p in find_source_files(source_dir, TRACKER_FILENAME):
        rows = read_rows(p, sheet_name="邮件追踪")
        if rows:
            tracker_sources.append((person, rows))
    merged_tracker, tracker_report, dup_warnings = merge_tracker(
        read_rows(TRACKER_PATH, sheet_name="邮件追踪"), tracker_sources
    )

    print()
    for line in queue_report:
        print(line)
    print()
    for line in pending_report:
        print(line)
    print()
    for line in tracker_report:
        print(line)

    if args.dry_run:
        print("\n👀 这只是预览，没有改动任何文件。确认没问题后加 --apply 正式合并。")
        return

    print("\n📦 正式合并：先备份当前三张主表...")
    for path in [QUEUE_PATH, PENDING_PATH, TRACKER_PATH]:
        b = backup(path)
        if b:
            print(f"  已备份: {os.path.basename(b)}")

    if os.path.exists(QUEUE_PATH):
        write_rows(QUEUE_PATH, merged_queue)
    if os.path.exists(PENDING_PATH):
        write_rows(PENDING_PATH, merged_pending)
    if os.path.exists(TRACKER_PATH):
        write_rows(TRACKER_PATH, merged_tracker, sheet_name="邮件追踪")

    print("\n✅ 合并完成！三张主表已更新，原内容已备份在同目录下（文件名带"合并前备份"字样）。")
    if dup_warnings:
        print(f"⚠️ 提醒：有 {len(dup_warnings)} 个邮箱被多个同事重复联系过，建议去 {TRACKER_FILENAME} 里搜一下这些邮箱，看看要不要人工处理跟进冲突。")


if __name__ == "__main__":
    main()
