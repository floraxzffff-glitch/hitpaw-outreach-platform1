"""
VikPea_交付前自检.py — check files, dependencies, queue, tracker, and common risks.

This script does not send email and does not modify outreach data, except that it
creates missing config/blacklist templates if needed.
"""

import importlib.util
import os
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，请先安装: pip install openpyxl")
    sys.exit(1)

from VikPea_common import (
    CONFIG_PATH, BLACKLIST_PATH, PREVIEW_PATH, create_default_workbooks,
    classify_bad_email, load_blacklist, blacklist_reason,
)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
QUEUE_PATH = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")

Q_NAME = 1
Q_EMAIL = 2
Q_LINK = 5
Q_NOTE = 6
Q_TYPE = 7
Q_SOURCE = 8
Q_DEEP_STATUS = 9

T_EMAIL = 4
T_KIND = 7
T_REPLY = 8
T_STATUS = 11
T_SOURCE = 18


def has_module(name):
    return importlib.util.find_spec(name) is not None


def check_files():
    print("\n文件检查")
    for filename in [
        "VikPea_发信名单.xlsx",
        "VikPea_邮件开发追踪.xlsx",
        "VikPea_配置.xlsx",
        "VikPea_黑名单.xlsx",
        "VikPea_搜索关键词.xlsx",
        "VikPea_文章搜索关键词.xlsx",
    ]:
        path = os.path.join(SCRIPT_DIR, filename)
        status = "OK" if os.path.exists(path) else "缺失"
        print(f"  {status:>4}  {filename}")


def check_dependencies():
    print("\n依赖检查")
    for mod in ["openpyxl", "requests", "bs4", "yt_dlp"]:
        print(f"  {'OK' if has_module(mod) else '缺失':>4}  {mod}")


def check_queue():
    if not os.path.exists(QUEUE_PATH):
        return
    wb = openpyxl.load_workbook(QUEUE_PATH, data_only=True)
    ws = wb.active
    blacklist = load_blacklist()
    ready = no_email = suspicious = blocked = deep_pending = 0
    type_counter = Counter()
    source_missing = 0
    samples = []
    for row in ws.iter_rows(min_row=2):
        name = str(row[Q_NAME - 1].value or "").strip()
        email = str(row[Q_EMAIL - 1].value or "").strip()
        link = str(row[Q_LINK - 1].value or "").strip()
        etype = str(row[Q_TYPE - 1].value or "YouTube").strip() or "YouTube"
        source = str(row[Q_SOURCE - 1].value or "").strip() if len(row) >= Q_SOURCE else ""
        deep_status = str(row[Q_DEEP_STATUS - 1].value or "").strip() if len(row) >= Q_DEEP_STATUS else ""
        if not name and not email and not link:
            continue
        type_counter[etype] += 1
        if not source:
            source_missing += 1
        if "@" in email:
            ready += 1
            reason = classify_bad_email(email) or blacklist_reason(name, email, link, blacklist)
            if reason:
                suspicious += 1
                if len(samples) < 8:
                    samples.append((row[0].row, name, email, reason))
            if blacklist_reason(name, email, link, blacklist):
                blocked += 1
        else:
            no_email += 1
            if not deep_status:
                deep_pending += 1

    print("\n发信名单检查")
    print(f"  可发邮箱行: {ready}")
    print(f"  无邮箱行: {no_email}")
    print(f"  深度未处理无邮箱行: {deep_pending}")
    print(f"  可疑邮箱/黑名单命中: {suspicious}")
    print(f"  黑名单命中: {blocked}")
    print(f"  来源关键词为空: {source_missing}")
    print(f"  类型分布: {dict(type_counter.most_common())}")
    if samples:
        print("  可疑样例:")
        for item in samples:
            print(f"    行{item[0]} {item[1]} {item[2]} -> {item[3]}")


def check_tracker():
    if not os.path.exists(TRACKER_PATH):
        return
    wb = openpyxl.load_workbook(TRACKER_PATH, read_only=True, data_only=True)
    ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
    status = Counter()
    kind = Counter()
    replied = sent = no_source_sent = no_email_unfinished = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = str((row[T_EMAIL - 1] if len(row) >= T_EMAIL else "") or "").strip()
        kd = str((row[T_KIND - 1] if len(row) >= T_KIND else "") or "").strip()
        rep = str((row[T_REPLY - 1] if len(row) >= T_REPLY else "") or "").strip()
        st = str((row[T_STATUS - 1] if len(row) >= T_STATUS else "") or "").strip()
        src = str((row[T_SOURCE - 1] if len(row) >= T_SOURCE else "") or "").strip()
        status[st or "(空)"] += 1
        kind[kd or "(空)"] += 1
        if "@" in email and "未发送" not in kd and st != "未获取邮箱":
            sent += 1
            if not src:
                no_source_sent += 1
        if rep in {"已回复", "是", "Y", "Yes", "yes", "TRUE", "True"}:
            replied += 1
        if "@" not in email and st not in {"未获取邮箱", "未发送"}:
            no_email_unfinished += 1

    print("\n追踪表检查")
    print(f"  主表行数: {ws.max_row}")
    print(f"  已发送口径行: {sent}")
    print(f"  已回复行: {replied}")
    print(f"  已发送但来源关键词为空: {no_source_sent}")
    print(f"  无邮箱但状态不清晰: {no_email_unfinished}")
    print(f"  状态Top: {status.most_common(8)}")
    print(f"  邮件类型Top: {kind.most_common(8)}")


def main():
    create_default_workbooks()
    print("=" * 64)
    print("VikPea 交付前自检")
    print("=" * 64)
    check_files()
    check_dependencies()
    check_queue()
    check_tracker()
    print("\n完成。这个自检不会发送邮件。")
    print(f"配置表: {CONFIG_PATH}")
    print(f"黑名单: {BLACKLIST_PATH}")
    print(f"发信/跟进预览默认输出: {PREVIEW_PATH}")


if __name__ == "__main__":
    main()
