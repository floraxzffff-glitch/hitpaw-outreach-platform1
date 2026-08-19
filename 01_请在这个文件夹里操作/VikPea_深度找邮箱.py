"""
VikPea_深度找邮箱.py — 深度查找 YouTube KOL 联系邮箱

比普通查邮箱脚本多做几步：
  1. 从视频/频道元数据拿频道描述、视频描述、频道主页
  2. 从描述里抽取官网/Linktree/Beacons/Carrd/社媒等外链
  3. 访问外链页面及常见联系页：contact/about/work-with-me/sponsor/media-kit
  4. 多组搜索引擎 query 兜底
  5. 找到邮箱写回发信名单；找不到写清楚失败原因
  6. 默认只查新行；已经深度查过/体检清过的空邮箱行不重复查

用法:
  python3 ~/Downloads/VikPea工作台/VikPea_深度找邮箱.py
"""

import os
import re
import sys
import time
import json
import shutil
import subprocess
import urllib.parse
import urllib.request
import socket
from html import unescape
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from VikPea_common import (
        apply_config, log_event, NO_EMAIL_POOL_PATH, ensure_no_email_pool,
        ensure_queue_headers, append_send_queue_row,
        email_relevance_reason,
        PENDING_EMAIL_REVIEW_PATH, ensure_pending_email_review_workbook, add_pending_email_review_row,
        cleanup_pending_email_review_known, archive_processed_pending_email_rows, fix_glued_tld,
    )
except ImportError:
    apply_config = None
    log_event = None
    NO_EMAIL_POOL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_无邮箱候选.xlsx")
    ensure_no_email_pool = None
    ensure_queue_headers = None
    append_send_queue_row = None
    email_relevance_reason = None
    PENDING_EMAIL_REVIEW_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_待确认邮箱.xlsx")
    ensure_pending_email_review_workbook = None
    add_pending_email_review_row = None
    fix_glued_tld = lambda e: e
    cleanup_pending_email_review_known = None
    archive_processed_pending_email_rows = None


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
QUEUE_PATH = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")

Q_NAME = 1
Q_EMAIL = 2
Q_HOME = 5
Q_VIDEO = 6
Q_NOTE = 7
Q_TYPE = 8
Q_SOURCE = 9
Q_DEEP_STATUS = 12
Q_DEEP_DATE = 13

MAX_ROWS = 80         # 单次最多查多少条；0 = 全部
NEWEST_FIRST = True   # 优先查表格底部最新搜索出来的黄行
PRIORITIZE_EMAIL_LIKELY = True  # 优先查更像商业/工具/教程账号的行，提高邮箱命中率
MAX_LINKS_PER_ROW = 8
MAX_CONTACT_PAGES_PER_SITE = 8
DELAY_SEC = 1.5
RECHECK_PROCESSED = False  # 想重查旧的"未找到邮箱"行时，改成 True
CHECKED_LINKS_PATH = os.path.join(SCRIPT_DIR, ".deep_email_checked_links.json")

GREEN = PatternFill("solid", start_color="E2EFDA")
YELLOW = PatternFill("solid", start_color="FFF2CC")
ORANGE = PatternFill("solid", start_color="FCE4D6")
BLUE = PatternFill("solid", start_color="DDEBF7")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
URL_RE = re.compile(r"https?://[^\s<>\")'\]]+", re.I)
BARE_DOMAIN_RE = re.compile(
    r"\b(?:www\.)?[a-zA-Z0-9][a-zA-Z0-9\-]{1,62}"
    r"(?:\.[a-zA-Z0-9][a-zA-Z0-9\-]{1,62})+"
    r"(?:/[^\s<>\")'\]]*)?",
    re.I,
)

SKIP_EMAIL_DOMAINS = {
    "google", "youtube", "youtu", "sentry", "w3.org", "example", "schema",
    "cloudflare", "facebook", "twitter", "instagram", "tiktok", "random",
    "localhost", "invalid", "domain.com", "email.com", "yourdomain.com",
    "duckduckgo", "bing", "yahoo",
    "googleapis", "googleusercontent", "googlevideo", "gstatic", "ytimg",
    "window.wiz", "wiz", "doubleclick", "googlesyndication",
    "hitpaw", "tenorshare", "topazlabs", "topaz", "vanceai", "avclabs",
    "wondershare", "aiarty", "patreon", "invideo", "openshot", "envato", "pxf.io", "o.market",
}

LOW_VALUE_EMAILS = {
    "support@wix.com", "abuse@cloudflare.com", "privacy@youtube.com",
    "support@linktr.ee", "support@beacons.ai", "error-lite@duckduckgo.com",
    "support@notegpt.io",
}

BAD_TLDS = {"png", "jpg", "jpeg", "gif", "webp", "svg", "avif", "css", "js", "json", "xml"}
BAD_LOCAL_PARTS = {"www", "http", "https", "mailto", "image", "img", "src", "href"}
SUSPICIOUS_LOCAL_FRAGMENTS = {
    "www.", "1.env", "6-months-l", "tunacoretalentcre", "ai-category-card",
}

SKIP_URL_DOMAINS = {
    "youtube.com", "youtu.be", "google.com", "gstatic.com", "ggpht.com",
    "ytimg.com", "schema.org", "duckduckgo.com", "googleapis.com",
    "googleusercontent.com", "googlevideo.com", "w3.org", "window.wiz",
    "doubleclick.net", "googlesyndication.com",
}

BARE_DOMAIN_SKIP_WORDS = {
    "youtube.com", "youtu.be", "google.com", "gmail.com", "facebook.com",
    "instagram.com", "twitter.com", "x.com", "tiktok.com", "linkedin.com",
    "schema.org", "gstatic.com", "ytimg.com",
}

CONTACT_PATHS = [
    "", "/contact", "/contact-us", "/about", "/about-us", "/work-with-me",
    "/work-with-us", "/collab", "/collaboration", "/sponsor", "/sponsorship",
    "/advertise", "/advertising", "/media-kit", "/mediakit", "/press",
    "/business", "/partnership", "/partnerships",
]

PROCESSED_NOTE_MARKERS = [
    "深度找到",
    "深度未找到",
    "yt-dlp无公开邮箱",
    "已清除",
    "假邮箱",
    "可疑邮箱",
    "竞品/平台邮箱",
    "DuckDuckGo假邮箱",
]

PROCESSED_STATUS_MARKERS = [
    "已查",
    "已找到",
    "已清除",
    "跳过",
    "查询中",
    "中断",
]

PENDING_STATUS_MARKERS = [
    "待深度查邮箱",
    "新线索待查",
]

HIGH_EMAIL_SIGNAL_WORDS = {
    "ai", "tech", "tool", "tools", "software", "review", "tutorial", "how to",
    "video", "editor", "editing", "creator", "youtube", "android", "iphone",
    "mobile", "app", "quality", "enhancer", "upscale", "restore", "photo",
    "camera", "content", "media", "digital", "guide", "tips", "fix",
}

LOW_EMAIL_SIGNAL_WORDS = {
    "archive", "home improvement", "furniture", "yacht", "binaural", "meditation",
    "healing", "locksley", "fridge", "manuscript", "money pit",
}


def is_processed_missing_row(note):
    if RECHECK_PROCESSED:
        return False
    note = str(note or "")
    # 批量搜索新加的黄行会写"需手动查邮箱"，这不是已处理。
    # 只有深度查过、体检清过、或明确标记为假邮箱/竞品邮箱的旧行才跳过。
    return any(marker in note for marker in PROCESSED_NOTE_MARKERS)


def normalize_link(link):
    link = str(link or "").strip()
    link = link.split("&pp=")[0]
    return link.rstrip("/")


def load_checked_links():
    if not os.path.exists(CHECKED_LINKS_PATH):
        return {}
    try:
        with open(CHECKED_LINKS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_checked_links(data):
    with open(CHECKED_LINKS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def ensure_deep_headers(ws):
    if not ws.cell(1, Q_DEEP_STATUS).value:
        ws.cell(1, Q_DEEP_STATUS).value = "处理状态"
    if not ws.cell(1, Q_DEEP_DATE).value:
        ws.cell(1, Q_DEEP_DATE).value = "最后处理日期"


def infer_status_from_note(note: str, current_status: str = "") -> str:
    note = str(note or "")
    current_status = str(current_status or "").strip()
    if "深度找到" in note:
        return "已找到/后续清理"
    if "已清除" in note or "假邮箱" in note or "竞品/平台邮箱" in note:
        return "已清除"
    if "深度未找到" in note:
        return "已查无邮箱"
    if "需手动查" in note or "无公开邮箱" in note:
        return current_status or "待深度查邮箱"
    return current_status or "待深度查邮箱"


def append_found_to_queue(queue_ws, name, email, link, note, source, email_type, subject="", opening="", video_link=""):
    if append_send_queue_row:
        return append_send_queue_row(
            queue_ws, name, email, subject, opening, link,
            f"深度找到: {note}", email_type or "YouTube", source, GREEN,
            video_link=video_link, home_link=link,
        )
    r = queue_ws.max_row + 1
    queue_ws.cell(r, Q_NAME).value = name
    queue_ws.cell(r, Q_EMAIL).value = email
    queue_ws.cell(r, 3).value = subject
    queue_ws.cell(r, 4).value = opening
    queue_ws.cell(r, Q_HOME).value = link
    queue_ws.cell(r, Q_VIDEO).value = video_link
    queue_ws.cell(r, Q_NOTE).value = f"深度找到: {note}"
    queue_ws.cell(r, Q_TYPE).value = email_type or "YouTube"
    queue_ws.cell(r, Q_SOURCE).value = source
    for c in range(1, 10):
        queue_ws.cell(r, c).fill = GREEN
    return r


def add_pending_review(name, email, confidence, source_label, link, source, email_type, note, video_link=""):
    if add_pending_email_review_row:
        return add_pending_email_review_row(
            name, email, confidence, source_label, link, source, email_type, note, PENDING_EMAIL_REVIEW_PATH,
            video_link=video_link, home_link=link,
        )
    return ""


def auto_move_manually_filled_rows(ws, queue_ws, queue_path):
    """无邮箱候选表里如果有人工补了邮箱，自动转主表。"""
    moved = 0
    today = datetime.now().strftime("%Y-%m-%d")
    existing_queue_keys = set()
    for row in queue_ws.iter_rows(min_row=2, values_only=True):
        q_name = str((row[Q_NAME - 1] if len(row) >= Q_NAME else "") or "").strip().lower()
        q_link = str((row[Q_HOME - 1] if len(row) >= Q_HOME else "") or "").strip().rstrip("/").lower()
        q_email = str((row[Q_EMAIL - 1] if len(row) >= Q_EMAIL else "") or "").strip().lower()
        existing_queue_keys.add((q_name, q_link, q_email))

    for rownum in range(2, ws.max_row + 1):
        name = str(ws.cell(rownum, Q_NAME).value or "").strip()
        email = str(ws.cell(rownum, Q_EMAIL).value or "").strip()
        link = str(ws.cell(rownum, Q_HOME).value or "").strip()
        video_link = str(ws.cell(rownum, Q_VIDEO).value or "").strip() if ws.max_column >= Q_VIDEO else ""
        note = str(ws.cell(rownum, Q_NOTE).value or "").strip()
        source = str(ws.cell(rownum, Q_SOURCE).value or "").strip()
        email_type = str(ws.cell(rownum, Q_TYPE).value or "").strip()
        status = str(ws.cell(rownum, Q_DEEP_STATUS).value or "").strip()
        if not name or not email or "@" not in email:
            continue
        if status in {"已找到", "已转主表", "已发送"}:
            continue
        key = (name.lower(), link.rstrip("/").lower(), email.lower())
        if key in existing_queue_keys:
            ws.cell(rownum, Q_DEEP_STATUS).value = "已转主表"
            ws.cell(rownum, Q_DEEP_DATE).value = today
            continue
        subject = str(ws.cell(rownum, 3).value or "").strip()
        opening = str(ws.cell(rownum, 4).value or "").strip()
        append_found_to_queue(queue_ws, name, email, link, "无邮箱候选人工补录", source, email_type, subject, opening, video_link=video_link)
        ws.cell(rownum, Q_NOTE).value = (note + " | " if note else "") + "无邮箱候选里的补录邮箱已自动转入主发信表"
        ws.cell(rownum, Q_DEEP_STATUS).value = "已转主表"
        ws.cell(rownum, Q_DEEP_DATE).value = today
        moved += 1
        existing_queue_keys.add(key)
    return moved


def is_processed_status(status):
    if RECHECK_PROCESSED:
        return False
    status = str(status or "")
    return any(marker in status for marker in PROCESSED_STATUS_MARKERS)


def email_likelihood_score(name, note, source):
    text = f"{name or ''} {note or ''} {source or ''}".lower()
    score = 0
    reasons = []
    for word in HIGH_EMAIL_SIGNAL_WORDS:
        if word in text:
            score += 1
            if len(reasons) < 4:
                reasons.append(word)
    for word in LOW_EMAIL_SIGNAL_WORDS:
        if word in text:
            score -= 2
            if len(reasons) < 4:
                reasons.append(f"-{word}")
    m = re.search(r"(\d+)\s*K粉", str(note or ""), flags=re.I)
    if m:
        subs_k = int(m.group(1))
        if 5 <= subs_k <= 250:
            score += 1
        elif subs_k < 2:
            score -= 1
    if "小博主信号" in str(note or ""):
        score += 1
    return score, ", ".join(reasons) or "普通"


def migrate_existing_processed_rows(ws, checked_links):
    """把旧备注里的深度查找结果迁移到独立状态列和链接去重文件。"""
    changed = False
    today = datetime.now().strftime("%Y-%m-%d")
    for row in ws.iter_rows(min_row=2):
        link = normalize_link(row[Q_HOME - 1].value)
        note = str(row[Q_NOTE - 1].value or "")
        status = str(row[Q_DEEP_STATUS - 1].value or "") if len(row) >= Q_DEEP_STATUS else ""
        if not link:
            continue
        inferred = infer_status_from_note(note, status)
        if inferred != status:
            row[Q_DEEP_STATUS - 1].value = inferred
            changed = True
        if inferred and not row[Q_DEEP_DATE - 1].value and inferred not in PENDING_STATUS_MARKERS:
            row[Q_DEEP_DATE - 1].value = today
            changed = True
        status = str(row[Q_DEEP_STATUS - 1].value or "")
        # 只有明确处理过的行才写入历史链接记录，避免误伤待查新线索。
        if status and status not in PENDING_STATUS_MARKERS:
            checked_links[link] = {"status": status, "date": str(row[Q_DEEP_DATE - 1].value or today)}
    return changed


def find_ytdlp():
    try:
        subprocess.run([sys.executable, "-m", "yt_dlp", "--version"],
                       capture_output=True, check=True)
        return [sys.executable, "-m", "yt_dlp"]
    except Exception:
        pass
    p = shutil.which("yt-dlp")
    if p:
        return [p]
    ver = f"{sys.version_info.major}.{sys.version_info.minor}"
    user_bin = os.path.expanduser(f"~/Library/Python/{ver}/bin/yt-dlp")
    if os.path.exists(user_bin):
        return [user_bin]
    return []


YTDLP_CMD = find_ytdlp()


def normalize_obfuscated(text):
    text = text or ""
    text = re.sub(r"\s*(?:\[at\]|\(at\)|\{at\}|\bat\b)\s*", "@", text, flags=re.I)
    text = re.sub(r"\s*(?:\[dot\]|\(dot\)|\{dot\}|\bdot\b)\s*", ".", text, flags=re.I)
    return text


def clean_emails(raw):
    seen, out = set(), []
    for email in raw:
        e = fix_glued_tld(email.strip(" .,:;()[]<>\"'").lower())
        if any(ch in e for ch in {'"', "'", " "}) or re.search(r"%(?:22|20|2b|40|2e|2d|5f|[0-9a-f]{2})", e, re.I):
            continue
        if re.search(r"\.(png|jpg|jpeg|gif|mp4|webp|svg|css|js)$", e, re.I):
            continue
        local, _, domain = e.partition("@")
        parts = domain.split(".")
        tld = parts[-1] if parts else ""
        main = parts[-2] if len(parts) >= 2 else ""
        if tld in BAD_TLDS:
            continue
        if local in BAD_LOCAL_PARTS or any(fragment in local for fragment in SUSPICIOUS_LOCAL_FRAGMENTS):
            continue
        if len(domain) > 80 or len(local) > 64:
            continue
        if re.search(r"\d+x\.", domain) or "compressed" in domain:
            continue
        if len(local) < 2 or len(main) < 2 or len(tld) < 2:
            continue
        if e in LOW_VALUE_EMAILS:
            continue
        if tld in SKIP_EMAIL_DOMAINS or any(skip in domain for skip in SKIP_EMAIL_DOMAINS):
            continue
        if e not in seen:
            seen.add(e)
            out.append(e)
    return out


def extract_emails(text, allow_obfuscated=True):
    text = unescape(text or "")
    candidates = EMAIL_RE.findall(text)
    if allow_obfuscated and "<html" not in text.lower():
        candidates += EMAIL_RE.findall(normalize_obfuscated(text))
    return clean_emails(candidates)


def domain_of(url):
    try:
        return urllib.parse.urlparse(url).netloc.lower().replace("www.", "")
    except Exception:
        return ""


def can_resolve_host(host: str) -> bool:
    try:
        socket.getaddrinfo(host, 443)
        return True
    except Exception:
        return False


def clean_url(url, base_url=""):
    if not url:
        return ""
    url = unescape(url).strip().strip(".,;()[]{}<>\"'")
    url = decode_search_redirect(url)
    if url.startswith("//"):
        url = "https:" + url
    if base_url and url.startswith("/"):
        url = urllib.parse.urljoin(base_url, url)
    url = decode_search_redirect(url)
    if not url.startswith(("http://", "https://")):
        return ""
    return url


def decode_search_redirect(url):
    """还原 Google/DuckDuckGo/Bing 搜索结果里的真实目标链接。"""
    url = unescape(str(url or ""))
    try:
        parsed = urllib.parse.urlparse(url)
        qs = urllib.parse.parse_qs(parsed.query)
        for key in ("uddg", "q", "url", "u"):
            if key in qs and qs[key]:
                candidate = urllib.parse.unquote(qs[key][0])
                if candidate.startswith(("http://", "https://")):
                    return candidate
        if "uddg=" in url:
            m = re.search(r"uddg=([^&]+)", url)
            if m:
                candidate = urllib.parse.unquote(m.group(1))
                if candidate.startswith(("http://", "https://")):
                    return candidate
    except Exception:
        pass
    return url


def extract_urls_from_text(text):
    urls = URL_RE.findall(text or "")
    for raw in BARE_DOMAIN_RE.findall(text or ""):
        domain = domain_of(raw if raw.startswith("http") else "https://" + raw)
        if not domain or domain in BARE_DOMAIN_SKIP_WORDS:
            continue
        if any(domain.endswith("." + item) or domain == item for item in BARE_DOMAIN_SKIP_WORDS):
            continue
        if "." not in domain:
            continue
        urls.append(raw if raw.startswith("http") else "https://" + raw)
    out, seen = [], set()
    for url in urls:
        url = clean_url(url)
        if url and url not in seen:
            seen.add(url)
            out.append(url)
    return out


def should_visit_url(url):
    d = domain_of(url)
    if not d:
        return False
    if any(skip in d for skip in SKIP_URL_DOMAINS):
        return False
    return True


def fetch_url(url, timeout=15):
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            ctype = resp.headers.get("Content-Type", "")
            if "text" not in ctype and "html" not in ctype and "json" not in ctype:
                return ""
            raw = resp.read(700_000)
            return raw.decode("utf-8", errors="replace")
    except Exception:
        return ""


def yt_dlp_json(url):
    if not YTDLP_CMD:
        return {}
    try:
        result = subprocess.run(
            YTDLP_CMD + ["--skip-download", "--no-warnings", "--dump-single-json", url],
            capture_output=True, text=True, timeout=45
        )
        if not result.stdout.strip():
            return {}
        return json.loads(result.stdout)
    except Exception:
        return {}


def yt_dlp_field(url, field):
    if not YTDLP_CMD:
        return ""
    try:
        result = subprocess.run(
            YTDLP_CMD + ["--skip-download", "--no-warnings", "--print", field, url],
            capture_output=True, text=True, timeout=30
        )
        return result.stdout.strip()
    except Exception:
        return ""


def build_channel_about_url(channel_url: str) -> str:
    url = str(channel_url or "").strip().rstrip("/")
    if not url:
        return ""
    url = re.sub(r"[?#].*$", "", url)
    url = re.sub(r"/(videos|shorts|streams|featured|playlists|community|about)$", "", url, flags=re.I)
    return url + "/about"


def normalize_channel_root_url(channel_url: str) -> str:
    url = str(channel_url or "").strip().rstrip("/")
    if not url:
        return ""
    url = re.sub(r"[?#].*$", "", url)
    url = re.sub(r"/(videos|shorts|streams|featured|playlists|community|about)$", "", url, flags=re.I)
    return url


def collect_youtube_text_and_links(target_url):
    texts, links, channel_url = [], [], ""
    info = yt_dlp_json(target_url)
    if info:
        for key in ["description", "channel_description", "title", "channel", "uploader"]:
            val = info.get(key)
            if val:
                texts.append(str(val))
        channel_url = str(info.get("channel_url") or "")
        for key in ["webpage_url", "original_url", "channel_url", "uploader_url"]:
            val = info.get(key)
            if val:
                links.append(str(val))
        for item in info.get("thumbnails") or []:
            pass

    if not channel_url:
        channel_url = yt_dlp_field(target_url, "%(channel_url)s")
    if channel_url:
        links.append(channel_url)
        root_url = normalize_channel_root_url(channel_url)
        if root_url:
            root_html = fetch_url(root_url, timeout=15)
            if root_html:
                texts.append(root_html)
                links.append(root_url)
                links.extend(extract_links_from_html(root_html, root_url))
        for field in ["%(description)s", "%(channel_description)s"]:
            txt = yt_dlp_field(channel_url, field)
            if txt:
                texts.append(txt)
        about_url = build_channel_about_url(channel_url)
        if about_url:
            about_html = fetch_url(about_url, timeout=15)
            if about_html:
                texts.append(about_html)
                links.append(about_url)
                links.extend(extract_links_from_html(about_html, about_url))

    for field in ["%(description)s", "%(channel_description)s"]:
        txt = yt_dlp_field(target_url, field)
        if txt:
            texts.append(txt)

    for text in texts:
        links.extend(extract_urls_from_text(text))

    unique_links = []
    seen = set()
    for url in links:
        url = clean_url(url)
        if url and url not in seen:
            seen.add(url)
            unique_links.append(url)
    return "\n".join(texts), unique_links, channel_url


def pick_relevant_email(emails, name, channel_url, source_label, text="", page_url=""):
    for email in emails:
        reason = ""
        if email_relevance_reason:
            reason = email_relevance_reason(
                email,
                channel_name=name,
                channel_url=channel_url,
                source_label=source_label,
                text=text,
                page_url=page_url,
            )
        if not reason:
            return email
    return ""


def extract_links_from_html(html, base_url):
    urls = extract_urls_from_text(html or "")
    hrefs = re.findall(r'href=["\']([^"\']+)["\']', html or "", flags=re.I)
    urls.extend(hrefs)
    out, seen = [], set()
    for url in urls:
        url = clean_url(url, base_url)
        if url and should_visit_url(url) and url not in seen:
            seen.add(url)
            out.append(url)
    return out


def expand_contact_pages(url):
    parsed = urllib.parse.urlparse(url)
    root = f"{parsed.scheme}://{parsed.netloc}"
    candidates = []
    for path in CONTACT_PATHS:
        candidates.append(root + path)
    if parsed.path and parsed.path != "/":
        candidates.insert(0, url)
    out, seen = [], set()
    for item in candidates:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out[:MAX_CONTACT_PAGES_PER_SITE]


def find_email_on_external_links(seed_links, name="", channel_url=""):
    notes = []
    visit_queue = []
    for link in seed_links:
        if should_visit_url(link):
            visit_queue.append(link)
    visit_queue = visit_queue[:MAX_LINKS_PER_ROW]

    visited = set()
    for link in visit_queue:
        for page in expand_contact_pages(link):
            if page in visited:
                continue
            visited.add(page)
            html = fetch_url(page)
            if not html:
                continue
            emails = extract_emails(html, allow_obfuscated=False)
            email = pick_relevant_email(emails, name, channel_url or link, "外链页面", html, page)
            if email:
                return email, f"外链页面: {page}", "B"
            more_links = extract_links_from_html(html, page)
            # Linktree/Beacons/Carrd often need one extra hop.
            if any(host in domain_of(link) for host in ["linktr.ee", "beacons.ai", "carrd.co", "bio.link", "solo.to"]):
                for more in more_links[:6]:
                    if more in visited:
                        continue
                    visited.add(more)
                    sub_html = fetch_url(more)
                    emails = extract_emails(sub_html, allow_obfuscated=False)
                    email = pick_relevant_email(emails, name, channel_url or link, "聚合页外链", sub_html, more)
                    if email:
                        return email, f"聚合页外链: {more}", "B"
        notes.append(domain_of(link))
    if visit_queue:
        return "", "外链无邮箱: " + ", ".join(notes[:5]), ""
    return "", "无可访问外链", ""


def search_engine_queries(name, channel_url):
    handle = ""
    m = re.search(r"/@([^/?#]+)", channel_url or "")
    if m:
        handle = m.group(1)
    base = [name]
    if handle:
        base.append(handle)
    queries = []
    for term in base:
        term = term.strip()
        if not term:
            continue
        queries.extend([
            f'"{term}" "business email"',
            f'"{term}" "sponsorship"',
            f'"{term}" "collab"',
            f'"{term}" "collaboration"',
            f'"{term}" "media kit"',
            f'"{term}" "work with me"',
            f'"{term}" "brand deals"',
            f'"{term}" "for business inquiries"',
            f'"{term}" "@gmail.com"',
            f'"{term}" "@outlook.com"',
            f'"{term}" "contact"',
        ])
    return queries


def search_email(name, channel_url):
    for query in search_engine_queries(name, channel_url)[:14]:
        for engine_url in [
            "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&num=8",
            "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(query),
            "https://www.bing.com/search?q=" + urllib.parse.quote(query),
        ]:
            html = fetch_url(engine_url, timeout=15)
            direct_emails = extract_emails(html, allow_obfuscated=False)
            email = pick_relevant_email(direct_emails, name, channel_url, "搜索引擎", html, engine_url)
            if email:
                return email, f"搜索结果摘要: {query}", "C"
            result_links = extract_links_from_html(html, engine_url)
            # 去掉搜索引擎/缓存/图片等无效链接，并限制同域重复。
            filtered_links = []
            seen_domains = set()
            for result_link in result_links:
                d = domain_of(result_link)
                if not d or any(skip in d for skip in SKIP_URL_DOMAINS):
                    continue
                if d in seen_domains:
                    continue
                seen_domains.add(d)
                filtered_links.append(result_link)
            for result_link in filtered_links[:8]:
                page = fetch_url(result_link, timeout=15)
                emails = extract_emails(page, allow_obfuscated=False)
                email = pick_relevant_email(emails, name, channel_url, "搜索引擎", page, result_link)
                if email:
                    return email, f"搜索结果页面: {result_link}", "C"
            time.sleep(0.4)
    return "", "搜索无结果", ""


def deep_find_email(name, target_url):
    text, links, channel_url = collect_youtube_text_and_links(target_url)
    emails = extract_emails(text, allow_obfuscated=True)
    email = pick_relevant_email(emails, name, channel_url or target_url, "YouTube描述/视频描述", text, channel_url or target_url)
    if email:
        return email, "YouTube描述/视频描述", "A", "YouTube描述/视频描述"

    email, note, confidence = find_email_on_external_links(links, name, channel_url or target_url)
    if email:
        return email, note, confidence or "B", "外链页面"

    email, search_note, confidence = search_email(name, channel_url or target_url)
    if email:
        return email, search_note, confidence or "C", "搜索引擎"

    details = []
    if note:
        details.append(note)
    if search_note:
        details.append(search_note)
    return "", "；".join(details) or "无公开邮箱", "", ""


def main():
    if apply_config:
        apply_config(globals(), {
            "DEEP_EMAIL_MAX_ROWS": "MAX_ROWS",
        })
    if not YTDLP_CMD:
        print("❌ 未找到 yt-dlp，请先安装 yt-dlp")
        sys.exit(1)

    dns_ok = can_resolve_host("www.google.com") or can_resolve_host("www.youtube.com")
    if not dns_ok:
        print("⚠️ 当前网络/DNS 解析异常：网页搜索和外链抓取可能失败。")
        print("↳ 脚本仍会继续处理无邮箱候选里的补录邮箱和本地表格状态，但网页查找成功率会很低。")

    if ensure_no_email_pool:
        wb, ws = ensure_no_email_pool(NO_EMAIL_POOL_PATH)
    else:
        if not os.path.exists(NO_EMAIL_POOL_PATH):
            print(f"❌ 找不到无邮箱候选表: {NO_EMAIL_POOL_PATH}")
            sys.exit(1)
        wb = openpyxl.load_workbook(NO_EMAIL_POOL_PATH)
        ws = wb.active
    ensure_deep_headers(ws)
    if ensure_pending_email_review_workbook:
        ensure_pending_email_review_workbook(PENDING_EMAIL_REVIEW_PATH)
    if cleanup_pending_email_review_known:
        cleaned_pending = cleanup_pending_email_review_known(PENDING_EMAIL_REVIEW_PATH)
        if cleaned_pending:
            print(f"🧹 已清理待确认邮箱表中 {cleaned_pending} 条已联系/已入队残留")
    if archive_processed_pending_email_rows:
        archived_pending = archive_processed_pending_email_rows(PENDING_EMAIL_REVIEW_PATH)
        if archived_pending:
            print(f"📦 已归档待确认邮箱表中 {archived_pending} 条已处理记录")
    checked_links = load_checked_links()
    if migrate_existing_processed_rows(ws, checked_links):
        wb.save(NO_EMAIL_POOL_PATH)
    save_checked_links(checked_links)

    if not os.path.exists(QUEUE_PATH):
        print(f"❌ 找不到发信名单: {QUEUE_PATH}")
        sys.exit(1)
    wb_q = openpyxl.load_workbook(QUEUE_PATH)
    ws_q = wb_q.active
    if ensure_queue_headers:
        ensure_queue_headers(ws_q)

    moved_manual = auto_move_manually_filled_rows(ws, ws_q, QUEUE_PATH)
    if moved_manual:
        wb_q.save(QUEUE_PATH)
        wb.save(NO_EMAIL_POOL_PATH)
        print(f"📥 检测到无邮箱候选里已补录的邮箱，已自动转入主发信表 {moved_manual} 条")

    candidates = []
    missing_email_rows = 0
    skipped_processed = 0
    skipped_checked_link = 0
    for row in ws.iter_rows(min_row=2):
        name = str(row[Q_NAME - 1].value or "").strip()
        email = str(row[Q_EMAIL - 1].value or "").strip()
        link = str(row[Q_HOME - 1].value or "").strip()
        video_link = str(row[Q_VIDEO - 1].value or "").strip() if len(row) >= Q_VIDEO else ""
        note = str(row[Q_NOTE - 1].value or "").strip()
        email_type = str(row[Q_TYPE - 1].value or "").strip() if len(row) >= Q_TYPE else ""
        source = str(row[Q_SOURCE - 1].value or "").strip() if len(row) >= Q_SOURCE else ""
        status = str(row[Q_DEEP_STATUS - 1].value or "").strip()
        link_key = normalize_link(link)
        if name and link and "@" not in email:
            missing_email_rows += 1
            if is_processed_status(status) or is_processed_missing_row(note):
                skipped_processed += 1
                continue
            checked = checked_links.get(link_key, {})
            checked_status = str(checked.get("status") or "")
            if (
                not RECHECK_PROCESSED
                and not status
                and checked_status
                and checked_status not in PENDING_STATUS_MARKERS
            ):
                skipped_checked_link += 1
                continue
            score, reason = email_likelihood_score(name, note, source)
            candidates.append((row[0].row, name, link, score, reason, source, email_type))

    if PRIORITIZE_EMAIL_LIKELY:
        candidates.sort(key=lambda item: (-item[3], -item[0] if NEWEST_FIRST else item[0], item[1].lower()))
    else:
        candidates.sort(key=lambda item: item[0], reverse=NEWEST_FIRST)
    targets = candidates[:MAX_ROWS] if MAX_ROWS else candidates

    print(
        f"无邮箱候选 {missing_email_rows} 条；"
        f"已处理旧行跳过 {skipped_processed} 条；"
        f"链接记录跳过 {skipped_checked_link} 条；"
        f"待查未标记 {len(candidates)} 条；"
        f"本次实际查 {len(targets)} 条"
    )
    if MAX_ROWS and len(candidates) > len(targets):
        direction = "高邮箱概率优先" if PRIORITIZE_EMAIL_LIKELY else ("最新行优先" if NEWEST_FIRST else "旧行优先")
        print(f"提示：为避免一次跑太久，当前每次最多查 {MAX_ROWS} 条（{direction}）。")
    if skipped_processed and not RECHECK_PROCESSED:
        print("提示：如需重查旧失败行，把脚本里的 RECHECK_PROCESSED 改成 True。")

    if not targets:
        print("✅ 没有需要深度查邮箱的行")
        return

    print(f"\n🔎 深度查邮箱：{len(targets)} 条\n{'─'*70}")
    found = 0
    pending_review = 0
    missed = 0
    found_rows_to_delete = []
    for idx, (rownum, name, link, score, score_reason, source, email_type) in enumerate(targets, 1):
        print(f"\n[{idx}/{len(targets)}] {name} | {link}")
        print(f"  目标优先级: {score} ({score_reason})" + (f" | 来源: {source}" if source else ""))
        now_str = datetime.now().strftime("%Y-%m-%d")
        ws.cell(rownum, Q_DEEP_STATUS).value = "查询中"
        ws.cell(rownum, Q_DEEP_DATE).value = now_str
        ws.cell(rownum, Q_NOTE).value = "深度查询中：如脚本中断且需要重查，请清空本行深度状态"
        checked_links[normalize_link(link)] = {
            "status": "查询中",
            "date": now_str,
            "name": name,
        }
        save_checked_links(checked_links)
        wb.save(NO_EMAIL_POOL_PATH)

        email, note, confidence, source_label = deep_find_email(name, link)
        if email:
            subject = str(ws.cell(rownum, 3).value or "").strip()
            opening = str(ws.cell(rownum, 4).value or "").strip()
            ws.cell(rownum, Q_EMAIL).value = email
            ws.cell(rownum, Q_DEEP_DATE).value = datetime.now().strftime("%Y-%m-%d")
            if confidence == "A":
                append_found_to_queue(ws_q, name, email, link, note, source, email_type, subject, opening, video_link=video_link)
                wb_q.save(QUEUE_PATH)
                ws.cell(rownum, Q_NOTE).value = f"深度找到并已转入主发信表: {note} | 置信度A"
                ws.cell(rownum, Q_DEEP_STATUS).value = "已找到-A已转主表"
                for c in range(1, ws.max_column + 1):
                    ws.cell(rownum, c).fill = GREEN
                found_rows_to_delete.append(rownum)
                found += 1
                print(f"  ✅ {email} ({note}) | 置信度A → 已转入主发信表")
            else:
                pending_result = add_pending_review(name, email, confidence or "B", source_label or "候选邮箱", link, source, email_type, note, video_link=video_link)
                if str(pending_result).startswith("SKIP:"):
                    ws.cell(rownum, Q_NOTE).value = f"深度找到候选邮箱但已联系过，跳过入池: {email} | {str(pending_result)[5:]}"
                    ws.cell(rownum, Q_DEEP_STATUS).value = "已发送"
                    ws.cell(rownum, Q_DEEP_DATE).value = datetime.now().strftime("%Y-%m-%d")
                    print(f"  ↳ 跳过：{email} 已联系过，不放入待确认邮箱")
                    continue
                ws.cell(rownum, Q_NOTE).value = f"深度找到候选邮箱，待人工确认: {email} | {note} | 置信度{confidence or 'B'}"
                ws.cell(rownum, Q_DEEP_STATUS).value = f"待人工确认-{confidence or 'B'}"
                for c in range(1, ws.max_column + 1):
                    ws.cell(rownum, c).fill = BLUE
                pending_review += 1
                print(f"  🔵 {email} ({note}) | 置信度{confidence or 'B'} → 已放入待确认邮箱表")
        else:
            ws.cell(rownum, Q_NOTE).value = f"深度未找到: {note}"
            ws.cell(rownum, Q_DEEP_STATUS).value = "已查无邮箱"
            ws.cell(rownum, Q_DEEP_DATE).value = datetime.now().strftime("%Y-%m-%d")
            for c in range(1, min(ws.max_column, Q_DEEP_DATE) + 1):
                ws.cell(rownum, c).fill = ORANGE
            missed += 1
            print(f"  🔶 {note}")
        checked_links[normalize_link(link)] = {
            "status": str(ws.cell(rownum, Q_DEEP_STATUS).value or ""),
            "date": str(ws.cell(rownum, Q_DEEP_DATE).value or ""),
            "name": name,
        }
        save_checked_links(checked_links)
        wb.save(NO_EMAIL_POOL_PATH)
        time.sleep(DELAY_SEC)

    for rownum in sorted(found_rows_to_delete, reverse=True):
        ws.delete_rows(rownum)
    if found_rows_to_delete:
        wb.save(NO_EMAIL_POOL_PATH)

    print(f"\n{'─'*70}")
    print(f"完成：✅ A级转主表 {found}  🔵 待确认 {pending_review}  🔶 未找到 {missed}")
    print("A级邮箱已转入主发信表；B/C级邮箱已进入待确认邮箱表；未找到的行留在无邮箱候选表并写明失败原因。")
    if log_event:
        log_event("深度找邮箱", f"本次查 {len(targets)} 条，A级 {found}，待确认 {pending_review}，未找到 {missed}")


if __name__ == "__main__":
    main()
