"""
VikPea_关键词聚类.py — 把文章/SEO关键词按主题粗聚类。

不需要额外 AI 模型，适合给同事先整理关键词池。
"""

import os
import re
import sys
from collections import defaultdict
from copy import copy
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError as exc:
    print(f"缺少依赖: {exc}")
    print("请先运行安装依赖，或执行: pip install openpyxl")
    sys.exit(1)

try:
    from VikPea_common import ARTICLE_KEYWORD_PATH, save_workbook_safe, set_column_widths
except ImportError:
    ARTICLE_KEYWORD_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_文章搜索关键词.xlsx")
    save_workbook_safe = None
    set_column_widths = None


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_关键词聚类.xlsx")

STOPWORDS = {
    "a", "an", "the", "to", "for", "of", "in", "on", "with", "and", "or",
    "best", "top", "free", "online", "software", "tool", "tools", "app", "apps",
    "review", "reviews", "comparison", "alternative", "alternatives", "2025", "2026",
}

THEME_HINTS = [
    ("video quality", {"video", "quality", "enhance", "enhancer", "improve"}),
    ("upscaling / 4K", {"upscale", "upscaler", "upscaling", "4k", "hd"}),
    ("restoration / repair", {"restore", "restoration", "repair", "old", "corrupted"}),
    ("denoise / sharpen", {"noise", "denoise", "sharpen", "sharpener", "blurry", "blur"}),
    ("AI video generator", {"generator", "generate", "sora", "veo", "runway", "midjourney"}),
    ("competitor / alternatives", {"topaz", "vanceai", "avclabs", "aiarty", "unifab", "alternative"}),
    ("tutorial / how-to", {"how", "tutorial", "guide", "fix"}),
]


def normalize_token(token):
    token = token.lower().strip()
    for suffix in ["ing", "ers", "er", "ed", "s"]:
        if len(token) > 5 and token.endswith(suffix):
            return token[: -len(suffix)]
    return token


def keyword_tokens(keyword):
    tokens = re.findall(r"[a-zA-Z0-9]{2,}", str(keyword or "").lower())
    return {normalize_token(t) for t in tokens if t not in STOPWORDS}


def load_keywords(path):
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        keyword = str((row[0] if row else "") or "").strip()
        enabled = str((row[1] if len(row) > 1 else "是") or "是").strip()
        note = str((row[2] if len(row) > 2 else "") or "").strip()
        if keyword and enabled not in {"否", "N", "n", "No", "no", "0"}:
            rows.append((keyword, note, keyword_tokens(keyword)))
    return rows


def theme_for(tokens):
    best_name = "其他"
    best_hit = 0
    for name, hints in THEME_HINTS:
        hit = len(tokens & hints)
        if hit > best_hit:
            best_name = name
            best_hit = hit
    return best_name


def cluster_keywords(rows):
    groups = defaultdict(list)
    for keyword, note, tokens in rows:
        theme = theme_for(tokens)
        core = tuple(sorted(list(tokens))[:4])
        key = (theme, core[:2] if theme == "其他" else theme)
        groups[key].append((keyword, note, tokens))

    merged = []
    for key, items in groups.items():
        token_count = defaultdict(int)
        for _keyword, _note, tokens in items:
            for token in tokens:
                token_count[token] += 1
        representative = max(items, key=lambda item: len(item[2]))[0]
        top_tokens = sorted(token_count, key=lambda t: (-token_count[t], t))[:8]
        theme = key[0]
        merged.append((theme, representative, len(items), ", ".join(top_tokens), items))
    return sorted(merged, key=lambda row: (-row[2], row[0], row[1]))


def main():
    print("\n" + "=" * 64)
    print("  VikPea 关键词聚类")
    print("=" * 64)
    rows = load_keywords(ARTICLE_KEYWORD_PATH)
    if not rows:
        print(f"没有可聚类的启用关键词: {ARTICLE_KEYWORD_PATH}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "关键词聚类"
    ws.append(["主题簇", "代表关键词", "关键词数", "核心词", "全部关键词", "生成日期"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    today = datetime.now().strftime("%Y-%m-%d")
    for theme, representative, count, top_tokens, items in cluster_keywords(rows):
        ws.append([
            theme,
            representative,
            count,
            top_tokens,
            "\n".join(item[0] for item in items),
            today,
        ])

    if set_column_widths:
        set_column_widths(ws, {"A": 24, "B": 46, "C": 12, "D": 44, "E": 78, "F": 14})
    for row in ws.iter_rows(min_row=2):
        alignment = copy(row[4].alignment)
        alignment.wrap_text = True
        row[4].alignment = alignment
    if save_workbook_safe:
        save_workbook_safe(wb, OUTPUT_PATH, "关键词聚类")
    else:
        wb.save(OUTPUT_PATH)
    print(f"完成：{len(rows)} 个关键词 -> {ws.max_row - 1} 个主题簇")
    print(f"输出文件: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
