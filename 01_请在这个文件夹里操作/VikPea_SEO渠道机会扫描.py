"""
VikPea_SEO渠道机会扫描.py — 从关键词搜索结果里筛选值得开发的 SEO 渠道机会。

它只生成分析表，不写入发信名单，也不会发送邮件。
"""

import os
import sys
import time
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError as exc:
    print(f"缺少依赖: {exc}")
    print("请先运行安装依赖，或执行: pip install openpyxl requests beautifulsoup4")
    sys.exit(1)

try:
    from VikPea_common import apply_config, set_column_widths, save_workbook_safe
except ImportError:
    apply_config = None
    set_column_widths = None
    save_workbook_safe = None

try:
    import VikPea_文章批量搜索 as article_search
    from VikPea_文章批量搜索 import (
        DELAY_SEARCH,
        load_queries,
        search_web,
        fetch,
        get_base_domain,
        root_of_domain,
        should_skip,
        score_article_candidate,
    )
except ImportError as exc:
    print(f"无法加载文章搜索模块: {exc}")
    sys.exit(1)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_SEO渠道机会扫描.xlsx")

SEO_OPPORTUNITY_RESULTS_PER_KEYWORD = 30
SEO_OPPORTUNITY_MIN_SCORE = 3

GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="F4CCCC")
BLUE = PatternFill("solid", fgColor="D9EAF7")


def classify_opportunity(score, title, url):
    text = f"{title or ''} {url or ''}".lower()
    if score >= 5:
        level = "A"
        action = "优先开发：评测/榜单/对比页，适合谈插链或工具收录"
    elif score >= SEO_OPPORTUNITY_MIN_SCORE:
        level = "B"
        action = "可复核：有一定相关性，建议人工看页面质量后再联系"
    else:
        level = "C"
        action = "暂不建议：页面相关性或可开发性偏弱"

    opportunity_type = "普通SEO机会"
    if any(word in text for word in ["best", "top", "rank", "tools", "software"]):
        opportunity_type = "榜单/工具合集"
    if any(word in text for word in ["review", "vs", "comparison", "alternative", "alternatives"]):
        opportunity_type = "评测/对比/替代品"
    if any(word in text for word in ["how to", "tutorial", "guide"]):
        opportunity_type = "教程/指南"
    return level, opportunity_type, action


def ensure_output_workbook(path):
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SEO机会扫描"

    headers = [
        "关键词", "排名", "机会等级", "机会类型", "站点分", "评分原因",
        "标题", "URL", "根域名", "建议动作", "搜索来源", "扫描日期",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = BLUE
    if set_column_widths:
        set_column_widths(ws, {
            "A": 34, "B": 10, "C": 12, "D": 18, "E": 10, "F": 42,
            "G": 54, "H": 70, "I": 28, "J": 48, "K": 28, "L": 14,
        })
    return wb, ws


def main():
    if apply_config:
        apply_config(globals(), {
            "SEO_OPPORTUNITY_RESULTS_PER_KEYWORD": "SEO_OPPORTUNITY_RESULTS_PER_KEYWORD",
            "SEO_OPPORTUNITY_MIN_SCORE": "SEO_OPPORTUNITY_MIN_SCORE",
            "RESPECT_ROBOTS_TXT": "RESPECT_ROBOTS_TXT",
            "CRAWL_DELAY_PER_DOMAIN": "CRAWL_DELAY_PER_DOMAIN",
            "SERP_PROVIDER": "SERP_PROVIDER",
            "SERPER_API_KEY": "SERPER_API_KEY",
            "SERPAPI_KEY": "SERPAPI_KEY",
            "DATAFORSEO_LOGIN": "DATAFORSEO_LOGIN",
            "DATAFORSEO_PASSWORD": "DATAFORSEO_PASSWORD",
        })
        for key in [
            "RESPECT_ROBOTS_TXT", "CRAWL_DELAY_PER_DOMAIN",
            "SERP_PROVIDER", "SERPER_API_KEY", "SERPAPI_KEY",
            "DATAFORSEO_LOGIN", "DATAFORSEO_PASSWORD",
        ]:
            setattr(article_search, key, globals().get(key, ""))

    print("\n" + "=" * 64)
    print("  VikPea SEO 渠道机会扫描")
    print("=" * 64)
    print("说明：只生成机会表，不写发信名单，不发送邮件。\n")

    queries = load_queries()
    if not queries:
        print("没有启用的文章/SEO关键词。请打开 VikPea_文章搜索关键词.xlsx，把第2列改成“是”。")
        return

    wb, ws = ensure_output_workbook(OUTPUT_PATH)
    seen_domains = set()
    added = 0
    scanned = 0
    today = datetime.now().strftime("%Y-%m-%d")

    print(f"关键词 {len(queries)} 个；每词取前 {SEO_OPPORTUNITY_RESULTS_PER_KEYWORD} 条；最低机会分 {SEO_OPPORTUNITY_MIN_SCORE}\n")
    for q_i, query in enumerate(queries, 1):
        print(f"[{q_i}/{len(queries)}] {query}")
        results, engine_info = search_web(query, n=SEO_OPPORTUNITY_RESULTS_PER_KEYWORD)
        print(f"  → {len(results)} 条结果 | {engine_info}")
        if not results:
            continue

        for rank, (title, url) in enumerate(results, 1):
            scanned += 1
            if should_skip(url):
                continue
            domain = root_of_domain(get_base_domain(url))
            if not domain or domain in seen_domains:
                continue

            pre_score, pre_reasons = score_article_candidate(title, url)
            soup, final_url = fetch(url)
            score, reasons = score_article_candidate(title, final_url or url, soup=soup) if soup else (pre_score, pre_reasons + ["页面抓取失败，仅按搜索结果评分"])
            level, opportunity_type, action = classify_opportunity(score, title, final_url or url)

            if score < SEO_OPPORTUNITY_MIN_SCORE:
                continue

            seen_domains.add(domain)
            ws.append([
                query, rank, level, opportunity_type, score, "；".join(reasons),
                title, final_url or url, domain, action, engine_info, today,
            ])
            row = ws.max_row
            fill = GREEN if level == "A" else YELLOW
            for col in range(1, 13):
                ws.cell(row, col).fill = fill
            added += 1
            print(f"    {level} {domain} | {score}分 | {opportunity_type}")
        time.sleep(DELAY_SEARCH)

    if save_workbook_safe:
        save_workbook_safe(wb, OUTPUT_PATH, "SEO渠道机会扫描")
    else:
        wb.save(OUTPUT_PATH)
    print("\n完成")
    print(f"扫描结果 {scanned} 条，输出机会 {added} 条")
    print(f"输出文件: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
