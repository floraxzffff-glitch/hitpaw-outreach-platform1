"""
VikPea_自动跟进.py — 从邮件追踪表筛选未回复对象，自动发送第1/第2次跟进

工作方式：
  · 读取 VikPea_邮件开发追踪.xlsx 的「邮件追踪」主表
  · 首封发出满5天且未回复 → 第1次跟进
  · 第1次跟进发出满7天且未回复 → 第2次跟进
  · 已发过第2次跟进，且跟进2后满7天仍未回复 → 标记暂不跟进
  · 人工可在「跟进优先级」列填：重点 / 不跟进 / 暂停

用法:
  python3 ~/Downloads/VikPea工作台/VikPea_自动跟进.py
"""

import os
import sys
import ssl
import smtplib
import time
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from VikPea_common import (
        apply_config, load_blacklist, safety_reason, export_send_preview,
        log_event,
    )
except ImportError:
    apply_config = None
    load_blacklist = None
    safety_reason = None
    export_send_preview = None
    log_event = None


# ── SMTP 配置：保持和 VikPea_读表发信.py 一致 ───────────────────
SMTP_SERVER = "smtp.qiye.aliyun.com"
SMTP_PORT   = 465
FROM_EMAIL  = "hannah@hitpaw.com"
FROM_NAME   = "Hannah"
PASSWORD    = ""  # 密码已迁移到 VikPea_配置.xlsx / 网页系统设置，不再硬编码
DELAY_SEC   = 8

TRACKER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "VikPea_邮件开发追踪.xlsx")

# 每天自动跟进上限，避免一次发太猛；需要更大量时改这里。
DAILY_LIMIT = 50
FOLLOWUP1_AFTER_DAYS = 5
FOLLOWUP2_AFTER_FOLLOWUP1_DAYS = 7
STOP_AFTER_DAYS = 14
STOP_AFTER_FOLLOWUP2_DAYS = 7
REQUIRE_SEND_CODE = "SEND"
SEND_CONFIRM_WORDS = {"y", "yes", "send", "发送", "确认"}

# 追踪表列（1-based）
T_NUM    = 1
T_DATE   = 2
T_NAME   = 3
T_EMAIL  = 4
T_TYPE   = 5
T_CHAN   = 6
T_KIND   = 7
T_REPLY  = 8
T_SUM    = 9
T_LINK   = 10
T_STATUS = 11
T_GRADE  = 12
T_INTENT = 19

# 新增辅助列
T_FU1_DATE = 13
T_FU2_DATE = 14
T_LAST_FU  = 15
T_PRIORITY = 16
T_FU_NOTE  = 17

BLUE   = PatternFill("solid", start_color="DDEBF7")
GREEN  = PatternFill("solid", start_color="E2EFDA")
YELLOW = PatternFill("solid", start_color="FFF2CC")
GRAY   = PatternFill("solid", start_color="D9E1F2")
BLACKLIST = None


def is_send_confirm(value):
    text = str(value or "").strip().lower()
    required = str(REQUIRE_SEND_CODE or "").strip().lower()
    return text in SEND_CONFIRM_WORDS or (required and text == required)


FOLLOWUP1_BODY = """Hi {name},

Just wanted to quickly follow up in case this got buried.

Would you be open to adding HitPaw VikPea as an alternative tool in your video description, or testing it for a future video?

Happy to send over a free license if helpful.

Best,
Hannah
HitPaw Team
https://www.hitpaw.com/hitpaw-video-enhancer.html"""


FOLLOWUP2_BODY = """Hi {name},

One last quick note from me.

I thought HitPaw VikPea might be relevant for your audience because it focuses on AI video upscaling, noise reduction, and face restoration, especially for creators comparing tools like Topaz Video AI.

If now isn't a fit, no worries at all.

Best,
Hannah
HitPaw Team
https://www.hitpaw.com/hitpaw-video-enhancer.html"""


def ensure_headers(ws):
    headers = {
        T_FU1_DATE: "跟进1日期",
        T_FU2_DATE: "跟进2日期",
        T_LAST_FU: "最后跟进日期",
        T_PRIORITY: "跟进优先级",
        T_FU_NOTE: "跟进备注",
    }
    for col, title in headers.items():
        cell = ws.cell(1, col)
        if not cell.value:
            cell.value = title
            cell.fill = BLUE
            cell.font = Font(bold=True)


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date; openpyxl sometimes returns raw serials for older rows.
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    return None


def is_replied(reply_value):
    text = str(reply_value or "").strip().lower()
    return text in {"已回复", "是", "y", "yes", "true", "replied", "回复"}


def has_reply_signal(row):
    """多重判断是否已有回复，避免只靠「是否回复」一列导致误跟进。"""
    reply = row[T_REPLY - 1].value
    summary = str(row[T_SUM - 1].value or "").strip()
    status = str(row[T_STATUS - 1].value or "").strip()
    grade = str(row[T_GRADE - 1].value or "").strip().upper()
    intent = ""
    if len(row) >= T_INTENT:
        intent = str(row[T_INTENT - 1].value or "").strip()

    if is_replied(reply):
        return True

    replied_status_words = [
        "待处理", "待分析", "不建议沟通", "已拒绝", "已合作", "沟通中",
        "已回复", "收到回复", "报价", "感兴趣", "合作中", "不合适",
    ]
    if any(word in status for word in replied_status_words):
        return True

    # ABC分级/回复意向/回复摘要只要有明显内容，也视为已经回复过。
    if grade in {"A", "B", "C"}:
        return True
    if intent:
        return True
    if summary and summary not in {"无", "none", "-", "n/a"}:
        return True

    return False


def is_blocked(priority, status):
    combined = f"{priority} {status}".strip()
    blocked_words = [
        "不跟进", "暂停", "已拒绝", "已合作", "沟通中", "待处理", "待分析",
        "超预算", "未合作", "不建议沟通", "暂不跟进", "未获取邮箱",
    ]
    return any(word in combined for word in blocked_words)


def build_subject(step, contact_type):
    if str(contact_type).strip() == "文章外链":
        base = "HitPaw VikPea for your article"
    else:
        base = "HitPaw VikPea collaboration"
    return f"Re: {base}" if step in (1, 2) else base


def make_email(name, to_email, contact_type, step):
    body_template = FOLLOWUP1_BODY if step == 1 else FOLLOWUP2_BODY
    body = body_template.format(name=name)
    subject = build_subject(step, contact_type)
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"] = to_email
    msg.attach(MIMEText(body, "plain", "utf-8"))
    return msg, subject, body


def classify_row(row, today):
    sent_date = parse_date(row[T_DATE - 1].value)
    if not sent_date:
        return None

    name = str(row[T_NAME - 1].value or "").strip()
    email = str(row[T_EMAIL - 1].value or "").strip()
    contact_type = str(row[T_TYPE - 1].value or "").strip()
    status = str(row[T_STATUS - 1].value or "").strip()
    priority = str(row[T_PRIORITY - 1].value or "").strip()
    fu1_date = parse_date(row[T_FU1_DATE - 1].value)
    fu2_date = parse_date(row[T_FU2_DATE - 1].value)
    if "已跟进1" in status and not fu1_date:
        fu1_date = sent_date
    if "已跟进2" in status and not fu2_date:
        fu2_date = sent_date

    if not name or not email or "@" not in email:
        return None
    reason = safety_reason(name, email, row[T_LINK - 1].value, BLACKLIST) if safety_reason else ""
    if reason:
        row[T_FU_NOTE - 1].value = f"自动跟进拦截: {reason}"
        return None
    if has_reply_signal(row) or is_blocked(priority, status):
        return None

    days = (today.date() - sent_date.date()).days
    days_since_fu2 = (today.date() - fu2_date.date()).days if fu2_date else None
    if fu2_date and days >= STOP_AFTER_DAYS and days_since_fu2 >= STOP_AFTER_FOLLOWUP2_DAYS:
        return {"action": "stop", "days": days, "row": row, "name": name, "email": email}
    days_since_fu1 = (today.date() - fu1_date.date()).days if fu1_date else None
    if fu1_date and not fu2_date and days_since_fu1 >= FOLLOWUP2_AFTER_FOLLOWUP1_DAYS:
        return {
            "action": "send", "step": 2, "days": days, "row": row,
            "name": name, "email": email, "type": contact_type, "priority": priority
        }
    if days >= FOLLOWUP1_AFTER_DAYS and not fu1_date:
        return {
            "action": "send", "step": 1, "days": days, "row": row,
            "name": name, "email": email, "type": contact_type, "priority": priority
        }
    return None


def collect_targets(ws, today):
    to_send = []
    to_stop = []
    for row in ws.iter_rows(min_row=2):
        item = classify_row(row, today)
        if not item:
            continue
        if item["action"] == "send":
            to_send.append(item)
        elif item["action"] == "stop":
            to_stop.append(item)

    # 重点优先，其次跟进2优先，因为这批更接近结束。
    def sort_key(item):
        priority_score = 0 if item.get("priority") == "重点" else 1
        step_score = 0 if item.get("step") == 2 else 1
        return (priority_score, step_score, -item["days"], item["name"].lower())

    return sorted(to_send, key=sort_key), to_stop


def update_sent_row(item, subject):
    row = item["row"]
    today_str = datetime.now().strftime("%Y-%m-%d")
    step = item["step"]
    if step == 1:
        row[T_FU1_DATE - 1].value = today_str
        row[T_STATUS - 1].value = "已跟进1"
        row[T_KIND - 1].value = "初次开发+跟进1"
    else:
        row[T_FU2_DATE - 1].value = today_str
        row[T_STATUS - 1].value = "已跟进2"
        row[T_KIND - 1].value = "初次开发+跟进2"
    row[T_LAST_FU - 1].value = today_str
    row[T_FU_NOTE - 1].value = f"已发送跟进{step}: {subject}"
    fill = GREEN if step == 1 else YELLOW
    for cell in row[:T_FU_NOTE]:
        cell.fill = fill


def mark_skip_replied(item):
    row = item["row"]
    row[T_FU_NOTE - 1].value = "自动跟进发送前复查：发现已有回复/待处理信号，已跳过"
    for cell in row[:T_FU_NOTE]:
        cell.fill = BLUE


def mark_stopped(item):
    row = item["row"]
    row[T_STATUS - 1].value = "暂不跟进"
    row[T_FU_NOTE - 1].value = f"首封后{item['days']}天未回复，已停止自动跟进"
    for cell in row[:T_FU_NOTE]:
        cell.fill = GRAY


def print_plan(to_send, to_stop):
    print(f"\n{'═'*70}")
    print("  VikPea 自动跟进队列")
    print(f"{'═'*70}")
    print(f"  今日可发送跟进: {len(to_send)}")
    print(f"  将标记暂不跟进: {len(to_stop)}")
    print(f"  每日发送上限: {DAILY_LIMIT}")
    print(f"{'─'*70}")

    preview_items = to_send[:min(20, len(to_send))]
    for i, item in enumerate(preview_items, 1):
        tag = "重点" if item.get("priority") == "重点" else "普通"
        print(f"  {i:>2}. 跟进{item['step']} | {tag} | {item['days']}天 | {item['name']} → {item['email']}")
    if len(to_send) > len(preview_items):
        print(f"  ... 还有 {len(to_send) - len(preview_items)} 个未展示")
    print(f"{'═'*70}")


def main():
    global BLACKLIST
    if apply_config:
        apply_config(globals(), {
            "SMTP_SERVER": "SMTP_SERVER",
            "SMTP_PORT": "SMTP_PORT",
            "FROM_EMAIL": "FROM_EMAIL",
            "FROM_NAME": "FROM_NAME",
            "PASSWORD": "PASSWORD",
            "DELAY_SEC": "DELAY_SEC",
            "FOLLOWUP_DAILY_LIMIT": "DAILY_LIMIT",
            "FOLLOWUP1_AFTER_DAYS": "FOLLOWUP1_AFTER_DAYS",
            "FOLLOWUP2_AFTER_DAYS": "FOLLOWUP2_AFTER_FOLLOWUP1_DAYS",
            "REQUIRE_SEND_CODE": "REQUIRE_SEND_CODE",
        })
    BLACKLIST = load_blacklist() if load_blacklist else {"emails": set(), "domains": set(), "names": set()}

    if not os.path.exists(TRACKER_PATH):
        print(f"❌ 找不到追踪表: {TRACKER_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
    ensure_headers(ws)

    today = datetime.now()
    to_send, to_stop = collect_targets(ws, today)
    print_plan(to_send, to_stop)

    if to_stop:
        confirm_stop = input(f"\n是否先把 {len(to_stop)} 条超期未回复记录标记为“暂不跟进”？(y/n): ").strip().lower()
        if confirm_stop == "y":
            for item in to_stop:
                mark_stopped(item)
            wb.save(TRACKER_PATH)
            print(f"✅ 已标记暂不跟进: {len(to_stop)} 条")

    if not to_send:
        wb.save(TRACKER_PATH)
        print("\n✅ 今天没有需要发送的跟进邮件")
        return

    batch = to_send[:DAILY_LIMIT]
    if export_send_preview and batch:
        export_send_preview([
            {
                "rownum": item["row"][0].row,
                "status": f"跟进{item['step']}",
                "name": item["name"],
                "email": item["email"],
                "type": item.get("type", ""),
                "source": "",
                "link": item["row"][T_LINK - 1].value,
                "subject": build_subject(item["step"], item.get("type", "")),
                "opening": f"首封后 {item['days']} 天，自动跟进候选",
            }
            for item in batch
        ])
        print("📋 已生成本次跟进预览: VikPea_本次发信预览.xlsx")
    first = batch[0]
    _, subject, body = make_email(first["name"], first["email"], first["type"], first["step"])
    print(f"\n【预览第1封】")
    print(f"To: {first['name']} <{first['email']}>")
    print(f"Subject: {subject}\n")
    print(body)

    confirm = input(f"\n确认发送以上 {len(batch)} 封跟进邮件？输入 y 或 send 才会真正发送: ").strip()
    if not is_send_confirm(confirm):
        wb.save(TRACKER_PATH)
        print("已取消发送；追踪表辅助列已保留")
        return

    sent = 0
    failed = []
    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=ctx) as server:
        server.login(FROM_EMAIL, PASSWORD)
        print("\n✅ SMTP 登录成功，开始发送跟进...\n")
        if log_event:
            log_event("跟进开始", f"计划发送 {len(batch)} 封")
        for item in batch:
            if has_reply_signal(item["row"]) or is_blocked(item.get("priority", ""), item["row"][T_STATUS - 1].value):
                mark_skip_replied(item)
                wb.save(TRACKER_PATH)
                print(f"  ⏭️ 跳过 {item['name']} → 已有回复/待处理信号")
                continue
            reason = safety_reason(item["name"], item["email"], item["row"][T_LINK - 1].value, BLACKLIST) if safety_reason else ""
            if reason:
                item["row"][T_FU_NOTE - 1].value = f"跟进发送前拦截: {reason}"
                wb.save(TRACKER_PATH)
                print(f"  ⏭️ 跳过 {item['name']} → {reason}")
                continue
            try:
                msg, subject, _ = make_email(item["name"], item["email"], item["type"], item["step"])
                server.sendmail(FROM_EMAIL, item["email"], msg.as_string())
                update_sent_row(item, subject)
                wb.save(TRACKER_PATH)
                sent += 1
                print(f"  ✅ [{sent}/{len(batch)}] 跟进{item['step']} {item['name']} → {item['email']}")
                if log_event:
                    log_event("跟进成功", f"跟进{item['step']} {item['name']} <{item['email']}>")
                time.sleep(DELAY_SEC)
            except Exception as e:
                failed.append((item["name"], item["email"], str(e)))
                item["row"][T_FU_NOTE - 1].value = f"跟进发送失败: {e}"
                wb.save(TRACKER_PATH)
                print(f"  ❌ {item['name']} → {item['email']} 错误: {e}")
                if log_event:
                    log_event("跟进失败", f"{item['name']} <{item['email']}> {e}")

    wb.save(TRACKER_PATH)
    print(f"\n{'─'*70}")
    print(f"完成: ✅ 已发送 {sent} 封跟进  ❌ 失败 {len(failed)} 封")
    if failed:
        for name, email, err in failed:
            print(f"  {name} ({email}): {err}")
    print(f"追踪表已更新: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
