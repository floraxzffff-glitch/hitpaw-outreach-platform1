"""
VikPea_文章批量搜索.py — 批量搜索文章/博客 → 提取联系邮箱 → 写入发信名单

流程：
  关键词 → DuckDuckGo 搜索 → 过滤无效域名 → 抓文章页/联系页
  → 提取邮箱 → 去重 → 写绿行（有邮箱）/ 黄行（无邮箱）→ VikPea_发信名单.xlsx

依赖（在你Mac终端运行）:
  pip3 install requests beautifulsoup4 openpyxl

用法: python3 ~/Downloads/VikPea工作台/VikPea_文章批量搜索.py
"""

import re, time, os, sys, urllib.parse, urllib.robotparser, html, ssl
from datetime import datetime

try:
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError as e:
    print(f"❌ 缺少依赖: {e}")
    print("   pip3 install requests beautifulsoup4 openpyxl")
    sys.exit(1)

try:
    from VikPea_common import (
        apply_config, log_event, add_no_email_candidate, existing_no_email_keys,
        classify_bad_email, root_domain, add_pending_email_review_row,
        ensure_pending_email_review_workbook, fix_glued_tld,
    )
except ImportError:
    apply_config = None
    log_event = None
    add_no_email_candidate = None
    existing_no_email_keys = None
    classify_bad_email = None
    root_domain = None
    add_pending_email_review_row = None
    ensure_pending_email_review_workbook = None
    fix_glued_tld = lambda e: e

# ── 配置 ────────────────────────────────────────────────────────
QUEUE_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_发信名单.xlsx")
TRACKER_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_邮件开发追踪.xlsx")
KEYWORD_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "VikPea_文章搜索关键词.xlsx")

RESULTS_PER_QUERY = 30  # 每个关键词取几条结果
DELAY_SEARCH = 4        # 两次搜索之间等待（秒）
DELAY_FETCH  = 2        # 两次页面抓取等待（秒）
REQUEST_TIMEOUT = 15    # 页面请求超时（秒）
CONTACT_SEARCH_PER_DOMAIN = 8  # 单域名补搜联系页时取前几条
ALLOW_INSECURE_SSL = True
ARTICLE_MIN_SITE_SCORE = 3
ARTICLE_MIN_EMAIL_CONFIDENCE = "A"  # A 直接进发信名单；B/C 进待确认邮箱表
RESPECT_ROBOTS_TXT = True
CRAWL_DELAY_PER_DOMAIN = 1.5
SERP_PROVIDER = ""
SERPER_API_KEY = ""
SERPAPI_KEY = ""
DATAFORSEO_LOGIN = ""
DATAFORSEO_PASSWORD = ""

# 搜索关键词（聚焦「评测/对比/推荐」类文章）
QUERIES = [
    # 榜单/推荐
    "best AI video enhancer software 2026 review",
    "best AI video enhancer tools 2026",
    "best AI video upscaler software 2026",
    "best video enhancer software for Windows 2026",
    "best video enhancer software for Mac 2026",
    "best video restoration software 2026",
    "best video quality enhancer 2025 2026",
    "best AI video repair software 2026",
    "best old video restoration software",
    "best software to improve video quality",
    "best tools to upscale video to 4K",
    "best AI tools for video creators 2026",

    # 替代品/竞品对比
    "topaz video AI alternative free 2026",
    "topaz video ai alternatives 2026",
    "best topaz video ai alternative",
    "topaz video ai vs hitpaw vikpea",
    "hitpaw video enhancer vs topaz 2026",
    "hitpaw vikpea alternatives",
    "hitpaw vikpea review 2026",
    "HitPaw VikPea review 2026",
    "unifab video enhancer review comparison",
    "unifab video enhancer alternatives",
    "aiarty video enhancer review 2026",
    "aiarty video enhancer alternatives",
    "avclabs video enhancer alternative",
    "vanceai video enhancer alternative",

    # 评测/对比
    "ai video upscaling software comparison 2026",
    "video upscaling tools ranked comparison 2026",
    "AI video enhancer software comparison",
    "AI video enhancer review",
    "video enhancer software review",
    "4k video upscaling tool comparison review",
    "AI video enhancer without watermark 2026",
    "free video upscaling AI software review",
    "free video enhancer software review",
    "online video enhancer review",

    # 功能场景
    "ai video noise reduction software review",
    "best video denoise software AI",
    "best AI video sharpening software",
    "best face restoration video software",
    "best blurry video fixer software",
    "best video repair software for corrupted videos",
    "best low resolution video enhancer",
    "best tool to make blurry video clear",
    "best app to enhance video quality",
    "how to improve video quality with AI software",
    "how to upscale video to 4K with AI",
    "how to fix blurry video using AI",
    "how to restore old videos with AI",

    # 免费/轻量需求
    "video enhancer software free alternative topaz",
    "free AI video enhancer no watermark",
    "free AI video upscaler without watermark",
    "free video quality enhancer for PC",
    "free video enhancer for creators",
    "affordable AI video enhancer software",

    # 文章站常见标题词
    "AI video enhancer roundup",
    "AI video enhancer tools list",
    "video enhancer tools list",
    "video upscaler tools list",
    "AI tools for video editing roundup",
    "creator tools AI video enhancer",
    "software to enhance video quality roundup",
    "video editing tools for creators AI enhancer",
]

# 跳过的域名（社交、视频、竞品、大媒体无法投放）
SKIP_DOMAINS = {
    # 社交/视频平台
    "youtube.com", "youtu.be", "vimeo.com", "tiktok.com",
    "twitter.com", "x.com", "facebook.com", "instagram.com",
    "reddit.com", "quora.com", "linkedin.com",
    "amazon.com", "ebay.com", "walmart.com",
    "google.com", "bing.com", "yahoo.com",
    "duckduckgo.com", "html.duckduckgo.com", "search.yahoo.com",
    "wikipedia.org", "wikihow.com",
    # 大媒体（预算不匹配）
    "techradar.com", "pcmag.com", "cnet.com", "tomsguide.com",
    "makeuseof.com", "theverge.com", "wired.com", "forbes.com",
    # 竞品工具（发了会尴尬）
    "topazlabs.com", "hitpaw.com", "aiarty.com", "unifab.com",
    "unifab.ai", "winxdvd.com", "aiseesoft.com", "videoproc.com",
    "vmake.ai", "magichour.ai", "pxz.ai", "imagera.ai",
    "gstory.ai", "myimageupscaler.com", "piximperfect.com",
    "podsqueeze.com", "contentmavericks.com",
    "opentools.ai", "futurepedia.io", "there'sanaiforthis.com",
    # 已直接联系，不需要脚本重复添加
    "fixthephoto.com",
}

LOW_VALUE_DOMAINS = {
    "zhihu.com", "ask.com", "geeksforgeeks.org", "shutterstock.com",
    "usebouncer.com", "emailchaser.com", "neverbounce.com",
    "apwg.org", "gov", "edu",
}

LOW_VALUE_EMAIL_DOMAINS = {
    "googleapis.com", "w3.org", "w3c.org", "window.wiz", "duckduckgo.com",
    "bing.com", "google.com", "youtube.com", "cloudflare.com", "sentry.io",
}

ARTICLE_POSITIVE_TERMS = {
    "best", "top", "review", "reviews", "alternative", "alternatives",
    "comparison", "compare", "vs", "tools", "software", "roundup",
    "guide", "tutorial", "how to", "ranked", "list",
}

ARTICLE_NEGATIVE_TERMS = {
    "login", "signup", "sign up", "pricing", "cart", "checkout", "download",
    "privacy policy", "terms", "support", "docs", "api", "forum", "community",
    "profile", "search", "tag", "category",
}

SEO_TITLE_MIN_LEN = 25
SEO_TITLE_MAX_LEN = 90
SEO_META_MIN_LEN = 70
SEO_META_MAX_LEN = 180

GREEN  = PatternFill("solid", start_color="E2EFDA")
YELLOW = PatternFill("solid", start_color="FFF2CC")
EMAIL_RE = re.compile(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}")
OBFUSCATED_EMAIL_RE = re.compile(
    r"([A-Za-z0-9._%+\-]{1,64})\s*(?:@|\[at\]|\(at\)|\sat\s)\s*"
    r"([A-Za-z0-9.\-]{1,80})\s*(?:\.|\[dot\]|\(dot\)|\sdot\s)\s*([A-Za-z]{2,24})",
    re.I,
)

Q_NAME=1; Q_EMAIL=2; Q_SUBJ=3; Q_OPEN=4; Q_LINK=5; Q_NOTE=6; Q_TYPE=7; Q_SOURCE=8

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
}
ROBOTS_CACHE = {}
DOMAIN_LAST_FETCH = {}
# ────────────────────────────────────────────────────────────────


def get_base_domain(url: str) -> str:
    try:
        parts = urllib.parse.urlparse(url)
        host = parts.netloc.lower().lstrip("www.")
        return host
    except Exception:
        return ""


def should_skip(url: str) -> bool:
    domain = get_base_domain(url)
    parsed = urllib.parse.urlparse(str(url or ""))
    path = parsed.path.lower()
    for skip in SKIP_DOMAINS:
        if domain == skip or domain.endswith("." + skip):
            return True
    if any(domain == item or domain.endswith("." + item) for item in LOW_VALUE_DOMAINS if "." in item):
        return True
    if domain.endswith(".gov") or domain.endswith(".edu"):
        return True
    if "/search" in path or "/login" in path or "/signup" in path:
        return True
    return False


def can_fetch_url(url: str) -> bool:
    if not RESPECT_ROBOTS_TXT:
        return True
    parsed = urllib.parse.urlparse(str(url or ""))
    if not parsed.scheme or not parsed.netloc:
        return False
    domain = parsed.netloc.lower()
    if domain in {"html.duckduckgo.com", "www.bing.com", "bing.com"}:
        return True
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    rp = ROBOTS_CACHE.get(robots_url)
    if rp is None:
        rp = urllib.robotparser.RobotFileParser()
        rp.set_url(robots_url)
        try:
            rp.read()
        except Exception:
            ROBOTS_CACHE[robots_url] = False
            return True
        ROBOTS_CACHE[robots_url] = rp
    if rp is False:
        return True
    try:
        return bool(rp.can_fetch(HEADERS["User-Agent"], url))
    except Exception:
        return True


def wait_for_domain(url: str):
    try:
        domain = urllib.parse.urlparse(str(url or "")).netloc.lower()
        if not domain:
            return
        now = time.time()
        last = DOMAIN_LAST_FETCH.get(domain, 0)
        wait = float(CRAWL_DELAY_PER_DOMAIN or 0) - (now - last)
        if wait > 0:
            time.sleep(wait)
        DOMAIN_LAST_FETCH[domain] = time.time()
    except Exception:
        return


def root_of_domain(domain: str) -> str:
    if root_domain:
        return root_domain(domain)
    parts = [p for p in str(domain or "").lower().replace("www.", "").split(".") if p]
    return ".".join(parts[-2:]) if len(parts) >= 2 else str(domain or "").lower()


def is_low_value_email(email: str) -> str:
    e = str(email or "").strip().lower()
    if classify_bad_email:
        bad = classify_bad_email(e)
        if bad:
            return bad
    if "@" not in e:
        return "邮箱格式异常"
    local, domain = e.split("@", 1)
    domain_root = root_of_domain(domain)
    if domain in LOW_VALUE_EMAIL_DOMAINS or domain_root in LOW_VALUE_EMAIL_DOMAINS:
        return "平台/搜索引擎/技术依赖邮箱"
    if any(fragment in domain for fragment in LOW_VALUE_EMAIL_DOMAINS):
        return "平台/搜索引擎/技术依赖邮箱"
    if re.search(r"^(noreply|no-reply|donotreply|abuse|privacy|security|reportphishing)$", local):
        return "非合作联系邮箱"
    if re.search(r"%[0-9a-f]{2}", e, re.I):
        return "URL编码碎片误判邮箱"
    return ""


def score_article_candidate(title: str, url: str, soup=None) -> tuple:
    """
    给文章结果打分，返回 (score, reasons)。
    分数越高越像可投放的评测/榜单/教程文章；低分直接跳过，不浪费邮箱抓取。
    """
    text = f"{title or ''} {url or ''}".lower()
    parsed = urllib.parse.urlparse(str(url or ""))
    domain = get_base_domain(url)
    path = parsed.path.lower()
    score = 0
    reasons = []

    if should_skip(url):
        return -9, ["黑名单/低价值域名"]
    if path in {"", "/"}:
        score -= 1
        reasons.append("首页结果")
    if any(term in text for term in ARTICLE_POSITIVE_TERMS):
        score += 2
        reasons.append("标题/URL像评测文章")
    if any(term in text for term in ARTICLE_NEGATIVE_TERMS):
        score -= 2
        reasons.append("标题/URL像低价值页面")
    if re.search(r"/(blog|article|review|reviews|best|top|tools|comparison|alternatives?|guide|tutorial|how-to|rank)", path):
        score += 2
        reasons.append("文章路径")
    if re.search(r"/(product|shop|pricing|download|support|docs|api|tag|category|search|login|signup)", path):
        score -= 2
        reasons.append("非文章路径")
    if domain.endswith(".gov") or domain.endswith(".edu"):
        score -= 4
        reasons.append("政府/学校站")

    if soup:
        page_title = soup.find("title")
        meta_desc = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
        canonical = soup.find("link", rel=lambda rel: rel and "canonical" in str(rel).lower())
        h1s = [h.get_text(" ", strip=True) for h in soup.find_all("h1")]
        h2s = [h.get_text(" ", strip=True) for h in soup.find_all("h2")]
        page_text = soup.get_text(" ", strip=True)
        page_lower = page_text[:3000].lower()
        word_count = len(re.findall(r"[a-zA-Z]{3,}", page_text))

        title_text = page_title.get_text(" ", strip=True) if page_title else ""
        meta_text = str(meta_desc.get("content", "")).strip() if meta_desc else ""

        if word_count >= 450:
            score += 1
            reasons.append("正文较完整")
        elif word_count < 160:
            score -= 2
            reasons.append("正文太少")
        if title_text and SEO_TITLE_MIN_LEN <= len(title_text) <= SEO_TITLE_MAX_LEN:
            score += 1
            reasons.append("标题长度正常")
        elif title_text:
            score -= 1
            reasons.append("标题长度异常")
        if meta_text and SEO_META_MIN_LEN <= len(meta_text) <= SEO_META_MAX_LEN:
            score += 1
            reasons.append("有正常描述")
        if h1s:
            score += 1
            reasons.append("有H1")
            if any(term in " ".join(h1s[:2]).lower() for term in ARTICLE_POSITIVE_TERMS):
                score += 1
                reasons.append("H1相关")
        else:
            score -= 1
            reasons.append("缺H1")
        if len(h2s) >= 2:
            score += 1
            reasons.append("有小标题结构")
        if canonical and canonical.get("href"):
            score += 1
            reasons.append("有canonical")
        if page_title and any(term in title_text.lower() for term in ARTICLE_POSITIVE_TERMS):
            score += 1
            reasons.append("页面标题相关")
        outbound_domains = set()
        internal_links = 0
        for a in soup.select("a[href]"):
            href = urllib.parse.urljoin(url, a.get("href", ""))
            out_domain = get_base_domain(href)
            if out_domain and out_domain != domain:
                outbound_domains.add(root_of_domain(out_domain))
            elif out_domain == domain:
                internal_links += 1
        if len(outbound_domains) >= 3:
            score += 1
            reasons.append("有多个外链")
        if len(outbound_domains) >= 8:
            score += 1
            reasons.append("外链丰富")
        if internal_links >= 5:
            score += 1
            reasons.append("站内结构正常")
        if any(term in page_lower for term in ["best ", "top ", "review", "comparison", "alternative", "tools"]):
            score += 1
            reasons.append("正文相关")
        if re.search(r"(updated|published|last updated|reviewed)", page_lower):
            score += 1
            reasons.append("有发布日期/更新痕迹")

    return score, reasons


def classify_article_email(email: str, article_url: str, source_url: str) -> tuple:
    """
    返回 (confidence, note)：
    A = 同站域名/联系页邮箱，可信，可以进主表
    B = 公共邮箱但来自同站页面，建议人工确认
    C = 搜索引擎/跨域/低可信邮箱，不直接发
    """
    bad = is_low_value_email(email)
    if bad:
        return "C", bad
    if "@" not in str(email or ""):
        return "C", "邮箱格式异常"
    email_domain = str(email).split("@", 1)[1].lower()
    email_root = root_of_domain(email_domain)
    article_root = root_of_domain(get_base_domain(article_url))
    source_root = root_of_domain(get_base_domain(source_url))
    public_domains = {"gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "icloud.com", "proton.me", "protonmail.com"}

    if email_root and email_root == article_root:
        return "A", f"同站邮箱: {source_url}"
    if source_root and source_root == article_root and email_domain in public_domains:
        return "B", f"同站页面里的公共邮箱: {source_url}"
    if source_root and source_root == article_root:
        return "B", f"同站页面候选邮箱: {source_url}"
    return "C", f"跨域/搜索结果邮箱: {source_url}"


def normalize_search_result_url(href: str) -> str:
    """
    DuckDuckGo HTML 结果里有时是跳转链接、相对链接或带 uddg 参数的中转链接。
    这里尽量还原出真实目标 URL。
    """
    text = str(href or "").strip()
    if not text:
        return ""
    if text.startswith("//"):
        text = "https:" + text
    if text.startswith("/"):
        text = urllib.parse.urljoin("https://html.duckduckgo.com", text)

    parsed = urllib.parse.urlparse(text)
    if "duckduckgo.com" in parsed.netloc:
        params = urllib.parse.parse_qs(parsed.query)
        uddg = params.get("uddg", [])
        if uddg and uddg[0]:
            return urllib.parse.unquote(uddg[0])
    return text


def unique_preserve(seq):
    seen = set()
    out = []
    for item in seq:
        key = item if isinstance(item, str) else repr(item)
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def fetch(url: str, timeout: int = REQUEST_TIMEOUT) -> tuple:
    """返回 (soup, final_url) 或 (None, url)"""
    if not can_fetch_url(url):
        return None, url
    wait_for_domain(url)
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        return soup, resp.url
    except requests.exceptions.SSLError:
        if not ALLOW_INSECURE_SSL:
            return None, url
        try:
            requests.packages.urllib3.disable_warnings()  # type: ignore[attr-defined]
        except Exception:
            pass
        try:
            resp = requests.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True, verify=False)
            resp.raise_for_status()
            soup = BeautifulSoup(resp.text, "html.parser")
            return soup, resp.url
        except Exception:
            return None, url
    except Exception:
        return None, url


def extract_result_links_from_duckduckgo(soup: BeautifulSoup) -> list:
    links = []
    selectors = [
        "a.result__a",
        "a.result-link",
        "a[data-testid='result-title-a']",
        ".result__body a[href]",
        ".links_main a[href]",
    ]
    for selector in selectors:
        for a in soup.select(selector):
            href = normalize_search_result_url(a.get("href", ""))
            title = a.get_text(" ", strip=True)
            if href.startswith("http") and title:
                links.append((title, href))
    return unique_preserve(links)


def _empty_search_debug(engine: str) -> dict:
    return {"engine": engine, "raw": 0, "kept": 0, "filtered": 0, "pages": 0}


def search_duckduckgo(query: str, n: int = 10) -> tuple:
    """
    DuckDuckGo 多页搜索。
    返回 ([(title, url), ...], debug)
    debug 里会带 raw / kept / filtered / pages，避免“其实有结果但全被过滤”时看起来像完全没搜到。
    """
    debug = _empty_search_debug("DuckDuckGo")
    results = []
    seen = set()

    for page in range(4):
        params = {
            "q": query,
            "kl": "us-en",
            "kp": "-1",
            "s": page * 30,
        }
        url = "https://html.duckduckgo.com/html/?" + urllib.parse.urlencode(params)
        soup, _ = fetch(url)
        debug["pages"] += 1
        if not soup:
            continue

        page_links = extract_result_links_from_duckduckgo(soup)
        if not page_links:
            for a in soup.select("a[href]"):
                href = normalize_search_result_url(a.get("href", ""))
                title = a.get_text(" ", strip=True)
                if href.startswith("http") and title:
                    page_links.append((title, href))

        page_links = unique_preserve(page_links)
        if not page_links and page > 0:
            break

        for title, href in page_links:
            clean = normalize_search_result_url(href).rstrip("/")
            if not clean.startswith("http") or not title:
                continue
            debug["raw"] += 1
            if should_skip(clean):
                debug["filtered"] += 1
                continue
            if clean in seen:
                continue
            seen.add(clean)
            results.append((title, clean))
            debug["kept"] = len(results)
            if len(results) >= n:
                return results, debug
    return results, debug


def search_bing(query: str, n: int = 10) -> tuple:
    debug = _empty_search_debug("Bing")
    results = []
    seen = set()

    for page in range(4):
        params = {"q": query, "setlang": "en-us", "first": page * 10 + 1}
        url = "https://www.bing.com/search?" + urllib.parse.urlencode(params)
        soup, _ = fetch(url)
        debug["pages"] += 1
        if not soup:
            continue

        page_links = []
        for a in soup.select("li.b_algo h2 a, .b_algo a[href], #b_results a[href]"):
            href = str(a.get("href", "")).strip()
            title = a.get_text(" ", strip=True)
            if href.startswith("http") and title:
                page_links.append((title, href))
        page_links = unique_preserve(page_links)
        if not page_links and page > 0:
            break

        for title, href in page_links:
            clean = normalize_search_result_url(href).rstrip("/")
            if not clean.startswith("http") or not title:
                continue
            debug["raw"] += 1
            if should_skip(clean):
                debug["filtered"] += 1
                continue
            if clean in seen:
                continue
            seen.add(clean)
            results.append((title, clean))
            debug["kept"] = len(results)
            if len(results) >= n:
                return unique_preserve(results), debug
    return unique_preserve(results), debug


def format_engine_debug(*debug_items) -> str:
    parts = []
    for item in debug_items:
        if not item:
            continue
        parts.append(
            f"{item['engine']}:保留{item['kept']}/原始{item['raw']}/过滤{item['filtered']}/页数{item['pages']}"
        )
    return " + ".join(parts) if parts else "搜索引擎无返回"


def search_web(query: str, n: int = 10) -> tuple:
    api_results, api_info = search_serp_api(query, n=n)
    if api_results:
        return api_results, api_info
    ddg, ddg_debug = search_duckduckgo(query, n=n)
    if ddg:
        return ddg, format_engine_debug(ddg_debug)
    bing, bing_debug = search_bing(query, n=n)
    if bing:
        return bing, format_engine_debug(ddg_debug, bing_debug)
    return [], format_engine_debug(ddg_debug, bing_debug)


def clean_api_result(title, link):
    title = str(title or "").strip()
    url = normalize_search_result_url(str(link or "").strip()).rstrip("/")
    if not title or not url.startswith("http"):
        return None
    if should_skip(url):
        return None
    return title, url


def search_serper(query: str, n: int = 10) -> tuple:
    if not SERPER_API_KEY:
        return [], "Serper未配置"
    try:
        resp = requests.post(
            "https://google.serper.dev/search",
            headers={
                "X-API-KEY": str(SERPER_API_KEY).strip(),
                "Content-Type": "application/json",
            },
            json={"q": query, "gl": "us", "hl": "en", "num": min(max(n, 10), 100)},
            timeout=REQUEST_TIMEOUT,
        )
        if resp.status_code >= 400:
            return [], f"Serper失败:{resp.status_code}"
        data = resp.json()
        rows = []
        for item in data.get("organic", []) or []:
            row = clean_api_result(item.get("title"), item.get("link"))
            if row:
                rows.append(row)
            if len(rows) >= n:
                break
        return unique_preserve(rows), f"Serper:{len(rows)}"
    except Exception as exc:
        return [], f"Serper异常:{str(exc)[:80]}"


def search_serpapi(query: str, n: int = 10) -> tuple:
    if not SERPAPI_KEY:
        return [], "SerpApi未配置"
    try:
        params = {
            "engine": "google",
            "q": query,
            "api_key": str(SERPAPI_KEY).strip(),
            "num": min(max(n, 10), 100),
            "gl": "us",
            "hl": "en",
        }
        resp = requests.get(
            "https://serpapi.com/search.json",
            params=params,
            headers=HEADERS,
            timeout=REQUEST_TIMEOUT,
        )
        if resp.status_code >= 400:
            return [], f"SerpApi失败:{resp.status_code}"
        data = resp.json()
        rows = []
        for item in data.get("organic_results", []) or []:
            row = clean_api_result(item.get("title"), item.get("link"))
            if row:
                rows.append(row)
            if len(rows) >= n:
                break
        return unique_preserve(rows), f"SerpApi:{len(rows)}"
    except Exception as exc:
        return [], f"SerpApi异常:{str(exc)[:80]}"


def search_serp_api(query: str, n: int = 10) -> tuple:
    provider = str(SERP_PROVIDER or "").strip().lower()
    if provider in {"serper", "google-serper"} or (not provider and SERPER_API_KEY):
        rows, info = search_serper(query, n=n)
        if rows:
            return rows, info
        if provider:
            return [], info
    if provider in {"serpapi", "serp-api"} or (not provider and SERPAPI_KEY):
        rows, info = search_serpapi(query, n=n)
        if rows:
            return rows, info
        if provider:
            return [], info
    if provider == "dataforseo":
        return [], "DataForSEO预留：当前文章搜索未启用"
    return [], ""


def load_queries() -> list:
    """优先读取文章关键词表；只有完全没有关键词表时才使用脚本内置关键词。"""
    queries = []
    has_keyword_file = os.path.exists(KEYWORD_PATH)
    if has_keyword_file:
        try:
            wb = openpyxl.load_workbook(KEYWORD_PATH, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                keyword = str((row[0] if row else "") or "").strip()
                enabled = str((row[1] if len(row) > 1 else "是") or "是").strip()
                if keyword and enabled not in {"否", "N", "n", "No", "no", "0"}:
                    queries.append(keyword)
        except Exception as e:
            print(f"⚠️ 读取文章关键词表失败，改用内置关键词: {e}")

    if not queries and not has_keyword_file:
        queries = QUERIES

    seen = set()
    unique = []
    for query in queries:
        key = query.lower()
        if key not in seen:
            seen.add(key)
            unique.append(query)
    return unique


def extract_emails_from_soup(soup: BeautifulSoup) -> list:
    """从页面提取邮箱（mailto链接 + 正文文本）"""
    found = []
    # 1. mailto 链接
    for a in soup.select("a[href^='mailto:']"):
        em = a["href"].replace("mailto:", "").split("?")[0].strip().lower()
        if em and EMAIL_RE.match(em):
            found.append(em)
    # 2. 页面全文 regex
    text = soup.get_text(" ", strip=True)
    text = html.unescape(text)
    found += [fix_glued_tld(e.lower()) for e in EMAIL_RE.findall(text)]
    for local, domain, suffix in OBFUSCATED_EMAIL_RE.findall(text):
        found.append(f"{local}@{domain}.{suffix}".lower())
    for script in soup.select("script[type='application/ld+json'], script"):
        script_text = html.unescape(script.get_text(" ", strip=True))
        found += [fix_glued_tld(e.lower()) for e in EMAIL_RE.findall(script_text)]
        for local, domain, suffix in OBFUSCATED_EMAIL_RE.findall(script_text):
            found.append(f"{local}@{domain}.{suffix}".lower())
    # 过滤无效/低价值邮箱
    found = [e for e in found
             if not re.search(r'\.(png|jpg|gif|mp4|webp|js|css)$', e, re.I)
             and "example.com" not in e
             and "sentry.io" not in e
             and "schema.org" not in e
             and "w3.org" not in e
             and len(e) < 80
             and not is_low_value_email(e)]
    found = unique_preserve(found)
    # 优先选 contact/info/press/media 开头的
    priority = [e for e in found
                if re.match(r'^(contact|info|press|media|hello|team|business|collab)', e)]
    return priority + [e for e in found if e not in priority]


def collect_contact_links(base_url: str, soup: BeautifulSoup) -> list:
    contact_hints = [
        "contact", "contact-us", "about", "about-us", "advertise", "advertising",
        "partner", "partners", "write-for-us", "guest-post", "guest-posts", "contribute",
        "submission", "submit", "editorial", "sponsorship", "sponsor", "collaborate",
        "collaboration", "work-with", "work-with-us", "hire", "team", "company",
        "press", "media-kit",
    ]
    parsed = urllib.parse.urlparse(base_url)
    root = f"{parsed.scheme}://{parsed.netloc}"
    candidates = []
    for a in soup.select("a[href]"):
        href_raw = a.get("href", "")
        href = href_raw.lower()
        text = a.get_text(" ", strip=True).lower()
        if any(h in href or h in text for h in contact_hints):
            full_url = urllib.parse.urljoin(root, href_raw)
            if get_base_domain(full_url) != get_base_domain(base_url):
                continue
            candidates.append(full_url)
    for path in [
        "/contact", "/contact-us", "/about", "/about-us", "/advertise", "/advertising",
        "/write-for-us", "/guest-post", "/contribute", "/sponsor", "/sponsorship",
        "/collaborate", "/work-with-us", "/team", "/editorial", "/press",
    ]:
        candidates.append(root + path)
    return unique_preserve(candidates)


def search_site_contact_pages(base_url: str) -> list:
    domain = get_base_domain(base_url)
    if not domain:
        return []
    query = f"site:{domain} contact OR advertise OR write for us OR guest post OR sponsorship"
    results, engine = search_web(query, n=CONTACT_SEARCH_PER_DOMAIN)
    if results:
        print(f"      ↳ 联系页补搜来源: {engine} {len(results)}条")
    return [url for _, url in results if get_base_domain(url) == domain]


def find_contact_email(base_url: str, soup: BeautifulSoup) -> tuple:
    """
    1. 从当前页找邮箱
    2. 找 /contact /about /advertise /work-with-us 链接并抓
    返回 (email, source_url)
    """
    emails = extract_emails_from_soup(soup)
    if emails:
        return emails[0], base_url

    checked = set()
    for contact_url in collect_contact_links(base_url, soup):
        if contact_url in checked:
            continue
        checked.add(contact_url)
        time.sleep(DELAY_FETCH)
        sub_soup, sub_url = fetch(contact_url)
        if sub_soup:
            sub_emails = extract_emails_from_soup(sub_soup)
            if sub_emails:
                return sub_emails[0], sub_url

    # 站内补搜 contact/advertise/write for us 页面
    for contact_url in search_site_contact_pages(base_url):
        if contact_url in checked:
            continue
        checked.add(contact_url)
        time.sleep(DELAY_FETCH)
        sub_soup, sub_url = fetch(contact_url)
        if sub_soup:
            sub_emails = extract_emails_from_soup(sub_soup)
            if sub_emails:
                return sub_emails[0], sub_url

    return "", base_url


def get_site_name(url: str, soup: BeautifulSoup) -> str:
    """获取网站名（title 或 og:site_name）"""
    og = soup.find("meta", property="og:site_name")
    if og and og.get("content"):
        return og["content"].strip()
    title = soup.find("title")
    if title:
        t = title.get_text(strip=True)
        # 去掉 "| Site" "- Site" 后缀
        t = re.split(r'[|\-–—]', t)[0].strip()
        return t[:50]
    return get_base_domain(url).replace("www.", "").split(".")[0].title()


def make_subject_opening(site_name: str, article_title: str, article_url: str) -> tuple:
    """生成个性化主题和开头"""
    title_short = article_title[:70] + ("..." if len(article_title) > 70 else "")
    domain = get_base_domain(article_url)

    if "topaz" in article_title.lower() or "alternative" in article_title.lower():
        subj = f"HitPaw VikPea — a Topaz alternative worth adding to your article"
    elif "best" in article_title.lower() or "top" in article_title.lower():
        subj = f"HitPaw VikPea for your '{article_title[:40]}...' roundup"
    elif "review" in article_title.lower():
        subj = f"HitPaw VikPea — would love a mention in your review"
    elif "comparison" in article_title.lower() or "vs" in article_title.lower():
        subj = f"HitPaw VikPea — missing from your AI video tool comparison?"
    else:
        subj = f"HitPaw VikPea — an AI video enhancer worth featuring on {site_name}"

    opening = (
        f'I came across your article "{title_short}" on {domain} — '
        f"really thorough breakdown for readers comparing AI video tools."
    )
    return subj, opening


def load_existing(tracker_path: str, queue_path: str) -> tuple:
    """返回 (existing_emails set, existing_domains set)"""
    emails = set()
    domains = set()
    for path in [tracker_path, queue_path]:
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    s = str(cell or "")
                    if "@" in s:
                        m = EMAIL_RE.search(s)
                        if m:
                            emails.add(m.group(0).lower())
                    if "http" in s:
                        d = get_base_domain(s)
                        if d:
                            domains.add(d)
    return emails, domains


def merge_search_results(query: str, limit: int) -> tuple:
    all_results = []
    engines = []
    api_results, api_info = search_serp_api(query, n=limit)
    if api_results:
        return api_results[:limit], api_info
    if api_info:
        engines.append(api_info)
    ddg, ddg_debug = search_duckduckgo(query, n=limit)
    if ddg:
        all_results.extend(ddg)
        engines.append(f"DDG:{len(ddg)}")
    bing, bing_debug = search_bing(query, n=limit)
    if bing:
        all_results.extend(bing)
        engines.append(f"Bing:{len(bing)}")

    unique = []
    seen_urls = set()
    for title, url in all_results:
        clean = normalize_search_result_url(url).rstrip("/")
        if not clean.startswith("http"):
            continue
        if should_skip(clean):
            continue
        if clean in seen_urls:
            continue
        seen_urls.add(clean)
        unique.append((title, clean))
        if len(unique) >= limit:
            break
    if engines:
        return unique, " + ".join(engines)
    return unique, format_engine_debug(ddg_debug, bing_debug)


def ensure_queue_headers(ws):
    headers = {
        Q_NAME: "频道名",
        Q_EMAIL: "邮箱",
        Q_SUBJ: "定制主题",
        Q_OPEN: "定制开头",
        Q_LINK: "主页链接",
        Q_NOTE: "备注",
        Q_TYPE: "类型",
        Q_SOURCE: "来源关键词",
    }
    for col, title in headers.items():
        if not ws.cell(1, col).value:
            ws.cell(1, col).value = title


def site_homepage(url: str) -> str:
    parsed = urllib.parse.urlparse(str(url or "").strip())
    if not parsed.scheme or not parsed.netloc:
        return str(url or "").strip()
    return f"{parsed.scheme}://{parsed.netloc}/"


def write_to_queue(ws, name, email, subj, opening, article_url, site_domain, found_email: bool, source_query=""):
    home_url = site_homepage(article_url)
    source_note = f"来源文章: {article_url}" if article_url else ""
    if not found_email:
        if add_no_email_candidate:
            add_no_email_candidate(
                name=name,
                link=home_url,
                note=f"需手动查联系方式 | {site_domain}" + (f" | {source_note}" if source_note else ""),
                email_type="文章外链",
                source=source_query,
                subject=subj,
                opening=opening,
                status="待深度查邮箱",
            )
        return None
    r = ws.max_row + 1
    ws.cell(r, Q_NAME).value  = name
    ws.cell(r, Q_EMAIL).value = email
    ws.cell(r, Q_SUBJ).value  = subj
    ws.cell(r, Q_OPEN).value  = opening
    ws.cell(r, Q_LINK).value  = home_url
    ws.cell(r, Q_NOTE).value  = source_note
    ws.cell(r, Q_TYPE).value  = "文章外链"
    ws.cell(r, Q_SOURCE).value = source_query
    for c in range(1, 9):
        ws.cell(r, c).fill = GREEN
    return r


def append_found_email_to_queue(name, email, subj, opening, article_url, site_domain, source_query=""):
    """重新读取最新主表后追加有邮箱行，避免长时间搜索覆盖主表。"""
    wb = openpyxl.load_workbook(QUEUE_PATH)
    ws = wb.active
    ensure_queue_headers(ws)
    rownum = write_to_queue(ws, name, email, subj, opening, article_url, site_domain, True, source_query)
    wb.save(QUEUE_PATH)
    return rownum


def add_article_pending_email(name, email, confidence, source_label, article_url, source_query, note):
    if not add_pending_email_review_row:
        return ""
    return add_pending_email_review_row(
        name=name,
        email=email,
        confidence=confidence,
        source_label=source_label,
        link=site_homepage(article_url),
        source=source_query,
        email_type="文章外链",
        note=note,
        video_link="",
        home_link=site_homepage(article_url),
    )


def main():
    if apply_config:
        apply_config(globals(), {
            "ARTICLE_RESULTS_PER_QUERY": "RESULTS_PER_QUERY",
            "ARTICLE_MIN_SITE_SCORE": "ARTICLE_MIN_SITE_SCORE",
            "RESPECT_ROBOTS_TXT": "RESPECT_ROBOTS_TXT",
            "CRAWL_DELAY_PER_DOMAIN": "CRAWL_DELAY_PER_DOMAIN",
            "SERP_PROVIDER": "SERP_PROVIDER",
            "SERPER_API_KEY": "SERPER_API_KEY",
            "SERPAPI_KEY": "SERPAPI_KEY",
            "DATAFORSEO_LOGIN": "DATAFORSEO_LOGIN",
            "DATAFORSEO_PASSWORD": "DATAFORSEO_PASSWORD",
        })
    print(f"\n{'═'*65}")
    print(f"  🚀 VikPea 文章站点批量搜索")
    print(f"{'═'*65}\n")

    # 检查依赖
    try:
        import requests, bs4
    except ImportError:
        print("❌ pip3 install requests beautifulsoup4")
        sys.exit(1)

    # 加载现有数据去重
    existing_emails, existing_domains = load_existing(TRACKER_PATH, QUEUE_PATH)
    no_email_keys = existing_no_email_keys() if existing_no_email_keys else set()
    print(f"已有数据：{len(existing_emails)} 邮箱 / {len(existing_domains)} 域名（去重用）\n")

    if not os.path.exists(QUEUE_PATH):
        print(f"❌ 找不到发信名单: {QUEUE_PATH}"); sys.exit(1)
    if ensure_pending_email_review_workbook:
        ensure_pending_email_review_workbook()

    added_green  = 0
    added_yellow = 0
    added_pending = 0
    skipped      = 0
    queries = load_queries()
    if not queries:
        print(f"⚠️ 当前没有启用的文章关键词。")
        print(f"↳ 请先到 {KEYWORD_PATH} 把需要使用的关键词第2列改成“是”，再重新运行。")
        return
    print(f"本次文章搜索关键词：{len(queries)} 个；每个关键词取前 {RESULTS_PER_QUERY} 条结果")
    if os.path.exists(KEYWORD_PATH):
        print(f"关键词来源：{KEYWORD_PATH}\n")
    else:
        print("关键词来源：脚本内置关键词（如需自定义，可新建 VikPea_文章搜索关键词.xlsx）\n")

    for qi, query in enumerate(queries):
        print(f"\n[{qi+1}/{len(queries)}] 搜索: {query}")
        results, engine_info = merge_search_results(query, limit=RESULTS_PER_QUERY)
        print(f"  → {len(results)} 条结果 | 来源: {engine_info}")
        if not results:
            print("    ↳ 当前关键词没有进候选池。常见原因：搜索引擎拦截、网络/证书问题，或搜索结果全被黑名单域名过滤。")
        time.sleep(DELAY_SEARCH)

        for title, url in results:
            domain = get_base_domain(url)

            # 跳过黑名单域名
            if should_skip(url):
                skipped += 1
                continue

            pre_score, pre_reasons = score_article_candidate(title, url)
            if pre_score < 0:
                skipped += 1
                print(f"  ⏭️  {domain:<35} 跳过：{'、'.join(pre_reasons)}")
                continue

            # 域名去重（同站已有了就跳过）
            if domain in existing_domains:
                skipped += 1
                continue

            print(f"  📄 {domain:<35} ", end="", flush=True)

            # 抓页面
            soup, final_url = fetch(url)
            time.sleep(DELAY_FETCH)
            if not soup:
                print("❌ 无法访问")
                continue

            score, score_reasons = score_article_candidate(title, final_url, soup)
            if score < int(ARTICLE_MIN_SITE_SCORE):
                skipped += 1
                print(f"⏭️  站点分{score}<{ARTICLE_MIN_SITE_SCORE}：{'、'.join(score_reasons[:3])}")
                continue

            site_name = get_site_name(final_url, soup)
            no_email_key = (site_name.lower(), final_url.strip().rstrip("/").lower())
            if no_email_key in no_email_keys:
                skipped += 1
                print("⏭️  无邮箱候选已存在")
                continue

            # 找邮箱
            email, source = find_contact_email(final_url, soup)

            # 邮箱去重
            if email and email in existing_emails:
                skipped += 1
                print(f"⏭️  邮箱已存在")
                continue

            # 生成主题/开头
            subj, opening = make_subject_opening(site_name, title, url)

            # 写入名单：有邮箱才进主表；无邮箱只进副表，不保存旧主表。
            if email:
                confidence, email_note = classify_article_email(email, final_url, source)
                if confidence == "A":
                    append_found_email_to_queue(site_name, email, subj, opening, url, domain, query)
                else:
                    pending_result = add_article_pending_email(
                        site_name, email, confidence, "文章站候选邮箱",
                        final_url, query, f"{email_note} | 站点分{score}: {'、'.join(score_reasons[:4])}"
                    )
                    if str(pending_result).startswith("SKIP:"):
                        skipped += 1
                        print(f"⏭️  {email} 已存在/已联系：{str(pending_result)[5:]}")
                        continue
                    added_pending += 1
            else:
                write_to_queue(None, site_name, email, subj, opening, url, domain, False, query)

            existing_domains.add(domain)
            if email:
                existing_emails.add(email)
                if confidence == "A":
                    added_green += 1
                    print(f"✅ {email} | A级 | 站点分{score}")
                else:
                    print(f"🔵 {email} | {confidence}级待确认 | {email_note}")
            else:
                no_email_keys.add(no_email_key)
                added_yellow += 1
                print(f"🔴 无公开邮箱，已放入无邮箱候选 | 站点分{score}")

    print(f"\n{'═'*65}")
    print(f"  ✅ 完成！主表有邮箱 +{added_green}  待确认 +{added_pending}  无邮箱候选 +{added_yellow}  跳过{skipped}")
    print(f"  → 运行 VikPea_读表发信.py 发送绿行邮件")
    print(f"{'═'*65}\n")
    if log_event:
        log_event("文章搜索", f"绿行 +{added_green} 待确认 +{added_pending} 黄行 +{added_yellow} 跳过 {skipped}")


if __name__ == "__main__":
    main()
