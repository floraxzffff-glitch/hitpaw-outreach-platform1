"""
VikPea_YouTube批量搜索.py — 批量搜索 YouTube KOL + 自动找邮箱 → 写入发信名单

流程：
  关键词 → yt-dlp 搜索视频 → 提取频道 → 过滤粉丝数 → 去重（对比追踪表）
  → 抓频道描述找邮箱 → 有邮箱写绿行 / 无邮箱写黄行 → VikPea_发信名单.xlsx

依赖（在你Mac终端运行）:
  pip3 install yt-dlp openpyxl

用法: python3 ~/Downloads/VikPea工作台/VikPea_YouTube批量搜索.py
"""

import subprocess, re, time, os, sys, json, shutil, warnings, ssl, socket
import urllib.error
import urllib.request, urllib.parse
from datetime import datetime, timedelta
from xml.etree import ElementTree as ET

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("⚠️ python-dotenv not installed, will use hardcoded config")

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("❌ pip3 install openpyxl"); sys.exit(1)

try:
    from VikPea_common import (
        apply_config, log_event, add_no_email_candidate, existing_no_email_keys,
        classify_bad_email, email_relevance_reason, linked_root_domains_from_text,
        add_pending_email_review_row, cleanup_pending_email_review_known,
        archive_processed_pending_email_rows, fix_glued_tld,
    )
except ImportError:
    apply_config = None
    log_event = None
    add_no_email_candidate = None
    existing_no_email_keys = None
    classify_bad_email = None
    email_relevance_reason = None
    linked_root_domains_from_text = None
    add_pending_email_review_row = None
    cleanup_pending_email_review_known = None
    archive_processed_pending_email_rows = None
    fix_glued_tld = lambda e: e

try:
    from VikPea_深度找邮箱 import deep_find_email
except ImportError:
    deep_find_email = None

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

# yt-dlp 可执行文件路径（兼容多种安装方式）
def find_ytdlp() -> list:
    # 1. 优先用当前 Python 环境里的 yt_dlp，避免误用旧版 3.9
    try:
        subprocess.run([sys.executable, "-m", "yt_dlp", "--version"],
                       capture_output=True, check=True)
        return [sys.executable, "-m", "yt_dlp"]
    except Exception:
        pass
    # 2. 系统 PATH 里找
    p = shutil.which("yt-dlp")
    if p:
        return [p]
    # 3. 当前 Python 对应的用户级安装路径（pip install --user）
    ver = f"{sys.version_info.major}.{sys.version_info.minor}"
    user_bin = os.path.expanduser(f"~/Library/Python/{ver}/bin/yt-dlp")
    if os.path.exists(user_bin):
        return [user_bin]
    return []

YTDLP_CMD = find_ytdlp()

# ── 配置 ────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR     = os.path.dirname(SCRIPT_DIR)
QUEUE_PATH   = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
KEYWORD_PATH = os.path.join(SCRIPT_DIR, "VikPea_搜索关键词.xlsx")
KEYWORD_METRICS_PATH = os.path.join(SCRIPT_DIR, "VikPea_关键词搜索记录.xlsx")
SEMRUSH_CANDIDATES = [
    os.path.join(SCRIPT_DIR, "video-quality_broad-match_us_2026-06-26.xlsx"),
    os.path.join(ROOT_DIR, "video-quality_broad-match_us_2026-06-26.xlsx"),
]

SUB_MIN     = 1_000      # 最少粉丝数；小博主模式放低门槛
SUB_MAX     = 250_000    # 最多粉丝数；优先更好谈的小中型频道
SEARCH_N    = 150        # 每个关键词取多少视频结果（扩大候选池：捕获关键词在简介/tag里的视频）
DELAY_SEARCH = 3         # 两次搜索间隔（秒）
DELAY_FETCH  = 2         # 两次频道描述抓取间隔（秒）
MIN_VIDEO_VIEWS = 800
MIN_SHORTS_VIEWS = 2500
MIN_RECENT_AVG_VIEWS = 1000
RECENT_VIDEO_COUNT = 5
ACTIVE_WITHIN_DAYS = 30
VIDEO_METRICS_TIMEOUT = 15
RECENT_CHANNEL_TIMEOUT = 20
YOUTUBE_API_KEY = ""
YTDLP_COOKIES_FROM_BROWSER = ""
YTDLP_RETRY_TIMES = 2
KEYWORD_FRESHNESS_SAMPLE = 10
KEYWORD_MIN_FRESH_HITS = 0
USE_SEMRUSH_FOR_YOUTUBE = False
YOUTUBE_KEYWORD_FILTER = True
SMALL_CREATOR_MARKET_MODE = True
MARKET_SCORE_MIN = 2      # 越高越严格；2=偏向发展中市场受众但不筛太死
YOUTUBE_API_DELAY_SEC = 0.35
YOUTUBE_API_RETRY_TIMES = 3
YOUTUBE_API_429_COOLDOWN = 8
SHORTS_MAX_SECONDS = 70
DEEPSEEK_API_KEY = ""
DEEPSEEK_API_BASE = ""
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
ANTHROPIC_API_BASE = os.getenv("ANTHROPIC_API_BASE", "https://api.vectorengine.ai/v1")
ANTHROPIC_TAG_MODEL = "claude-3-5-sonnet-20241022"

# 频道综合评分权重配置
CHANNEL_SCORE_WEIGHTS = {
    "content_relevance": 0.5,
    "commercial_history": 0.2,
    "audience_fit": 0.15,
    "avg_views_normalized": 0.15,
}

# 默认关键词。若同目录存在 VikPea_搜索关键词.xlsx，会优先使用表格第一列。
BASE_KEYWORDS = [
    "topaz video ai alternative 2026",
    "best ai video enhancer comparison 2026",
    "ai video upscaling software review",
    "video quality enhancer tutorial 2026",
    "hitpaw video enhancer review",
    "unifab video enhancer vs topaz",
    "aiarty video enhancer review",
    "ai video upscaler free 2026",
    "video noise reduction ai tool",
    "4k video upscaling software comparison",
    "ai video restoration tutorial",
    "video enhance ai review",
    "best video upscaling software 2026",
    "topaz video ai tutorial 2026",
    "ai video quality improvement tool",
    "free video enhancer android",
    "capcut video quality enhancer",
    "kinemaster video quality tutorial",
    "vn editor video quality settings",
    "how to improve video quality in capcut",
    "low end pc video enhancer",
    "video quality enhancer no watermark",
    "best free video enhancer app",
    "how to make blurry video clear android",
    "mobile video editing quality tutorial",
]

GREEN  = PatternFill("solid", start_color="E2EFDA")
YELLOW = PatternFill("solid", start_color="FFF2CC")
EMAIL_RE = re.compile(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}")
SKIP_EMAIL_DOMAINS = {
    "example", "schema", "sentry", "w3.org", "google", "youtube",
    "facebook", "twitter", "instagram", "tiktok", "cloudflare",
    "random", "localhost", "invalid", "duckduckgo", "bing", "yahoo",
    "hitpaw", "tenorshare", "topazlabs", "topaz", "vanceai", "avclabs",
    "wondershare", "aiarty", "patreon", "invideo", "openshot", "envato", "pxf.io", "o.market"
}
BAD_TLDS = {"png", "jpg", "jpeg", "gif", "webp", "svg", "avif", "css", "js", "json", "xml"}
BAD_LOCAL_PARTS = {"www", "http", "https", "mailto", "image", "img", "src", "href"}
SUSPICIOUS_LOCAL_FRAGMENTS = {"www.", "1.env", "6-months-l", "tunacoretalentcre", "ai-category-card"}
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

_SSL_FALLBACK_WARNED = False
YOUTUBE_API_LAST_ERROR = ""
YOUTUBE_API_LAST_TS = 0.0
YOUTUBE_API_RATE_LIMITED_UNTIL = 0.0


def safe_urlopen(req_or_url, timeout=20):
    """
    某些本机 Python（尤其新装的 python.org 版本）会缺系统证书链，
    导致 urllib 请求 Google/YouTube 直接 SSL 校验失败。
    这里做一个更稳的兜底，避免 API 明明配好了却全部静默失败。
    """
    global _SSL_FALLBACK_WARNED
    try:
        return urllib.request.urlopen(req_or_url, timeout=timeout)
    except ssl.SSLCertVerificationError:
        pass
    except urllib.error.URLError as exc:
        if not isinstance(getattr(exc, "reason", None), ssl.SSLCertVerificationError):
            raise

    contexts = []
    try:
        import certifi
        contexts.append(ssl.create_default_context(cafile=certifi.where()))
    except Exception:
        pass
    contexts.append(ssl._create_unverified_context())

    last_exc = None
    for ctx in contexts:
        try:
            if not _SSL_FALLBACK_WARNED:
                print("  ⚠️ 本机 Python 证书链异常，已启用 HTTPS 兼容兜底")
                _SSL_FALLBACK_WARNED = True
            return urllib.request.urlopen(req_or_url, timeout=timeout, context=ctx)
        except Exception as exc:
            last_exc = exc
            continue
    if last_exc:
        raise last_exc
    raise RuntimeError("safe_urlopen 未能建立 HTTPS 连接")


def get_channel_bio(channel_id: str) -> str:
    if not YOUTUBE_API_KEY or not channel_id:
        return ""
    data = youtube_api_get("channels", {"part": "snippet", "id": channel_id})
    items = data.get("items") or []
    if not items:
        return ""
    return str((items[0].get("snippet") or {}).get("description") or "").strip()[:500]


def get_video_top_comments(video_id: str, max_comments: int = 10) -> list:
    """获取视频的高赞评论"""
    if not YOUTUBE_API_KEY or not video_id:
        return []
    try:
        data = youtube_api_get("commentThreads", {
            "part": "snippet",
            "videoId": video_id,
            "order": "relevance",  # 按相关性（通常是高赞）排序
            "maxResults": max_comments,
            "textFormat": "plainText"
        })
        items = data.get("items") or []
        comments = []
        for item in items:
            snippet = (item.get("snippet") or {}).get("topLevelComment", {}).get("snippet", {})
            text = (snippet.get("textDisplay") or "").strip()
            likes = snippet.get("likeCount", 0)
            if text:
                comments.append({"text": text[:200], "likes": likes})
        return comments
    except Exception as exc:
        print(f"      ↳ 获取评论失败：{exc}")
        return []


def call_deepseek_for_tag(channel_name: str, bio: str, titles: list) -> str:
    if not DEEPSEEK_API_KEY:
        return ""
    if not titles and not bio:
        return ""
    titles_text = "\n".join(f"- {t}" for t in titles[:10]) if titles else "（无）"
    bio_text = bio[:300] if bio else "（无）"
    prompt = (
        f"YouTube频道名：{channel_name}\n"
        f"频道简介：{bio_text}\n"
        f"最近视频标题：\n{titles_text}\n\n"
        "请用1-3个中文词（逗号分隔）描述这个频道的内容类别，"
        "例如：科技测评、手机摄影、游戏解说、教程、Vlog、美妆、健身等。"
        "只输出分类词，不要任何解释或标点之外的内容。"
    )
    payload = json.dumps({
        "model": "deepseek-chat",
        "max_tokens": 30,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    api_base = DEEPSEEK_API_BASE if DEEPSEEK_API_BASE else "https://api.deepseek.com"
    req = urllib.request.Request(
        f"{api_base}/chat/completions",
        data=payload,
        headers={
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
            "content-type": "application/json",
        },
        method="POST",
    )
    try:
        resp = safe_urlopen(req, timeout=12)
        body = json.loads(resp.read().decode("utf-8"))
        text = ((body.get("choices") or [{}])[0].get("message", {}).get("content") or "").strip()
        return text[:60]
    except Exception as exc:
        print(f"      ↳ AI标签获取失败：{exc}")
        return ""


def call_claude_for_channel_scoring(channel_name: str, bio: str, titles: list,
                                     subs: int, hit_video_title: str, product_context: str = "HitPaw VikPea 视频增强工具") -> dict:
    """
    频道综合画像 AI 评分 - 评估整个频道是否适合合作
    返回: {
        "content_relevance": int (0-100),
        "commercial_history": int (0-100),
        "audience_fit": int (0-100),
        "verdict": str ("推荐" | "待人工确认" | "不推荐"),
        "reasoning": str
    }
    """
    if not ANTHROPIC_API_KEY:
        return {
            "content_relevance": 0,
            "commercial_history": 0,
            "audience_fit": 0,
            "verdict": "待人工确认",
            "reasoning": "AI评分未启用（缺少API密钥）"
        }

    titles_text = "\n".join(f"- {t}" for t in titles[:10]) if titles else "（无）"
    bio_text = bio[:500] if bio else "（无）"

    prompt = f"""你是一个 KOL 筛选专家。请评估以下 YouTube 频道是否适合推广"{product_context}"。

频道信息：
- 频道名：{channel_name}
- 订阅数：{subs:,}
- 频道简介：{bio_text}
- 近期视频标题：
{titles_text}
- 命中的视频标题：{hit_video_title}

评分标准：
1. **内容相关度 (0-100分)**：
   - 重要：不要只看命中的那一条视频，要综合判断整个频道的内容定位
   - 即使频道从未直接提到产品关键词，只要整体内容定位与视频增强/视频编辑/视频质量提升相关，也应给高分
   - 例如：视频编辑教程频道、视频后期制作频道、摄影摄像频道、内容创作者工具测评频道等都高度相关
   - 泛娱乐、游戏实况、生活Vlog等不相关内容给低分

2. **商业化历史 (0-100分)**：
   - 基于近期视频标题判断是否有商业合作痕迹
   - 关键词：sponsored、ad、partnership、review（产品测评）、vs（产品对比）、affiliate、promo code
   - 提到其他软件品牌名（Topaz、Aiarty、DemoCreator、Adobe、CapCut等）通常说明有商业合作经验
   - 有明显商业化痕迹给高分(80-100)，偶尔商业化给中分(40-70)，无商业化痕迹给低分(0-30)

3. **受众匹配度 (0-100分)**：
   - 基于内容风格判断受众是否偏工具型/技术型（而非泛娱乐型）
   - 教程类、测评类、对比类、工作流分享类内容的受众更匹配 → 高分
   - 纯娱乐、搞笑、剧情类内容的受众不匹配 → 低分

请严格按以下JSON格式输出（不要添加任何markdown格式符号或额外文本）：
{{
  "content_relevance": <0-100的整数>,
  "commercial_history": <0-100的整数>,
  "audience_fit": <0-100的整数>,
  "verdict": "<推荐|待人工确认|不推荐>",
  "reasoning": "<一句话说明判断依据，50字以内>"
}}"""

    payload = json.dumps({
        "model": ANTHROPIC_TAG_MODEL,
        "max_tokens": 300,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")

    api_base = ANTHROPIC_API_BASE if ANTHROPIC_API_BASE else "https://api.anthropic.com"

    # 重试逻辑：3次重试，指数退避
    for attempt in range(3):
        try:
            req = urllib.request.Request(
                f"{api_base}/messages",
                data=payload,
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                method="POST",
            )

            resp = safe_urlopen(req, timeout=20)
            body = json.loads(resp.read().decode("utf-8"))
            content = body.get("content", [])
            if not content or len(content) == 0:
                if attempt < 2:
                    time.sleep(1 << attempt)  # 1s, 2s, 4s
                    continue
                return {
                    "content_relevance": 0,
                    "commercial_history": 0,
                    "audience_fit": 0,
                    "verdict": "待人工确认",
                    "reasoning": "AI评分调用失败（响应为空）"
                }

            text = content[0].get("text", "").strip()

            # 尝试解析JSON
            # 去掉可能的markdown代码块标记
            text = re.sub(r"^```json\s*", "", text)
            text = re.sub(r"\s*```$", "", text)

            try:
                result = json.loads(text)
                # 验证必需字段
                if all(k in result for k in ["content_relevance", "commercial_history", "audience_fit", "verdict", "reasoning"]):
                    # 确保数值在0-100范围内
                    result["content_relevance"] = max(0, min(100, int(result["content_relevance"])))
                    result["commercial_history"] = max(0, min(100, int(result["commercial_history"])))
                    result["audience_fit"] = max(0, min(100, int(result["audience_fit"])))
                    # 确保verdict是有效值
                    if result["verdict"] not in ["推荐", "待人工确认", "不推荐"]:
                        result["verdict"] = "待人工确认"
                    return result
            except (json.JSONDecodeError, ValueError, KeyError):
                if attempt < 2:
                    time.sleep(1 << attempt)
                    continue
                pass

            # JSON解析失败，返回默认值
            return {
                "content_relevance": 0,
                "commercial_history": 0,
                "audience_fit": 0,
                "verdict": "待人工确认",
                "reasoning": "AI评分调用失败（JSON解析失败）"
            }

        except Exception as exc:
            print(f"      ↳ AI频道评分失败（尝试 {attempt+1}/3）：{exc}")
            if attempt < 2:
                time.sleep(1 << attempt)
                continue
            return {
                "content_relevance": 0,
                "commercial_history": 0,
                "audience_fit": 0,
                "verdict": "待人工确认",
                "reasoning": f"AI评分调用失败：{str(exc)[:30]}"
            }

    # 所有重试都失败
    return {
        "content_relevance": 0,
        "commercial_history": 0,
        "audience_fit": 0,
        "verdict": "待人工确认",
        "reasoning": "AI评分调用失败（重试3次后仍失败）"
    }


def call_claude_for_analysis(channel_name: str, bio: str, titles: list, comments: list) -> dict:
    """
    使用 Claude API (via VectorEngine) 进行完整频道分析
    返回: {
        "tags": str,
        "vertical_score": int (1-10),
        "competitor_mentioned": bool,
        "competitor_names": list,
        "collaboration_type": str ("插链接" or "Dedicated")
    }
    """
    if not ANTHROPIC_API_KEY:
        return {}

    titles_text = "\n".join(f"- {t}" for t in titles[:5]) if titles else "（无）"
    bio_text = bio[:400] if bio else "（无）"
    comments_text = "\n".join(f"- {c.get('text', '')} ({c.get('likes', 0)}👍)"
                               for c in comments[:10]) if comments else "（无）"

    prompt = f"""分析以下 YouTube 频道信息：

频道名：{channel_name}

频道简介：
{bio_text}

最近5个视频标题：
{titles_text}

评论区高赞评论：
{comments_text}

请按以下格式输出分析结果（严格遵守格式，不要添加任何解释）：

频道标签：[2-4个中文词，逗号分隔]
内容垂直度评分：[1-10的数字]
是否推过竞品：[是/否]
竞品名称：[如果推过，列出产品名，如 Topaz/Aiarty/DemoCreator 等；如果没推过，写"无"]
建议合作方式：[插链接/Dedicated]

评分标准：
- 内容垂直度：10分=专注单一领域，8-9分=主要聚焦某领域但偶尔跨界，5-7分=内容较分散，1-4分=杂乱无章
- 竞品：检查是否提到过 Topaz Video AI, Aiarty, DemoCreator, HitPaw, AVCLabs, UniFab, VanceAI 等视频增强/编辑软件
- 合作方式：如果是专业测评频道或经常做产品对比，建议"Dedicated"；如果是教程类或偶尔提及工具，建议"插链接"
"""

    payload = json.dumps({
        "model": ANTHROPIC_TAG_MODEL,
        "max_tokens": 500,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")

    api_base = ANTHROPIC_API_BASE if ANTHROPIC_API_BASE else "https://api.anthropic.com"
    req = urllib.request.Request(
        f"{api_base}/messages",
        data=payload,
        headers={
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        method="POST",
    )

    try:
        resp = safe_urlopen(req, timeout=15)
        body = json.loads(resp.read().decode("utf-8"))
        content = body.get("content", [])
        if not content or len(content) == 0:
            return {}

        text = content[0].get("text", "").strip()

        # 解析返回的结构化文本
        result = {
            "tags": "",
            "vertical_score": 0,
            "competitor_mentioned": False,
            "competitor_names": [],
            "collaboration_type": ""
        }

        for line in text.split("\n"):
            line = line.strip()
            if "频道标签" in line:
                tags = re.sub(r"^[^：:]+[：:]", "", line).strip()
                result["tags"] = tags[:60]
            elif "内容垂直度评分" in line or "垂直度评分" in line:
                match = re.search(r"(\d+)", line)
                if match:
                    result["vertical_score"] = min(10, max(1, int(match.group(1))))
            elif "是否推过竞品" in line or "是否提到竞品" in line:
                result["competitor_mentioned"] = "是" in line
            elif "竞品名称" in line:
                names = re.sub(r"^[^：:]+[：:]", "", line).strip()
                if names and names != "无":
                    result["competitor_names"] = [n.strip() for n in names.split("/") if n.strip()]
            elif "建议合作方式" in line or "合作方式" in line:
                if "Dedicated" in line:
                    result["collaboration_type"] = "Dedicated"
                elif "插链接" in line:
                    result["collaboration_type"] = "插链接"

        return result

    except Exception as exc:
        print(f"      ↳ AI分析失败：{exc}")
        return {}


Q_NAME  = 1; Q_EMAIL = 2; Q_SUBJ = 3; Q_OPEN = 4
Q_HOME  = 5; Q_VIDEO = 6; Q_NOTE  = 7; Q_TYPE = 8; Q_SOURCE = 9; Q_TAG = 10
Q_VERTICAL_SCORE = 11; Q_COMPETITOR = 12; Q_COLLAB_TYPE = 13
# 新增：频道综合评分相关列
Q_CONTENT_RELEVANCE = 14; Q_COMMERCIAL_HISTORY = 15; Q_AUDIENCE_FIT = 16
Q_COMPOSITE_SCORE = 17; Q_AI_VERDICT = 18; Q_AI_REASONING = 19
Q_AVG_RECENT_VIEWS = 20
# ────────────────────────────────────────────────────────────────

MARKET_POSITIVE_SIGNALS = {
    "review": 3, "tutorial": 2, "comparison": 2, "alternative": 3,
    "vs": 2, "best": 1, "software": 1,
    "topaz": 3, "hitpaw": 3, "avclabs": 3, "vanceai": 3,
    "unifab": 3, "aiarty": 3, "video enhancer": 2, "video upscaler": 2,
    "video enhancement": 2,
    "free": 2, "android": 3, "mobile": 2, "phone": 2, "app": 1,
    "capcut": 4, "kinemaster": 4, "vn editor": 4, "inshot": 3,
    "alight motion": 3, "no watermark": 3, "without watermark": 3,
    "low end": 4, "low-end": 4, "2gb ram": 4, "4gb ram": 3,
    "beginner": 2, "how to": 1, "fix": 1,
    "hindi": 4, "urdu": 4, "tagalog": 4, "filipino": 4,
    "bahasa": 4, "indonesia": 3, "bangla": 4, "bengali": 4,
    "tamil": 4, "telugu": 4, "malayalam": 4, "vietnamese": 4,
    "thai": 4, "arabic": 3, "spanish": 2, "portuguese": 2,
}

MARKET_NEGATIVE_SIGNALS = {
    "enterprise": -4, "agency": -3, "official": -3, "news": -2,
    "conference": -3, "webinar": -2, "course": -2,
    "top 10": -1, "best software 2026": -1,
}

BAD_YOUTUBE_KEYWORD_PATTERNS = [
    r"\bwhy (does|is|are|do|did|can)\b",
    r"\blower\b",
    r"\breduce\b",
    r"\breducer\b",
    r"\bdecrease\b",
    r"\bdecreaser\b",
    r"\bdownload(er)?\b",
    r"\binternet archive\b",
    r"\bsound quality\b",
    r"\baudio quality\b",
    r"\bquality of service\b",
    r"\bfcp\b",
    r"\bchrome store\b",
    r"\bextension\b",
]


def youtube_keyword_skip_reason(keyword: str) -> str:
    """过滤掉更像用户问题/下载/压缩/音频的词，避免拿它们找KOL。"""
    text = str(keyword or "").strip().lower()
    if not text:
        return "空关键词"
    if len(text.split()) <= 2 and "topaz" not in text and "capcut" not in text:
        return "过宽泛"
    for pattern in BAD_YOUTUBE_KEYWORD_PATTERNS:
        if re.search(pattern, text):
            return "非KOL合作意图"
    return ""


def clean_emails(text: str) -> list:
    """从文本里提取更像真实联系邮箱的地址。"""
    seen, out = set(), []
    for email in EMAIL_RE.findall(text or ""):
        e = fix_glued_tld(email.strip(" .,:;()[]<>").lower())
        if e == "error-lite@duckduckgo.com":
            continue
        if any(ch in e for ch in {'"', "'", " "}) or re.search(r"%(?:22|20|2b|40|2e|2d|5f|[0-9a-f]{2})", e, re.I):
            continue
        if re.search(r"\.(png|jpg|jpeg|gif|mp4|webp|svg)$", e, re.I):
            continue
        domain = e.split("@", 1)[-1]
        local = e.split("@", 1)[0]
        domain_parts = domain.split(".")
        tld = domain_parts[-1] if domain_parts else ""
        main_domain = domain_parts[-2] if len(domain_parts) >= 2 else ""
        if tld in BAD_TLDS:
            continue
        if local in BAD_LOCAL_PARTS or any(fragment in local for fragment in SUSPICIOUS_LOCAL_FRAGMENTS):
            continue
        if len(domain) > 80 or len(local) > 64:
            continue
        if re.search(r"\d+x\.", domain) or "compressed" in domain:
            continue
        if len(local) < 2 or len(main_domain) < 2:
            continue
        if tld in SKIP_EMAIL_DOMAINS or any(skip in domain for skip in SKIP_EMAIL_DOMAINS):
            continue
        if e not in seen:
            seen.add(e)
            out.append(e)
    return out


def normalize_obfuscated_emails(text: str) -> str:
    text = text or ""
    text = re.sub(r"\s*(?:\[at\]|\(at\)|\{at\}|\bat\b)\s*", "@", text, flags=re.I)
    text = re.sub(r"\s*(?:\[dot\]|\(dot\)|\{dot\}|\bdot\b)\s*", ".", text, flags=re.I)
    return text


def extract_emails(text: str) -> list:
    combined = text or ""
    if "<html" not in combined.lower():
        combined += "\n" + normalize_obfuscated_emails(text or "")
    return clean_emails(combined)


def load_keyword_workbook(path: str, limit: int = 200) -> list:
    """从 VikPea_搜索关键词.xlsx 读取关键词；第二列=是否启用。"""
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        keywords = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            kw = str(row[0] or "").strip()
            status = str((row[1] if len(row) > 1 else "是") or "是").strip().lower()
            if not kw or kw.lower() in {"keyword", "keywords", "关键词"}:
                continue
            if status in {"停用", "禁用", "否", "no", "n", "0", "false"}:
                continue
            if kw not in keywords:
                keywords.append(kw)
            if len(keywords) >= limit:
                break
        print(f"  📄 从关键词表加载 {len(keywords)} 个关键词: {path}")
        return keywords
    except Exception as e:
        print(f"  ⚠️  关键词表读取失败: {e}")
        return []


def load_semrush_keywords(paths: list, limit: int = 30) -> list:
    """从 SEMrush xlsx 加载关键词"""
    path = next((p for p in paths if os.path.exists(p)), "")
    if not path:
        print("  ⚠️  未找到 SEMrush 关键词表，跳过追加关键词")
        return []
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        keywords = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            kw = str(row[0] or "").strip()
            if kw and len(keywords) < limit:
                keywords.append(kw)
        print(f"  📄 从 SEMrush 加载 {len(keywords)} 个关键词: {path}")
        return keywords
    except Exception as e:
        print(f"  ⚠️  SEMrush 读取失败: {e}")
        return []


def ensure_queue_headers(ws):
    if str(ws.cell(1, 6).value or "").strip() == "备注":
        ws.insert_cols(6, 1)
    if not str(ws.cell(1, 6).value or "").strip() and str(ws.cell(1, 7).value or "").strip() == "备注":
        ws.cell(1, 6).value = "视频链接"
    headers = {
        Q_NAME: "频道名",
        Q_EMAIL: "邮箱",
        Q_SUBJ: "定制主题",
        Q_OPEN: "定制开头",
        Q_HOME: "主页链接",
        Q_VIDEO: "视频链接",
        Q_NOTE: "备注",
        Q_TYPE: "类型",
        Q_SOURCE: "来源关键词",
        Q_TAG: "频道标签",
        Q_VERTICAL_SCORE: "垂直度",
        Q_COMPETITOR: "推过竞品",
        Q_COLLAB_TYPE: "合作方式",
        Q_CONTENT_RELEVANCE: "内容相关度",
        Q_COMMERCIAL_HISTORY: "商业化历史",
        Q_AUDIENCE_FIT: "受众匹配度",
        Q_COMPOSITE_SCORE: "综合分数",
        Q_AI_VERDICT: "AI判断",
        Q_AI_REASONING: "评分理由",
        Q_AVG_RECENT_VIEWS: "近期均播",
    }
    for col, title in headers.items():
        ws.cell(1, col).value = title


def update_keyword_metrics(path: str, keyword: str, found: int, eligible: int,
                           added_green: int, added_yellow: int, skipped: int):
    headers = [
        "日期", "关键词", "搜索到频道数", "符合粉丝范围数",
        "新增有邮箱数", "新增无邮箱数", "跳过数", "邮箱命中率"
    ]
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "搜索记录"
        ws.append(headers)

    if ws.max_row == 0:
        ws.append(headers)

    total_added = added_green + added_yellow
    email_rate = added_green / total_added if total_added else 0
    ws.append([
        datetime.now().strftime("%Y-%m-%d"),
        keyword,
        found,
        eligible,
        added_green,
        added_yellow,
        skipped,
        round(email_rate, 4),
    ])
    wb.save(path)


def load_existing_emails(tracker_path: str, queue_path: str) -> set:
    """加载已有联系邮箱（追踪表 + 发信名单），用于去重"""
    emails = set()
    for path in [tracker_path, queue_path]:
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    s = str(cell or "")
                    if "@" in s and "." in s:
                        m = EMAIL_RE.search(s)
                        if m:
                            emails.add(m.group(0).lower())
    return emails


def normalize_name_key(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"\s*[（(][^)）]*[)）]\s*$", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def header_index_map(ws) -> dict:
    mapping = {}
    for idx in range(1, ws.max_column + 1):
        title = str(ws.cell(1, idx).value or "").strip()
        if title:
            mapping[title] = idx - 1
    return mapping


def load_existing_channel_names(tracker_path: str, queue_path: str) -> set:
    """加载已有频道名（模糊去重），兼容旧表/空行/短行。"""
    names = set()
    for path in [tracker_path, queue_path]:
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path)
        for ws in wb.worksheets:
            headers = header_index_map(ws)
            if path == tracker_path:
                preferred_cols = [
                    headers.get("联系人/平台"),
                    headers.get("频道/文章/平台"),
                    2,
                    5,
                ]
            else:
                preferred_cols = [
                    headers.get("频道名"),
                    headers.get("名称"),
                    0,
                ]
            for row in ws.iter_rows(min_row=2, values_only=True):
                raw = ""
                for col in preferred_cols:
                    if col is None:
                        continue
                    if col < len(row):
                        candidate = row[col]
                        if str(candidate or "").strip():
                            raw = candidate
                            break
                n = normalize_name_key(raw)
                if n:
                    names.add(n)
    return names


def yt_dlp_base_args() -> list:
    args = ["--no-warnings", "--no-color"]
    # 对 YouTube 元数据抓取更稳一些，减少网页层请求。
    args += ["--extractor-args", "youtube:player_client=tv,mweb,web;player_skip=webpage,configs"]
    args += ["--extractor-args", "youtubetab:skip=webpage"]
    browser = str(YTDLP_COOKIES_FROM_BROWSER or "").strip()
    if browser:
        args += ["--cookies-from-browser", browser]
    return args


def can_resolve_host(host: str) -> bool:
    try:
        socket.getaddrinfo(host, 443)
        return True
    except Exception:
        return False


def network_ready_for_youtube() -> tuple:
    checks = [
        ("www.youtube.com", "YouTube"),
        ("www.googleapis.com", "YouTube API"),
    ]
    failed = [label for host, label in checks if not can_resolve_host(host)]
    return (not failed), failed


def should_retry_ytdlp(stderr_text: str) -> bool:
    text = str(stderr_text or "").lower()
    retry_signals = [
        "the page needs to be reloaded",
        "try again later",
        "timed out",
        "temporarily unavailable",
        "unable to download api page",
        "remote end closed connection",
        "http error 429",
    ]
    return any(signal in text for signal in retry_signals)


def is_auth_tab_error(stderr_text: str) -> bool:
    text = str(stderr_text or "").lower()
    signals = [
        "playlists that require authentication",
        "successful webpage download",
        "[youtube:tab]",
        "cookies are only for the first account and channel",
    ]
    return any(signal in text for signal in signals)


def normalize_ytdlp_error(stderr_text: str) -> str:
    text = short_error(stderr_text)
    if is_auth_tab_error(text):
        return "频道页抓取受限（YouTube 认证限制）"
    return text


def run_yt_dlp_result(*args, timeout=60):
    env = os.environ.copy()
    env["PYTHONWARNINGS"] = "ignore"
    attempts = max(int(YTDLP_RETRY_TIMES or 1), 1)
    last_result = None
    for i in range(attempts):
        try:
            result = subprocess.run(
                YTDLP_CMD + yt_dlp_base_args() + list(args),
                capture_output=True, text=True, timeout=timeout, env=env
            )
            last_result = result
            if result.returncode == 0 or not should_retry_ytdlp(result.stderr):
                return result
            if i < attempts - 1:
                time.sleep(2 + i)
        except subprocess.TimeoutExpired as exc:
            last_result = exc
            if i < attempts - 1:
                time.sleep(2 + i)
                continue
            return None
        except Exception:
            return None
    return last_result


def run_yt_dlp(*args, timeout=60) -> str:
    r = run_yt_dlp_result(*args, timeout=timeout)
    if not r:
        return ""
    return r.stdout.strip()


def run_yt_dlp_detailed(*args, timeout=60) -> tuple:
    """
    返回 (stdout, error_reason)。
    error_reason 为空表示命令本身成功执行，不代表业务上一定取到数据。
    """
    try:
        env = os.environ.copy()
        env["PYTHONWARNINGS"] = "ignore"
        attempts = max(int(YTDLP_RETRY_TIMES or 1), 1)
        last_error = ""
        for i in range(attempts):
            r = subprocess.run(
                YTDLP_CMD + yt_dlp_base_args() + list(args),
                capture_output=True, text=True, timeout=timeout, env=env
            )
            stdout = (r.stdout or "").strip()
            stderr = normalize_ytdlp_error(r.stderr)
            if stdout:
                return stdout, ""
            last_error = stderr or f"yt-dlp 返回码 {r.returncode}"
            if r.returncode == 0 or not should_retry_ytdlp(stderr) or i >= attempts - 1:
                break
            time.sleep(2 + i)
        return "", last_error or "yt-dlp 返回空结果"
    except subprocess.TimeoutExpired:
        return "", f"超时（>{timeout}s）"
    except Exception as exc:
        return "", short_error(str(exc))


def short_error(text: str) -> str:
    text = " ".join(str(text or "").split())
    if len(text) > 260:
        return text[:260] + "..."
    return text


def normalize_video_url(value: str) -> str:
    value = str(value or "").strip()
    if not value or value == "NA":
        return ""
    if value.startswith("http"):
        return value
    if value.startswith("/watch"):
        return "https://www.youtube.com" + value
    if re.fullmatch(r"[A-Za-z0-9_-]{8,}", value):
        return "https://www.youtube.com/watch?v=" + value
    return value


def normalize_channel_root_url(channel_url: str) -> str:
    url = str(channel_url or "").strip().rstrip("/")
    if not url:
        return ""
    url = re.sub(r"[?#].*$", "", url)
    url = re.sub(r"/(videos|shorts|streams|featured|playlists|community|about)$", "", url, flags=re.I)
    return url


def normalize_channel_url(value: str, channel_id: str = "", uploader_id: str = "") -> str:
    value = str(value or "").strip()
    if value and value != "NA":
        if value.startswith("http"):
            return value
        if value.startswith("/@") or value.startswith("/channel/") or value.startswith("/c/"):
            return "https://www.youtube.com" + value
        if value.startswith("@"):
            return "https://www.youtube.com/" + value
    channel_id = str(channel_id or "").strip()
    if channel_id and channel_id != "NA":
        return "https://www.youtube.com/channel/" + channel_id
    uploader_id = str(uploader_id or "").strip()
    if uploader_id and uploader_id.startswith("@"):
        return "https://www.youtube.com/" + uploader_id
    return ""


def looks_like_video_url(url: str) -> bool:
    text = str(url or "").strip().lower()
    return "/watch?v=" in text or "youtu.be/" in text


def parse_int(value) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def parse_upload_date(value: str):
    text = str(value or "").strip()
    if not text or len(text) != 8 or not text.isdigit():
        return None
    try:
        return datetime.strptime(text, "%Y%m%d")
    except Exception:
        return None


def parse_iso8601_duration_seconds(value: str) -> int:
    text = str(value or "").strip().upper()
    if not text.startswith("PT"):
        return 0
    total = 0
    for num, unit in re.findall(r"(\d+)([HMS])", text):
        n = int(num)
        if unit == "H":
            total += n * 3600
        elif unit == "M":
            total += n * 60
        elif unit == "S":
            total += n
    return total


def is_probable_short(video_url: str = "", duration_seconds: int = 0) -> bool:
    url = str(video_url or "").lower()
    if "/shorts/" in url:
        return True
    return 0 < int(duration_seconds or 0) <= SHORTS_MAX_SECONDS


def fetch_video_page_metadata(video_url: str) -> dict:
    """
    兜底从公开视频页 HTML 里提取 uploadDate / channelId / interactionCount。
    """
    result = {"upload_date": "", "channel_id": "", "view_count": 0, "duration_seconds": 0}
    url = normalize_video_url(video_url)
    if not url:
        return result
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return result

    upload_patterns = [
        r'"uploadDate":"(\d{4}-\d{2}-\d{2})"',
        r'"datePublished":"(\d{4}-\d{2}-\d{2})"',
    ]
    for pattern in upload_patterns:
        found = re.search(pattern, html)
        if found:
            result["upload_date"] = found.group(1).replace("-", "")
            break

    channel_patterns = [
        r'"channelId":"(UC[A-Za-z0-9_-]+)"',
        r'"externalChannelId":"(UC[A-Za-z0-9_-]+)"',
    ]
    for pattern in channel_patterns:
        found = re.search(pattern, html)
        if found:
            result["channel_id"] = found.group(1)
            break

    view_patterns = [
        r'"interactionCount":"(\d+)"',
        r'"viewCount":"(\d+)"',
    ]
    for pattern in view_patterns:
        found = re.search(pattern, html)
        if found:
            result["view_count"] = parse_int(found.group(1))
            break

    duration_patterns = [
        r'"lengthSeconds":"(\d+)"',
        r'"duration":"PT([^"]+)"',
    ]
    for pattern in duration_patterns:
        found = re.search(pattern, html)
        if not found:
            continue
        if pattern.endswith(r'(\d+)"'):
            result["duration_seconds"] = parse_int(found.group(1))
        else:
            result["duration_seconds"] = parse_iso8601_duration_seconds("PT" + found.group(1))
        break

    return result


def enrich_video_metrics(video_url: str) -> tuple:
    """补抓单条视频播放和日期。返回 (view_count, upload_date_str)。"""
    view_count, upload_date, _duration_seconds = enrich_video_metrics_full(video_url)
    return view_count, upload_date


def enrich_video_metrics_full(video_url: str) -> tuple:
    """补抓单条视频播放、日期、时长。返回 (view_count, upload_date_str, duration_seconds)。"""
    if not video_url:
        return 0, "", 0
    raw = run_yt_dlp(
        "--skip-download",
        "--print", "%(view_count)s|||%(upload_date)s|||%(duration)s",
        video_url,
        timeout=VIDEO_METRICS_TIMEOUT
    )
    if not raw:
        meta = fetch_video_page_metadata(video_url)
        return meta["view_count"], meta["upload_date"], meta["duration_seconds"]
    parts = raw.split("|||")
    if len(parts) < 3:
        meta = fetch_video_page_metadata(video_url)
        return meta["view_count"], meta["upload_date"], meta["duration_seconds"]
    view_count = parse_int(parts[0])
    upload_date = str(parts[1] or "").strip()
    duration_seconds = parse_int(parts[2])
    if (not view_count or not upload_date):
        meta = fetch_video_page_metadata(video_url)
        if meta["view_count"] > 0:
            view_count = meta["view_count"]
        if meta["upload_date"]:
            upload_date = meta["upload_date"]
        if meta["duration_seconds"] > 0:
            duration_seconds = meta["duration_seconds"]
    return view_count, upload_date, duration_seconds


def format_sub_label(subs: int) -> str:
    subs = parse_int(subs)
    if subs >= 1_000_000:
        return f"{subs / 1_000_000:.1f}M"
    if subs >= 10_000:
        return f"{subs // 1000}K"
    if subs >= 1_000:
        return f"{subs / 1000:.1f}K"
    return str(subs)


def fetch_channel_page_subs(channel_url: str) -> int:
    url = normalize_channel_url(channel_url)
    if not url or looks_like_video_url(url):
        return 0
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return 0

    patterns = [
        r'"subscriberCountText":\{"accessibility":\{"accessibilityData":\{"label":"([\d,\.]+)\s+subscribers?"',
        r'"subscriberCountText":\{"simpleText":"([\d,\.A-Za-z ]+)\s+subscribers?"',
        r'"subscriberCount":"(\d+)"',
    ]
    for pattern in patterns:
        found = re.search(pattern, html, re.I)
        if not found:
            continue
        value = found.group(1).strip().replace(",", "")
        if value.isdigit():
            return parse_int(value)
        m = re.match(r"([\d.]+)\s*([KMB])", value, re.I)
        if m:
            num = float(m.group(1))
            unit = m.group(2).upper()
            factor = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}.get(unit, 1)
            return int(num * factor)
    return 0


def enrich_channel_subs(channel_url: str = "", video_url: str = "", channel_id: str = "") -> int:
    """
    粉丝数多路补抓：
    1) YouTube API channels.statistics.subscriberCount
    2) yt-dlp 直接读频道/视频 channel_follower_count
    3) 频道页 HTML subscriberCountText
    """
    cid = str(channel_id or "").strip() or extract_channel_id(channel_url)
    if YOUTUBE_API_KEY and cid:
        payload = youtube_api_get("channels", {
            "part": "statistics",
            "id": cid,
            "maxResults": 1,
        })
        for item in payload.get("items", []) or []:
            stats = item.get("statistics", {}) or {}
            subs = parse_int(stats.get("subscriberCount"))
            if subs > 0:
                return subs

    targets = []
    channel_url = normalize_channel_url(channel_url, cid, "")
    video_url = normalize_video_url(video_url)
    if channel_url:
        targets.append(channel_url)
    if video_url:
        targets.append(video_url)

    for target in targets:
        raw, _ = run_yt_dlp_detailed(
            "--skip-download",
            "--print", "%(channel_follower_count)s|||%(channel_id)s",
            target,
            timeout=VIDEO_METRICS_TIMEOUT
        )
        if not raw:
            continue
        first = raw.splitlines()[0].strip()
        parts = first.split("|||")
        subs = parse_int(parts[0] if len(parts) > 0 else "")
        if subs > 0:
            return subs
        if not cid and len(parts) > 1:
            cid = str(parts[1] or "").strip()

    if channel_url:
        subs = fetch_channel_page_subs(channel_url)
        if subs > 0:
            return subs
    return 0


def resolve_channel_url(channel_url: str, video_url: str = "", cache=None) -> str:
    """
    某些 yt-dlp 搜索结果只有视频链接，没有稳定的频道链接。
    这时从视频页反查频道地址，避免后续 /videos 抓取大面积失败。
    """
    cache = cache if cache is not None else {}
    key = ((channel_url or "").strip(), (video_url or "").strip())
    if key in cache:
        return cache[key]

    normalized = normalize_channel_url(channel_url)
    if normalized and not looks_like_video_url(normalized):
        cache[key] = normalized
        return normalized

    video = normalize_video_url(video_url or channel_url)
    if not video:
        cache[key] = normalized
        return normalized

    raw, _ = run_yt_dlp_detailed(
        "--skip-download",
        "--print", "%(channel_url)s|||%(channel_id)s|||%(uploader_id)s",
        video,
        timeout=VIDEO_METRICS_TIMEOUT
    )
    if raw:
        first = raw.splitlines()[0].strip()
        parts = first.split("|||")
        resolved = normalize_channel_url(
            parts[0] if len(parts) > 0 else "",
            parts[1] if len(parts) > 1 else "",
            parts[2] if len(parts) > 2 else "",
        )
        if resolved and not looks_like_video_url(resolved):
            cache[key] = resolved
            return resolved

    meta = fetch_video_page_metadata(video)
    if meta["channel_id"]:
        resolved = normalize_channel_url("", meta["channel_id"], "")
        if resolved:
            cache[key] = resolved
            return resolved

    cache[key] = normalized
    return normalized


def extract_channel_id(channel_url: str) -> str:
    text = normalize_channel_root_url(channel_url) or str(channel_url or "").strip()
    m = re.search(r"/channel/([A-Za-z0-9_-]+)", text)
    if m:
        return m.group(1)
    # 优先直接抓频道页 HTML，避免 yt-dlp 对频道页/playlist 认证过敏
    try:
        req = urllib.request.Request(text, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
        patterns = [
            r'"externalId":"(UC[A-Za-z0-9_-]+)"',
            r'"channelId":"(UC[A-Za-z0-9_-]+)"',
            r'channel_id=(UC[A-Za-z0-9_-]+)',
            r'/feeds/videos\.xml\?channel_id=(UC[A-Za-z0-9_-]+)',
        ]
        for pattern in patterns:
            found = re.search(pattern, html)
            if found:
                return found.group(1)
    except Exception:
        pass
    raw, _ = run_yt_dlp_detailed(
        "--skip-download",
        "--print", "%(channel_id)s",
        text,
        timeout=VIDEO_METRICS_TIMEOUT
    )
    if raw:
        first = raw.splitlines()[0].strip()
        if first and first != "NA":
            return first
    return ""


def fetch_recent_videos_from_feed(channel_url: str, count: int = 5) -> list:
    """
    频道 /videos 被 YouTube 拦截时，退回 RSS feed 拿最近视频。
    返回 [{title, video_url, upload_date, view_count}]
    """
    channel_id = extract_channel_id(channel_url)
    if not channel_id:
        return []
    feed_url = f"https://www.youtube.com/feeds/videos.xml?channel_id={channel_id}"
    try:
        req = urllib.request.Request(feed_url, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            xml_text = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return []
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []

    ns = {
        "atom": "http://www.w3.org/2005/Atom",
        "yt": "http://www.youtube.com/xml/schemas/2015",
    }
    rows = []
    for entry in root.findall("atom:entry", ns)[: max(count * 3, count)]:
        title = (entry.findtext("atom:title", default="", namespaces=ns) or "").strip()
        video_id = (entry.findtext("yt:videoId", default="", namespaces=ns) or "").strip()
        published = (entry.findtext("atom:published", default="", namespaces=ns) or "").strip()
        if not video_id:
            continue
        video_url = f"https://www.youtube.com/watch?v={video_id}"
        upload_date = ""
        if len(published) >= 10:
            upload_date = published[:10].replace("-", "")
        view_count, metric_upload_date, duration_seconds = enrich_video_metrics_full(video_url)
        if metric_upload_date:
            upload_date = metric_upload_date
        is_short = is_probable_short(video_url, duration_seconds)
        rows.append({
            "upload_date": parse_upload_date(upload_date),
            "view_count": view_count,
            "title": title,
            "video_url": video_url,
            "video_id": video_id,
            "duration_seconds": duration_seconds,
            "is_short": is_short,
        })
    return rows[: max(count, 1) * 3]


def get_channel_latest_date(channel_url: str, cache=None):
    cache = cache if cache is not None else {}
    key = ("latest_date", (channel_url or "").strip().rstrip("/"))
    if key in cache:
        return cache[key]
    if YOUTUBE_API_KEY:
        channel_id = extract_channel_id(channel_url)
        if channel_id:
            payload = youtube_api_get("channels", {
                "part": "contentDetails",
                "id": channel_id,
                "maxResults": 1,
            })
            items = payload.get("items", []) or []
            if items:
                uploads = ((((items[0].get("contentDetails", {}) or {}).get("relatedPlaylists", {}) or {}).get("uploads")) or "").strip()
                if uploads:
                    rows = get_recent_videos_from_uploads_playlist(uploads, count=8, cache=cache)
                    non_short_rows = [row for row in rows if not row.get("is_short")]
                    if non_short_rows:
                        latest_dates = [row.get("upload_date") for row in non_short_rows if row.get("upload_date")]
                        latest_date = max(latest_dates) if latest_dates else None
                        cache[key] = latest_date
                        return latest_date
    rows = fetch_recent_videos_from_feed(channel_url, 8)
    non_short_rows = [row for row in rows if not row.get("is_short")]
    latest_date = non_short_rows[0]["upload_date"] if non_short_rows else None
    cache[key] = latest_date
    return latest_date


def get_latest_date_from_uploads_playlist(uploads_playlist: str, cache=None):
    cache = cache if cache is not None else {}
    playlist_id = str(uploads_playlist or "").strip()
    if not playlist_id:
        return None
    key = ("uploads_latest_date", playlist_id)
    if key in cache:
        return cache[key]
    if not YOUTUBE_API_KEY:
        cache[key] = None
        return None
    rows = get_recent_videos_from_uploads_playlist(playlist_id, count=8, cache=cache)
    non_short_rows = [row for row in rows if not row.get("is_short")]
    if not non_short_rows:
        cache[key] = None
        return None
    latest_dates = [row.get("upload_date") for row in non_short_rows if row.get("upload_date")]
    latest_date = max(latest_dates) if latest_dates else None
    cache[key] = latest_date
    return latest_date


def get_recent_videos_from_uploads_playlist(uploads_playlist: str, count: int = 5, cache=None) -> list:
    cache = cache if cache is not None else {}
    playlist_id = str(uploads_playlist or "").strip()
    if not playlist_id:
        return []
    key = ("uploads_recent_rows", playlist_id, int(count))
    if key in cache:
        return cache[key]
    if not YOUTUBE_API_KEY:
        cache[key] = []
        return []

    payload = youtube_api_get("playlistItems", {
        "part": "snippet,contentDetails",
        "playlistId": playlist_id,
        "maxResults": min(max(int(count) * 3, 1), 50),
    })
    items = payload.get("items", []) or []
    if not items:
        cache[key] = []
        return []

    video_ids = []
    base_rows = []
    for item in items:
        snippet = item.get("snippet", {}) or {}
        resource = snippet.get("resourceId", {}) or {}
        video_id = (resource.get("videoId") or (item.get("contentDetails", {}) or {}).get("videoId") or "").strip()
        published = str(snippet.get("publishedAt") or "")[:10].replace("-", "")
        title = str(snippet.get("title") or "").strip()
        if not video_id:
            continue
        video_ids.append(video_id)
        base_rows.append({
            "video_id": video_id,
            "video_url": normalize_video_url(video_id),
            "upload_date": parse_upload_date(published),
            "title": title,
            "view_count": 0,
        })

    if not base_rows:
        cache[key] = []
        return []

    view_map = {}
    for batch in chunked(video_ids, 50):
        video_payload = youtube_api_get("videos", {
            "part": "statistics,snippet,contentDetails",
            "id": ",".join(batch),
            "maxResults": 50,
        })
        for item in video_payload.get("items", []) or []:
            vid = item.get("id", "")
            stats = item.get("statistics", {}) or {}
            snippet = item.get("snippet", {}) or {}
            content = item.get("contentDetails", {}) or {}
            view_map[vid] = {
                "view_count": parse_int(stats.get("viewCount")),
                "upload_date": parse_upload_date(str(snippet.get("publishedAt") or "")[:10].replace("-", "")),
                "duration_seconds": parse_iso8601_duration_seconds(content.get("duration")),
            }

    rows = []
    for row in base_rows:
        meta = view_map.get(row["video_id"], {})
        if meta.get("view_count", 0) > 0:
            row["view_count"] = meta["view_count"]
        if meta.get("upload_date"):
            row["upload_date"] = meta["upload_date"]
        row["duration_seconds"] = meta.get("duration_seconds", 0)
        row["is_short"] = is_probable_short(row.get("video_url", ""), row.get("duration_seconds", 0))
        rows.append(row)

    cache[key] = rows[: max(count, 1) * 3]
    return cache[key]


def get_recent_channel_metrics(channel_url: str, count: int = 5, cache=None, uploads_playlist: str = "") -> dict:
    """
    读取频道最近 N 条视频的播放和日期。
    返回:
      {
        ok, avg_views, latest_date, recent_count, reason
      }
    """
    cache = cache if cache is not None else {}
    channel_key = (channel_url or "").strip().rstrip("/")
    cache_key = (channel_key, int(count), str(uploads_playlist or "").strip())
    if cache_key in cache:
        return cache[cache_key]

    result = {
        "ok": False,
        "avg_views": 0,
        "latest_date": None,
        "recent_count": 0,
        "reason": "频道地址为空",
        "titles": [],
    }
    if not channel_key:
        cache[cache_key] = result
        return result

    rows = []
    uploads_playlist = str(uploads_playlist or "").strip()
    if uploads_playlist:
        rows = get_recent_videos_from_uploads_playlist(uploads_playlist, count=count, cache=cache)
        if rows:
            regular_rows = [row for row in rows if not row.get("is_short")]
            view_rows = [row["view_count"] for row in regular_rows if row["view_count"] > 0]
            latest_dates = [row["upload_date"] for row in regular_rows if row["upload_date"]]
            if len(regular_rows) >= count and len(view_rows) >= count:
                latest_date = max(latest_dates) if latest_dates else None
                avg_views = int(sum(view_rows) / len(view_rows)) if view_rows else 0
                result = {
                    "ok": True,
                    "avg_views": avg_views,
                    "latest_date": latest_date,
                    "recent_count": len(regular_rows),
                    "reason": "",
                    "titles": [row["title"] for row in regular_rows if row.get("title")],
                    "video_ids": [row["video_id"] for row in regular_rows if row.get("video_id")],
                }
                cache[cache_key] = result
                return result

    rows = fetch_recent_videos_from_feed(channel_key, count)
    if rows:
        regular_rows = [row for row in rows if not row.get("is_short")]
        view_rows = [row["view_count"] for row in regular_rows if row["view_count"] > 0]
        latest_dates = [row["upload_date"] for row in regular_rows if row["upload_date"]]
        if len(regular_rows) >= count and len(view_rows) >= count:
            latest_date = max(latest_dates) if latest_dates else None
            avg_views = int(sum(view_rows) / len(view_rows)) if view_rows else 0
            result = {
                "ok": True,
                "avg_views": avg_views,
                "latest_date": latest_date,
                "recent_count": len(regular_rows),
                "reason": "",
                "titles": [row["title"] for row in regular_rows if row.get("title")],
                "video_ids": [row["video_id"] for row in regular_rows if row.get("video_id")],
            }
            cache[cache_key] = result
            return result

    urls_to_try = [channel_key]
    if "/videos" not in channel_key:
        urls_to_try.insert(0, channel_key + "/videos")

    raw = ""
    fail_notes = []
    for url in urls_to_try:
        raw, fail_reason = run_yt_dlp_detailed(
            "--skip-download",
            "--playlist-end", str(count),
            "--print", "%(upload_date)s|||%(view_count)s|||%(title)s",
            url,
            timeout=RECENT_CHANNEL_TIMEOUT
        )
        if raw:
            break
        concise = normalize_ytdlp_error(fail_reason or "空结果")
        fail_notes.append(f"{url} -> {concise}")

    if not raw:
        if fail_notes:
            result["reason"] = "最近视频数据抓取失败：" + "；".join(fail_notes[:2])
        else:
            result["reason"] = "最近视频数据抓取失败"
        cache[cache_key] = result
        return result
    else:
        rows = []
        for line in raw.splitlines():
            parts = line.split("|||")
            if len(parts) < 3:
                continue
            upload_date = parse_upload_date(parts[0])
            view_count = parse_int(parts[1])
            title = parts[2].strip()
            if not title:
                continue
            rows.append({
                "upload_date": upload_date,
                "view_count": view_count,
                "title": title,
                "duration_seconds": 0,
                "is_short": False,
            })
            if len(rows) >= max(count, 1) * 3:
                break

    regular_rows = [row for row in rows if not row.get("is_short")]
    if len(regular_rows) < count:
        result["reason"] = f"最近常规视频不足{count}条（Shorts不计）"
        result["recent_count"] = len(regular_rows)
        cache[cache_key] = result
        return result

    view_rows = [row["view_count"] for row in regular_rows if row["view_count"] > 0]
    if len(view_rows) < count:
        result["reason"] = "最近常规视频播放量抓取不足"
        result["recent_count"] = len(regular_rows)
        cache[cache_key] = result
        return result

    latest_dates = [row["upload_date"] for row in regular_rows if row["upload_date"]]
    latest_date = max(latest_dates) if latest_dates else None
    avg_views = int(sum(view_rows) / len(view_rows)) if view_rows else 0
    result = {
        "ok": True,
        "avg_views": avg_views,
        "latest_date": latest_date,
        "recent_count": len(regular_rows),
        "reason": "",
        "titles": [row["title"] for row in regular_rows if row.get("title")],
        "video_ids": [],  # yt-dlp 方式无法获取 video_id，留空
    }
    cache[cache_key] = result
    return result


def search_youtube(keyword: str, n: int = 20) -> list:
    """
    用 yt-dlp 搜索关键词，返回视频元数据列表
    每条: {title, channel, channel_url, subs, video_url, view_count, upload_date}
    """
    print(f"  🔍 搜索: {keyword}")
    api_results = search_youtube_api(keyword, n=n)
    if api_results:
        return api_results
    search_expr = f"ytsearch{n}:{keyword}"

    results = []
    seen_channels = set()

    # 优先用扁平JSON搜索：更快，也不容易触发 YouTube 单视频页面报错。
    primary = run_yt_dlp_result(
        "--skip-download",
        "--flat-playlist",
        "--dump-json",
        search_expr,
        timeout=120
    )
    raw_json = primary.stdout.strip() if primary else ""
    if primary and primary.returncode != 0 and not raw_json:
        print(f"  ⚠️ yt-dlp 搜索报错: {short_error(primary.stderr)}")
    for line in raw_json.splitlines():
        try:
            item = json.loads(line)
        except Exception:
            continue
        title = str(item.get("title") or "").strip()
        channel = str(item.get("channel") or item.get("uploader") or "").strip()
        ch_url = normalize_channel_url(
            item.get("channel_url") or item.get("uploader_url") or "",
            item.get("channel_id") or "",
            item.get("uploader_id") or "",
        )
        video_url = normalize_video_url(item.get("url") or item.get("webpage_url") or item.get("id") or "")
        key = ch_url or channel.lower() or video_url
        if not key or key in seen_channels:
            continue
        seen_channels.add(key)
        subs = item.get("channel_follower_count") or 0
        try:
            subs = int(subs or 0)
        except Exception:
            subs = 0
        results.append({
            "title": title,
            "channel": channel or "Unknown Channel",
            "channel_url": ch_url or video_url,
            "subs": subs,
            "video_url": video_url,
            "view_count": parse_int(item.get("view_count")),
            "upload_date": str(item.get("upload_date") or ""),
        })
        if len(results) >= n:
            break
    if results:
        return results

    # 兜底：旧版 --print 方式。有些环境JSON字段缺失时仍可取到频道。
    fallback = run_yt_dlp_result(
        "--skip-download",
        "--print", "%(title)s|||%(channel)s|||%(channel_url)s|||%(channel_id)s|||%(uploader_id)s|||%(channel_follower_count)s|||%(webpage_url)s",
        search_expr,
        timeout=120
    )
    raw = fallback.stdout.strip() if fallback else ""
    if fallback and fallback.returncode != 0 and not raw:
        print(f"  ⚠️ yt-dlp 兜底搜索也失败: {short_error(fallback.stderr)}")
    for line in raw.splitlines():
        parts = line.split("|||")
        if len(parts) < 7:
            continue
        title, channel, ch_url, channel_id, uploader_id, subs_raw, video_url = parts[:7]
        ch_url = normalize_channel_url(ch_url, channel_id, uploader_id)
        video_url = normalize_video_url(video_url)
        if not ch_url:
            # 没有频道URL时不直接丢掉；用频道名做去重，并保留视频URL给后续查邮箱。
            ch_url = video_url
        key = ch_url or channel.strip().lower()
        if not key or key in seen_channels:
            continue
        seen_channels.add(key)
        try:
            subs = int(subs_raw) if str(subs_raw).isdigit() else 0
        except Exception:
            subs = 0
        results.append({
            "title": title.strip(),
            "channel": channel.strip() or "Unknown Channel",
            "channel_url": ch_url.strip(),
            "subs": subs,
            "video_url": video_url.strip(),
            "view_count": 0,
            "upload_date": "",
        })
    if not results and not raw and not raw_json:
        print("  ⚠️ 没有拿到 YouTube 搜索结果。若连续为 0，请检查网络/yt-dlp 是否可访问 YouTube。")
    return results


def market_signal_score(channel: str, title: str, keyword: str, subs: int) -> tuple:
    """
    不用YouTube地区字段，改看公开内容信号：
    移动剪辑、免费工具、低配设备、本地语言教程、小粉丝量等。
    """
    text = f"{channel} {title} {keyword}".lower()
    score = 0
    hits = []
    for signal, value in MARKET_POSITIVE_SIGNALS.items():
        if signal in text:
            score += value
            hits.append(signal)
    for signal, value in MARKET_NEGATIVE_SIGNALS.items():
        if signal in text:
            score += value
            hits.append(f"-{signal}")

    if 1_000 <= subs <= 50_000:
        score += 3
        hits.append("small-channel")
    elif 50_000 < subs <= 150_000:
        score += 2
        hits.append("mid-small-channel")
    elif 150_000 < subs <= 250_000:
        score += 1
        hits.append("upper-small-channel")
    elif subs > 250_000:
        score -= 3
        hits.append("-large-channel")

    return score, ", ".join(hits[:6]) if hits else "no market signal"


def google_search_email(channel: str, channel_url: str) -> str:
    """用搜索引擎结果页补查邮箱。可能会被 Google 风控，失败时返回空。"""
    handle = ""
    m = re.search(r"/@([^/?#]+)", channel_url or "")
    if m:
        handle = m.group(1)
    query = f'"{channel}" YouTube contact email'
    if handle:
        query += f' OR "{handle}" email'
    url = "https://www.google.com/search?q=" + urllib.parse.quote(query) + "&num=8"
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return ""
    emails = extract_emails(html)
    if emails:
        return emails[0]

    ddg_url = "https://html.duckduckgo.com/html/?q=" + urllib.parse.quote(query)
    try:
        req = urllib.request.Request(ddg_url, headers=HEADERS)
        with safe_urlopen(req, timeout=15) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return ""
    emails = extract_emails(html)
    return emails[0] if emails else ""


def build_channel_about_url(channel_url: str) -> str:
    url = normalize_channel_root_url(channel_url)
    if not url:
        return ""
    return url + "/about"


def fetch_page_html(url: str, timeout: int = 15) -> str:
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with safe_urlopen(req, timeout=timeout) as resp:
            return resp.read().decode("utf-8", errors="replace")
    except Exception:
        return ""


def collect_direct_email_candidates(channel_url: str, video_url: str = "") -> list:
    candidates = []
    pages = []
    if video_url:
        pages.append(("视频页面HTML", video_url))
    root_url = normalize_channel_root_url(channel_url)
    if root_url:
        pages.append(("频道主页HTML", root_url))
    about_url = build_channel_about_url(channel_url)
    if about_url:
        pages.append(("频道About页面", about_url))
    for label, url in pages:
        html = fetch_page_html(url, timeout=15)
        if not html:
            continue
        emails = extract_emails(html)
        for email in emails[:5]:
            candidates.append({
                "email": email,
                "label": label,
                "text": html,
                "page_url": url,
            })
    return candidates


def pick_best_email_candidate(candidates, channel: str, channel_url: str) -> tuple:
    ranked = {
        "人工补充邮箱": 0,
        "频道About页面": 1,
        "频道主页HTML": 1,
        "视频频道描述": 1,
        "视频描述": 2,
        "频道描述": 3,
        "频道频道描述": 4,
        "视频页面HTML": 4,
        "外链页面": 5,
        "聚合页外链": 6,
        "搜索引擎": 9,
    }
    for item in sorted(candidates, key=lambda x: ranked.get(x["label"], 99)):
        reason = ""
        if email_relevance_reason:
            reason = email_relevance_reason(
                item["email"],
                channel_name=channel,
                channel_url=channel_url,
                source_label=item["label"],
                text=item.get("text", ""),
                page_url=item.get("page_url", ""),
            )
        elif classify_bad_email:
            reason = classify_bad_email(item["email"])
        if not reason:
            return item["email"], item["label"]
    return "", "无可信公开邮箱"


def get_channel_email(channel_url: str, video_url: str = "", channel: str = "") -> tuple:
    """多来源提取邮箱。返回 (email, note, confidence)"""
    sources = []
    if video_url:
        sources.append(("视频频道描述", video_url, "%(channel_description)s"))
        sources.append(("视频描述", video_url, "%(description)s"))
    if channel_url:
        sources.append(("频道描述", channel_url, "%(description)s"))
        sources.append(("频道频道描述", channel_url, "%(channel_description)s"))

    candidates = []
    for label, url, field in sources:
        desc = run_yt_dlp(url, "--skip-download", "--print", field, timeout=45)
        emails = extract_emails(desc)
        if emails:
            for email in emails[:5]:
                candidates.append({
                    "email": email,
                    "label": label,
                    "text": desc,
                    "page_url": channel_url or video_url,
                })

    candidates.extend(collect_direct_email_candidates(channel_url, video_url))

    email = google_search_email(channel, channel_url)
    if email:
        candidates.append({
            "email": email,
            "label": "搜索引擎",
            "text": "",
            "page_url": "",
        })
    if candidates:
        email, note = pick_best_email_candidate(candidates, channel, channel_url)
        if email:
            return email, note, "A"

    if deep_find_email:
        email, note, confidence, _source_label = deep_find_email(channel, video_url or channel_url)
        if email:
            return email, note or "深度补查邮箱", confidence or "B"
    return "", "无公开邮箱", ""


def make_subject_opening(channel: str, title: str, subs: int) -> tuple:
    """根据频道信息生成个性化主题和开头"""
    lower = str(title or "").lower()
    if any(w in lower for w in ["capcut", "premiere", "davinci", "editing", "editor"]):
        subj = "Quick idea for your video editing audience"
    elif any(w in lower for w in ["upscale", "4k", "resolution", "quality", "enhance"]):
        subj = "Possible VikPea fit for your video quality content?"
    elif any(w in lower for w in ["restore", "restoration", "denoise", "deblur", "blurry"]):
        subj = "Quick idea for your video restoration audience"
    else:
        subj = "Quick collaboration idea for your channel"
    opening = (
        "I came across your channel and liked that you cover practical creator-facing video content. "
        "It felt relevant for viewers looking at ways to improve video quality and workflow."
    )
    return subj, opening


def write_to_queue(ws, name, email, subj, opening, home_url, subs, found_email: bool,
                   source_keyword: str, market_score: int = 0, market_reason: str = "",
                   video_url: str = "", tag: str = "", analysis: dict = None,
                   channel_score: dict = None, avg_recent_views: int = 0):
    """写一行到发信名单"""
    sub_label = format_sub_label(subs)
    market_note = f" | 小博主信号{market_score}: {market_reason}" if market_reason else ""
    hit_note = f" | 命中视频: {video_url}" if video_url else ""
    note = f"{sub_label}粉{market_note}{hit_note}"
    if not found_email:
        if add_no_email_candidate:
            add_no_email_candidate(
                name=name,
                link=home_url,
                note=note + " | 无公开邮箱",
                email_type="YouTube",
                source=source_keyword,
                subject=subj,
                opening=opening,
                status="待深度查邮箱",
                video_link=video_url,
            )
        return None
    r = ws.max_row + 1
    ws.cell(r, Q_NAME).value  = name
    ws.cell(r, Q_EMAIL).value = email
    ws.cell(r, Q_SUBJ).value  = subj
    ws.cell(r, Q_OPEN).value  = opening
    ws.cell(r, Q_HOME).value  = home_url
    ws.cell(r, Q_VIDEO).value = video_url
    ws.cell(r, Q_NOTE).value  = note
    ws.cell(r, Q_TYPE).value  = ""  # 留空即 YouTube；发信脚本会按 YouTube 处理
    ws.cell(r, Q_SOURCE).value = source_keyword
    ws.cell(r, Q_TAG).value   = tag

    # 写入 AI 分析结果（旧版）
    if analysis:
        vertical_score = analysis.get("vertical_score", 0)
        ws.cell(r, Q_VERTICAL_SCORE).value = vertical_score if vertical_score > 0 else ""

        competitor_mentioned = analysis.get("competitor_mentioned", False)
        competitor_names = analysis.get("competitor_names", [])
        if competitor_mentioned and competitor_names:
            ws.cell(r, Q_COMPETITOR).value = "是 (" + "/".join(competitor_names) + ")"
        elif competitor_mentioned:
            ws.cell(r, Q_COMPETITOR).value = "是"
        else:
            ws.cell(r, Q_COMPETITOR).value = "否"

        collab_type = analysis.get("collaboration_type", "")
        ws.cell(r, Q_COLLAB_TYPE).value = collab_type

    # 写入频道综合评分（新增）
    if channel_score:
        ws.cell(r, Q_CONTENT_RELEVANCE).value = channel_score.get("content_relevance", "")
        ws.cell(r, Q_COMMERCIAL_HISTORY).value = channel_score.get("commercial_history", "")
        ws.cell(r, Q_AUDIENCE_FIT).value = channel_score.get("audience_fit", "")
        ws.cell(r, Q_COMPOSITE_SCORE).value = channel_score.get("composite_score", "")
        ws.cell(r, Q_AI_VERDICT).value = channel_score.get("verdict", "")
        ws.cell(r, Q_AI_REASONING).value = channel_score.get("reasoning", "")

    # 写入近期均播
    if avg_recent_views > 0:
        ws.cell(r, Q_AVG_RECENT_VIEWS).value = avg_recent_views

    for c in range(1, Q_AVG_RECENT_VIEWS + 1):
        ws.cell(r, c).fill = GREEN
    return r


def append_found_email_to_queue(name, email, subj, opening, home_url, subs,
                                source_keyword, market_score=0, market_reason="", video_url="",
                                tag="", analysis=None, channel_score=None, avg_recent_views=0):
    """重新读取最新主表后再追加，避免长时间搜索脚本用旧工作簿覆盖清理结果。"""
    wb = openpyxl.load_workbook(QUEUE_PATH)
    ws = wb.active
    ensure_queue_headers(ws)
    rownum = write_to_queue(
        ws, name, email, subj, opening, home_url, subs, True,
        source_keyword, market_score, market_reason, video_url, tag=tag, analysis=analysis,
        channel_score=channel_score, avg_recent_views=avg_recent_views
    )
    wb.save(QUEUE_PATH)
    return rownum


def calculate_composite_score(content_relevance: int, commercial_history: int,
                               audience_fit: int, avg_views: int, max_avg_views: int = 1000000) -> float:
    """
    计算频道综合分数
    max_avg_views: 用于归一化播放量的最大值（默认100万）
    """
    # 归一化播放量到0-100范围
    normalized_views = min(100, (avg_views / max_avg_views) * 100) if max_avg_views > 0 else 0

    # 应用权重计算综合分数
    weights = CHANNEL_SCORE_WEIGHTS
    composite = (
        content_relevance * weights["content_relevance"] +
        commercial_history * weights["commercial_history"] +
        audience_fit * weights["audience_fit"] +
        normalized_views * weights["avg_views_normalized"]
    )

    return round(composite, 2)


def short_channel_name(name: str, limit: int = 28) -> str:
    text = str(name or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def youtube_api_get(endpoint: str, params: dict) -> dict:
    global YOUTUBE_API_LAST_ERROR, YOUTUBE_API_LAST_TS, YOUTUBE_API_RATE_LIMITED_UNTIL
    if not YOUTUBE_API_KEY:
        YOUTUBE_API_LAST_ERROR = "未配置 YOUTUBE_API_KEY"
        return {}
    now = time.time()
    if YOUTUBE_API_RATE_LIMITED_UNTIL and now < YOUTUBE_API_RATE_LIMITED_UNTIL:
        wait_left = max(0, int(YOUTUBE_API_RATE_LIMITED_UNTIL - now))
        YOUTUBE_API_LAST_ERROR = f"API 限流冷却中，还需等待 {wait_left}s"
        return {}
    gap = float(YOUTUBE_API_DELAY_SEC or 0)
    if YOUTUBE_API_LAST_TS and gap > 0:
        sleep_for = gap - (now - YOUTUBE_API_LAST_TS)
        if sleep_for > 0:
            time.sleep(sleep_for)
    query = dict(params or {})
    query["key"] = YOUTUBE_API_KEY
    url = "https://www.googleapis.com/youtube/v3/" + endpoint + "?" + urllib.parse.urlencode(query)
    retries = max(int(YOUTUBE_API_RETRY_TIMES or 1), 1)
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with safe_urlopen(req, timeout=20) as resp:
                YOUTUBE_API_LAST_TS = time.time()
                payload = json.loads(resp.read().decode("utf-8", errors="replace"))
                api_error = (payload.get("error", {}) or {}).get("message") if isinstance(payload, dict) else ""
                YOUTUBE_API_LAST_ERROR = str(api_error or "").strip()
                return payload if isinstance(payload, dict) else {}
        except urllib.error.HTTPError as exc:
            YOUTUBE_API_LAST_TS = time.time()
            body = ""
            try:
                body = exc.read().decode("utf-8", errors="replace")
            except Exception:
                body = ""
            err_text = short_error(body or str(exc))
            if exc.code == 429:
                cooldown = max(int(YOUTUBE_API_429_COOLDOWN or 1), 1) * (attempt + 1)
                YOUTUBE_API_RATE_LIMITED_UNTIL = time.time() + cooldown
                YOUTUBE_API_LAST_ERROR = f"HTTP 429 Too Many Requests，冷却 {cooldown}s"
                if attempt < retries - 1:
                    time.sleep(cooldown)
                    continue
                return {}
            if exc.code in {403, 500, 503} and attempt < retries - 1:
                wait_sec = 2 * (attempt + 1)
                YOUTUBE_API_LAST_ERROR = f"HTTP {exc.code}，重试中（{wait_sec}s）"
                time.sleep(wait_sec)
                continue
            YOUTUBE_API_LAST_ERROR = err_text or f"HTTP {exc.code}"
            return {}
        except Exception as exc:
            YOUTUBE_API_LAST_TS = time.time()
            YOUTUBE_API_LAST_ERROR = short_error(str(exc))
            return {}
    return {}


def chunked(seq, size: int):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def search_youtube_api(keyword: str, n: int = 20) -> list:
    if not YOUTUBE_API_KEY:
        return []
    print("  ↳ 正在使用 YouTube API 拉取候选…")
    results = []
    page_token = ""
    seen_channels = set()
    while len(results) < n:
        payload = youtube_api_get("search", {
            "part": "snippet",
            "q": keyword,
            "type": "video",
            "maxResults": min(50, n - len(results)),
            "order": "relevance",
            "pageToken": page_token,
        })
        items = payload.get("items", []) or []
        if not items:
            if YOUTUBE_API_LAST_ERROR:
                print(f"  ↳ API 搜索失败：{YOUTUBE_API_LAST_ERROR}")
            else:
                print("  ↳ API 没返回结果，回退 yt-dlp 搜索")
            break
        for item in items:
            snippet = item.get("snippet", {}) or {}
            video_id = ((item.get("id", {}) or {}).get("videoId") or "").strip()
            channel_id = snippet.get("channelId", "") or ""
            if not video_id or not channel_id or channel_id in seen_channels:
                continue
            seen_channels.add(channel_id)
            results.append({
                "title": str(snippet.get("title") or "").strip(),
                "channel": str(snippet.get("channelTitle") or "").strip() or "Unknown Channel",
                "channel_url": normalize_channel_url("", channel_id, ""),
                "channel_id": channel_id,
                "subs": 0,
                "video_url": normalize_video_url(video_id),
                "video_id": video_id,
                "view_count": 0,
                "upload_date": str(snippet.get("publishedAt") or "")[:10].replace("-", ""),
            })
            if len(results) >= n:
                break
        page_token = payload.get("nextPageToken", "")
        if not page_token:
            break

    if not results:
        return []
    print(f"  ↳ API 候选 {len(results)} 个频道，继续补粉丝/播放/发布时间")

    channel_ids = [item["channel_id"] for item in results if item.get("channel_id")]
    video_ids = [item["video_id"] for item in results if item.get("video_id")]

    channel_meta = {}
    for batch in chunked(channel_ids, 50):
        payload = youtube_api_get("channels", {
            "part": "statistics,contentDetails",
            "id": ",".join(batch),
            "maxResults": 50,
        })
        for item in payload.get("items", []) or []:
            cid = item.get("id", "")
            stats = item.get("statistics", {}) or {}
            details = item.get("contentDetails", {}) or {}
            uploads = ((details.get("relatedPlaylists", {}) or {}).get("uploads") or "").strip()
            channel_meta[cid] = {
                "subs": parse_int(stats.get("subscriberCount")),
                "uploads": uploads,
            }

    video_meta = {}
    for batch in chunked(video_ids, 50):
        payload = youtube_api_get("videos", {
            "part": "statistics,snippet,contentDetails",
            "id": ",".join(batch),
            "maxResults": 50,
        })
        for item in payload.get("items", []) or []:
            vid = item.get("id", "")
            stats = item.get("statistics", {}) or {}
            snippet = item.get("snippet", {}) or {}
            content = item.get("contentDetails", {}) or {}
            video_meta[vid] = {
                "view_count": parse_int(stats.get("viewCount")),
                "upload_date": str(snippet.get("publishedAt") or "")[:10].replace("-", ""),
                "duration_seconds": parse_iso8601_duration_seconds(content.get("duration")),
            }

    for item in results:
        cid = item.get("channel_id", "")
        vid = item.get("video_id", "")
        meta = channel_meta.get(cid, {})
        item["subs"] = meta.get("subs", item["subs"])
        item["uploads_playlist"] = meta.get("uploads", "")
        vmeta = video_meta.get(vid, {})
        if vmeta.get("view_count", 0) > 0:
            item["view_count"] = vmeta["view_count"]
        if vmeta.get("upload_date"):
            item["upload_date"] = vmeta["upload_date"]
        item["duration_seconds"] = vmeta.get("duration_seconds", 0)
        if item.get("subs", 0) <= 0:
            item["subs"] = enrich_channel_subs(
                channel_url=item.get("channel_url", ""),
                video_url=item.get("video_url", ""),
                channel_id=cid,
            )
    return results


def keyword_has_enough_fresh_results(results: list, sample_size: int, min_hits: int, active_days: int) -> tuple:
    if min_hits <= 0:
        return True, 0, 0
    sample = results[:max(sample_size, 1)]
    if not sample:
        return True, 0, 0
    fresh_hits = 0
    cutoff = datetime.now() - timedelta(days=active_days)
    for item in sample:
        if is_probable_short(item.get("video_url", ""), parse_int(item.get("duration_seconds"))):
            continue
        upload_date = parse_upload_date(item.get("upload_date") or "")
        if upload_date and upload_date >= cutoff:
            fresh_hits += 1
    return fresh_hits >= min_hits, fresh_hits, len(sample)


def main():
    config = {}
    if apply_config:
        config = apply_config(globals(), {
            "YOUTUBE_RESULTS_PER_KEYWORD": "SEARCH_N",
            "YOUTUBE_MIN_VIDEO_VIEWS": "MIN_VIDEO_VIEWS",
            "YOUTUBE_MIN_SHORTS_VIEWS": "MIN_SHORTS_VIEWS",
            "YOUTUBE_MIN_RECENT_AVG_VIEWS": "MIN_RECENT_AVG_VIEWS",
            "YOUTUBE_RECENT_VIDEO_COUNT": "RECENT_VIDEO_COUNT",
            "YOUTUBE_ACTIVE_WITHIN_DAYS": "ACTIVE_WITHIN_DAYS",
            "YOUTUBE_VIDEO_METRICS_TIMEOUT": "VIDEO_METRICS_TIMEOUT",
            "YOUTUBE_RECENT_CHANNEL_TIMEOUT": "RECENT_CHANNEL_TIMEOUT",
            "YOUTUBE_KEYWORD_FRESHNESS_SAMPLE": "KEYWORD_FRESHNESS_SAMPLE",
            "YOUTUBE_KEYWORD_MIN_FRESH_HITS": "KEYWORD_MIN_FRESH_HITS",
            "YOUTUBE_API_KEY": "YOUTUBE_API_KEY",
            "YOUTUBE_API_DELAY_SEC": "YOUTUBE_API_DELAY_SEC",
            "YOUTUBE_API_RETRY_TIMES": "YOUTUBE_API_RETRY_TIMES",
            "YOUTUBE_API_429_COOLDOWN": "YOUTUBE_API_429_COOLDOWN",
            "YTDLP_COOKIES_FROM_BROWSER": "YTDLP_COOKIES_FROM_BROWSER",
            "YTDLP_RETRY_TIMES": "YTDLP_RETRY_TIMES",
            "USE_SEMRUSH_FOR_YOUTUBE": "USE_SEMRUSH_FOR_YOUTUBE",
            "YOUTUBE_KEYWORD_FILTER": "YOUTUBE_KEYWORD_FILTER",
            "YOUTUBE_SUB_MIN": "SUB_MIN",
            "YOUTUBE_SUB_MAX": "SUB_MAX",
            "YOUTUBE_MARKET_SCORE_MIN": "MARKET_SCORE_MIN",
            "ANTHROPIC_API_KEY": "ANTHROPIC_API_KEY",
            "ANTHROPIC_API_BASE": "ANTHROPIC_API_BASE",
            "ANTHROPIC_TAG_MODEL": "ANTHROPIC_TAG_MODEL",
            "DEEPSEEK_API_KEY": "DEEPSEEK_API_KEY",
            "DEEPSEEK_API_BASE": "DEEPSEEK_API_BASE",
        })
    # 检查 yt-dlp
    if not YTDLP_CMD:
        print("❌ 未找到 yt-dlp，请运行:")
        print(f"   {sys.executable} -m pip install yt-dlp")
        sys.exit(1)
    print(f"  ✅ yt-dlp 路径: {' '.join(YTDLP_CMD)}")

    print(f"\n{'═'*65}")
    print(f"  🚀 VikPea YouTube KOL 批量搜索")
    print(f"{'═'*65}\n")
    # 加载关键词：YouTube 找博主默认只用合作型关键词表。
    # SEMrush 更适合文章/SEO，不默认混入，避免搜到用户问题词。
    keywords = load_keyword_workbook(KEYWORD_PATH)
    if USE_SEMRUSH_FOR_YOUTUBE:
        semrush = load_semrush_keywords(SEMRUSH_CANDIDATES, limit=30)
        for kw in semrush:
            if kw not in keywords:
                keywords.append(kw)
    else:
        print("  📌 YouTube 搜索默认不追加 SEMrush 词；如需开启，在 VikPea_配置.xlsx 设置 USE_SEMRUSH_FOR_YOUTUBE=True")
    if not keywords:
        print("  ⚠️  当前没有启用的 YouTube 关键词。")
        print(f"  ↳ 请先到 {KEYWORD_PATH} 把需要使用的关键词第2列改成“是”，再重新运行。")
        return

    network_ok, failed_hosts = network_ready_for_youtube()
    if not network_ok:
        print("  ❌ 当前网络/DNS 解析异常，YouTube 搜索无法正常开始。")
        print(f"  ↳ 解析失败: {', '.join(failed_hosts)}")
        print("  ↳ 请先确认代理、DNS、或网络权限，再重新运行。")
        sys.exit(2)
    if sys.version_info < (3, 10):
        print("  ⚠️ 当前 Python 是 3.9，yt-dlp 对 YouTube 抓取稳定性会变差；建议后续升级到 Python 3.10+")
    if not YTDLP_COOKIES_FROM_BROWSER:
        print("  📌 如频繁出现 'page needs to be reloaded'，可在 VikPea_配置.xlsx 设置 YTDLP_COOKIES_FROM_BROWSER=chrome")
    if YOUTUBE_API_KEY:
        print("  📡 已启用 YouTube API 模式（优先用 API 取搜索结果、播放量、频道最新发布时间）")
    else:
        print("  📌 未启用 YouTube API；如需更稳的最新发布时间，请在 VikPea_配置.xlsx 填写 YOUTUBE_API_KEY")

    if YOUTUBE_KEYWORD_FILTER:
        kept = []
        skipped_keywords = []
        for kw in keywords:
            reason = youtube_keyword_skip_reason(kw)
            if reason:
                skipped_keywords.append((kw, reason))
            elif kw not in kept:
                kept.append(kw)
        keywords = kept
        if skipped_keywords:
            print(f"  🧹 已跳过 {len(skipped_keywords)} 个不适合找KOL的关键词")
            for kw, reason in skipped_keywords[:8]:
                print(f"     - {kw} ({reason})")
            if len(skipped_keywords) > 8:
                print(f"     ... 还有 {len(skipped_keywords) - 8} 个")
    print(f"共 {len(keywords)} 个关键词待搜索\n")

    # 加载已有数据（去重用）
    if cleanup_pending_email_review_known:
        cleaned_pending = cleanup_pending_email_review_known()
        if cleaned_pending:
            print(f"🧹 已清理待确认邮箱表中 {cleaned_pending} 条已联系/已入队残留")
    if archive_processed_pending_email_rows:
        archived_pending = archive_processed_pending_email_rows()
        if archived_pending:
            print(f"📦 已归档待确认邮箱表中 {archived_pending} 条已处理记录")
    existing_emails   = load_existing_emails(TRACKER_PATH, QUEUE_PATH)
    existing_names    = load_existing_channel_names(TRACKER_PATH, QUEUE_PATH)
    no_email_keys = existing_no_email_keys() if existing_no_email_keys else set()
    print(f"已有邮箱 {len(existing_emails)} 条（用于去重）\n")

    if not os.path.exists(QUEUE_PATH):
        print(f"❌ 找不到发信名单: {QUEUE_PATH}"); sys.exit(1)

    added_green  = 0
    added_yellow = 0
    skipped      = 0
    recent_stats_cache = {}
    resolved_channel_cache = {}
    latest_date_cache = {}

    for ki, keyword in enumerate(keywords):
        print(f"\n[{ki+1}/{len(keywords)}] ", end="")
        results = search_youtube(keyword, n=SEARCH_N)
        print(f"  → {len(results)} 个频道")
        fresh_ok, fresh_hits, fresh_checked = keyword_has_enough_fresh_results(
            results, KEYWORD_FRESHNESS_SAMPLE, KEYWORD_MIN_FRESH_HITS, ACTIVE_WITHIN_DAYS
        )
        if results and not fresh_ok:
            print(f"  ⏭️ 关键词偏旧：前{fresh_checked}个结果里只有 {fresh_hits} 个在最近{ACTIVE_WITHIN_DAYS}天内更新，跳过这个关键词")
            skipped += len(results)
            update_keyword_metrics(
                KEYWORD_METRICS_PATH,
                keyword,
                found=len(results),
                eligible=0,
                added_green=0,
                added_yellow=0,
                skipped=len(results),
            )
            continue
        keyword_eligible = 0
        keyword_green = 0
        keyword_yellow = 0
        keyword_skipped = 0
        keyword_low_signal = 0
        keyword_low_video = 0
        keyword_shorts = 0
        keyword_low_recent = 0
        keyword_inactive = 0

        for ci, ch in enumerate(results, start=1):
            name     = ch["channel"]
            ch_url   = ch["channel_url"]
            subs     = ch["subs"]
            channel_id = ch.get("channel_id", "")
            title    = ch["title"]
            vid_url  = ch["video_url"]
            view_count = parse_int(ch.get("view_count"))
            upload_date = str(ch.get("upload_date") or "")
            duration_seconds = parse_int(ch.get("duration_seconds"))
            short_name = short_channel_name(name)

            # 频道名去重放在最前面——一拿到名字就先查，命中就直接跳过，
            # 不用再浪费网络请求去解析链接、补粉丝数、补播放数据、查最近视频均播。
            normalized_name = normalize_name_key(name)
            if normalized_name in existing_names:
                skipped += 1
                keyword_skipped += 1
                print(f"    · [{ci}/{len(results)}] {short_name} → 跳过：频道名已存在")
                continue

            resolved_channel_url = resolve_channel_url(ch_url, vid_url, resolved_channel_cache)
            if resolved_channel_url and resolved_channel_url != ch_url:
                print(f"      ↳ 已修正频道链接")
            ch_url = resolved_channel_url

            if subs <= 0:
                subs = enrich_channel_subs(ch_url, vid_url, channel_id)
                ch["subs"] = subs

            # 粉丝数过滤
            if subs > 0 and not (SUB_MIN <= subs <= SUB_MAX):
                continue

            market_score, market_reason = market_signal_score(name, title, keyword, subs)
            if SMALL_CREATOR_MARKET_MODE and market_score < MARKET_SCORE_MIN:
                keyword_low_signal += 1
                skipped += 1
                keyword_skipped += 1
                continue

            print(f"    · [{ci}/{len(results)}] 检查 {short_name}")

            if view_count < MIN_VIDEO_VIEWS or not upload_date or not duration_seconds:
                extra_view_count, extra_upload_date, extra_duration_seconds = enrich_video_metrics_full(vid_url)
                if extra_view_count > 0:
                    view_count = extra_view_count
                if extra_upload_date:
                    upload_date = extra_upload_date
                if extra_duration_seconds > 0:
                    duration_seconds = extra_duration_seconds

            # 命中的这条视频是 Shorts 不代表整个频道不要——Shorts 播放量跟长视频不是一回事，
            # 不能拿长视频的播放门槛去卡它。这里只是不用这条 Shorts 的播放量做判断，
            # 频道到底行不行，交给下面"最近活跃度 + 最近长视频均播"这两道频道级筛选去判断
            # （那两道筛选本来就只看长视频，会自动把 Shorts 排除在统计之外）。
            is_short_video = is_probable_short(vid_url, duration_seconds)
            if is_short_video:
                keyword_shorts += 1
                print("      ↳ 命中视频是 Shorts，不计入单条播放量门槛，改看频道近期长视频表现")
            elif view_count < MIN_VIDEO_VIEWS:
                keyword_low_video += 1
                skipped += 1
                keyword_skipped += 1
                print(f"      ↳ 跳过：常规视频播放不足 {MIN_VIDEO_VIEWS}")
                continue

            uploads_playlist = str(ch.get("uploads_playlist") or "").strip()
            latest_date = None
            if uploads_playlist:
                latest_date = get_latest_date_from_uploads_playlist(uploads_playlist, latest_date_cache)
            if not latest_date:
                latest_date = get_channel_latest_date(ch_url, latest_date_cache)
            if not latest_date and not is_short_video:
                latest_date = parse_upload_date(upload_date)
            if not latest_date or latest_date < (datetime.now() - timedelta(days=ACTIVE_WITHIN_DAYS)):
                keyword_inactive += 1
                skipped += 1
                keyword_skipped += 1
                latest_text = latest_date.strftime("%Y-%m-%d") if latest_date else "未拿到日期"
                print(f"      ↳ 跳过：最近{ACTIVE_WITHIN_DAYS}天未更新（latest_date={latest_text}）")
                continue

            if MIN_RECENT_AVG_VIEWS > 0:
                recent_metrics = get_recent_channel_metrics(
                    ch_url,
                    count=RECENT_VIDEO_COUNT,
                    cache=recent_stats_cache,
                    uploads_playlist=uploads_playlist,
                )
                if not recent_metrics.get("ok"):
                    keyword_low_recent += 1
                    skipped += 1
                    keyword_skipped += 1
                    print(f"      ↳ 跳过：最近{RECENT_VIDEO_COUNT}条数据不足（{recent_metrics.get('reason') or '未知原因'}）")
                    continue
                recent_avg_views = parse_int(recent_metrics.get("avg_views"))
                if recent_avg_views < MIN_RECENT_AVG_VIEWS:
                    keyword_low_recent += 1
                    skipped += 1
                    keyword_skipped += 1
                    print(f"      ↳ 跳过：最近{RECENT_VIDEO_COUNT}条均播 {recent_avg_views} < {MIN_RECENT_AVG_VIEWS}")
                    continue

            keyword_eligible += 1

            # 频道名去重已经在最前面做过了，这里只需要再查一遍"无邮箱候选池"
            # （这个要用解析后的链接才准，所以留在这里，不能挪到最前面）。
            no_email_key = (name.lower(), (ch_url or vid_url).strip().rstrip("/").lower())
            if no_email_key in no_email_keys:
                skipped += 1
                keyword_skipped += 1
                print("      ↳ 跳过：无邮箱候选里已存在")
                continue

            # 频道综合评分（新增）
            _channel_score = {}
            _ai_tag = ""
            _analysis = {}
            _avg_recent_views = recent_metrics.get("avg_views", 0) if MIN_RECENT_AVG_VIEWS > 0 else 0

            if ANTHROPIC_API_KEY:
                _ai_titles = recent_metrics.get("titles", []) if MIN_RECENT_AVG_VIEWS > 0 else []
                _bio = get_channel_bio(channel_id) if channel_id else ""

                # 调用频道综合评分函数（替代旧的单视频判断逻辑）
                _channel_score = call_claude_for_channel_scoring(
                    channel_name=name,
                    bio=_bio,
                    titles=_ai_titles,
                    subs=subs,
                    hit_video_title=title,
                    product_context="HitPaw VikPea 视频增强工具"
                )

                if _channel_score:
                    # 计算综合分数
                    composite_score = calculate_composite_score(
                        content_relevance=_channel_score.get("content_relevance", 0),
                        commercial_history=_channel_score.get("commercial_history", 0),
                        audience_fit=_channel_score.get("audience_fit", 0),
                        avg_views=_avg_recent_views,
                        max_avg_views=1000000
                    )
                    _channel_score["composite_score"] = composite_score

                    print(f"      ↳ AI频道评分: 内容相关{_channel_score.get('content_relevance', 0)} | "
                          f"商业化{_channel_score.get('commercial_history', 0)} | "
                          f"受众{_channel_score.get('audience_fit', 0)} | "
                          f"综合{composite_score:.1f}")
                    print(f"      ↳ AI判断: {_channel_score.get('verdict', '')} - {_channel_score.get('reasoning', '')}")

                # 仍然保留旧版AI分析（用于兼容，提供额外标签信息）
                _comments = []
                if _ai_titles and recent_metrics.get("video_ids"):
                    recent_video_id = recent_metrics.get("video_ids", [])[0] if recent_metrics.get("video_ids") else None
                    if recent_video_id:
                        _comments = get_video_top_comments(recent_video_id, max_comments=10)

                _analysis = call_claude_for_analysis(name, _bio, _ai_titles, _comments)
                if _analysis:
                    _ai_tag = _analysis.get("tags", "")
                    if _ai_tag:
                        print(f"      ↳ 频道标签：{_ai_tag}")

            # 生成主题/开头
            subj, opening = make_subject_opening(name, title, subs)

            # 抓邮箱
            email, source, confidence = get_channel_email(ch_url, vid_url, name)
            time.sleep(DELAY_FETCH)

            # 邮箱去重
            if email and email in existing_emails:
                skipped += 1
                keyword_skipped += 1
                print(f"      ↳ 跳过：邮箱已存在 {email}")
                continue

            # 写入名单：有邮箱才进主表；无邮箱只进副表，不保存旧主表。
            if email and confidence == "A":
                append_found_email_to_queue(
                    name, email, subj, opening, ch_url or vid_url, subs,
                    keyword, market_score, market_reason, vid_url, tag=_ai_tag, analysis=_analysis,
                    channel_score=_channel_score, avg_recent_views=_avg_recent_views
                )
            elif email:
                pending_result = ""
                if add_pending_email_review_row:
                    _pending_note = f"{format_sub_label(subs)}粉 | 命中视频: {vid_url}"
                    if _ai_tag:
                        _pending_note = f"[标签:{_ai_tag}] " + _pending_note
                    pending_result = add_pending_email_review_row(
                        name=name,
                        email=email,
                        confidence=confidence or "B",
                        source_label=source,
                        link=ch_url or vid_url,
                        source=keyword,
                        email_type="YouTube",
                        note=_pending_note,
                        home_link=ch_url or vid_url,
                        video_link=vid_url,
                    )
                    if str(pending_result).startswith("SKIP:"):
                        skipped += 1
                        keyword_skipped += 1
                        print(f"      ↳ 跳过：{str(pending_result)[5:]}")
                        continue
            else:
                write_to_queue(
                    None, name, email, subj, opening, ch_url or vid_url, subs, False,
                    keyword, market_score, market_reason, vid_url, tag=_ai_tag,
                    channel_score=_channel_score, avg_recent_views=_avg_recent_views
                )

            existing_names.add(normalized_name)
            if email and confidence == "A":
                existing_emails.add(email)
                added_green += 1
                keyword_green += 1
                composite_label = f"综合{_channel_score.get('composite_score', 0):.0f}" if _channel_score else ""
                print(f"    ✅ {name:<28} ({format_sub_label(subs)}粉/S{market_score}/{composite_label}) → {email} [{source}]")
            elif email:
                added_yellow += 1
                keyword_yellow += 1
                print(f"    🔵 {name:<28} ({format_sub_label(subs)}粉/S{market_score}) → {email} [{source}|置信度{confidence}] 已放入待确认邮箱")
            else:
                no_email_keys.add(no_email_key)
                added_yellow += 1
                keyword_yellow += 1
                print(f"    🔴 {name:<28} ({format_sub_label(subs)}粉/S{market_score}) → 无邮箱，已放入无邮箱候选")
        if keyword_low_signal:
            print(f"  · 低小博主/发展中市场受众信号跳过 {keyword_low_signal} 个")
        if keyword_low_video:
            print(f"  · 单条视频播放不足{MIN_VIDEO_VIEWS}跳过 {keyword_low_video} 个")
        if keyword_shorts:
            print(f"  · Shorts直接跳过 {keyword_shorts} 个")
        if keyword_low_recent:
            print(f"  · 最近{RECENT_VIDEO_COUNT}条均播不足{MIN_RECENT_AVG_VIEWS}或数据不足 跳过 {keyword_low_recent} 个")
        if keyword_inactive:
            print(f"  · 最近{ACTIVE_WITHIN_DAYS}天未更新 跳过 {keyword_inactive} 个")

        update_keyword_metrics(
            KEYWORD_METRICS_PATH,
            keyword,
            found=len(results),
            eligible=keyword_eligible,
            added_green=keyword_green,
            added_yellow=keyword_yellow,
            skipped=keyword_skipped,
        )

        if ki < len(keywords) - 1:
            time.sleep(DELAY_SEARCH)

    print(f"\n{'═'*65}")
    print(f"  ✅ 完成！主表有邮箱 +{added_green}  无邮箱候选 +{added_yellow}  跳过{skipped}")
    print(f"  → 运行 VikPea_读表发信.py 发送绿行邮件")
    print(f"{'═'*65}\n")
    if log_event:
        log_event("YouTube搜索", f"绿行 +{added_green} 黄行 +{added_yellow} 跳过 {skipped}")


if __name__ == "__main__":
    main()
