"""
VikPea_关键词复盘.py — 汇总关键词搜索/发信/回复表现，反哺下一轮搜索

输入：
  · VikPea_关键词搜索记录.xlsx  搜索脚本自动生成
  · VikPea_邮件开发追踪.xlsx    发信/回复脚本维护

输出：
  · VikPea_关键词效果复盘.xlsx

用法：
  python3 ~/Downloads/VikPea工作台/VikPea_关键词复盘.py
"""

import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
METRICS_PATH = os.path.join(SCRIPT_DIR, "VikPea_关键词搜索记录.xlsx")
TRACKER_PATH = os.path.join(SCRIPT_DIR, "VikPea_邮件开发追踪.xlsx")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "VikPea_关键词效果复盘.xlsx")

T_EMAIL = 4
T_KIND = 8
T_REPLY = 9
T_STATUS = 12
T_GRADE = 13
T_SOURCE = 19

MIN_SENT_FOR_STRONG_ACTION = 5
MIN_REPLIES_FOR_STRONG_ACTION = 2

EXCLUDED_SOURCE_KEYWORDS = {
    "已发送邮箱补录",
    "发送补录",
    "已发送补录",
    "补录已发送",
    "历史补录",
    "无关键词",
    "(无关键词)",
}

UNSENT_KINDS = {
    "未发送",
    "未发送-未获取邮箱",
}

UNSENT_STATUSES = {
    "",
    "未发送",
    "未联系",
    "未获取邮箱",
}

SENT_STATUS_WORDS = [
    "已发送", "已跟进1", "已跟进2", "暂不跟进",
    "待处理", "待分析", "不建议沟通", "已拒绝", "沟通中", "已合作",
]

HEADER_FILL = PatternFill("solid", start_color="2F5496")
GOOD_FILL = PatternFill("solid", start_color="E2EFDA")
MID_FILL = PatternFill("solid", start_color="FFF2CC")
BAD_FILL = PatternFill("solid", start_color="FCE4D6")


def pct(num, den):
    return round(num / den, 4) if den else 0


def is_replied(value):
    return str(value or "").strip() in {"已回复", "是", "Y", "Yes", "yes", "TRUE", "True"}


def normalize_source(value):
    source = str(value or "").strip()
    if not source:
        return ""
    if source in EXCLUDED_SOURCE_KEYWORDS:
        return ""
    if "补录" in source and " " not in source:
        return ""
    return source


def is_sent_row(email, kind, status):
    email = str(email or "").strip()
    kind = str(kind or "").strip()
    status = str(status or "").strip()
    if "@" not in email:
        return False
    if kind in UNSENT_KINDS:
        return False
    if status in UNSENT_STATUSES:
        return False
    return any(word in status for word in SENT_STATUS_WORDS) or "开发" in kind or "跟进" in kind or "补录" in kind


def has_reply_signal(reply, summary, status, grade):
    if is_replied(reply):
        return True
    if str(grade or "").strip().upper() in {"A", "B", "C"}:
        return True
    if str(summary or "").strip():
        return True
    status = str(status or "").strip()
    return any(word in status for word in ["待处理", "待分析", "不建议沟通", "已拒绝", "沟通中", "已合作"])


def load_search_metrics():
    data = defaultdict(lambda: {
        "search_runs": 0,
        "found_channels": 0,
        "eligible_channels": 0,
        "new_email": 0,
        "new_no_email": 0,
        "skipped": 0,
        "last_search_date": "",
    })
    if not os.path.exists(METRICS_PATH):
        return data

    wb = openpyxl.load_workbook(METRICS_PATH, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        date, keyword, found, eligible, green, yellow, skipped = (list(row) + [None] * 7)[:7]
        keyword = str(keyword or "").strip()
        if not keyword:
            continue
        item = data[keyword]
        item["search_runs"] += 1
        item["found_channels"] += int(found or 0)
        item["eligible_channels"] += int(eligible or 0)
        item["new_email"] += int(green or 0)
        item["new_no_email"] += int(yellow or 0)
        item["skipped"] += int(skipped or 0)
        item["last_search_date"] = str(date or item["last_search_date"])
    return data


def load_tracker_metrics(data):
    if not os.path.exists(TRACKER_PATH):
        return data

    wb = openpyxl.load_workbook(TRACKER_PATH, data_only=True)
    ws = wb["邮件追踪"] if "邮件追踪" in wb.sheetnames else wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_value = row[T_SOURCE - 1] if len(row) >= T_SOURCE else ""
        source = normalize_source(source_value)
        if not source:
            continue
        item = data[source]
        item.setdefault("search_runs", 0)
        item.setdefault("found_channels", 0)
        item.setdefault("eligible_channels", 0)
        item.setdefault("new_email", 0)
        item.setdefault("new_no_email", 0)
        item.setdefault("skipped", 0)
        item.setdefault("last_search_date", "")

        email = str(row[T_EMAIL - 1] or "").strip()
        kind_value = row[T_KIND - 1] if len(row) >= T_KIND else ""
        kind = str(kind_value or "").strip()
        reply = row[T_REPLY - 1]
        summary_value = row[9] if len(row) >= 10 else ""  # 回复摘要（第10列）
        status = str(row[T_STATUS - 1] or "").strip()
        grade_value = row[T_GRADE - 1] if len(row) >= T_GRADE else ""
        grade = str(grade_value or "").strip()

        sent_row = is_sent_row(email, kind, status)
        if sent_row:
            item["sent"] = item.get("sent", 0) + 1
        if sent_row and has_reply_signal(reply, summary_value, status, grade):
            item["replied"] = item.get("replied", 0) + 1
        if sent_row and grade == "A":
            item["grade_a"] = item.get("grade_a", 0) + 1
        elif sent_row and grade == "B":
            item["grade_b"] = item.get("grade_b", 0) + 1
        elif sent_row and grade == "C":
            item["grade_c"] = item.get("grade_c", 0) + 1
        if sent_row and "已合作" in status:
            item["cooperated"] = item.get("cooperated", 0) + 1
    return data


def suggest(item):
    sent = item.get("sent", 0)
    replied = item.get("replied", 0)
    grade_a = item.get("grade_a", 0)
    added = item.get("new_email", 0) + item.get("new_no_email", 0)
    email_rate = pct(item.get("new_email", 0), added)
    reply_rate = pct(replied, sent)
    a_rate = pct(grade_a, sent)

    if sent and sent < MIN_SENT_FOR_STRONG_ACTION:
        return "样本太小"
    if sent >= MIN_SENT_FOR_STRONG_ACTION and (
        replied >= MIN_REPLIES_FOR_STRONG_ACTION and reply_rate >= 0.12 or a_rate >= 0.08
    ):
        return "加大搜索"
    if grade_a > 0:
        return "保留观察"
    if added >= 10 and email_rate < 0.15:
        return "减少搜索"
    if sent >= 10 and replied == 0:
        return "暂停观察"
    return "继续测试"


def build_output(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "关键词复盘"

    headers = [
        "关键词", "搜索次数", "搜索到频道数", "符合粉丝范围数",
        "新增有邮箱数", "新增无邮箱数", "邮箱命中率",
        "已发送数", "已回复数", "回复率",
        "A级回复数", "A类回复率", "B级", "C级",
        "已合作数", "最近搜索日期", "建议动作", "数据备注"
    ]
    ws.append(headers)

    rows = []
    for keyword, item in data.items():
        added = item.get("new_email", 0) + item.get("new_no_email", 0)
        sent = item.get("sent", 0)
        replied = item.get("replied", 0)
        grade_a = item.get("grade_a", 0)
        note = ""
        if sent == 0:
            note = "还没发信，先看邮箱命中率"
        elif sent < MIN_SENT_FOR_STRONG_ACTION:
            note = f"已发送{sent}封，样本太小"
        rows.append([
            keyword,
            item.get("search_runs", 0),
            item.get("found_channels", 0),
            item.get("eligible_channels", 0),
            item.get("new_email", 0),
            item.get("new_no_email", 0),
            pct(item.get("new_email", 0), added),
            sent,
            replied,
            pct(replied, sent),
            grade_a,
            pct(grade_a, sent),
            item.get("grade_b", 0),
            item.get("grade_c", 0),
            item.get("cooperated", 0),
            item.get("last_search_date", ""),
            suggest(item),
            note,
        ])

    rows.sort(key=lambda r: (r[16] != "加大搜索", r[16] == "样本太小", -r[11], -r[9], -r[4], r[0].lower()))
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2):
        action = row[16].value
        fill = GOOD_FILL if action == "加大搜索" else MID_FILL if action in {"保留观察", "继续测试"} else BAD_FILL
        for cell in row:
            cell.fill = fill

    for col in ["G", "J", "L"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0%"

    widths = {
        "A": 42, "B": 10, "C": 14, "D": 14, "E": 12, "F": 12, "G": 12,
        "H": 10, "I": 10, "J": 10, "K": 10, "L": 10, "M": 8, "N": 8,
        "O": 10, "P": 14, "Q": 12, "R": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_PATH)
    return len(rows)


def main():
    data = load_search_metrics()
    data = load_tracker_metrics(data)
    count = build_output(data)
    print(f"✅ 已生成关键词复盘: {OUTPUT_PATH}")
    print(f"   共 {count} 个关键词")
    if not os.path.exists(METRICS_PATH):
        print("   提醒：还没有搜索记录表。下次运行 YouTube 批量搜索后，数据会更完整。")


if __name__ == "__main__":
    main()
