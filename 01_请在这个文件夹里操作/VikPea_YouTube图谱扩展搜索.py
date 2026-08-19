"""
VikPea_YouTube图谱扩展搜索.py — 从"已经合作过的博主"出发，顺着 YouTube 频道页的
"推荐频道"（Featured Channels）往外扩散找同类创作者，不是靠关键词搜索猜。

思路：
  合作跟踪表里已经验证过效果的博主 → 抓他们频道主页上的"推荐频道"列表（博主自己精选的
  同类/合作过的创作者，比关键词搜索猜出来的更准）→ 抓到的候选频道跑一遍跟
  VikPea_YouTube批量搜索.py 完全一样的筛选 + 找邮箱 + 查重流程（直接复用那边的函数，
  不重写一遍，保证两种搜索方式的筛选标准、查重逻辑完全一致）
  → 写入 VikPea_发信名单.xlsx / VikPea_待确认邮箱.xlsx（跟关键词搜索共用同一批产出表）

依赖：
  · 需要在 VikPea_配置.xlsx / 网页系统设置里配置好 YOUTUBE_API_KEY
    （推荐频道数据只能通过官方 API 拿，yt-dlp 抓不到）
  · 种子数据来自团队的"合作跟踪表.xlsx"（默认在桌面 数据/ 文件夹下，
    可以在 VikPea_配置.xlsx 加一行 COLLAB_TRACKER_PATH 改路径）

用法：
  python3 VikPea_YouTube图谱扩展搜索.py
"""

import os
import sys
import time
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

try:
    from VikPea_common import apply_config, log_event
except ImportError:
    apply_config = None
    log_event = None

# 直接复用关键词搜索脚本里的函数（同一份筛选/查重/找邮箱逻辑，不重写一遍）
import VikPea_YouTube批量搜索 as yt

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
COLLAB_TRACKER_PATH_DEFAULT = os.path.expanduser("~/Desktop/数据/合作跟踪表.xlsx")
MAX_SEED_CHANNELS = 40          # 单次最多用多少个已合作博主当种子，避免 API 配额爆
MAX_CANDIDATES_PER_RUN = 150    # 扩展出的候选频道上限


def load_seed_channel_urls(path: str) -> list:
    """从合作跟踪表「已合作」sheet 的"主页链接"列读种子频道。"""
    if not os.path.exists(path):
        print(f"  ⚠️ 找不到合作跟踪表: {path}")
        print(f"     可以在 VikPea_配置.xlsx 加一行 COLLAB_TRACKER_PATH 指定正确路径")
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    if "已合作" not in wb.sheetnames:
        print("  ⚠️ 合作跟踪表里没有找到「已合作」这个 sheet")
        return []
    ws = wb["已合作"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    if "主页链接" not in headers:
        print("  ⚠️ 「已合作」sheet 里没有找到「主页链接」这一列")
        return []
    col = headers.index("主页链接")
    urls = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col < len(row) and row[col]:
            url = str(row[col]).strip()
            if url.lower().startswith("http") and "youtube.com" in url.lower():
                urls.append(url)
    # 去重保序，最新的排前面（合作跟踪表通常越往下越新，倒序取更有参考价值）
    seen = set()
    ordered = []
    for url in reversed(urls):
        if url not in seen:
            seen.add(url)
            ordered.append(url)
    return ordered[:MAX_SEED_CHANNELS]


def get_featured_channel_ids(channel_id: str) -> list:
    """调用 channels.list?part=brandingSettings 拿这个频道主页上的"推荐频道" ID 列表。"""
    if not yt.YOUTUBE_API_KEY or not channel_id:
        return []
    payload = yt.youtube_api_get("channels", {
        "part": "brandingSettings",
        "id": channel_id,
        "maxResults": 1,
    })
    for item in payload.get("items", []) or []:
        branding = (item.get("brandingSettings") or {}).get("channel") or {}
        urls = branding.get("featuredChannelsUrls") or []
        return [str(u).strip() for u in urls if u]
    return []


def get_channel_snippet(channel_id: str) -> dict:
    """拿候选频道的基本信息（名字/粉丝数），用来决定要不要往下走完整筛选流程。"""
    if not yt.YOUTUBE_API_KEY or not channel_id:
        return {}
    payload = yt.youtube_api_get("channels", {
        "part": "snippet,statistics",
        "id": channel_id,
        "maxResults": 1,
    })
    items = payload.get("items", []) or []
    if not items:
        return {}
    item = items[0]
    snippet = item.get("snippet", {}) or {}
    stats = item.get("statistics", {}) or {}
    return {
        "channel_id": channel_id,
        "name": snippet.get("title", "") or channel_id,
        "channel_url": f"https://www.youtube.com/channel/{channel_id}",
        "subs": yt.parse_int(stats.get("subscriberCount")),
    }


def main():
    print("\n" + "=" * 64)
    print("  VikPea YouTube 图谱扩展搜索（Featured Channels）")
    print("=" * 64)

    collab_tracker_path = COLLAB_TRACKER_PATH_DEFAULT
    if apply_config:
        config = apply_config(yt.__dict__, {
            "YOUTUBE_SUB_MIN": "SUB_MIN",
            "YOUTUBE_SUB_MAX": "SUB_MAX",
            "YOUTUBE_ACTIVE_WITHIN_DAYS": "ACTIVE_WITHIN_DAYS",
            "YOUTUBE_MIN_RECENT_AVG_VIEWS": "MIN_RECENT_AVG_VIEWS",
            "YOUTUBE_RECENT_VIDEO_COUNT": "RECENT_VIDEO_COUNT",
            "YOUTUBE_API_KEY": "YOUTUBE_API_KEY",
            "YOUTUBE_API_DELAY_SEC": "YOUTUBE_API_DELAY_SEC",
            "YOUTUBE_API_RETRY_TIMES": "YOUTUBE_API_RETRY_TIMES",
            "YOUTUBE_API_429_COOLDOWN": "YOUTUBE_API_429_COOLDOWN",
        })
        custom_path = str((config or {}).get("COLLAB_TRACKER_PATH") or "").strip()
        if custom_path:
            collab_tracker_path = os.path.expanduser(custom_path)

    if not yt.YOUTUBE_API_KEY:
        print("❌ 需要先配置 YOUTUBE_API_KEY（网页「系统设置」或 VikPea_配置.xlsx）")
        print("   推荐频道数据只有官方 API 能拿到，yt-dlp 抓不到这个信息。")
        return

    print(f"📄 种子来源: {collab_tracker_path}")
    seed_urls = load_seed_channel_urls(collab_tracker_path)
    if not seed_urls:
        print("没有可用的种子频道，先确认合作跟踪表路径和数据。")
        return
    print(f"读到 {len(seed_urls)} 个已合作频道作为种子\n")

    seed_ids = []
    for url in seed_urls:
        cid = yt.extract_channel_id(url)
        if cid:
            seed_ids.append(cid)
    seed_ids = list(dict.fromkeys(seed_ids))
    print(f"成功解析 {len(seed_ids)} 个种子频道 ID\n")
    if not seed_ids:
        return

    existing_emails = yt.load_existing_emails(yt.TRACKER_PATH, yt.QUEUE_PATH)
    existing_names = yt.load_existing_channel_names(yt.TRACKER_PATH, yt.QUEUE_PATH)
    no_email_keys = yt.existing_no_email_keys() if yt.existing_no_email_keys else set()
    print(f"已有邮箱 {len(existing_emails)} 条（用于去重）\n")

    visited_ids = set(seed_ids)
    candidate_ids = []
    for i, seed_id in enumerate(seed_ids, 1):
        print(f"[{i}/{len(seed_ids)}] 抓取推荐频道: {seed_id}")
        featured = get_featured_channel_ids(seed_id)
        print(f"  → 找到 {len(featured)} 个推荐频道")
        for fid in featured:
            if fid not in visited_ids:
                visited_ids.add(fid)
                candidate_ids.append(fid)
        if len(candidate_ids) >= MAX_CANDIDATES_PER_RUN:
            candidate_ids = candidate_ids[:MAX_CANDIDATES_PER_RUN]
            break
        time.sleep(float(yt.YOUTUBE_API_DELAY_SEC or 0.5))

    print(f"\n扩展到 {len(candidate_ids)} 个候选频道（已去重、已排除种子本身）\n")
    if not candidate_ids:
        print("没有扩展到新候选，可能种子频道都没填「推荐频道」。")
        return

    added_green = added_yellow = skipped = 0
    for ci, cid in enumerate(candidate_ids, 1):
        info = get_channel_snippet(cid)
        if not info or not info.get("name"):
            skipped += 1
            continue
        name = info["name"]
        ch_url = info["channel_url"]
        subs = info["subs"]
        short_name = yt.short_channel_name(name)

        if subs > 0 and not (yt.SUB_MIN <= subs <= yt.SUB_MAX):
            continue

        normalized_name = yt.normalize_name_key(name)
        if normalized_name in existing_names:
            continue

        print(f"  · [{ci}/{len(candidate_ids)}] 检查 {short_name}")

        latest_date = yt.get_channel_latest_date(ch_url, {})
        if not latest_date or latest_date < (datetime.now() - timedelta(days=yt.ACTIVE_WITHIN_DAYS)):
            skipped += 1
            print(f"      ↳ 跳过：最近{yt.ACTIVE_WITHIN_DAYS}天未更新")
            continue

        if yt.MIN_RECENT_AVG_VIEWS > 0:
            recent_metrics = yt.get_recent_channel_metrics(ch_url, count=yt.RECENT_VIDEO_COUNT)
            if not recent_metrics.get("ok"):
                skipped += 1
                print(f"      ↳ 跳过：最近{yt.RECENT_VIDEO_COUNT}条数据不足")
                continue
            if yt.parse_int(recent_metrics.get("avg_views")) < yt.MIN_RECENT_AVG_VIEWS:
                skipped += 1
                print(f"      ↳ 跳过：近期均播不达标")
                continue

        no_email_key = (name.lower(), ch_url.strip().rstrip("/").lower())
        if no_email_key in no_email_keys:
            skipped += 1
            print("      ↳ 跳过：无邮箱候选里已存在")
            continue

        subj, opening = yt.make_subject_opening(name, "", subs)
        email, source, confidence = yt.get_channel_email(ch_url, "", name)
        time.sleep(yt.DELAY_FETCH if hasattr(yt, "DELAY_FETCH") else 2)

        if email and email in existing_emails:
            skipped += 1
            print(f"      ↳ 跳过：邮箱已存在 {email}")
            continue

        if email and confidence == "A":
            yt.append_found_email_to_queue(
                name, email, subj, opening, ch_url, subs,
                "图谱扩展（Featured Channels）", 0, "", ""
            )
            existing_emails.add(email)
            added_green += 1
            print(f"    ✅ {short_name:<28} → {email} [{source}]")
        elif email:
            pending_result = yt.add_pending_email_review_row(
                name=name, email=email, confidence=confidence or "B",
                source_label=source, link=ch_url, source="图谱扩展（Featured Channels）",
                email_type="YouTube", note=f"{yt.format_sub_label(subs)}粉 | Featured Channels 扩展",
                home_link=ch_url, video_link="",
            )
            if str(pending_result).startswith("SKIP:"):
                skipped += 1
                print(f"      ↳ 跳过：{str(pending_result)[5:]}")
                continue
            added_yellow += 1
            print(f"    🔵 {short_name:<28} → {email} [{source}|置信度{confidence}] 已放入待确认邮箱")
        else:
            skipped += 1
            print(f"    🔴 {short_name:<28} → 无邮箱")

        existing_names.add(normalized_name)

    print(f"\n{'═'*64}")
    print(f"  ✅ 完成！主表有邮箱 +{added_green}  待确认 +{added_yellow}  跳过 {skipped}")
    print(f"{'═'*64}\n")
    if log_event:
        log_event("图谱扩展搜索", f"绿行 +{added_green} 黄行 +{added_yellow} 跳过 {skipped}")


if __name__ == "__main__":
    main()
