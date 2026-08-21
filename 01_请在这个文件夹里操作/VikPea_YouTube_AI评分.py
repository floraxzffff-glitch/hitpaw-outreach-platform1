"""
VikPea_YouTube_AI评分.py - AI两阶段判断脚本

流程：
1. 读取 VikPea_发信名单.xlsx（YouTube搜索结果）
2. 阶段1：AI视频相关性初筛（扩大候选池）
   - 获取视频完整元数据（标题+简介+tags）
   - 判断视频是否与关键词真正相关
3. 阶段2：AI KOL适配度判断
   - 对相关视频的频道进行深度评估
   - 评分和推荐度判断
4. 导出结果到Excel（分表）：
   - Sheet1：推荐+待确认（主表）
   - Sheet2：不推荐（存档）

依赖：
  pip3 install anthropic openpyxl yt-dlp

用法：
  export ANTHROPIC_API_KEY="your-api-key"
  python3 VikPea_YouTube_AI评分.py
"""

import os
import sys
import json
import subprocess
import shutil
from typing import List, Dict, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("❌ pip3 install openpyxl")
    sys.exit(1)

try:
    from ai_video_relevance import (
        batch_judge_video_relevance,
        KeywordQuery,
        RelevanceVerdict
    )
    from ai_kol_scorer import batch_score_kol_fit, FitVerdict
except ImportError:
    print("❌ 缺少 ai_video_relevance.py 或 ai_kol_scorer.py")
    sys.exit(1)


# ── 配置 ────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(SCRIPT_DIR, "VikPea_发信名单.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "VikPea_YouTube_AI评分结果.xlsx")

# yt-dlp 可执行文件路径
def find_ytdlp() -> list:
    try:
        subprocess.run([sys.executable, "-m", "yt_dlp", "--version"],
                       capture_output=True, check=True)
        return [sys.executable, "-m", "yt_dlp"]
    except Exception:
        pass
    p = shutil.which("yt-dlp")
    if p:
        return [p]
    ver = f"{sys.version_info.major}.{sys.version_info.minor}"
    user_bin = os.path.expanduser(f"~/Library/Python/{ver}/bin/yt-dlp")
    if os.path.exists(user_bin):
        return [user_bin]
    return []

YTDLP_CMD = find_ytdlp()


def fetch_video_metadata(video_url: str) -> Dict:
    """
    获取视频的完整元数据（标题+简介+tags）

    Returns:
        {title, description, tags, view_count, upload_date}
    """
    if not YTDLP_CMD:
        print("  ⚠️ yt-dlp 未安装")
        return {"title": "", "description": "", "tags": [], "view_count": 0, "upload_date": ""}

    try:
        result = subprocess.run(
            YTDLP_CMD + [
                "--skip-download",
                "--dump-json",
                video_url
            ],
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode != 0:
            return {"title": "", "description": "", "tags": [], "view_count": 0, "upload_date": ""}

        data = json.loads(result.stdout.strip())
        return {
            "title": data.get("title", ""),
            "description": data.get("description", ""),
            "tags": data.get("tags", []),
            "view_count": data.get("view_count", 0),
            "upload_date": data.get("upload_date", "")
        }
    except Exception as e:
        print(f"  ⚠️ 获取视频元数据失败: {video_url} - {e}")
        return {"title": "", "description": "", "tags": [], "view_count": 0, "upload_date": ""}


def fetch_channel_info(channel_url: str) -> Dict:
    """
    获取频道信息（简介、最近视频）

    Returns:
        {channel_description, recent_videos: [{title, views, upload_date}, ...]}
    """
    if not YTDLP_CMD:
        return {"channel_description": "", "recent_videos": []}

    try:
        # 获取频道信息
        result = subprocess.run(
            YTDLP_CMD + [
                "--skip-download",
                "--playlist-end", "5",  # 只获取最近5个视频
                "--dump-json",
                channel_url
            ],
            capture_output=True,
            text=True,
            timeout=60
        )

        if result.returncode != 0:
            return {"channel_description": "", "recent_videos": []}

        lines = result.stdout.strip().split("\n")
        channel_description = ""
        recent_videos = []

        for line in lines:
            try:
                data = json.loads(line)
                if not channel_description and data.get("channel"):
                    channel_description = data.get("description", "")

                recent_videos.append({
                    "title": data.get("title", ""),
                    "views": data.get("view_count", 0),
                    "upload_date": data.get("upload_date", "")
                })
            except json.JSONDecodeError:
                continue

        return {
            "channel_description": channel_description,
            "recent_videos": recent_videos[:5]
        }
    except Exception as e:
        print(f"  ⚠️ 获取频道信息失败: {channel_url} - {e}")
        return {"channel_description": "", "recent_videos": []}


def load_search_results() -> List[Dict]:
    """
    从 VikPea_发信名单.xlsx 读取搜索结果

    Returns:
        [{channel_name, channel_url, video_url, keyword, subscriber_count}, ...]
    """
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 未找到输入文件: {INPUT_FILE}")
        print("   请先运行 VikPea_YouTube批量搜索.py 生成搜索结果")
        sys.exit(1)

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:  # 跳过空行
            continue

        # 假设列顺序：频道名、频道URL、视频URL、关键词、订阅数
        # 实际顺序可能不同，需要根据实际表格调整
        channel_name = str(row[0]) if row[0] else ""
        channel_url = str(row[1]) if len(row) > 1 and row[1] else ""
        video_url = str(row[2]) if len(row) > 2 and row[2] else ""
        keyword = str(row[3]) if len(row) > 3 and row[3] else ""
        subscriber_count = int(row[4]) if len(row) > 4 and row[4] and str(row[4]).isdigit() else 0

        if channel_name and video_url:
            results.append({
                "channel_name": channel_name,
                "channel_url": channel_url,
                "video_url": video_url,
                "keyword": keyword,
                "subscriber_count": subscriber_count
            })

    wb.close()
    print(f"✅ 读取到 {len(results)} 条搜索结果")
    return results


def export_results_to_excel(results: List[Dict], output_path: str):
    """
    导出结果到Excel（分表）

    Sheet1: 推荐+待确认（主表）
    Sheet2: 不推荐（存档）

    Args:
        results: [
            {
                channel_name, channel_url, video_url, keyword, subscriber_count,
                relevance_verdict, relevance_reason,
                fit_score, fit_verdict, fit_reason, suggested_angle
            },
            ...
        ]
    """
    wb = openpyxl.Workbook()

    # Sheet1: 推荐+待确认
    ws_main = wb.active
    ws_main.title = "推荐和待确认"

    # Sheet2: 不推荐
    ws_archive = wb.create_sheet("AI判定不推荐-存档")

    # 表头
    headers = [
        "频道名", "频道URL", "视频URL", "关键词", "订阅数",
        "适配度评分", "推荐度", "评分理由", "建议合作角度",
        "视频相关性", "相关性理由"
    ]

    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 写入表头（两个sheet都要）
    for sheet in [ws_main, ws_archive]:
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font

    # 分类写入数据
    main_row = 2
    archive_row = 2

    for result in results:
        fit_verdict = result.get("fit_verdict", FitVerdict.UNCERTAIN)

        # 根据verdict决定写入哪个sheet
        if fit_verdict == FitVerdict.NOT_RECOMMENDED:
            ws = ws_archive
            row_idx = archive_row
            archive_row += 1
            row_fill = red_fill
        else:
            ws = ws_main
            row_idx = main_row
            main_row += 1

            # 主表根据verdict上色
            if fit_verdict == FitVerdict.RECOMMENDED:
                row_fill = green_fill
            else:  # UNCERTAIN
                row_fill = yellow_fill

        # 写入数据
        ws.cell(row_idx, 1, result.get("channel_name", ""))
        ws.cell(row_idx, 2, result.get("channel_url", ""))
        ws.cell(row_idx, 3, result.get("video_url", ""))
        ws.cell(row_idx, 4, result.get("keyword", ""))
        ws.cell(row_idx, 5, result.get("subscriber_count", 0))
        ws.cell(row_idx, 6, result.get("fit_score", 0))
        ws.cell(row_idx, 7, result.get("fit_verdict", ""))
        ws.cell(row_idx, 8, result.get("fit_reason", ""))
        ws.cell(row_idx, 9, result.get("suggested_angle", ""))
        ws.cell(row_idx, 10, result.get("relevance_verdict", ""))
        ws.cell(row_idx, 11, result.get("relevance_reason", ""))

        # 给整行上色
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row_idx, col_idx).fill = row_fill

    # 调整列宽
    for sheet in [ws_main, ws_archive]:
        sheet.column_dimensions['A'].width = 25  # 频道名
        sheet.column_dimensions['B'].width = 40  # 频道URL
        sheet.column_dimensions['C'].width = 40  # 视频URL
        sheet.column_dimensions['D'].width = 20  # 关键词
        sheet.column_dimensions['E'].width = 12  # 订阅数
        sheet.column_dimensions['F'].width = 12  # 适配度评分
        sheet.column_dimensions['G'].width = 12  # 推荐度
        sheet.column_dimensions['H'].width = 50  # 评分理由
        sheet.column_dimensions['I'].width = 40  # 建议合作角度
        sheet.column_dimensions['J'].width = 12  # 视频相关性
        sheet.column_dimensions['K'].width = 50  # 相关性理由

    wb.save(output_path)

    # 统计
    recommended_count = sum(1 for r in results if r.get("fit_verdict") == FitVerdict.RECOMMENDED)
    uncertain_count = sum(1 for r in results if r.get("fit_verdict") == FitVerdict.UNCERTAIN)
    not_recommended_count = sum(1 for r in results if r.get("fit_verdict") == FitVerdict.NOT_RECOMMENDED)

    print(f"\n✅ 结果已导出: {output_path}")
    print(f"   主表（推荐+待确认）: {recommended_count + uncertain_count} 条")
    print(f"     ├─ 推荐: {recommended_count} 条")
    print(f"     └─ 待确认: {uncertain_count} 条")
    print(f"   存档表（不推荐）: {not_recommended_count} 条")


def main():
    print("=" * 60)
    print("VikPea YouTube AI 两阶段判断")
    print("=" * 60)

    # 检查API密钥
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("❌ 未设置 ANTHROPIC_API_KEY")
        print("   请执行: export ANTHROPIC_API_KEY='your-api-key'")
        sys.exit(1)

    # 检查yt-dlp
    if not YTDLP_CMD:
        print("❌ 未找到 yt-dlp，请安装: pip3 install yt-dlp")
        sys.exit(1)

    # 1. 读取搜索结果
    print("\n[1/4] 读取搜索结果...")
    search_results = load_search_results()

    if not search_results:
        print("❌ 没有搜索结果可处理")
        sys.exit(1)

    # 2. 阶段1：视频相关性初筛
    print(f"\n[2/4] 阶段1：AI视频相关性初筛（扩大候选池）...")
    print(f"   处理 {len(search_results)} 个视频...")

    enriched_videos = []
    for idx, result in enumerate(search_results, 1):
        print(f"   [{idx}/{len(search_results)}] 获取视频元数据: {result['channel_name']}")

        video_meta = fetch_video_metadata(result["video_url"])

        enriched_videos.append({
            **result,
            "video_title": video_meta["title"] or result.get("channel_name", ""),
            "video_description": video_meta["description"],
            "video_tags": video_meta["tags"]
        })

    # 批量AI判断相关性
    print("\n   开始AI相关性判断...")

    def relevance_progress(current, total, title):
        print(f"   [{current}/{total}] 判断: {title[:50]}")

    videos_for_relevance = [
        {
            "video_url": v["video_url"],
            "title": v["video_title"],
            "description": v["video_description"],
            "tags": v["video_tags"]
        }
        for v in enriched_videos
    ]

    # 按关键词分组判断（假设所有视频用同一个关键词，实际可能需要分组）
    # 这里简化处理，实际应该按result["keyword"]分组
    keyword = enriched_videos[0]["keyword"] if enriched_videos else "video"

    relevance_results = batch_judge_video_relevance(
        videos=videos_for_relevance,
        keyword=keyword,
        api_key=api_key,
        progress_callback=relevance_progress
    )

    # 合并相关性结果
    for i, relevance in enumerate(relevance_results):
        enriched_videos[i]["relevance_verdict"] = relevance.verdict.value
        enriched_videos[i]["relevance_reason"] = relevance.reason
        enriched_videos[i]["matched_fields"] = relevance.matched_fields

    # 筛选相关的视频进入下一阶段
    relevant_videos = [
        v for v in enriched_videos
        if v["relevance_verdict"] == RelevanceVerdict.RELEVANT.value
    ]

    print(f"\n   ✅ 相关性初筛完成")
    print(f"      相关: {len(relevant_videos)} 个")
    print(f"      不相关: {len([v for v in enriched_videos if v['relevance_verdict'] == RelevanceVerdict.IRRELEVANT.value])} 个")
    print(f"      待确认: {len([v for v in enriched_videos if v['relevance_verdict'] == RelevanceVerdict.UNCERTAIN.value])} 个")

    # 3. 阶段2：KOL适配度判断
    print(f"\n[3/4] 阶段2：AI KOL适配度判断...")
    print(f"   处理 {len(relevant_videos)} 个相关频道...")

    channels_for_scoring = []
    for idx, video in enumerate(relevant_videos, 1):
        print(f"   [{idx}/{len(relevant_videos)}] 获取频道信息: {video['channel_name']}")

        channel_info = fetch_channel_info(video["channel_url"])

        channels_for_scoring.append({
            "channel_name": video["channel_name"],
            "channel_url": video["channel_url"],
            "channel_description": channel_info["channel_description"],
            "recent_videos": channel_info["recent_videos"],
            "subscriber_count": video["subscriber_count"]
        })

    # 批量AI评分
    print("\n   开始AI适配度评分...")

    def scoring_progress(current, total, name):
        print(f"   [{current}/{total}] 评分: {name[:50]}")

    fit_results = batch_score_kol_fit(
        channels=channels_for_scoring,
        product_name="HitPaw",
        product_category="视频编辑软件",
        api_key=api_key,
        progress_callback=scoring_progress
    )

    # 合并适配度结果
    final_results = []

    # 先添加相关且已评分的
    for i, fit_result in enumerate(fit_results):
        video = relevant_videos[i]
        final_results.append({
            "channel_name": video["channel_name"],
            "channel_url": video["channel_url"],
            "video_url": video["video_url"],
            "keyword": video["keyword"],
            "subscriber_count": video["subscriber_count"],
            "relevance_verdict": video["relevance_verdict"],
            "relevance_reason": video["relevance_reason"],
            "fit_score": fit_result.fit_score,
            "fit_verdict": fit_result.verdict.value,
            "fit_reason": fit_result.reason,
            "suggested_angle": fit_result.suggested_angle
        })

    # 添加不相关的（直接标记为"不推荐"）
    for video in enriched_videos:
        if video["relevance_verdict"] != RelevanceVerdict.RELEVANT.value:
            final_results.append({
                "channel_name": video["channel_name"],
                "channel_url": video["channel_url"],
                "video_url": video["video_url"],
                "keyword": video["keyword"],
                "subscriber_count": video["subscriber_count"],
                "relevance_verdict": video["relevance_verdict"],
                "relevance_reason": video["relevance_reason"],
                "fit_score": 0,
                "fit_verdict": FitVerdict.NOT_RECOMMENDED.value,
                "fit_reason": "视频内容与关键词不相关",
                "suggested_angle": ""
            })

    # 4. 导出Excel
    print(f"\n[4/4] 导出结果到Excel...")
    export_results_to_excel(final_results, OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("✅ 处理完成")
    print("=" * 60)


if __name__ == "__main__":
    main()
